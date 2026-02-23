"""
eXtra AC Price Scraper v4
=========================
eXtra.com.sa 에어컨 가격 스크래퍼
Prices DB 시트에 최신 가격 데이터를 업데이트합니다.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import re
import os
import json
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── 설정 ──────────────────────────────────────────────────────────────────────
EXCEL_FILE = "eXtra_ac_Prices_Tracking_Master.xlsx"
PRICES_DB_SHEET = "Prices DB"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

AC_CATEGORIES = [
    {
        "name": "Split AC",
        "url": "https://www.extra.com.sa/en/c/air-conditioners/split-air-conditioners",
    },
    {
        "name": "Window AC",
        "url": "https://www.extra.com.sa/en/c/air-conditioners/window-air-conditioners",
    },
    {
        "name": "Portable AC",
        "url": "https://www.extra.com.sa/en/c/air-conditioners/portable-air-conditioners",
    },
]

# Prices DB 시트 컬럼 정의
PRICES_DB_COLUMNS = [
    "스크랩 날짜",       # A
    "카테고리",          # B
    "브랜드",            # C
    "모델명",            # D
    "SKU",               # E
    "현재가 (SAR)",      # F
    "원가 (SAR)",        # G
    "할인율 (%)",        # H
    "재고 상태",         # I
    "URL",               # J
]


def get_session() -> requests.Session:
    """요청 세션 생성"""
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def parse_price(price_str: str) -> float:
    """가격 문자열에서 숫자 추출"""
    if not price_str:
        return 0.0
    cleaned = re.sub(r"[^\d.]", "", str(price_str).replace(",", ""))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def scrape_category(session: requests.Session, category: dict) -> list:
    """카테고리별 에어컨 목록 스크랩"""
    products = []
    page = 1

    while True:
        url = f"{category['url']}?page={page}"
        logger.info(f"  [{category['name']}] 페이지 {page} 스크랩 중: {url}")

        try:
            resp = session.get(url, timeout=15)
            resp.raise_for_status()
        except requests.RequestException as e:
            logger.warning(f"  요청 실패: {e}")
            break

        soup = BeautifulSoup(resp.text, "html.parser")

        # 상품 카드 파싱 (eXtra 사이트 구조에 맞게 조정)
        product_cards = soup.select("div.product-item, article.product, li.product-item")

        if not product_cards:
            # JSON-LD 데이터 시도
            json_ld_tags = soup.find_all("script", type="application/ld+json")
            for tag in json_ld_tags:
                try:
                    data = json.loads(tag.string)
                    if isinstance(data, dict) and data.get("@type") == "ItemList":
                        for item in data.get("itemListElement", []):
                            product = item.get("item", {})
                            offer = product.get("offers", {})
                            name = product.get("name", "")
                            brand_info = product.get("brand", {})
                            brand = (
                                brand_info.get("name", "")
                                if isinstance(brand_info, dict)
                                else str(brand_info)
                            )
                            sku = product.get("sku", "")
                            current_price = parse_price(offer.get("price", 0))
                            product_url = product.get("url", url)

                            if name and current_price > 0:
                                products.append({
                                    "카테고리": category["name"],
                                    "브랜드": brand,
                                    "모델명": name,
                                    "SKU": sku,
                                    "현재가 (SAR)": current_price,
                                    "원가 (SAR)": current_price,
                                    "할인율 (%)": 0.0,
                                    "재고 상태": offer.get("availability", "").split("/")[-1],
                                    "URL": product_url,
                                })
                except (json.JSONDecodeError, AttributeError):
                    continue

            if not products:
                logger.info(f"  더 이상 상품 없음. 종료.")
            break

        for card in product_cards:
            try:
                # 상품명
                name_tag = card.select_one(
                    "h2.product-name, h3.product-title, a.product-name, span.product-name"
                )
                name = name_tag.get_text(strip=True) if name_tag else ""

                # 브랜드
                brand_tag = card.select_one("span.brand, div.brand, a.brand")
                brand = brand_tag.get_text(strip=True) if brand_tag else ""
                if not brand and name:
                    brand = name.split()[0] if name else ""

                # SKU
                sku = card.get("data-sku", "") or card.get("data-product-id", "")

                # 가격
                price_tag = card.select_one(
                    "span.price, span.current-price, div.price, "
                    "span.special-price, p.price"
                )
                current_price = parse_price(price_tag.get_text() if price_tag else "0")

                # 원가
                old_price_tag = card.select_one(
                    "span.old-price, span.original-price, del.price, s.price"
                )
                original_price = parse_price(
                    old_price_tag.get_text() if old_price_tag else "0"
                )
                if original_price == 0:
                    original_price = current_price

                # 할인율
                discount = 0.0
                if original_price > 0 and current_price < original_price:
                    discount = round(
                        (original_price - current_price) / original_price * 100, 1
                    )

                # 재고 상태
                stock_tag = card.select_one(
                    "span.availability, span.stock, div.stock-status"
                )
                stock = (
                    stock_tag.get_text(strip=True)
                    if stock_tag
                    else "In Stock"
                )

                # URL
                link_tag = card.select_one("a[href]")
                product_url = ""
                if link_tag:
                    href = link_tag.get("href", "")
                    product_url = (
                        href
                        if href.startswith("http")
                        else f"https://www.extra.com.sa{href}"
                    )

                if name and current_price > 0:
                    products.append({
                        "카테고리": category["name"],
                        "브랜드": brand,
                        "모델명": name,
                        "SKU": sku,
                        "현재가 (SAR)": current_price,
                        "원가 (SAR)": original_price,
                        "할인율 (%)": discount,
                        "재고 상태": stock,
                        "URL": product_url,
                    })

            except Exception as e:
                logger.debug(f"  상품 파싱 오류: {e}")
                continue

        # 다음 페이지 확인
        next_btn = soup.select_one("a.next, a[rel='next'], li.next a")
        if not next_btn:
            break

        page += 1
        time.sleep(1.5)

    logger.info(f"  [{category['name']}] 총 {len(products)}개 상품 수집")
    return products


def ensure_workbook() -> None:
    """Excel 파일과 Prices DB 시트가 없으면 생성"""
    if not os.path.exists(EXCEL_FILE):
        logger.info(f"'{EXCEL_FILE}' 파일 생성 중...")
        wb = Workbook()
        ws = wb.active
        ws.title = PRICES_DB_SHEET
        ws.append(PRICES_DB_COLUMNS)
        wb.save(EXCEL_FILE)
        logger.info("파일 생성 완료.")
        return

    wb = load_workbook(EXCEL_FILE)
    if PRICES_DB_SHEET not in wb.sheetnames:
        logger.info(f"'{PRICES_DB_SHEET}' 시트 생성 중...")
        ws = wb.create_sheet(PRICES_DB_SHEET, 0)
        ws.append(PRICES_DB_COLUMNS)
        wb.save(EXCEL_FILE)
    wb.close()


def save_to_excel(products: list) -> None:
    """스크랩 데이터를 Prices DB 시트에 추가"""
    if not products:
        logger.warning("저장할 데이터가 없습니다.")
        return

    ensure_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[PRICES_DB_SHEET]

    today_str = datetime.now().strftime("%Y-%m-%d")
    appended = 0

    for p in products:
        row = [
            today_str,
            p.get("카테고리", ""),
            p.get("브랜드", ""),
            p.get("모델명", ""),
            p.get("SKU", ""),
            p.get("현재가 (SAR)", 0),
            p.get("원가 (SAR)", 0),
            p.get("할인율 (%)", 0),
            p.get("재고 상태", ""),
            p.get("URL", ""),
        ]
        ws.append(row)
        appended += 1

    wb.save(EXCEL_FILE)
    wb.close()
    logger.info(f"Prices DB에 {appended}개 행 저장 완료 → '{EXCEL_FILE}'")


def main() -> None:
    logger.info("=" * 60)
    logger.info("eXtra AC 가격 스크래퍼 v4 시작")
    logger.info("=" * 60)

    session = get_session()
    all_products = []

    for category in AC_CATEGORIES:
        logger.info(f"\n카테고리 처리 중: {category['name']}")
        products = scrape_category(session, category)
        all_products.extend(products)
        time.sleep(2)

    logger.info(f"\n총 {len(all_products)}개 상품 수집 완료.")
    save_to_excel(all_products)

    logger.info("\n[완료] Prices DB 업데이트 완료.")
    logger.info("다음 단계: extra_ac_dashboard_builder.py 를 실행하세요.")


if __name__ == "__main__":
    main()
