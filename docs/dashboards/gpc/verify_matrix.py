#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""계층3-확장: 필터 조합을 바꿔가며 화면 렌더값 ↔ 원본 엑셀 재계산 전수 대조.
Official + 채널 부분선택은 안분 추정이므로 안분식까지 여기서 독립 재현해 검증한다."""
import asyncio, collections, glob, json, re, sys
import openpyxl
from playwright.async_api import async_playwright

ACC=sorted(glob.glob("/home/ubuntu/2026/02. Operation Team/01. GPC Management/01. Monthly/GPC_Accrual*.xlsx"))[-1]
OFFX=sorted(glob.glob("/home/ubuntu/2026/10. Automation/03. Operation/00. GPC/02. 2026/04. Official GPC/*GPC official*.xlsx"))[-1]
PROV="/home/ubuntu/Shaker-MD-App/docs/dashboards/gpc/gpc_provisional.json"
AMT={"gsv":19,"yed":21,"adc":22,"vpd":25,"dsi":15,"cogs":12,"inv":27,"vsp":23}
MON={n:i for i,n in enumerate(["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],1)}
S2C={"inverter":"Split Inverter","on/off":"Split On/Off","window":"Window AC","free stand":"Floor Standing AC",
     "free standing":"Floor Standing AC","cassette":"Cassette AC","concealed":"Concealed Set","package":"Unitary Package"}
O2C={"window":"Window AC","split (on / off)":"Split On/Off","split inverter":"Split Inverter",
     "free standing":"Floor Standing AC","cassette":"Cassette AC","concealed (cac)":"Concealed Set",
     "package (cac)":"Unitary Package","convertible (cac)":"CAC Ducted","air purifier":"Accessory/Others",
     "multi v":"Multi-V","others":"Accessory/Others"}
OM={"GSV":("gsv",1),"YED":("yed",-1),"ADC":("adc",-1),"VPD":("vpd",-1),"DSI":("dsi",-1),
    "COGS":("cogs",1),"INV":("inv",-1),"VSP":("vsp",1),"NSV":("nsv",1),"GM":("gm",1)}
nk=lambda x: re.sub(r"[\s\-]+"," ",re.sub(r"\s+"," ",str(x or "")).strip().lower()).strip()
def num(v):
    try: return float(v)
    except (TypeError,ValueError): return 0.0
def cn(s):
    s=re.sub(r"\s+"," ",str(s or "")).strip(); return S2C.get(s.lower(),s)

# ── 원본을 최소 입도로 한 번만 읽어둔다 ──
A=collections.defaultdict(lambda: collections.defaultdict(float))     # (y,m,ch,cat)
wb=openpyxl.load_workbook(ACC,read_only=True,data_only=True)
for sh in wb.sheetnames:
    m0=re.fullmatch(r"Raw (20\d\d)",sh)
    if not m0: continue
    y=int(m0.group(1))
    for r in wb[sh].iter_rows(min_row=2,values_only=True):
        if r[28] is None and r[19] is None and r[12] is None: continue
        d=r[8]; m=d.month if hasattr(d,"month") else None
        if m is None:
            try:
                mi=int(float(r[30])); m=mi if 1<=mi<=12 else None
            except (TypeError,ValueError): m=MON.get(str(r[30]).strip()[:3].title())
        c=str(r[29] or ""); ch="IR" if "IR" in c else ("OR" if "OR" in c else None)
        if not m or not ch: continue
        for k,ci in AMT.items(): A[(y,m,ch,cn(r[28]))][k]+=num(r[ci])
wb.close()
pj=json.load(open(PROV,encoding="utf-8"))
for r in pj["rows"]:
    for k in AMT: A[(pj["year"],pj["month"],r["ch"],r["cat"])][k]+=r.get(k,0.0) or 0.0
PKEY=(pj["year"],pj["month"])

O=collections.defaultdict(lambda: collections.defaultdict(float))     # (y,m,cat) 실적만
V=list(openpyxl.load_workbook(OFFX,read_only=True,data_only=True).active.iter_rows(values_only=True))
lab,band=V[2],V[1]
for r in V[3:279]:
    met=str(r[5])
    if met not in OM or r[4] is None: continue
    cat=O2C[nk(r[4])]; k,sg=OM[met]
    for j in range(6,98):
        mm=re.fullmatch(r"(?:\((A|B)\))?(\d{4})(\d{2})",str(lab[j] or ""))
        if not mm or r[j] in (None,"") or mm.group(1)=="B" or str(band[j])=="FCST": continue
        O[(int(mm.group(2)),int(mm.group(3)),cat)][k]+=float(r[j])*sg

def exp_acc(y,months,chans,cats):
    o=dict.fromkeys(AMT,0.0)
    for (yy,m,ch,c),v in A.items():
        if yy!=y or m not in months or ch not in chans or c not in cats: continue
        for k in AMT: o[k]+=v[k]
    o["nsv"]=o["gsv"]+o["yed"]+o["adc"]+o["vpd"]+o["dsi"]
    o["gp"]=o["nsv"]-o["cogs"]+o["inv"]+o["vsp"]; return o

# Official 엔 'Applied' 축이 없다 — 그 매출은 Official 의 Others 안에 있다.
# 따라서 Accessory/Others 셀의 채널 비중 분모는 Accrual 의 (Accessory/Others + Applied) 여야 한다.
ABS={"Applied":"Accessory/Others"}
SHARE=collections.defaultdict(lambda: {"IR":0.0,"OR":0.0})
for (yy,m,ch,c),v in A.items():
    if (yy,m)==PKEY: continue                     # 가마감은 실적이 아니므로 비중 분모에서 제외
    SHARE[(yy,m,ABS.get(c,c))][ch]+=v["gsv"]

def exp_off(y,months,chans,cats):
    """Official. 채널 부분선택이면 (연·월·카테고리) Accrual GSV 비중으로 안분(가마감 제외)."""
    whole=chans=={"IR","OR"}
    o=collections.defaultdict(float)
    for (yy,m,c),v in O.items():
        if yy!=y or m not in months or c not in cats: continue
        w=1.0
        if not whole:
            b=SHARE.get((yy,m,c))
            t=(abs(b["IR"])+abs(b["OR"])) if b else 0.0
            if t<1: continue                      # 안분 불가 → 제외(화면도 동일)
            w=sum(abs(b[x]) for x in chans)/t
        for k,val in v.items(): o[k]+=val*w
    # Official 은 가마감을 내지 않는다 → 공시 없는 월은 Accrual 값을 그대로 (안분 없음)
    offM={m for (yy,m,_) in O if yy==y}
    for m in months:
        if m in offM: continue
        f=dict.fromkeys(AMT,0.0); any_=False
        for (yy,mm,ch,c),v in A.items():
            if yy!=y or mm!=m or ch not in chans or c not in cats: continue
            any_=True
            for k in AMT: f[k]+=v[k]
        if not any_: continue
        fn=f["gsv"]+f["yed"]+f["adc"]+f["vpd"]+f["dsi"]
        for k in AMT: o[k]+=f[k]
        o["nsv"]+=fn; o["gm"]+=fn-f["cogs"]+f["inv"]+f["vsp"]
    o["gp"]=o.get("gm",0.0); return o

CASES=[("기본 1~8월·전채널", "month",[1,2,3,4,5,6,7,8],{"IR","OR"},None),
       ("단월 3월",          "month",[3],{"IR","OR"},None),
       ("단월 7월",          "month",[7],{"IR","OR"},None),
       ("분기 Q2",           "quarter",["Q2"],{"IR","OR"},None),
       ("반기 H1",           "half",["H1"],{"IR","OR"},None),
       ("연간",              "year",["Y"],{"IR","OR"},None),
       ("1~8월 · IR만",      "month",[1,2,3,4,5,6,7,8],{"IR"},None),
       ("1~8월 · OR만",      "month",[1,2,3,4,5,6,7,8],{"OR"},None),
       ("단월 5월 · OR만",   "month",[5],{"OR"},None),
       ("1~8월 · Split Inverter만","month",[1,2,3,4,5,6,7,8],{"IR","OR"},["Split Inverter"]),
       ("Q3 · IR · Cassette","quarter",["Q3"],{"IR"},["Cassette AC"])]
MO={"month":lambda p:[p],"quarter":lambda p:[int(p[1])*3-2,int(p[1])*3-1,int(p[1])*3],
    "half":lambda p:[1,2,3,4,5,6] if p=="H1" else [7,8,9,10,11,12],"year":lambda p:list(range(1,13))}
LAB=["GSV","YED","ADC","VPD","DSI","NSV","COGS","INV","VSP","GP"]
KEY=["gsv","yed","adc","vpd","dsi","nsv","cogs","inv","vsp","gp"]

async def main():
    fails=[];errs=[];checked=0
    async with async_playwright() as p:
        b=await p.chromium.launch(); pg=await b.new_page(viewport={"width":1600,"height":1100})
        pg.on("console",lambda m: errs.append(m.text) if m.type=="error" else None)
        pg.on("pageerror",lambda e: errs.append(str(e)))
        await pg.goto("http://127.0.0.1:8899/index.html",wait_until="networkidle"); await pg.wait_for_timeout(600)
        for name,unit,periods,chans,cats in CASES:
            months=set(m for pp in periods for m in MO[unit](pp))
            cs=set(cats) if cats else None
            for basis,bl in [("Accrual","acc"),("Official","off")]:
                st=await pg.evaluate("""([u,ps,ch,ca,bs])=>{
                    S.basis=bs;S.unit=u;S.periods=new Set(ps);S.chans=new Set(ch);
                    S.cats=ca?new Set(ca):new Set(GPC_META.cats);
                    buildBar();buildPeriods();buildCat();render();
                    return [...S.periods].length;}""",[unit,periods,list(chans),cats,bl])
                await pg.wait_for_timeout(450)
                rows=await pg.eval_on_selector_all("#tblLadder tbody tr.mrow",
                    "es=>es.map(e=>[...e.querySelectorAll('td')].map(t=>t.textContent.trim()))")
                yrs=await pg.evaluate("()=>selYears()")
                allcats=cs or set(await pg.evaluate("()=>GPC_META.cats"))
                bad=0
                for i,lb in enumerate(LAB):
                    r=next((x for x in rows if x[0].split()[0]==lb),None)
                    if r is None: fails.append(f"{name}/{basis} {lb} 행없음"); continue
                    for j,y in enumerate(yrs):
                        shown=int(r[1+j].replace(",","").replace("−","-"))
                        e=(exp_acc if bl=="acc" else exp_off)(y,months,chans,allcats)
                        exp=round(e[KEY[i]]/1000); checked+=1
                        if shown!=exp:
                            bad+=1; fails.append(f"{name}/{basis} {lb} {y}: 화면={shown:,} 엑셀={exp:,} Δ={shown-exp:,}")
                print(f"  {name:<26} {basis:<9} {'✅' if bad==0 else '❌ %d건'%bad}")
        await b.close()
    print(f"\n총 대조 {checked:,}개 값 / 불일치 {len(fails)}건")
    for f in fails[:20]: print("   ❌",f)
    if errs: print("   콘솔에러:",errs[:3])
    print("="*62); print("필터 전 조합 렌더 완전 일치 ✅" if not (fails or errs) else "❌ 불일치")
    sys.exit(1 if (fails or errs) else 0)
asyncio.run(main())
