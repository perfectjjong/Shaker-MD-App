#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""계층3: 실제 화면에 렌더된 숫자 ↔ 원본 엑셀 직접 재계산 대조.
데이터파일이 맞아도 렌더 로직이 틀릴 수 있으므로 DOM 텍스트를 읽어 비교한다."""
import asyncio, collections, glob, json, re, sys
import openpyxl
from playwright.async_api import async_playwright

ACC=sorted(glob.glob("/home/ubuntu/2026/02. Operation Team/01. GPC Management/01. Monthly/GPC_Accrual*.xlsx"))[-1]
OFFX=sorted(glob.glob("/home/ubuntu/2026/10. Automation/03. Operation/00. GPC/02. 2026/04. Official GPC/*GPC official*.xlsx"))[-1]
PROV="/home/ubuntu/Shaker-MD-App/docs/dashboards/gpc/gpc_provisional.json"
URL="http://127.0.0.1:8899/index.html"
AMT={"qty":18,"gsv":19,"yed":21,"adc":22,"vpd":25,"dsi":15,"cogs":12,"inv":27,"vsp":23}
MON={n:i for i,n in enumerate(["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],1)}
S2C={"inverter":"Split Inverter","on/off":"Split On/Off","window":"Window AC","free stand":"Floor Standing AC",
     "free standing":"Floor Standing AC","cassette":"Cassette AC","concealed":"Concealed Set","package":"Unitary Package"}
O2C={"window":"Window AC","split (on / off)":"Split On/Off","split inverter":"Split Inverter",
     "free standing":"Floor Standing AC","cassette":"Cassette AC","concealed (cac)":"Concealed Set",
     "package (cac)":"Unitary Package","convertible (cac)":"CAC Ducted","air purifier":"Accessory/Others",
     "multi v":"Multi-V","others":"Accessory/Others"}
OM={"GSV":("gsv",1),"YED":("yed",-1),"ADC":("adc",-1),"VPD":("vpd",-1),"DSI":("dsi",-1),
    "COGS":("cogs",1),"INV":("inv",-1),"VSP":("vsp",1),"NSV":("nsv",1),"GM":("gm",1)}
nk=lambda x: re.sub(r"[\s\-]+"," ",re.sub(r"\s+"," ",str(x or "")).strip().lower()).strip()
def num(v):
    try: return float(v)
    except (TypeError,ValueError): return 0.0
def cn(s):
    s=re.sub(r"\s+"," ",str(s or "")).strip(); return S2C.get(s.lower(),s)
def mo_of(r):
    d=r[8]
    if hasattr(d,"month"): return d.month
    try:
        mi=int(float(r[30]));  return mi if 1<=mi<=12 else None
    except (TypeError,ValueError): pass
    return MON.get(str(r[30]).strip()[:3].title())

def accrual_expect(months):
    """대시보드 Accrual 모드가 보여야 할 값 = 엑셀 + 가마감 JSON"""
    a=collections.defaultdict(lambda: collections.defaultdict(float))
    wb=openpyxl.load_workbook(ACC,read_only=True,data_only=True)
    for sh in wb.sheetnames:
        m0=re.fullmatch(r"Raw (20\d\d)",sh)
        if not m0: continue
        y=int(m0.group(1))
        for r in wb[sh].iter_rows(min_row=2,values_only=True):
            if r[28] is None and r[19] is None and r[12] is None: continue
            m=mo_of(r); c=str(r[29] or "")
            ch="IR" if "IR" in c else ("OR" if "OR" in c else None)
            if not m or not ch or m not in months: continue
            for k,ci in AMT.items(): a[(y,cn(r[28]))][k]+=num(r[ci])
    wb.close()
    pj=json.load(open(PROV,encoding="utf-8"))
    if pj["month"] in months:
        for r in pj["rows"]:
            for k in AMT: a[(pj["year"],r["cat"])][k]+=r.get(k,0.0) or 0.0
    return a

def official_expect(months):
    V=list(openpyxl.load_workbook(OFFX,read_only=True,data_only=True).active.iter_rows(values_only=True))
    lab,band=V[2],V[1]
    a=collections.defaultdict(lambda: collections.defaultdict(float))
    for r in V[3:279]:
        met=str(r[5])
        if met not in OM or r[4] is None: continue
        cat=O2C[nk(r[4])]; k,sg=OM[met]
        for j in range(6,98):
            mm=re.fullmatch(r"(?:\((A|B)\))?(\d{4})(\d{2})",str(lab[j] or ""))
            if not mm or r[j] in (None,""): continue
            if mm.group(1)=="B" or str(band[j])=="FCST": continue      # 실적만
            if int(mm.group(3)) not in months: continue
            a[(int(mm.group(2)),cat)][k]+=float(r[j])*sg
    return a

def ladder_from(cells, official):
    out=collections.defaultdict(dict)
    for (y,c),v in cells.items():
        for k,val in v.items(): out[y][k]=out[y].get(k,0.0)+val
    for y,d in out.items():
        if official:
            d["gp"]=d.get("gm",0.0)
        else:
            d["nsv"]=d["gsv"]+d["yed"]+d["adc"]+d["vpd"]+d["dsi"]
            d["gp"]=d["nsv"]-d["cogs"]+d["inv"]+d["vsp"]
    return out

LAB=["GSV","YED","ADC","VPD","DSI","NSV","COGS","INV","VSP","GP"]
KEY=["gsv","yed","adc","vpd","dsi","nsv","cogs","inv","vsp","gp"]

async def main():
    months=set(range(1,9))                   # 화면 기본 = YOY 기준월 1~8
    EXP={"acc":ladder_from(accrual_expect(months),False),
         "off":ladder_from(official_expect(months),True)}
    EXPC={"acc":accrual_expect(months),"off":official_expect(months)}
    fails=[];errs=[]
    async with async_playwright() as p:
        b=await p.chromium.launch(); pg=await b.new_page(viewport={"width":1600,"height":1100})
        pg.on("console",lambda m: errs.append(m.text) if m.type=="error" else None)
        pg.on("pageerror",lambda e: errs.append(str(e)))
        await pg.goto(URL,wait_until="networkidle"); await pg.wait_for_timeout(700)
        for basis,label in [("Accrual","acc"),("Official","off")]:
            await pg.click(f'#cBasis button:has-text("{basis}")'); await pg.wait_for_timeout(800)
            yrs=await pg.evaluate("()=>selYears()")
            # 사다리표
            rows=await pg.eval_on_selector_all("#tblLadder tbody tr.mrow",
                "es=>es.map(e=>[...e.querySelectorAll('td')].map(t=>t.textContent.trim()))")
            print(f"\n■ [{basis}] 사다리표 (화면 vs 엑셀, 천 SAR)")
            for i,lb in enumerate(LAB):
                r=next((x for x in rows if x[0].split()[0]==lb),None)
                if r is None: fails.append(f"{basis} {lb} 행 없음"); continue
                for j,y in enumerate(yrs):
                    shown=int(r[1+j].replace(",","").replace("−","-"))
                    exp=round(EXP[label][y][KEY[i]]/1000)
                    ok=shown==exp
                    if not ok: fails.append(f"{basis} {lb} {y}: 화면={shown} 기대={exp}")
                    if i in (0,5,9) or not ok:
                        print(f"   {lb:<5}{y}  화면 {shown:>11,}  엑셀 {exp:>11,}  {'✅' if ok else '❌'}")
            # 카테고리표 GSV 열
            await pg.click('.tabs button[data-tab="category"]'); await pg.wait_for_timeout(800)
            crows=await pg.eval_on_selector_all("#tblCat tbody tr",
                "es=>es.map(e=>[...e.querySelectorAll('td')].map(t=>t.textContent.trim()))")
            ncol=6
            bad=0;chk=0
            for cr in crows:
                cat=cr[0].replace("※","").strip()
                if cat=="합계": continue
                for j,y in enumerate(yrs):
                    shown=int(cr[1+j*ncol].replace(",",""))
                    exp=round(EXPC[label].get((y,cat),{}).get("gsv",0.0)/1000)
                    chk+=1
                    if shown!=exp: bad+=1; fails.append(f"{basis} cat {cat} {y}: 화면={shown} 기대={exp}")
            print(f"   카테고리표 GSV: {chk}셀 / 불일치 {bad}건 {'✅' if bad==0 else '❌'}")
            await pg.click('.tabs button[data-tab="exec"]'); await pg.wait_for_timeout(400)
        await b.close()
    print("\n"+"="*60)
    if errs: print("콘솔에러:",errs[:4])
    print("화면 렌더 완전 일치 ✅" if not (fails or errs) else "❌ 불일치 %d건"%len(fails))
    for f in fails[:15]: print("  ",f)
    print("="*60)
    sys.exit(1 if (fails or errs) else 0)
asyncio.run(main())
