#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""GPC 전 계층 대조: 원본 엑셀 ↔ gpc_data.js.
빌더 코드를 재사용하지 않고 엑셀을 처음부터 다시 읽는다(같은 버그를 공유하지 않기 위함).
컬럼 좌표는 Accrual 워크북의 'Notes' 시트(2. RAW SHEET COLUMN MAP)로 교차 확인한 값이다."""
import collections, glob, json, re, sys
import openpyxl

sys.path.insert(0, "/home/ubuntu/2026/10. Automation")
from shared_classification import IR_CHANNEL_MAP, OR_CHANNEL_MAP, channel_from_name

# 계정 축은 SSOT 가 정의 그 자체이므로 같은 맵을 쓴다(여기서 재구현하면 그게 오답이 된다)
MAIN_ID = {**OR_CHANNEL_MAP,
           **{k: v for k, v in IR_CHANNEL_MAP.items() if v != "IR_Others"},
           1110000360: "Box Appliance"}
MAIN_SET = ({"eXtra", "Al Manea", "SWS", "Black Box", "Al Khunizan"} |
            {"BH", "Al Shathri", "BM", "Tamkeen", "Star Appliance",
             "Al Ghanem", "Dhamin", "Zagzoog", "Box Appliance"})


def account_of(row):
    try:
        cid = int(float(row[4]))
    except (TypeError, ValueError):
        cid = None
    ac = MAIN_ID.get(cid)
    if not ac:
        r = channel_from_name(row[5])
        ac = r if r in MAIN_SET else None
    return ac or "Others"

ACC_DIR = "/home/ubuntu/2026/02. Operation Team/01. GPC Management/01. Monthly"
OFF_DIR = ("/home/ubuntu/2026/10. Automation/03. Operation/00. GPC/02. 2026/04. Official GPC")
JS = "/home/ubuntu/Shaker-MD-App/docs/dashboards/gpc/gpc_data.js"
TOL = 0.01          # SAR. 빌더 반올림을 6자리로 올린 뒤의 실질 정밀도

# Notes 시트 기준 컬럼(0-idx): M=Cost(12) P=DSI(15) S=Qty(18) T=Value(19)
# V=YED(21) W=ADC(22) X=VSP(23) Z=EVPD(25) AB=Inv(27) AC=Category(28) AD=Chanel(29) AE=Months(30)
AMT = {"qty": 18, "gsv": 19, "yed": 21, "adc": 22, "vpd": 25, "dsi": 15,
       "cogs": 12, "inv": 27, "vsp": 23}
MON = {n: i for i, n in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}
SRC2CANON = {"inverter": "Split Inverter", "on/off": "Split On/Off", "window": "Window AC",
             "free stand": "Floor Standing AC", "free standing": "Floor Standing AC",
             "cassette": "Cassette AC", "concealed": "Concealed Set",
             "package": "Unitary Package"}
OFF2CANON = {"window": "Window AC", "split (on / off)": "Split On/Off",
             "split inverter": "Split Inverter", "free standing": "Floor Standing AC",
             "cassette": "Cassette AC", "concealed (cac)": "Concealed Set",
             "package (cac)": "Unitary Package", "convertible (cac)": "CAC Ducted",
             "air purifier": "Accessory/Others", "multi v": "Multi-V",
             "others": "Accessory/Others"}
OFF_MET = {"GSV": ("gsv", 1), "YED": ("yed", -1), "ADC": ("adc", -1), "VPD": ("vpd", -1),
           "DSI": ("dsi", -1), "COGS": ("cogs", 1), "INV": ("inv", -1), "VSP": ("vsp", 1),
           "NSV": ("nsv", 1), "GM": ("gm", 1)}

FAIL = []


def nk(x):
    return re.sub(r"[\s\-]+", " ", re.sub(r"\s+", " ", str(x or "")).strip().lower()).strip()


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def canon(s):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    return SRC2CANON.get(s.lower(), s)


def load_js():
    t = open(JS, encoding="utf-8").read()
    g = lambda p: json.loads(re.search(p, t, re.S).group(1))
    return (g(r"const GPC_META = (\{.*?\});"), g(r"const GPC_DATA = (\[.*?\]);"),
            g(r"const GPC_OFFICIAL = (\[.*?\]);"))


def month_of(row):
    d = row[8]
    if hasattr(d, "month"):
        return d.month
    m = row[30]
    try:
        mi = int(float(m))
        if 1 <= mi <= 12:
            return mi
    except (TypeError, ValueError):
        pass
    return MON.get(str(m).strip()[:3].title())


def cmp_block(title, xl, js, keys):
    """xl/js: dict[key] -> dict[metric] -> value"""
    bad = 0
    allk = set(xl) | set(js)
    for k in sorted(allk, key=str):
        a, b = xl.get(k, {}), js.get(k, {})
        for mk in keys:
            va, vb = a.get(mk, 0.0), b.get(mk, 0.0)
            if abs(va - vb) > TOL:
                bad += 1
                if bad <= 12:
                    print(f"    ❌ {k} {mk}: 엑셀={va:,.2f} 대시보드={vb:,.2f} Δ={vb-va:,.2f}")
    n = len(allk) * len(keys)
    print(f"  {title}: {n:,}셀 / 불일치 {bad}건 {'✅' if bad == 0 else '❌'}")
    if bad:
        FAIL.append(f"{title} {bad}건")
    return bad


def main():
    META, DATA, OFF = load_js()
    prov = META.get("provisional")
    pkey = (prov["year"], prov["month"]) if prov else None
    print(f"소스(Accrual) : {META['source']}")
    print(f"소스(Official): {META['official']['source']}")
    print(f"가마감 제외   : {pkey}\n")

    # ── 1. Accrual 원본 재집계 ──────────────────────────────
    src = sorted(glob.glob(f"{ACC_DIR}/GPC_Accrual*.xlsx"))[-1]
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    xl_full = collections.defaultdict(lambda: collections.defaultdict(float))
    rows_seen = 0
    for sh in wb.sheetnames:
        mo = re.fullmatch(r"Raw (20\d\d)", sh)
        if not mo:
            continue
        y = int(mo.group(1))
        for row in wb[sh].iter_rows(min_row=2, values_only=True):
            if row[28] is None and row[19] is None and row[12] is None:
                continue
            m = month_of(row)
            ch = str(row[29] or "")
            ch = "IR" if "IR" in ch else ("OR" if "OR" in ch else None)
            if not m or not ch:
                continue
            rows_seen += 1
            k = (y, m, ch, account_of(row), canon(row[28]))
            for mk, ci in AMT.items():
                xl_full[k][mk] += num(row[ci])
    wb.close()
    print(f"[1] Accrual 원본 재집계: 유효행 {rows_seen:,}건")

    KEYS = list(AMT)
    js_full = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in DATA:
        if pkey and (r["y"], r["m"]) == pkey:
            continue                       # 가마감은 엑셀에 없다(별도 JSON)
        k = (r["y"], r["m"], r["ch"], r["ac"], r["cat"])
        for mk in KEYS:
            js_full[k][mk] += r.get(mk, 0.0)
    cmp_block("연×월×채널×계정×카테고리 (9지표, 최소입도)", xl_full, js_full, KEYS)

    # 축을 접어가며 재확인 (집계 경로가 달라도 같아야 한다)
    for name, fold in [("연×월", lambda k: (k[0], k[1])),
                       ("연×월×채널", lambda k: (k[0], k[1], k[2])),
                       ("연×월×카테고리", lambda k: (k[0], k[1], k[4])),
                       ("연×채널", lambda k: (k[0], k[2])),
                       ("연×계정", lambda k: (k[0], k[3])),
                       ("연×월×계정", lambda k: (k[0], k[1], k[3])),
                       ("연×카테고리", lambda k: (k[0], k[4])),
                       ("연×채널×카테고리", lambda k: (k[0], k[2], k[4])),
                       ("연 합계", lambda k: (k[0],))]:
        A = collections.defaultdict(lambda: collections.defaultdict(float))
        B = collections.defaultdict(lambda: collections.defaultdict(float))
        for k, v in xl_full.items():
            for mk in KEYS:
                A[fold(k)][mk] += v[mk]
        for k, v in js_full.items():
            for mk in KEYS:
                B[fold(k)][mk] += v[mk]
        cmp_block(f"  ↳ {name}", A, B, KEYS)

    # ── 2. Official 원본 재집계 ────────────────────────────
    off_src = sorted(glob.glob(f"{OFF_DIR}/*GPC official*.xlsx"))[-1]
    V = list(openpyxl.load_workbook(off_src, read_only=True, data_only=True).active.iter_rows(values_only=True))
    lab, band = V[2], V[1]
    xl_off = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in V[3:279]:
        met = str(r[5])
        if met not in OFF_MET or r[4] is None:
            continue
        cat = OFF2CANON[nk(r[4])]
        mk, sg = OFF_MET[met]
        for j in range(6, 98):
            mo = re.fullmatch(r"(?:\((A|B)\))?(\d{4})(\d{2})", str(lab[j] or ""))
            if not mo or r[j] in (None, ""):
                continue
            b = "B" if mo.group(1) == "B" else ("F" if str(band[j]) == "FCST" else "A")
            xl_off[(int(mo.group(2)), int(mo.group(3)), b, cat)][mk] += float(r[j]) * sg
    xl_off = {k: v for k, v in xl_off.items() if any(abs(x) > TOL for x in v.values())}
    js_off = {(r["y"], r["m"], r["b"], r["cat"]): r for r in OFF}
    print(f"\n[2] Official 원본 재집계: 유효 셀그룹 {len(xl_off):,}")
    cmp_block("연×월×기준×카테고리 (10지표)", xl_off, js_off, list({v[0] for v in OFF_MET.values()}))

    print("\n" + "=" * 64)
    print("전 계층 일치 ✅" if not FAIL else "불일치 ❌ " + " / ".join(FAIL))
    print("=" * 64)
    sys.exit(1 if FAIL else 0)


main()
