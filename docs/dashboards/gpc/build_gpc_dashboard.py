#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""GPC 대시보드 데이터 빌더 v2: 인보이스 라인아이템 Accrual 원본 -> gpc_data.js

소스 (2026-08-16 형님 지시로 전면 교체):
  01. Monthly/GPC_Accrual Only_*.xlsx 최신본 — Raw 2024/2025/2026 시트 (35컬럼 라인아이템)

집계 축: 연 x 월 x 채널(IR/OR) x 계정(메인 14 + Others) x 카테고리
지표(금액 SAR, 파일 부호 그대로 — 할인 4종/INV 음수, VSP 양수):
  qty, gsv, yed, adc, vpd(EVPD), dsi, cogs, inv, vsp
파생(프론트 계산): NSV = GSV+YED+ADC+VPD+DSI / GP = NSV - COGS + INV + VSP

메인 계정 = SSOT shared_classification ID맵 (OR 5 + IR 8) + Box Appliance(1110000360),
ID 미매핑은 channel_from_name 이름 폴백. 채널 축 자체는 파일 AD(Chanel) 컬럼 기준.

검증: 빌더 집계 vs 원본 Summary Value 시트 실측값(YOY 사다리) 자동 대사.
"""
import glob
import json
import re
import sys

import openpyxl

sys.path.insert(0, "/home/ubuntu/2026/10. Automation")
from shared_classification import IR_CHANNEL_MAP, OR_CHANNEL_MAP, channel_from_name

SRC_DIR = "/home/ubuntu/2026/02. Operation Team/01. GPC Management/01. Monthly"
OUT = "/home/ubuntu/Shaker-MD-App/docs/dashboards/gpc/gpc_data.js"

MAIN_ID = {**OR_CHANNEL_MAP,
           **{k: v for k, v in IR_CHANNEL_MAP.items() if v != "IR_Others"},
           1110000360: "Box Appliance"}
OR_MAINS = ["eXtra", "Al Manea", "SWS", "Black Box", "Al Khunizan"]
IR_MAINS = ["BH", "Al Shathri", "BM", "Tamkeen", "Star Appliance",
            "Al Ghanem", "Dhamin", "Zagzoog", "Box Appliance"]
MAIN_SET = set(OR_MAINS) | set(IR_MAINS)

# 2026-08-26 형님 지시로 카테고리 정본화. 원본 Accrual 엑셀(AC컬럼)은 구 라벨을 쓰지만
# 대시보드 산출물은 전사 정본 라벨로 통일한다 — 다른 대시보드와 필터·비교가 어긋나던 문제.
# ⚠️ **1:1 대응만** 바꾼다. GPC 고유 축(CAC Ducted / Applied / Accessory/Others)은
#    정본에 대응 항목이 없어 병합하면 손익 분석 정보가 사라지므로 그대로 둔다.
#    (예: CAC Ducted 를 Concealed Set 에 합치면 덕트형 구분이 소실)
CATS = ["Split Inverter", "Split On/Off", "Window AC", "Floor Standing AC", "Cassette AC",
        "Concealed Set", "CAC Ducted", "Unitary Package", "Multi-V", "Applied",
        "Accessory/Others"]

# 원본 엑셀 라벨 → 정본 라벨 (norm_cat 에서 최우선 적용)
_SRC_TO_CANON = {
    "inverter": "Split Inverter",
    "on/off": "Split On/Off",
    "window": "Window AC",
    "free stand": "Floor Standing AC",
    "free standing": "Floor Standing AC",
    "cassette": "Cassette AC",
    "concealed": "Concealed Set",
    "package": "Unitary Package",
}
MONTH_NAMES = {n: i for i, n in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}

# Raw 시트 컬럼(0-idx): 4=Customer 5=CustomerName 8=InvoiceDate 12=Cost 15=DSI
# 18=Qty 19=Value(GSV) 21=YED 22=ADC 23=VSP 25=EVPD 27=Inv 28=Category 29=Chanel 30=Months
AMT = {"gsv": 19, "yed": 21, "adc": 22, "vpd": 25, "dsi": 15,
       "cogs": 12, "inv": 27, "vsp": 23}
METRIC_KEYS = ["qty", "gsv", "yed", "adc", "vpd", "dsi", "cogs", "inv", "vsp"]


def latest_src():
    files = sorted(glob.glob(f"{SRC_DIR}/GPC_Accrual*.xlsx"))
    if not files:
        sys.exit(f"❌ {SRC_DIR}에 GPC_Accrual*.xlsx 없음")
    return files[-1]  # 파일명 날짜(YYMMDD) 정렬 = 최신


def norm_cat(s):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    canon = _SRC_TO_CANON.get(s.lower())
    if canon:
        return canon
    for c in CATS:
        if s.lower() == c.lower():
            return c
    return s or "Accessory/Others"


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def row_month(row):
    d = row[8]
    if hasattr(d, "month"):
        return d.month
    m = row[30]
    try:
        mi = int(float(m))
        if 1 <= mi <= 12:
            return mi
    except (TypeError, ValueError):
        pass
    return MONTH_NAMES.get(str(m).strip()[:3].title())


def account_of(row):
    try:
        cid = int(float(row[4]))
    except (TypeError, ValueError):
        cid = None
    ac = MAIN_ID.get(cid)
    if not ac:
        r = channel_from_name(row[5])
        ac = r if r in MAIN_SET else None
    return ac or "Others"


def main():
    src = latest_src()
    print(f"소스: {src.split('/')[-1]}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    agg = {}  # (y,m,ch,ac,cat) -> [metrics]
    skipped = 0
    for sheet in wb.sheetnames:
        mo = re.fullmatch(r"Raw (20\d\d)", sheet)
        if not mo:
            continue
        year = int(mo.group(1))
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[28] is None and row[19] is None and row[12] is None:
                continue
            m = row_month(row)
            chan = str(row[29] or "")
            ch = "IR" if "IR" in chan else ("OR" if "OR" in chan else None)
            if not m or not ch:
                skipped += 1
                continue
            key = (year, m, ch, account_of(row), norm_cat(row[28]))
            a = agg.setdefault(key, [0.0] * len(METRIC_KEYS))
            a[0] += num(row[18])
            for i, k in enumerate(METRIC_KEYS[1:], 1):
                a[i] += num(row[AMT[k]])

    records = []
    for (y, m, ch, ac, cat), v in sorted(agg.items()):
        rec = {"y": y, "m": m, "ch": ch, "ac": ac, "cat": cat}
        rec.update({k: round(v[i], 1) for i, k in enumerate(METRIC_KEYS)})
        records.append(rec)

    years = sorted({r["y"] for r in records})
    year_months = {str(y): sorted({r["m"] for r in records if r["y"] == y}) for y in years}
    yoy_months = year_months[str(years[-1])]

    # ---- 정합성 검증: 원본 Summary Value 시트(YOY 사다리) vs 빌더 집계 ----
    def tot(y, key, months):
        return sum(r[key] for r in records if r["y"] == y and r["m"] in months)

    ws = wb["Summary Value"]
    sheet_vals = {}  # (metric, year) -> value
    ycols = {}
    for row in ws.iter_rows(min_row=3, max_row=25, min_col=5, max_col=11, values_only=True):
        label = str(row[0] or "").strip()
        if label == "Metric":
            ycols = {int(v): i for i, v in enumerate(row[1:4], 1) if isinstance(v, (int, float))}
        elif label in ("GSV", "YED", "ADC", "VPD", "DSI", "NSV", "COGS", "INV", "VSP", "GP"):
            for yy, ci in ycols.items():
                if isinstance(row[ci], (int, float)):
                    sheet_vals[(label, yy)] = float(row[ci])
    wb.close()

    def derived(y, months):
        d = {k: tot(y, k, months) for k in METRIC_KEYS[1:]}
        d["NSV"] = d["gsv"] + d["yed"] + d["adc"] + d["vpd"] + d["dsi"]
        d["GP"] = d["NSV"] - d["cogs"] + d["inv"] + d["vsp"]
        return {"GSV": d["gsv"], "YED": d["yed"], "ADC": d["adc"], "VPD": d["vpd"],
                "DSI": d["dsi"], "NSV": d["NSV"], "COGS": d["cogs"], "INV": d["inv"],
                "VSP": d["vsp"], "GP": d["GP"]}

    # 허용오차 10 SAR: 레코드별 0.1 SAR 반올림이 ~2천건 누적될 때의 상한 (실측 최대 Δ=1.6)
    print("=== 정합성 검증 (Summary Value 시트 YOY vs 빌더, 허용오차 10 SAR) ===")
    bad = 0
    yset = set(yoy_months)
    for (label, yy), sv in sorted(sheet_vals.items(), key=lambda x: (x[0][1], x[0][0])):
        cv = derived(yy, yset)[label]
        if abs(cv - sv) > 10:
            print(f"  ❌ {yy} {label}: sheet={sv:,.1f} builder={cv:,.1f} Δ={cv - sv:,.1f}")
            bad += 1
    print(f"  검증 셀: {len(sheet_vals)} / 불일치: {bad} (skipped rows: {skipped})")

    # ── 가마감(provisional) 주입 ──────────────────────────────────
    # 월마감 전 잠정 시뮬레이션을 별도 JSON으로 얹는다. 실적(Accrual)과 구분 표기.
    # 생성: 03. Operation/00. GPC/_engine/build_provisional.py
    prov_meta = None
    PROV = OUT.replace("gpc_data.js", "gpc_provisional.json")
    import os
    if os.path.exists(PROV):
        with open(PROV, encoding="utf-8") as pf:
            pj = json.load(pf)
        py, pm = int(pj["year"]), int(pj["month"])
        exists = any(r["y"] == py and r["m"] == pm for r in records)
        if exists:
            print(f"  ⏭  {py}-{pm}월은 이미 실적이 있어 가마감을 얹지 않음")
        else:
            records.extend(pj["rows"])
            ys = str(py)
            if ys in year_months and pm not in year_months[ys]:
                year_months[ys] = sorted(year_months[ys] + [pm])
            if pm not in yoy_months:
                yoy_months = sorted(yoy_months + [pm])   # 화면 선택 가능하도록 (검증은 위에서 실적만으로 완료)
            prov_meta = {"year": py, "month": pm,
                         "basis": pj.get("basis", "가마감"),
                         "built": pj.get("built", "")}
            print(f"  🟡 가마감 주입: {py}-{pm}월 {len(pj['rows'])}건 ({pj.get('basis','')})")

    # ── Official (Finance 공시 최종본) 레이어 ────────────────────
    # Accrual 값은 건드리지 않는다. 별도 배열로 얹어 화면에서 대사만 한다.
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))  # cwd 무관하게 import
    import gpc_official
    print("=== Official 레이어 ===")
    off_rows, off_meta = gpc_official.load(records)

    meta = {
        "cats": CATS,
        "channels": ["IR", "OR"],
        "or_mains": OR_MAINS,
        "ir_mains": IR_MAINS,
        "metric_keys": METRIC_KEYS,
        "years": years,
        "year_months": year_months,
        "yoy_months": yoy_months,
        "source": src.split("/")[-1],
        "basis": "Accrual (발생주의 잠정)",
        "provisional": prov_meta,
        "official": off_meta,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// GPC 대시보드 데이터 v2 (build_gpc_dashboard.py 자동생성 — 직접 수정 금지)\n")
        f.write("const GPC_META = " + json.dumps(meta, ensure_ascii=False) + ";\n")
        f.write("const GPC_DATA = " + json.dumps(records, ensure_ascii=False) + ";\n")
        f.write("const GPC_OFFICIAL = " + json.dumps(off_rows, ensure_ascii=False) + ";\n")

    print(f"\n✅ 생성: {OUT}")
    print(f"   레코드 {len(records)}건 · 연도 {years} · YOY기준월 {yoy_months}")
    for y in years:
        mm = year_months[str(y)]
        print(f"   {y}: {len(mm)}개월 ({mm[0]}~{mm[-1]}) GSV {tot(y, 'gsv', set(mm)) / 1e6:,.1f}M")


if __name__ == "__main__":
    main()
