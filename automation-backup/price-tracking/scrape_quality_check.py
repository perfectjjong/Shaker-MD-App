#!/usr/bin/env python3
"""
Price Tracking 스크래퍼 품질 검증 + 자동 재실행 공통 모듈.

각 채널 wrapper(_run/_pipeline/_run_all)에서 import하여 사용.
스크래퍼 일시 누락(페이지 fail 등)으로 SKU 카운트가 -10%+ 급감 시
자동 재실행 1회 + telegram 알림.

사용 (각 wrapper의 main 함수에서):
    from scrape_quality_check import check_scrape_quality, notify_telegram, run_with_retry

    CHANNEL = {
        'name': 'eXtra',
        'master': SCRIPT_DIR / 'extra_ac_Prices_Tracking_Master.xlsx',
        'sku_col': 'SKU',
        'date_col': 'Scraped_At',
    }

    # Scraper 실행 후
    ok = run_step(...)
    if ok:
        run_with_retry(CHANNEL, scraper_runner=lambda: run_step('Scraper Retry', SCRAPER), threshold=10.0)
"""
import os
import sys
from pathlib import Path
from collections import defaultdict


def check_scrape_quality(
    master_path,
    sku_col: str,
    date_col: str,
    sheet_name=None,
    retry_threshold_pct: float = 10.0,
) -> tuple[bool, str]:
    """
    스크래퍼 품질 점검: master xlsx에서 latest vs prev SKU 카운트 비교.

    Returns:
        (ok, msg) — ok=True면 정상, False면 재실행 필요
    """
    try:
        import openpyxl
        master_path = Path(master_path)
        if not master_path.exists():
            return True, f"master 파일 없음 ({master_path.name}) — 첫 실행 또는 검증 스킵"

        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        headers = [str(c.value or '').strip() for c in ws[1]]

        if sku_col not in headers:
            wb.close()
            return True, f"SKU 컬럼('{sku_col}') 없음 — 검증 스킵"
        if date_col not in headers:
            wb.close()
            return True, f"date 컬럼('{date_col}') 없음 — 검증 스킵"
        sku_idx = headers.index(sku_col)
        date_idx = headers.index(date_col)

        date_skus = defaultdict(set)
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(row[sku_idx]) if row[sku_idx] is not None else None
            raw_date = row[date_idx]
            if raw_date is None:
                continue
            # date 형식 통일 (datetime/str 모두 → YYYY-MM-DD)
            date = str(raw_date)[:10]
            if sku and date:
                date_skus[date].add(sku)
        wb.close()

        dates = sorted(date_skus.keys())
        if len(dates) < 2:
            return True, "비교할 전일 데이터 없음"

        latest, prev = dates[-1], dates[-2]
        cur_count, prev_count = len(date_skus[latest]), len(date_skus[prev])
        if prev_count == 0:
            return True, "전일 데이터 0건"

        diff_pct = (cur_count - prev_count) / prev_count * 100
        msg = f"latest({latest})={cur_count} vs prev({prev})={prev_count} → {diff_pct:+.1f}%"

        if diff_pct < -retry_threshold_pct:
            return False, f"⚠️ SKU -{abs(diff_pct):.1f}% 급감 → 스크래퍼 누락 의심. {msg}"
        return True, f"정상. {msg}"
    except Exception as e:
        return True, f"검증 오류 (스킵): {e}"


def notify_telegram(msg: str):
    """telegram 알림."""
    try:
        sys.path.insert(0, '/home/ubuntu/sonolbot')
        from telegram_sender import send_message_sync
        env_path = '/home/ubuntu/sonolbot/.env'
        chat_id = None
        if os.path.isfile(env_path):
            with open(env_path) as f:
                for line in f:
                    if line.startswith('TELEGRAM_ALLOWED_USERS='):
                        chat_id = int(line.split('=', 1)[1].strip().split(',')[0])
                        break
        if chat_id:
            send_message_sync(chat_id, msg)
    except Exception as e:
        print(f"[Telegram notify failed] {e}")


def check_scrape_quality_filebased(
    directory,
    pattern: str,
    sku_col: str,
    sheet_name=None,
    retry_threshold_pct: float = 10.0,
    exclude_substr: str = '_partial',
) -> tuple[bool, str]:
    """
    파일 기반 품질 점검: 가장 최근 2개 파일 SKU 카운트 비교.
    Tamkeen 같이 매일 별도 파일로 저장되는 채널용.
    """
    try:
        import openpyxl, glob
        files = sorted([f for f in glob.glob(f"{directory}/{pattern}") if exclude_substr not in f],
                       key=os.path.getmtime, reverse=True)
        if len(files) < 2:
            return True, f"비교할 전일 파일 없음 (찾음: {len(files)}건)"
        latest_f, prev_f = files[0], files[1]

        def _count_skus(f):
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            headers = [str(c.value or '').strip() for c in ws[1]]
            if sku_col not in headers:
                wb.close()
                return None
            idx = headers.index(sku_col)
            skus = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[idx] is not None:
                    skus.add(str(row[idx]))
            wb.close()
            return len(skus)

        cur_count = _count_skus(latest_f)
        prev_count = _count_skus(prev_f)
        if cur_count is None or prev_count is None:
            return True, f"SKU 컬럼('{sku_col}') 없음 — 검증 스킵"
        if prev_count == 0:
            return True, "전일 파일 SKU 0건"
        diff_pct = (cur_count - prev_count) / prev_count * 100
        msg = f"latest({os.path.basename(latest_f)})={cur_count} vs prev({os.path.basename(prev_f)})={prev_count} → {diff_pct:+.1f}%"
        if diff_pct < -retry_threshold_pct:
            return False, f"⚠️ SKU -{abs(diff_pct):.1f}% 급감. {msg}"
        return True, f"정상. {msg}"
    except Exception as e:
        return True, f"검증 오류 (스킵): {e}"


def run_with_retry(channel_config: dict, scraper_runner, log_func=print, threshold: float = 10.0):
    """
    스크래퍼 실행 후 품질 검증 + 자동 1회 재실행.

    Args:
        channel_config: {'name': str, 'master': Path, 'sku_col': str, 'date_col': str, 'sheet_name': str(optional)}
        scraper_runner: callable() → bool (재실행 함수)
        log_func: 로그 출력 함수 (default print)
        threshold: 급감 임계 (%)
    """
    name = channel_config['name']
    sheet = channel_config.get('sheet_name')
    ok, msg = check_scrape_quality(
        channel_config['master'],
        channel_config['sku_col'],
        channel_config['date_col'],
        sheet_name=sheet,
        retry_threshold_pct=threshold,
    )
    log_func(f"[Quality Check 1] {name}: {msg}")
    if ok:
        return True

    log_func(f"[Auto Retry] {name}: 1회 자동 재실행 시작...")
    notify_telegram(f"[{name} Auto Retry] 스크래퍼 SKU 누락 감지 → 자동 재실행 1회\n{msg}")

    retry_ok = scraper_runner()
    ok2, msg2 = check_scrape_quality(
        channel_config['master'],
        channel_config['sku_col'],
        channel_config['date_col'],
        sheet_name=sheet,
        retry_threshold_pct=threshold,
    )
    log_func(f"[Quality Check 2 — Retry] {name}: {msg2}")
    if not ok2:
        notify_telegram(f"[{name} Retry 실패] 자동 재실행 후에도 누락 지속\n{msg2}\n수동 점검 필요")
        return False
    return True
