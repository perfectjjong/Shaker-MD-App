#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Najm Store AC — 통합 실행 파일 (스케줄러용)
  1단계: 스크래퍼  (najm_scraper.py)          → Master Excel 누적 저장
  2단계: 대시보드  (najm_ac_html_dashboard.py) → HTML 생성 + 배포

스케줄러 등록 예시 (Windows 작업 스케줄러):
  python "C:/Users/J_park/Documents/2026/01. Work/06. Price Tracking/03. Najm Store/najm_ac_run.py"
"""

import subprocess
import sys
import os
from pathlib import Path
from datetime import datetime

SCRIPT_DIR = Path(__file__).parent
SCRAPER    = SCRIPT_DIR / 'najm_scraper.py'
DASHBOARD  = SCRIPT_DIR / 'najm_ac_html_dashboard.py'
LOG_FILE   = SCRIPT_DIR / 'najm_ac_run.log'

# stdout → UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass


def log(msg: str):
    """콘솔 + 로그 파일에 동시 출력"""
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def run_step(label: str, script: Path) -> bool:
    log(f"{'='*60}")
    log(f"START: {label}")
    log(f"{'='*60}")

    if not script.exists():
        log(f"[ERROR] Script not found: {script}")
        return False

    STEP_TIMEOUT = 1500  # 25분: 이 시간 초과 시 스크래퍼 강제 종료 후 대시보드 단계 진행
    try:
        result = subprocess.run(
            [sys.executable, str(script)],
            cwd=str(SCRIPT_DIR),
            stdin=subprocess.DEVNULL,  # input() 대기 없이 즉시 통과 → 자동화 필수
            timeout=STEP_TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        log(f"[TIMEOUT] {label} — {STEP_TIMEOUT}s 초과, 강제 종료")
        return False

    if result.returncode == 0:
        log(f"[OK] {label} completed successfully.")
        return True
    else:
        log(f"[ERROR] {label} failed — returncode={result.returncode}")
        return False


def check_scrape_quality(retry_threshold_pct: float = 10.0) -> tuple[bool, str]:
    """
    스크래퍼 결과 품질 점검: 최신 일자 SKU 카운트 vs 전일 비교.
    -10%+ 누락이면 (False, 사유) 반환.
    """
    try:
        import openpyxl
        from collections import defaultdict
        master = SCRIPT_DIR / "najm_ac_master.xlsx"
        if not master.exists():
            return True, "master 파일 없음 (첫 실행)"
        wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in ws[1]]
        try:
            sku_idx = headers.index('product_id')
            date_idx = headers.index('run_date')
        except ValueError:
            wb.close()
            return True, "컬럼 구조 변경 — 검증 스킵"
        date_skus = defaultdict(set)
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(row[sku_idx]) if row[sku_idx] else None
            date = str(row[date_idx])[:10] if row[date_idx] else None
            if sku and date:
                date_skus[date].add(sku)
        wb.close()
        dates = sorted(date_skus.keys())
        if len(dates) < 2:
            return True, "비교할 전일 데이터 없음"
        latest, prev = dates[-1], dates[-2]
        cur_count, prev_count = len(date_skus[latest]), len(date_skus[prev])
        if prev_count == 0:
            return True, "전일 데이터 0건"
        diff_pct = (cur_count - prev_count) / prev_count * 100
        msg = f"latest({latest})={cur_count} vs prev({prev})={prev_count} → {diff_pct:+.1f}%"
        if diff_pct < -retry_threshold_pct:
            return False, f"⚠️ SKU 카운트 -{abs(diff_pct):.1f}% 급감 → 스크래퍼 누락 의심. {msg}"
        return True, f"정상. {msg}"
    except Exception as e:
        return True, f"검증 오류 (스킵): {e}"


def notify_telegram(msg: str):
    """telegram 알림 (.env 토큰)."""
    try:
        sys.path.insert(0, '/home/ubuntu/sonolbot')
        from telegram_sender import send_message_sync
        env_path = '/home/ubuntu/sonolbot/.env'
        chat_id = None
        if os.path.isfile(env_path):
            with open(env_path) as f:
                for line in f:
                    if line.startswith('TELEGRAM_ALLOWED_USERS='):
                        chat_id = int(line.split('=', 1)[1].strip().split(',')[0])
                        break
        if chat_id:
            send_message_sync(chat_id, msg)
    except Exception as e:
        log(f"[Telegram notify failed] {e}")


def main():
    started = datetime.now()
    log(f"{'#'*60}")
    log(f"Najm Store AC Runner started")
    log(f"{'#'*60}")

    # ── Step 1: Scraper ──────────────────────────────────────────
    ok_scraper = run_step("Scraper (najm_scraper)", SCRAPER)

    # ── Step 1.5: 품질 검증 + 자동 재실행 (최대 1회) ──────────────
    # 스크래퍼 일시 누락(페이지 fail 등)으로 SKU 카운트 -10%+ 급감 시 자동 재시도.
    # drop_duplicates(keep='last')라 동일 일자 데이터 안전하게 덮어씀.
    quality_ok, quality_msg = check_scrape_quality(retry_threshold_pct=10.0)
    log(f"[Quality Check 1] {quality_msg}")
    if not quality_ok and ok_scraper:
        log(f"[Auto Retry] 1회 자동 재실행 시작...")
        notify_telegram(f"[Najm Auto Retry] 스크래퍼 SKU 누락 감지 → 자동 재실행 1회\n{quality_msg}")
        ok_scraper2 = run_step("Scraper Retry (najm_scraper)", SCRAPER)
        quality_ok2, quality_msg2 = check_scrape_quality(retry_threshold_pct=10.0)
        log(f"[Quality Check 2 — Retry] {quality_msg2}")
        if not quality_ok2:
            notify_telegram(f"[Najm Retry 실패] 자동 재실행 후에도 누락 지속\n{quality_msg2}\n수동 점검 필요")

    # ── Step 2: Dashboard (스크래퍼 성공 여부 관계없이 실행) ──────
    #   Excel에 데이터가 쌓여 있다면 대시보드는 항상 최신 기준으로 생성
    ok_dashboard = run_step("Dashboard (najm_ac_html_dashboard)", DASHBOARD)

    # ── 결과 요약 ────────────────────────────────────────────────
    elapsed = (datetime.now() - started).seconds
    log(f"{'#'*60}")
    log(f"DONE — elapsed: {elapsed}s  |  scraper={'OK' if ok_scraper else 'FAIL'}  |  dashboard={'OK' if ok_dashboard else 'FAIL'}")
    log(f"{'#'*60}\n")

    # 둘 다 실패 시 exit code 1 (스케줄러에서 실패 감지용)
    if not ok_scraper and not ok_dashboard:
        sys.exit(1)


if __name__ == '__main__':
    main()
