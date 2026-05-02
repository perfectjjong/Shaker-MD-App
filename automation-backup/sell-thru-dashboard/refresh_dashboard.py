"""
Sell-Thru Dashboard Auto Refresh Script
========================================
이 스크립트를 실행하면:
1. 00. Daily Sell Thru 폴더에서 최신 파일을 읽어 2026 Raw data에 자동 반영
2. Excel(Sell_Thru_Dashboard_Data.xlsx) + HTML(Sell_Thru_Dashboard.html) 자동 갱신

- 같은 월 데이터끼리만 비교 (월 단위 마감)
- 이미 있는 날짜 데이터는 건너뜀

사용법:
  python refresh_dashboard.py
"""

import openpyxl
import pandas as pd
import json
import os
import re
import calendar
import hashlib
import pickle
import time
from collections import Counter
from datetime import datetime, date, timedelta

# ============================================================
# CONFIG
# ============================================================
BASE = os.path.dirname(os.path.abspath(__file__))
RAW_DATA_BASE = os.path.join(os.path.dirname(BASE), '00. Raw Data')

FILES = {
    '2024_raw': os.path.join(RAW_DATA_BASE, '00. 2024', '2024 Sell thru Raw data.xlsx'),
    '2025_raw': os.path.join(RAW_DATA_BASE, '01. 2025', '2025 Sell thru Raw data.xlsx'),
    '2026_raw': os.path.join(RAW_DATA_BASE, '02. 2026', '2026 Sell thru Raw data.xlsx'),
    'classification': os.path.join(RAW_DATA_BASE, '01. Classfication.xlsx'),
}

# IR Excel Mar Sell-Thru Target
IR_TARGET_FILE = os.path.join(
    os.path.dirname(os.path.dirname(BASE)),
    '01. Sales', '01. Sell out', '02. Monthly', '01. 2026-Jan', '01. IR',
    'All IR - STK & SO as of 28 Feb 2026_Dashboard.xlsx'
)

DAILY_DIR = os.path.join(RAW_DATA_BASE, '02. 2026', '00. Daily Sell Thru')
DEALER_MAPPING_2026 = os.path.join(RAW_DATA_BASE, '02. 2026 Dealer Mapping.xlsx')

# 2026 Dealer Mapping C열 → 대시보드 팀명 변환
_DEALER_MAP_EXCEL_TO_TEAM = {
    'B2B':                       'Projects',
    'AFS':                       'AFS',
    'Dealer - IR':               'IR_Others',
    'Dealer - OR':               'OR_Others',
    'AMC':                       'AMC',
    'Projects':                  'Projects',
    'Workshops':                 'Workshops',
    'Dealers - SP':              'AFS',
    'Spart Part - Counter Sales':'Spart Part - Counter Sales',
    'SME':                       'SME',
    'Showroom':                  'OR_Others',
    'ESCO':                      'ESCO',
    'Online':                    'E-Commerce',
}

def _load_dealer_map_2026():
    """02. 2026 Dealer Mapping.xlsx → {account_id: team}"""
    if not os.path.exists(DEALER_MAPPING_2026):
        return {}
    try:
        df = pd.read_excel(DEALER_MAPPING_2026)
        result = {}
        for _, row in df.iterrows():
            cid = normalize_id(row.get('Customer ID (Payer)'))
            ctype = str(row.get('Customer Type Description', '')).strip()
            team = _DEALER_MAP_EXCEL_TO_TEAM.get(ctype)
            if cid and team:
                result[cid] = team
        print(f"   Dealer Mapping 2026 로드: {len(result)}개 계정")
        return result
    except Exception as e:
        print(f"   [WARN] Dealer Mapping 로드 실패: {e}")
        return {}

# OUD: try user-confirmed path first, then original path
_OUD_USER = os.path.join(os.path.dirname(os.path.dirname(BASE)), '01. Sales', '02. Sell Thru', '06. OUD')
_OUD_ORIG = os.path.join(RAW_DATA_BASE, '02. 2026', '05. OUD')
OUD_DIR = _OUD_USER if os.path.isdir(_OUD_USER) else _OUD_ORIG

OUTPUT_EXCEL = os.path.join(BASE, 'Sell_Thru_Dashboard_Data.xlsx')
# 배포 폴더: 03. Reporting 경로의 실제 git 레포 사용 (10. Automation 경로에는 .git 없음)
# BASE = .../01. Work/10. Automation/00. Sell Thru Dashboard/01. Python Code
# dirname x3 → .../01. Work → + 03. Reporting/00. Sell Thru/03. github-deploy
DEPLOY_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(BASE))),
                          '03. Reporting', '00. Sell Thru', '03. github-deploy')
OUTPUT_HTML = os.path.join(DEPLOY_DIR, 'index.html')
OUTPUT_DATA_JSON = os.path.join(DEPLOY_DIR, 'data.json')

CACHE_DIR = os.path.join(BASE, '.cache')


def _file_signature(filepath):
    """Return mtime+size string for change detection."""
    if not os.path.exists(filepath):
        return None
    st = os.stat(filepath)
    return f"{st.st_mtime}_{st.st_size}"


def _load_cache(key):
    """Load cached data if exists. Returns (signature, data) or (None, None)."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_file = os.path.join(CACHE_DIR, f'{key}.pkl')
    if os.path.exists(cache_file):
        with open(cache_file, 'rb') as f:
            return pickle.load(f)
    return None, None


def _save_cache(key, signature, data):
    """Save data to cache with file signature."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_file = os.path.join(CACHE_DIR, f'{key}.pkl')
    with open(cache_file, 'wb') as f:
        pickle.dump((signature, data), f)

# Employee Numbers → force SME
SME_EMPLOYEES = {24, 158}

# 2026 OR 5-channel account IDs (Sell Out managed channels)
# These keep their 'OR' team label; all other OR/OR_Others → 'OR_Others'
OR_5_CHANNEL_IDS = {
    1110000001, 1110000002, 1110000003,  # Al Manea
    1110000004, 1110000005,               # SWS
    1110000006,                           # Black Box
    1110000007,                           # Al Khunizan
    1120000000,                           # eXtra (United Electronics) — 1110000369 Extra Value Est는 OR_Others로 분리 (msg_post_9611, 2026-05-02)
}

# 2026 IR 8-channel account IDs (Sell Out managed channels)
# These are reclassified to 'IR'; all other IR/IR_Others/SME → 'IR_Others'
IR_8_CHANNEL_IDS = {
    1110000000,   # BH  – Saudi Bin Hamood Co.
    1110000009,   # BM  – Bin Momen Trading Co.
    1110000010,   # Tamkeen International Co.
    1110000299,   # Zagzoog – Al-Zaqzouq Home Appliances
    1110000015,   # Dhamin – Abdullah Ali Al-Dhamin & Sons
    1110000101,   # Star Appliance – Star Appliances
    1110000253,   # Al Ghanem – Alghanim Trading Company
    1110000065,   # Al Shathri – Abdulaziz Al-Shathery
}

# Manual Team Override (account_id → team) — 2025/2026 only
TEAM_OVERRIDE = {
    1400000008: 'AFS',
    1110000019: 'Projects',
    1500000046: 'AFS',
    1160000004: 'SDA',
    1180000123: 'Projects',
    1180000363: 'Projects',
    1500000024: 'AMC',
    1500000045: 'AFS',
    1500000047: 'AMC',
    1500000058: 'AFS',
    1800005448: 'SME',    # M. R. Abu-Nayyan Trad. & Invest. — DM: SME
    1180001803: 'Projects', # Delta Horizon General Contracting — DM: B2B
    1180004093: 'Projects', # Advanced Gulf Cooling EST — DM: B2B
    1800000780: 'Projects', # Hisham Nabil Ali Reza TRD COMP — DM: B2B
    1180004273: 'Showrooms', # Afaq Air Conditioning Co. — DM: Showroom
    1180000455: 'Showrooms', # Maintenance Triangle Est. — DM: Showroom
    1180004274: 'Showrooms', # FANCO HVAC — DM: Showroom
    1110000031: 'IR_Others', # Future Of Electrical Devi — DM: Home Appliances (was Showrooms)
    1110000010: 'IR',        # Tamkeen International Co. — OR_Others → IR
    1110000299: 'IR',        # Al-Zaqzouq (Zagzoog) — OR_Others → IR (msg_post_9611, 2026-05-02 형님 확정)
    1110000369: 'OR_Others', # Extra Value Est — OR → OR_Others (eXtra 본채널은 1120000000 United Electronics)
}

# Manual Status Override (account_id → status)
STATUS_OVERRIDE = {
    1110000101: 'Re-active_2025',  # Star Appliances — 자동분류 Active → Re-active_2025
    1110000009: 'Active',          # Bin Momen Trading Co. — 자동분류 Re-active_2025 → Active
}

# Category Mapping
CATEGORY_MAP = {
    'Split Inverter': 'Split Inverter', 'Split on/off': 'Split on/off',
    'Cassette': 'Cassette', 'Concealed': 'Concealed',
    'Convertible (CAC)': 'Convertible (CAC)', 'Free Standing': 'Free Standing',
    'Window': 'Window', 'Multi-V': 'Multi-V', 'AHU': 'AHU',
    'Unitary Package': 'Unitary Package', 'Accessories': 'Accessories',
    'installation': 'installation', 'Others': 'Others',
    'CAC Ducted': 'Concealed', 'Window On/Off': 'Window', 'Window SEEC': 'Window',
    'Miscellaneous': 'Miscellaneous',
}
EXCLUDE_CATS = {'Cooker', 'Dishwasher', 'Laundry', 'SDA', 'RAC', 'FHD'}
# installation: SAR 요금이 Qty 컬럼에 잘못 입력되는 데이터 오염 패턴 (qty==val) — Miscellaneous와 동일하게 0처리
ZERO_QTY_CATS = {'Miscellaneous', 'installation'}

# 서비스 요금 자재: Qty 컬럼에 SAR 금액이 입력되는 오염 패턴 → qty 강제 0
ZERO_QTY_MATERIALS = {'ENERGY_SERVICE', 'INSTALLATION CHRGS'}

# Quantity ÷ 2 categories (indoor+outdoor set)
HALF_QTY_CATS = {'Cassette', 'Concealed', 'Convertible (CAC)', 'Free Standing', 'Split Inverter', 'Split on/off'}

# Material-level category override (align with B2C model mapping)
MATERIAL_CAT_OVERRIDE = {
    'AUUQ34GT6': 'Cassette',
    'AUUQ40GT6': 'Concealed',
    'ATNQ21GPLTA': 'Cassette',
    'ATNW28GPLTA': 'Cassette',
    'ATNW40GYLTA': 'Cassette',
}


# Account ID Alias: 동일 거래선의 다중 SAP ID를 하나로 통합 (msg_post_9611, 2026-05-02)
ACCOUNT_ALIAS = {
    1110000129: 1110000009,   # Bin Momen 서브 1 → 메인
    1110000130: 1110000009,   # Bin Momen 서브 2 → 메인
    1800002288: 1110000009,   # Bin Momen 서브 3 → 메인
}

def normalize_id(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        v = int(val)
        if v == 0: return None
        return ACCOUNT_ALIAS.get(v, v)
    s = str(val).strip()
    if s in ('', 'None', '#N/A'): return None
    try:
        v = int(s)
        if v == 0: return None
        return ACCOUNT_ALIAS.get(v, v)
    except ValueError:
        return s


def map_category(raw_cat):
    if raw_cat in EXCLUDE_CATS: return None
    m = CATEGORY_MAP.get(raw_cat)
    if m: return m
    return 'Others' if raw_cat and raw_cat != '---' else 'Others'


def normalize_emp(val):
    """Normalize employee number to int"""
    if val is None: return None
    try: return int(val)
    except: return None


# ============================================================
# STEP 1: Load Classification
# ============================================================
def load_classification():
    print("1. Loading Classification...")
    wb = openpyxl.load_workbook(FILES['classification'], read_only=True, data_only=True)

    c24 = {}
    for row in wb['2024'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        sap = normalize_id(r[0])
        if sap: c24[sap] = {'name': r[1], 'team': r[5]}

    c25 = {}
    for row in wb['2025'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        cid = normalize_id(r[1])
        if cid: c25[cid] = {'name': r[2], 'team': r[3]}

    wb.close()
    print(f"   2024: {len(c24)} / 2025: {len(c25)} accounts")
    return c24, c25


# ============================================================
# STEP 2: Load Raw Data
# ============================================================
def load_2024(c24, c25=None):
    """Load 2024 monthly data and distribute evenly across days in each month."""
    print("2. Loading 2024 Raw...")
    wb = openpyxl.load_workbook(FILES['2024_raw'], read_only=True, data_only=True)
    ws = wb['2024_Raw']
    monthly_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        cat = map_category(r[3])
        if cat is None: continue
        sap = normalize_id(r[4])
        # Team: TEAM_OVERRIDE → Raw col 22 → c25 → c24 fallback
        if sap in TEAM_OVERRIDE:
            team = TEAM_OVERRIDE[sap]
        else:
            team = r[21] if len(r) > 21 and r[21] else None
            if not team and c25: team = c25.get(sap, {}).get('team')
            if not team: team = c24.get(sap, {}).get('team')
        qty = int(r[7]) if r[7] else 0
        if cat in HALF_QTY_CATS: qty = qty // 2
        if cat in ZERO_QTY_CATS: qty = 0
        monthly_rows.append({
            'Year': 2024, 'Month': str(r[1]).zfill(2) if r[1] else '00',
            'Account_ID': sap, 'Account_Name': r[5],
            'Team': team, 'Category': cat,
            'Value': float(r[6] or 0), 'Quantity': qty,
        })
    wb.close()

    # Distribute monthly totals evenly across days
    rows = []
    for mr in monthly_rows:
        mm = int(mr['Month']) if mr['Month'] != '00' else 1
        days_in_month = calendar.monthrange(2024, mm)[1]
        daily_val = mr['Value'] / days_in_month
        daily_qty = mr['Quantity'] / days_in_month
        for d in range(1, days_in_month + 1):
            rows.append({
                'Year': 2024,
                'Month': mr['Month'],
                'Day': f"2024-{mr['Month']}-{str(d).zfill(2)}",
                'Account_ID': mr['Account_ID'], 'Account_Name': mr['Account_Name'],
                'Team': mr['Team'], 'Category': mr['Category'],
                'Value': round(daily_val, 2), 'Quantity': round(daily_qty, 2),
            })
    print(f"   {len(monthly_rows)} monthly rows → {len(rows)} daily rows")
    return rows


def load_raw_2025_2026(path, sheet, c24, c25, year_label, dealer_map=None):
    print(f"3. Loading {year_label} Raw...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows, skipped = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        if len(r) < 101: continue
        cat = map_category(r[77])
        if cat is None:
            skipped += 1
            continue

        material = str(r[70]).strip() if r[70] else ''
        mat_clean = re.sub(r'\.[A-Z0-9]+$', '', material.rstrip('.'))
        if mat_clean in MATERIAL_CAT_OVERRIDE:
            cat = MATERIAL_CAT_OVERRIDE[mat_clean]

        inv_date = r[1]
        month = str(inv_date.month).zfill(2) if isinstance(inv_date, datetime) else '00'
        cid = normalize_id(r[38])
        emp_txn = normalize_emp(r[27])  # Sales Employee Number (Transaction)
        classification = r[98]

        # Team logic: Manual override → SME override → Classification file → raw → Dealer Mapping
        if cid in TEAM_OVERRIDE:
            team = TEAM_OVERRIDE[cid]
        elif emp_txn in SME_EMPLOYEES:
            team = 'SME'
        else:
            team = c25.get(cid, {}).get('team')
            if not team: team = c24.get(cid, {}).get('team')
            if not team: team = classification
            if not team and dealer_map: team = dealer_map.get(cid)

        # 2026 grouping: re-classify OR/IR based on Sell Out managed channels
        # SME는 직원 번호 기반 분류이므로 IR 채널 재분류 대상에서 제외
        if int(year_label) == 2026:
            if team in ('OR', 'OR_Others'):
                team = 'OR' if cid in OR_5_CHANNEL_IDS else 'OR_Others'
            elif team in ('IR', 'IR_Others'):
                team = 'IR' if cid in IR_8_CHANNEL_IDS else 'IR_Others'

        qty = int(r[5]) if isinstance(r[5], (int, float)) else 0
        if cat in HALF_QTY_CATS: qty = qty // 2
        if cat in ZERO_QTY_CATS: qty = 0
        if mat_clean in ZERO_QTY_MATERIALS: qty = 0

        day_str = inv_date.strftime('%Y-%m-%d') if isinstance(inv_date, datetime) else f'{year_label}-{month}-01'
        rows.append({
            'Year': int(year_label), 'Month': month, 'Day': day_str,
            'Account_ID': cid, 'Account_Name': r[39],
            'Team': team, 'Category': cat,
            'Value': float(r[6]) if r[6] else 0, 'Quantity': qty,
        })
    wb.close()
    print(f"   {len(rows)} rows loaded, {skipped} excluded")
    return rows


# ============================================================
# STEP 3: Account Status
# ============================================================
def classify_accounts(df, c24):
    print("4. Classifying Account Status...")
    yearly = df.groupby(['Year', 'Account_ID']).agg(
        Total_Value=('Value', 'sum')).reset_index()

    a24_txn = set(yearly[yearly['Year'] == 2024]['Account_ID'].unique())
    a25_txn = set(yearly[yearly['Year'] == 2025]['Account_ID'].unique())
    a26_txn = set(yearly[yearly['Year'] == 2026]['Account_ID'].unique())
    a24_all = a24_txn | set(c24.keys())

    t24 = yearly[yearly['Year'] == 2024].set_index('Account_ID')['Total_Value'].to_dict()
    t25 = yearly[yearly['Year'] == 2025].set_index('Account_ID')['Total_Value'].to_dict()
    t26 = yearly[yearly['Year'] == 2026].set_index('Account_ID')['Total_Value'].to_dict()

    all_accts = a24_all | a25_txn | a26_txn
    status = {}

    for a in all_accts:
        i24, i25, i26 = a in a24_all, a in a25_txn, a in a26_txn
        v24, v25 = t24.get(a, 0), t25.get(a, 0)
        growth = (v25 - v24) / v24 if v24 > 0 else None

        if i26:
            if not i24 and not i25: s = 'New_2026'
            elif i24 and not i25: s = 'Re-active from 2024'
            elif i25 and not i24: s = 'New_2025'
            elif i25 and i24 and growth is not None and growth >= 0.5: s = 'Re-active_2025'
            else: s = 'Active'
        elif i25:
            if not i24: s = 'New_2025'
            elif growth is not None and growth >= 0.5: s = 'Re-active_2025'
            else: s = 'Active'
        elif i24:
            s = 'Need to re-active'
        else:
            s = 'Unknown'
        status[a] = s

    for acct_id, forced_status in STATUS_OVERRIDE.items():
        if acct_id in status:
            status[acct_id] = forced_status

    dist = Counter(status.values())
    print("   " + ", ".join(f"{k}:{v}" for k, v in sorted(dist.items())))
    return status, t24, t25, t26, all_accts


# ============================================================
# STEP 4: Save Excel
# ============================================================
def save_excel(df, status, t24, t25, t26, all_accts, c24, c25):
    print("5. Saving Excel...")
    df['Account_Status'] = df['Account_ID'].map(status)
    df_final = df[['Year', 'Month', 'Day', 'Account_ID', 'Account_Name', 'Team',
                    'Account_Status', 'Category', 'Value', 'Quantity']].copy()
    df_final = df_final.sort_values(['Year', 'Day', 'Account_ID', 'Category'])

    # Account names & teams (latest first)
    names, teams = {}, {}
    for _, r in df_final.sort_values('Year', ascending=False).iterrows():
        a = r['Account_ID']
        if a not in names and r['Account_Name']: names[a] = r['Account_Name']
        if a not in teams and r['Team']: teams[a] = r['Team']
    for a, i in c24.items():
        names.setdefault(a, i.get('name', ''))
        teams.setdefault(a, i.get('team', ''))
    for a, i in c25.items():
        names.setdefault(a, i.get('name', ''))
        teams.setdefault(a, i.get('team', ''))

    # Manual TEAM_OVERRIDE는 master에서도 강제 우선 적용 (transaction fallback 방지)
    for a, t in TEAM_OVERRIDE.items():
        teams[a] = t

    master = []
    for a in sorted(all_accts, key=lambda x: str(x)):
        if not a: continue
        v24, v25, v26 = t24.get(a, 0), t25.get(a, 0), t26.get(a, 0)
        master.append({
            'Account_ID': a, 'Account_Name': names.get(a, ''),
            'Team': teams.get(a, ''), 'Account_Status': status.get(a, ''),
            'Value_2024': v24, 'Value_2025': v25, 'Value_2026': v26,
            'Growth_24_25': round((v25 - v24) / v24 * 100, 1) if v24 > 0 else None,
            'Growth_25_26': round((v26 - v25) / v25 * 100, 1) if v25 > 0 else None,
        })

    dm = pd.DataFrame(master)
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as w:
        df_final.to_excel(w, sheet_name='Transactions', index=False)
        dm.to_excel(w, sheet_name='Account_Master', index=False)

    print(f"   {OUTPUT_EXCEL}")
    print(f"   Transactions: {len(df_final)} / Master: {len(dm)}")
    return df_final, dm


# ============================================================
# STEP 4b: Load IR Excel Mar Sell-Thru Target
# ============================================================
def load_ir_target():
    """Read IR Excel and return per-account Mar Sell-Thru Target.
    Returns dict: {str(account_id): {'mar_tq': int, 'mar_tv': float}}
    Uses sheet 1 (All IR Stock and SO 31 Feb 2026):
      Col 7 (idx 6):  Dealer ACC No  -> Account ID (join key)
      Col 32 (idx 31): Sell thru Target_Mar 2026 Qty
      Col 33 (idx 32): Sell thru Target_Mar 2026 Value
    """
    if not os.path.exists(IR_TARGET_FILE):
        print(f"   WARNING: IR Target file not found: {IR_TARGET_FILE}")
        return {}
    print(f"   Loading IR Mar Target: {os.path.basename(IR_TARGET_FILE)}")
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(IR_TARGET_FILE, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        result = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue  # skip empty row 1 + header row 2
            if not row[0]:
                continue  # skip empty rows
            acc_no = str(row[6]).strip() if row[6] else ''
            if not acc_no or acc_no == 'None':
                continue
            try:
                tq = int(row[31]) if row[31] else 0
            except (ValueError, TypeError):
                tq = 0
            try:
                tv = float(row[32]) if row[32] else 0.0
            except (ValueError, TypeError):
                tv = 0.0
            nm = str(row[5]).strip() if row[5] else ''
            if acc_no in result:
                result[acc_no]['mar_tq'] += tq
                result[acc_no]['mar_tv'] += tv
                if nm and not result[acc_no].get('nm'):
                    result[acc_no]['nm'] = nm
            else:
                result[acc_no] = {'mar_tq': tq, 'mar_tv': round(tv, 0), 'nm': nm}
        wb.close()
        # Round values
        for v in result.values():
            v['mar_tv'] = round(v['mar_tv'], 0)
        print(f"   IR Target loaded: {len(result)} accounts")
        return result
    except Exception as e:
        print(f"   WARNING: Failed to load IR target: {e}")
        return {}


# ============================================================
# STEP 4c: Load RSM FCST (Monthly Excel → rsm_fcst_data.json)
# ============================================================
# Account ID corrections: RSM FCST file에 잘못 입력된 계정 ID → 올바른 ID 보정
_RSM_ID_CORRECTIONS = {
    '111000000':  '1110000000',  # Saudi Bin Hamood (0 하나 누락)
    '111000079':  '1110000079',  # SAMA AL KHER (0 하나 누락)
    '11140000400': '1110000400', # OCEAN AIR CONDITIONING (자릿수 오류)
}
# 유효하지 않은 계정 ID (스킵)
_RSM_SKIP_IDS = {'New', '(blank)', 'None', ''}

# OR 채널명 → 계정 ID 매핑 (00. OR RSM FCST 파일의 채널명 기준)
_OR_CHANNEL_ACCOUNTS = {
    'eXtra':          '1120000000',  # United Electronics Company
    'Al Manea':       '1110000002',  # Hamad A. Al-Manea Trade Co.
    'SWS':            '1110000005',  # ALSHETA & ALSAIF FOR TRADING CO.
    'B.BOX':          '1110000006',  # Black Box Co.
    'AlKhunaizan':    '1110000007',  # Alkhunaizan Trading Co.
    'Lulu & Other':   '1120000002',  # Lulu Saudi Hypermarkets L.L.C.
}
# OR 채널명 → 계정 이름
_OR_CHANNEL_NAMES = {
    'eXtra':          'United Electronics Company (eXtra)',
    'Al Manea':       'Hamad A. Al-Manea Trade Co.',
    'SWS':            'ALSHETA & ALSAIF FOR TRADING CO. (SWS)',
    'B.BOX':          'Black Box Co.',
    'AlKhunaizan':    'Alkhunaizan Trading Co.',
    'Lulu & Other':   'Lulu Saudi Hypermarkets L.L.C.',
}


def _load_or_rsm_fcst_data():
    """00. OR 폴더의 OR 채널 RSM FCST 로딩 (Apr. 시트 기준).

    OR RSM FCST 파일 구조 (Apr. 시트):
      Row 0: Category, Class, PrevModel, [None], [None], [eXtra ratio], ...
      Row 1: None, None, None, 'New Model', 'APR', 'eXtra', 'Al Manea', 'SWS', 'B.BOX', 'AlKhunaizan', 'Lulu & Other', [Value cols...]
      Row 2+: 모델별 데이터
        Col 3: New Model (mapping code)
        Col 4: APR Total Qty
        Col 5-10: 채널별 Qty (eXtra, Al Manea, SWS, B.BOX, AlKhunaizan, Lulu & Other)
        Col 11-16: 채널별 Value (eXtra, Al Manea, SWS, B.BOX, AlKhunaizan, Lulu & Other)

    Returns: {채널명: {'qty': int, 'val': float}} or {}
    """
    or_dir = os.path.join(RAW_DATA_BASE, '02. 2026', '06. RSM FCST', '00. OR')
    if not os.path.isdir(or_dir):
        print(f"   INFO: OR RSM FCST dir not found: {or_dir}")
        return {}

    files = [f for f in os.listdir(or_dir)
             if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files:
        print(f"   INFO: No OR RSM FCST xlsx in {or_dir}")
        return {}

    # 최신 파일 선택
    files_with_mtime = [(f, os.path.getmtime(os.path.join(or_dir, f))) for f in files]
    files_with_mtime.sort(key=lambda x: x[1], reverse=True)
    or_file = os.path.join(or_dir, files_with_mtime[0][0])
    print(f"   Loading OR RSM FCST: {os.path.basename(or_file)}")

    try:
        wb = openpyxl.load_workbook(or_file, read_only=True, data_only=True)

        # Apr. 시트 우선 (없으면 마지막 시트)
        target_ws = None
        for sname in wb.sheetnames:
            if 'apr' in sname.lower():
                target_ws = wb[sname]
                break
        if target_ws is None:
            target_ws = wb.worksheets[-1]
        print(f"   OR Sheet: {target_ws.title}")

        rows = list(target_ws.iter_rows(values_only=True))
        if len(rows) < 3:
            wb.close()
            return {}

        # OR 채널 순서 (Row 1 Col 5~10)
        channels = ['eXtra', 'Al Manea', 'SWS', 'B.BOX', 'AlKhunaizan', 'Lulu & Other']
        qty_cols = [5, 6, 7, 8, 9, 10]
        val_cols = [11, 12, 13, 14, 15, 16]

        ch_qty = {c: 0.0 for c in channels}
        ch_val = {c: 0.0 for c in channels}

        for row in rows[2:]:  # 헤더 2행 스킵
            if not row or not row[0]:
                continue
            cat = str(row[0]).strip() if row[0] else ''
            if not cat or cat == 'None':
                continue
            model = str(row[3]).strip() if row[3] else ''
            if not model or 'TTL' in model or model.lower().startswith('sub') or model.lower() == 'total':
                continue  # 소계/합계 행 스킵 (model 없는 합계/소계 행 포함)

            for i, ch in enumerate(channels):
                try:
                    q = float(row[qty_cols[i]]) if row[qty_cols[i]] is not None else 0.0
                    ch_qty[ch] += q
                except (ValueError, TypeError):
                    pass
                try:
                    v = float(row[val_cols[i]]) if row[val_cols[i]] is not None else 0.0
                    ch_val[ch] += v
                except (ValueError, TypeError):
                    pass

        wb.close()

        result = {}
        for ch in channels:
            if ch_qty[ch] > 0 or ch_val[ch] > 0:
                result[ch] = {'qty': int(round(ch_qty[ch])), 'val': round(ch_val[ch], 0)}

        total_q = sum(v['qty'] for v in result.values())
        total_v = sum(v['val'] for v in result.values())
        print(f"   OR RSM FCST: {len(result)} channels / Qty={total_q:,.0f} / Val={total_v:,.0f} SAR")
        return result

    except Exception as e:
        print(f"   WARNING: Failed to load OR RSM FCST: {e}")
        import traceback
        traceback.print_exc()
        return {}


def _find_latest_rsm_fcst_file(rsm_dir):
    """06. RSM FCST/01. IR/ 폴더에서 최신 xlsx 파일 반환."""
    if not os.path.isdir(rsm_dir):
        return None
    files = [f for f in os.listdir(rsm_dir)
             if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files:
        return None
    # 파일명에서 날짜(DDMMYYYY) 또는 수정일 기준으로 최신 선택
    files_with_mtime = [(f, os.path.getmtime(os.path.join(rsm_dir, f))) for f in files]
    files_with_mtime.sort(key=lambda x: x[1], reverse=True)
    return os.path.join(rsm_dir, files_with_mtime[0][0])


def load_rsm_fcst():
    """RSM FCST 월별 Excel → 계정별/모델별 집계 후 rsm_fcst_data.json 저장.

    출력 구조:
    {
      "value":  {str(acc_no): float, ...},   # 계정별 FCST Value (SAR)
      "qty":    {str(acc_no): int,   ...},   # 계정별 FCST Qty
      "models": {mapping_code: {"qty":int, "val":float, "category":str, "class":str}, ...},
      "_month": "April 2026",
      "_updated": "ISO datetime"
    }

    채널 맵핑:
      - 계정 ID → master data의 team (IR / IR_Others) 으로 자동 분류
      - ID 오류 계정은 _RSM_ID_CORRECTIONS 테이블로 보정
      - 'New', '(blank)' 등 무효 계정은 스킵

    모델 맵핑:
      - Mapping 컬럼(6자리 모델코드)으로 집계
      - Category / Class 정보 포함 (Inverter, ON/OFF, CAC 등)
    """
    rsm_dir = os.path.join(RAW_DATA_BASE, '02. 2026', '06. RSM FCST', '01. IR')
    rsm_file = _find_latest_rsm_fcst_file(rsm_dir)
    if not rsm_file:
        print(f"   WARNING: RSM FCST file not found in {rsm_dir}")
        return None

    print(f"   Loading RSM FCST: {os.path.basename(rsm_file)}")
    try:
        wb = openpyxl.load_workbook(rsm_file, read_only=True, data_only=True)

        # 데이터 시트 선택: 'All IR' 포함 시트 우선, 없으면 마지막 시트
        target_ws = None
        for sname in wb.sheetnames:
            if 'All IR' in sname or 'all ir' in sname.lower():
                target_ws = wb[sname]
                break
        if target_ws is None:
            target_ws = wb.worksheets[-1]

        print(f"   Sheet: {target_ws.title}")

        rows = list(target_ws.iter_rows(values_only=True))
        # 헤더 행 탐색 (Region 컬럼이 있는 행)
        header_idx = None
        for i, row in enumerate(rows):
            if row and str(row[0]).strip().lower() == 'region':
                header_idx = i
                break
        if header_idx is None:
            print("   WARNING: RSM FCST header row not found")
            wb.close()
            return None

        # 월 이름 추출 (시트명에서, 예: "All IR April FCST 13042026" → "April 2026")
        month_label = ''
        import re as _re
        title_str = target_ws.title
        # 월 이름 추출
        m_month = _re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)', title_str, _re.IGNORECASE)
        # 연도 추출: DDMMYYYY(8자리 날짜) 형태에서 마지막 4자리 우선, 없으면 독립된 4자리 연도
        m_year_dmy = _re.search(r'\d{2,4}(20\d{2})\b', title_str)  # "13042026" → 2026
        m_year_standalone = _re.search(r'\b(20\d{2})\b', title_str)  # 독립된 연도
        if m_month:
            month_name = m_month.group(1).capitalize()
            if m_year_dmy:
                month_label = f"{month_name} {m_year_dmy.group(1)}"
            elif m_year_standalone:
                month_label = f"{month_name} {m_year_standalone.group(1)}"
            else:
                # 파일명에서 연도 추출
                m_file_dmy = _re.search(r'\d{2,4}(20\d{2})\b', os.path.basename(rsm_file))
                m_file_yr = _re.search(r'\b(20\d{2})\b', os.path.basename(rsm_file))
                year_str = (m_file_dmy.group(1) if m_file_dmy else
                            m_file_yr.group(1) if m_file_yr else '')
                month_label = f"{month_name} {year_str}".strip()
        else:
            # 파일명에서 연도만 추출
            m2 = _re.search(r'(20\d{2})', os.path.basename(rsm_file))
            month_label = m2.group(1) if m2 else ''

        # 컬럼 인덱스 (헤더에서 동적으로 탐색)
        hdr = [str(c).strip().lower() if c else '' for c in rows[header_idx]]

        def _col(candidates):
            for c in candidates:
                for i, h in enumerate(hdr):
                    if c.lower() in h:
                        return i
            return None

        col_region    = _col(['region'])                            # 0
        col_dealer_nm = _col(['dealer name'])                      # 5
        col_acc_no    = _col(['dealer acc no', 'acc no'])          # 6
        col_mapping   = _col(['mapping'])                          # 8
        col_category  = _col(['category'])                         # 9
        col_class     = _col(['class'])                            # 10
        col_qty       = _col(['sell-thru plan', 'sell thru plan']) # 20
        col_val       = _col(['sell -thru value', 'sell-thru value', 'sell thru value'])  # 21

        # fallback to positional if header detection fails
        if col_acc_no is None:    col_acc_no    = 6
        if col_dealer_nm is None: col_dealer_nm = 5
        if col_mapping is None:   col_mapping   = 8
        if col_category is None:  col_category  = 9
        if col_class is None:     col_class     = 10
        if col_qty is None:       col_qty       = 20
        if col_val is None:       col_val       = 21

        acc_val  = {}   # {str(acc_no): float value total}
        acc_qty  = {}   # {str(acc_no): int qty total}
        acc_name = {}   # {str(acc_no): dealer name}
        model_data = {} # {mapping_code: {qty, val, category, class}}
        skip_count = 0

        for row in rows[header_idx + 1:]:
            if not row or not row[0]:
                continue  # 빈 행 스킵

            raw_acc = str(row[col_acc_no]).strip() if row[col_acc_no] else ''
            if raw_acc in _RSM_SKIP_IDS or raw_acc.lower() in {'new', 'none', ''}:
                skip_count += 1
                continue

            # ID 보정
            acc_no = _RSM_ID_CORRECTIONS.get(raw_acc, raw_acc)

            mapping  = str(row[col_mapping]).strip()  if row[col_mapping]  else 'Unknown'
            category = str(row[col_category]).strip() if row[col_category] else ''
            class_   = str(row[col_class]).strip()    if row[col_class]    else ''
            dealer_nm = str(row[col_dealer_nm]).strip() if row[col_dealer_nm] else ''

            try:
                qty = float(row[col_qty]) if row[col_qty] and str(row[col_qty]) not in ['None', ''] else 0.0
            except (ValueError, TypeError):
                qty = 0.0
            try:
                val = float(row[col_val]) if row[col_val] and str(row[col_val]) not in ['None', ''] else 0.0
            except (ValueError, TypeError):
                val = 0.0

            # 계정별 집계
            if acc_no not in acc_val:
                acc_val[acc_no]  = 0.0
                acc_qty[acc_no]  = 0.0
                acc_name[acc_no] = dealer_nm
            acc_val[acc_no] += val
            acc_qty[acc_no] += qty

            # 모델별 집계
            if mapping not in model_data:
                model_data[mapping] = {'qty': 0.0, 'val': 0.0, 'category': category, 'class': class_}
            model_data[mapping]['qty'] += qty
            model_data[mapping]['val'] += val

        wb.close()

        # 정수/반올림 처리
        out_val   = {k: round(v, 0) for k, v in acc_val.items() if v > 0}
        out_qty   = {k: int(round(v))  for k, v in acc_qty.items() if v > 0}
        out_models = {k: {'qty': int(round(v['qty'])), 'val': round(v['val'], 0),
                          'category': v['category'], 'class': v['class']}
                      for k, v in model_data.items() if v['qty'] > 0 or v['val'] > 0}

        # 계정 이름 정보 (마스터에 없는 계정 추가 시 활용)
        out_names = {k: acc_name.get(k, '') for k in out_qty.keys()}

        result = {
            'value':   out_val,
            'qty':     out_qty,
            'models':  out_models,
            'names':   out_names,   # 계정 이름 (마스터 보완용)
            '_month':  month_label,
            '_updated': datetime.utcnow().isoformat() + 'Z',
            '_clear_local': True,   # 브라우저 localStorage 구버전 캐시 무효화
        }

        total_acc = len(out_val)
        total_qty = sum(out_qty.values())
        total_val = sum(out_val.values())
        print(f"   RSM FCST loaded: {total_acc} accounts / Qty={total_qty:,.0f} / Val={total_val:,.0f} SAR"
              f" / Models={len(out_models)} / Skipped={skip_count} rows [{month_label}]")

        # rsm_fcst_data.json은 OR 병합 후 main()에서 저장
        return result

    except Exception as e:
        print(f"   WARNING: Failed to load RSM FCST: {e}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================
# STEP 4d: Load OR RSM FCST (00. OR folder → retailer-based Excel)
# ============================================================
# OR 채널 리테일러 → 계정 ID 매핑
_OR_RETAILER_ACC = {
    'extra':         '1120000000',   # United Electronics Company (eXtra)
    'al manea':      '1110000002',   # Hamad A. Al-Manea Trade Co. (Active)
    'sws':           '1110000005',   # Alsheta & Alsaif For Trading Co. (Active)
    'b.box':         '1110000006',   # Black Box Co.
    'alkhunaizan':   '1110000007',   # Alkhunaizan Trading Co.
    'lulu & other':  '1120000002',   # Lulu Saudi Hypermarkets
}
_OR_RETAILER_NAMES = {
    '1120000000': 'United Electronics Company (eXtra)',
    '1110000001': 'Hamad A. Al-Manea Trade Co.',
    '1110000004': 'Alsheta & Alsaif For Trading Co.',
    '1110000006': 'Black Box Co.',
    '1110000007': 'Alkhunaizan Trading Co.',
    '1120000002': 'Lulu Saudi Hypermarkets',
}

# 현재 월 이름 → 시트명 약칭 매핑 (대소문자 무관 탐색)
_MONTH_SHEET_ALIASES = {
    'january': ['jan', 'jan.', 'january'],
    'february': ['feb', 'feb.', 'february'],
    'march': ['mar', 'mar.', 'march'],
    'april': ['apr', 'apr.', 'april'],
    'may': ['may'],
    'june': ['jun', 'jun.', 'june'],
    'july': ['jul', 'jul.', 'july'],
    'august': ['aug', 'aug.', 'august'],
    'september': ['sep', 'sep.', 'sept', 'september'],
    'october': ['oct', 'oct.', 'october'],
    'november': ['nov', 'nov.', 'november'],
    'december': ['dec', 'dec.', 'december'],
}


def load_or_rsm_fcst():
    """OR RSM FCST 월별 Excel (00. OR 폴더) → 리테일러별 집계 후 계정-레벨 반환.

    OR 파일 구조 (모델 × 리테일러):
      Col 0: Category, Col 1: Class, Col 2: Prev Model, Col 3: New Model
      Col 4: 월 합계 Qty, Col 5-10: 리테일러별 Qty, Col 11-16: 리테일러별 Value
      Header row 0에 리테일러 공유 비율, row 1에 리테일러 이름
    """
    import re as _re
    or_dir = os.path.join(RAW_DATA_BASE, '02. 2026', '06. RSM FCST', '00. OR')
    if not os.path.isdir(or_dir):
        print(f"   WARNING: OR RSM FCST folder not found: {or_dir}")
        return None

    files = [f for f in os.listdir(or_dir)
             if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files:
        print(f"   WARNING: No OR RSM FCST xlsx in {or_dir}")
        return None

    files.sort(key=lambda f: os.path.getmtime(os.path.join(or_dir, f)), reverse=True)
    or_file = os.path.join(or_dir, files[0])
    print(f"   Loading OR RSM FCST: {os.path.basename(or_file)}")

    try:
        wb = openpyxl.load_workbook(or_file, read_only=True, data_only=True)

        # 현재 월에 맞는 시트 탐색 (예: April → 'Apr.', 'April' 등)
        cur_month = datetime.utcnow().strftime('%B').lower()   # 'april'
        aliases = _MONTH_SHEET_ALIASES.get(cur_month, [cur_month])
        target_ws = None
        for sname in wb.sheetnames:
            if sname.lower().rstrip('.') in [a.rstrip('.') for a in aliases]:
                target_ws = wb[sname]
                break
        if target_ws is None:
            # fallback: 마지막 시트
            target_ws = wb.worksheets[-1]
            print(f"   OR RSM FCST: sheet for '{cur_month}' not found, using '{target_ws.title}'")
        else:
            print(f"   OR RSM FCST Sheet: {target_ws.title}")

        rows = list(target_ws.iter_rows(values_only=True))
        if len(rows) < 3:
            print("   WARNING: OR RSM FCST sheet has too few rows")
            wb.close()
            return None

        # 헤더: row 0 = 공유 비율, row 1 = 리테일러 이름
        # 리테일러 순서: col 5-10 (qty), col 11-16 (value)
        hdr1 = [str(c).strip() if c else '' for c in rows[1]]
        retailers = []
        for col_idx in range(5, 11):
            name = hdr1[col_idx] if col_idx < len(hdr1) else ''
            acc = _OR_RETAILER_ACC.get(name.lower().strip(), None)
            retailers.append({'name': name, 'acc': acc, 'qty_col': col_idx, 'val_col': col_idx + 6})

        acc_val  = {}
        acc_qty  = {}
        model_data = {}

        skip_keywords = {'ttl', 'total', 'subtotal', 'grand total', ''}
        for row in rows[2:]:
            if not row or not row[0]:
                continue
            new_model = str(row[3]).strip() if row[3] else ''
            if new_model.lower().replace(' ', '') in skip_keywords:
                continue
            if new_model.upper().endswith('TTL') or new_model.upper() == 'TOTAL':
                continue

            category = str(row[0]).strip() if row[0] else ''
            class_   = str(row[1]).strip() if row[1] else ''
            mapping  = new_model[:6] if len(new_model) >= 6 else new_model

            for r in retailers:
                if r['acc'] is None:
                    continue
                acc = r['acc']
                try:
                    qty = float(row[r['qty_col']]) if r['qty_col'] < len(row) and row[r['qty_col']] and str(row[r['qty_col']]) not in ['None', ''] else 0.0
                except (ValueError, TypeError):
                    qty = 0.0
                try:
                    val = float(row[r['val_col']]) if r['val_col'] < len(row) and row[r['val_col']] and str(row[r['val_col']]) not in ['None', ''] else 0.0
                except (ValueError, TypeError):
                    val = 0.0

                if acc not in acc_val:
                    acc_val[acc] = 0.0
                    acc_qty[acc] = 0.0
                acc_val[acc] += val
                acc_qty[acc] += qty

                if qty > 0 or val > 0:
                    if mapping not in model_data:
                        model_data[mapping] = {'qty': 0.0, 'val': 0.0, 'category': category, 'class': class_}
                    model_data[mapping]['qty'] += qty
                    model_data[mapping]['val'] += val

        wb.close()

        # 월 레이블 추출
        month_label = ''
        m_month = _re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)',
                             target_ws.title, _re.IGNORECASE)
        if m_month:
            month_label = m_month.group(1).capitalize()
            m_yr = _re.search(r'(20\d{2})', os.path.basename(or_file))
            if m_yr:
                month_label += f' {m_yr.group(1)}'
        else:
            month_label = datetime.utcnow().strftime('%B %Y')

        out_val   = {k: round(v, 0) for k, v in acc_val.items() if v > 0}
        out_qty   = {k: int(round(v))  for k, v in acc_qty.items() if v > 0}
        out_models = {k: {'qty': int(round(v['qty'])), 'val': round(v['val'], 0),
                          'category': v['category'], 'class': v['class']}
                      for k, v in model_data.items() if v['qty'] > 0 or v['val'] > 0}

        total_acc = len(out_val)
        total_qty = sum(out_qty.values())
        total_val = sum(out_val.values())
        print(f"   OR RSM FCST loaded: {total_acc} retailers / Qty={total_qty:,.0f} / Val={total_val:,.0f} SAR"
              f" / Models={len(out_models)} [{month_label}]")

        return {
            'value':   out_val,
            'qty':     out_qty,
            'models':  out_models,
            'names':   {k: _OR_RETAILER_NAMES.get(k, '') for k in out_qty.keys()},
            '_month':  month_label,
        }

    except Exception as e:
        print(f"   WARNING: Failed to load OR RSM FCST: {e}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================
# STEP 5: Save HTML Dashboard
# ============================================================
def save_html(df_final, dm, oud_data=None, ar_data=None, col_data=None, pgi_data=None, remain_data=None, open_data=None, sell_thru_date=None, ir_target=None, rsm_fcst=None):
    print("6. Saving HTML Dashboard...")
    txn = []
    for _, r in df_final.iterrows():
        txn.append([
            int(r['Year']), str(r['Month']).zfill(2),
            str(r['Day']) if pd.notna(r['Day']) else '',
            r['Account_ID'] if pd.notna(r['Account_ID']) else None,
            str(r['Account_Name']) if pd.notna(r['Account_Name']) else '',
            str(r['Team']) if pd.notna(r['Team']) else '',
            str(r['Account_Status']) if pd.notna(r['Account_Status']) else '',
            str(r['Category']) if pd.notna(r['Category']) else '',
            round(float(r['Value']), 2) if pd.notna(r['Value']) else 0,
            round(float(r['Quantity']), 2) if pd.notna(r['Quantity']) else 0
        ])
    master = []
    for _, r in dm.iterrows():
        master.append({
            'id': r['Account_ID'] if pd.notna(r['Account_ID']) else None,
            'name': str(r['Account_Name']) if pd.notna(r['Account_Name']) else '',
            'team': str(r['Team']) if pd.notna(r['Team']) else '',
            'status': str(r['Account_Status']) if pd.notna(r['Account_Status']) else '',
            'v24': round(float(r['Value_2024']), 0) if pd.notna(r['Value_2024']) else 0,
            'v25': round(float(r['Value_2025']), 0) if pd.notna(r['Value_2025']) else 0,
            'v26': round(float(r['Value_2026']), 0) if pd.notna(r['Value_2026']) else 0,
        })

    # Build OUD JSON
    oud_json = {'current': None, 'prev': None}
    if oud_data:
        for key in ['current', 'prev']:
            if key in oud_data and oud_data[key]:
                d = oud_data[key]
                accts = {}
                for aid, info in d['accounts'].items():
                    acct_data = {
                        'nm': info['name'],
                        'v': round(info['value'], 0),
                        'q': round(info['qty'], 0),
                    }
                    # Per-category breakdown
                    if 'cats' in info and info['cats']:
                        cats = {}
                        for cat, cv in info['cats'].items():
                            cats[cat] = {'v': round(cv['value'], 0), 'q': round(cv['qty'], 0)}
                        acct_data['c'] = cats
                    accts[str(aid)] = acct_data
                oud_json[key] = {
                    'date': d['date'],
                    'tv': round(d['total_value'], 0),
                    'tq': round(d['total_qty'], 0),
                    'accts': accts,
                }

    # Build AR JSON
    ar_json = {'current': None, 'prev': None}
    if ar_data:
        for key in ['current', 'prev']:
            if key in ar_data and ar_data[key]:
                d = ar_data[key]
                accts = {}
                for aid, info in d['accounts'].items():
                    accts[str(aid)] = {
                        'nm': info['name'], 'rg': info['region'],
                        'cd': info['credit_days'], 'cl': round(info['credit_limit'], 0),
                        'bal': round(info['balance'], 0),
                        'ovd': round(info['overdue'], 0),
                        'o1': round(info['ovd_0_30'], 0),
                        'o2': round(info['ovd_30_60'], 0),
                        'o3': round(info['ovd_60_plus'], 0),
                    }
                ar_json[key] = {
                    'date': d['date'],
                    'tb': round(d['total_balance'], 0),
                    'to': round(d['total_overdue'], 0),
                    'accts': accts,
                }

    # Build Collection JSON
    col_json = None
    if col_data:
        accts = {}
        for aid, info in col_data['accounts'].items():
            accts[str(aid)] = {
                'nm': info['name'],
                'mtd': round(info['mtd'], 0),
                'ytd': round(info['ytd'], 0),
            }
        col_json = {
            'date': col_data['date'],
            'tm': round(col_data['total_mtd'], 0),
            'ty': round(col_data['total_ytd'], 0),
            'accts': accts,
        }

    # Build PGI/Remain/Open JSON (current + prev)
    def _build_pro_json(data):
        """Convert PGI/Remain/Open data → JSON with current + prev + snapshots."""
        if not data:
            return None
        result = {'current': None, 'prev': None}
        for key in ['current', 'prev']:
            if key in data and data[key]:
                d = data[key]
                accts = {}
                for aid, info in d['accounts'].items():
                    entry = {'v': round(info['value'], 0), 'q': round(info['qty'], 0)}
                    if info.get('cats'):
                        entry['c'] = {cat: {'v': round(cv['v'], 0), 'q': round(cv['q'], 0)} for cat, cv in info['cats'].items()}
                    accts[str(aid)] = entry
                result[key] = {
                    'date': d['date'],
                    'tv': round(d['total_value'], 0),
                    'tq': round(d['total_qty'], 0),
                    'accts': accts,
                }
        # Add all snapshots for date-filter support
        if data.get('snapshots'):
            snaps_json = {}
            dates = sorted(data['snapshots'].keys())
            for mm_dd in dates:
                snap = data['snapshots'][mm_dd]
                accts = {}
                for aid, info in snap['accounts'].items():
                    entry = {'v': round(info['value'], 0), 'q': round(info['qty'], 0)}
                    if info.get('cats'):
                        entry['c'] = {cat: {'v': round(cv['v'], 0), 'q': round(cv['q'], 0)} for cat, cv in info['cats'].items()}
                    accts[str(aid)] = entry
                snaps_json[mm_dd] = {
                    'date': snap['date'],
                    'tv': round(snap['total_value'], 0),
                    'tq': round(snap['total_qty'], 0),
                    'accts': accts,
                }
            result['snapshots'] = snaps_json
            result['dates'] = dates
        return result if result['current'] else None

    pgi_json = _build_pro_json(pgi_data)
    remain_json = _build_pro_json(remain_data)
    open_json = _build_pro_json(open_data)

    # OUD/AR/Collection이 비어있으면 기존 data.json에서 보존
    if (not oud_data or not ar_data or not col_data) and os.path.exists(OUTPUT_DATA_JSON):
        try:
            with open(OUTPUT_DATA_JSON, 'r', encoding='utf-8') as f:
                old_data = json.load(f)
            if not oud_data and old_data.get('oud'):
                oud_json = old_data['oud']
                print("   Preserving existing OUD data from data.json")
            if not ar_data and old_data.get('ar'):
                ar_json = old_data['ar']
                print("   Preserving existing AR data from data.json")
            if not col_data and old_data.get('col'):
                col_json = old_data['col']
                print("   Preserving existing Collection data from data.json")
        except Exception as e:
            print(f"   WARNING: Could not preserve existing data: {e}")

    raw_dict = {'txn': txn, 'master': master, 'oud': oud_json, 'ar': ar_json, 'col': col_json, 'pgi': pgi_json, 'remain': remain_json, 'open': open_json}
    if sell_thru_date:
        raw_dict['sell_thru_date'] = sell_thru_date
    if ir_target:
        raw_dict['ir_target'] = ir_target
    # RSM FCST: embed in data.json to avoid GitHub API rate limiting
    if rsm_fcst:
        raw_dict['rsm_fcst'] = rsm_fcst
    elif os.path.exists(OUTPUT_DATA_JSON):
        try:
            with open(OUTPUT_DATA_JSON, 'r', encoding='utf-8') as f:
                old_d = json.load(f)
            if old_d.get('rsm_fcst'):
                raw_dict['rsm_fcst'] = old_d['rsm_fcst']
                print("   Preserving existing RSM FCST data from data.json")
        except Exception:
            pass
    json_str = json.dumps(raw_dict, ensure_ascii=False, separators=(',', ':'))

    # Save data to separate JSON file
    with open(OUTPUT_DATA_JSON, 'w', encoding='utf-8') as f:
        f.write(json_str)
    print(f"   {OUTPUT_DATA_JSON} ({len(json_str):,} bytes)")

    if not os.path.exists(OUTPUT_HTML):
        print(f"   WARNING: HTML file not found: {OUTPUT_HTML}")
        print("   Skipping HTML update (Excel was saved successfully)")
        return

    with open(OUTPUT_HTML, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Replace the RAW data line with empty placeholder (data loaded from data.json)
    new_lines = []
    replaced = False
    for line in lines:
        if (line.strip().startswith('const RAW') or line.strip().startswith('var RAW')) and not replaced:
            new_lines.append('var RAW; // loaded asynchronously from data.json\n')
            replaced = True
        else:
            new_lines.append(line)

    if not replaced:
        print("   WARNING: Could not find RAW data line to replace!")

    html = ''.join(new_lines)

    # Update date in header
    today = datetime.now().strftime('%Y-%m-%d')
    html = re.sub(r'Updated</span><span class="val">\d{4}-\d{2}-\d{2}',
                  f'Updated</span><span class="val">{today}', html)

    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"   {OUTPUT_HTML} ({len(html):,} bytes)")


# ============================================================
# STEP 6: Merge Duplicate Account Names
# ============================================================
def merge_duplicate_accounts(df):
    """
    동일 Account_Name에 복수 Account_ID가 있는 경우,
    총 매출(Value) 기준 가장 큰 ID를 대표 ID로 선정하고 나머지를 합산.
    """
    print("   Merging duplicate Account Names...")

    # Build name → {id: total_value}
    name_ids = {}
    for _, r in df.iterrows():
        nm = str(r['Account_Name']).strip() if r['Account_Name'] else ''
        if not nm:
            continue
        aid = r['Account_ID']
        val = float(r['Value']) if r['Value'] else 0
        if nm not in name_ids:
            name_ids[nm] = {}
        name_ids[nm][aid] = name_ids[nm].get(aid, 0) + val

    # Find names with multiple IDs → pick representative (highest total value)
    id_remap = {}  # old_id → representative_id
    merge_count = 0
    for nm, ids_vals in name_ids.items():
        if len(ids_vals) <= 1:
            continue
        # Representative = ID with highest total value
        rep_id = max(ids_vals, key=lambda x: ids_vals[x])
        for aid in ids_vals:
            if aid != rep_id:
                id_remap[aid] = rep_id
        merge_count += 1

    if id_remap:
        df['Account_ID'] = df['Account_ID'].map(lambda x: id_remap.get(x, x))
        print(f"   → {merge_count} account names merged ({len(id_remap)} IDs remapped)")
    else:
        print("   → No duplicates found")

    return df


# ============================================================
# STEP 7: Load OUD Data
# ============================================================
def parse_oud_date(filename):
    """Extract date from filename like '14-MAR-2026 HVAC.xlsx' or '11-April-2026 HVAC.xlsx' → '2026-03-14'"""
    import locale
    # Try short month (e.g., MAR) or full month (e.g., March, April)
    m = re.match(r'(\d{2})-([A-Za-z]+)-(\d{4})', filename)
    if not m:
        return None
    day, mon_str, year = m.group(1), m.group(2).upper(), m.group(3)
    months = {
        'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
        'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12',
        'JANUARY':'01','FEBRUARY':'02','MARCH':'03','APRIL':'04','MAY':'05','JUNE':'06',
        'JULY':'07','AUGUST':'08','SEPTEMBER':'09','OCTOBER':'10','NOVEMBER':'11','DECEMBER':'12',
    }
    mm = months.get(mon_str)
    if not mm:
        return None
    return f"{year}-{mm}-{day}"


def load_oud(id_remap=None):
    """
    Load OUD files from 06. OUD folder.
    Returns: {
        'current': {'date': '2026-03-14', 'accounts': {account_id: {name, value, qty}}, 'total_value', 'total_qty'},
        'prev':    {'date': '2026-03-07', 'accounts': {account_id: {name, value, qty}}, 'total_value', 'total_qty'},
    }
    """
    print("7. Loading OUD Data...")

    if not os.path.isdir(OUD_DIR):
        print(f"   OUD folder not found: {OUD_DIR}")
        return None

    # Find all OUD xlsx files and sort by date
    oud_files = []
    for f in os.listdir(OUD_DIR):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            dt = parse_oud_date(f)
            if dt:
                oud_files.append((dt, f))
    oud_files.sort(key=lambda x: x[0])

    if not oud_files:
        print("   No OUD files found")
        return None

    print(f"   Found {len(oud_files)} OUD files: {[f[1] for f in oud_files]}")

    # OUD Group → Sell-Thru category mapping (OUD names are truncated)
    OUD_GROUP_MAP = {
        'Split Inve': 'Split Inverter',
        'Split on/o': 'Split on/off',
        'Cassette':   'Cassette',
        'Concealed':  'Concealed',
        'Free Stand': 'Free Standing',
        'CAC Ducted': 'Convertible (CAC)',
        'Window SEE': 'Window',
        'Window SEEC':'Window',
        'AHU':        'AHU',
        'Multi-V':    'Multi-V',
        'Unitary Pa': 'Unitary Package',
        'Accessorie': 'Accessories',
    }

    def read_oud_file(filepath):
        """Read a single OUD file → {account_id: {name, value, qty}}"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find header row (row with 'Customer Name')
        header_idx = None
        for i, r in enumerate(rows):
            if r and r[0] == 'Customer Name':
                header_idx = i
                break
        if header_idx is None:
            return {}

        headers = [str(h).strip() if h else '' for h in rows[header_idx]]

        # Find column indices dynamically
        def find_col(name):
            for i, h in enumerate(headers):
                if h == name:
                    return i
            return None

        col_name = find_col('Customer Name')
        col_acc = find_col('Dealer ACC No') or find_col('Column1')
        col_rqty = find_col('R-Qty')
        col_tval = find_col('Total Value')
        col_group = find_col('Group')

        accounts = {}
        for r in rows[header_idx + 1:]:
            if not r or not r[col_name]:
                continue
            name = str(r[col_name]).strip()
            if 'Total' in name or name == 'Grand Total':
                continue

            acc_id = normalize_id(r[col_acc]) if col_acc is not None else name
            # Apply id_remap if available
            if id_remap and acc_id in id_remap:
                acc_id = id_remap[acc_id]

            rqty = float(r[col_rqty]) if col_rqty is not None and r[col_rqty] else 0
            tval = float(r[col_tval]) if col_tval is not None and r[col_tval] else 0

            # Apply HALF_QTY_CATS logic based on Group column
            if col_group is not None and r[col_group]:
                grp = str(r[col_group]).strip()
                mapped_cat = OUD_GROUP_MAP.get(grp, grp)
                if mapped_cat in HALF_QTY_CATS:
                    rqty = rqty / 2

            if acc_id not in accounts:
                accounts[acc_id] = {'name': name, 'value': 0, 'qty': 0, 'cats': {}}
            accounts[acc_id]['value'] += tval
            accounts[acc_id]['qty'] += rqty
            # Per-category breakdown
            if col_group is not None and r[col_group]:
                grp = str(r[col_group]).strip()
                cat_name = OUD_GROUP_MAP.get(grp, grp)
            else:
                cat_name = 'Others'
            if cat_name not in accounts[acc_id]['cats']:
                accounts[acc_id]['cats'][cat_name] = {'value': 0, 'qty': 0}
            accounts[acc_id]['cats'][cat_name]['value'] += tval
            accounts[acc_id]['cats'][cat_name]['qty'] += rqty

        return accounts

    # Build Customer Name → Dealer ACC No mapping from ALL files that have the column
    name_to_accno = {}
    for _, fname in oud_files:
        fpath = os.path.join(OUD_DIR, fname)
        wb_tmp = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws_tmp = wb_tmp[wb_tmp.sheetnames[0]]
        rows_tmp = list(ws_tmp.iter_rows(values_only=True))
        wb_tmp.close()
        hdr_idx = None
        for i, r in enumerate(rows_tmp):
            if r and r[0] == 'Customer Name':
                hdr_idx = i
                break
        if hdr_idx is None:
            continue
        hdrs = [str(h).strip() if h else '' for h in rows_tmp[hdr_idx]]
        acc_col = None
        for i, h in enumerate(hdrs):
            if h in ('Dealer ACC No', 'Column1'):
                acc_col = i
                break
        if acc_col is None:
            continue
        name_col = 0
        for r in rows_tmp[hdr_idx + 1:]:
            if not r or not r[name_col]:
                continue
            cname = str(r[name_col]).strip()
            if 'Total' in cname or cname == 'Grand Total':
                continue
            if r[acc_col]:
                name_to_accno[cname] = normalize_id(r[acc_col])
    if name_to_accno:
        print(f"   Built name→ID mapping: {len(name_to_accno)} entries")

    result = {}

    # Current = latest file
    cur_date, cur_file = oud_files[-1]
    cur_accounts = read_oud_file(os.path.join(OUD_DIR, cur_file))
    # Remap customer-name keys to Dealer ACC No using the mapping
    cur_accounts = _remap_oud_keys(cur_accounts, name_to_accno)
    result['current'] = {
        'date': cur_date,
        'accounts': cur_accounts,
        'total_value': sum(a['value'] for a in cur_accounts.values()),
        'total_qty': sum(a['qty'] for a in cur_accounts.values()),
    }
    print(f"   Current: {cur_file} → {len(cur_accounts)} accounts, Value={result['current']['total_value']:,.0f}")

    # Previous = second to last file (if exists)
    if len(oud_files) >= 2:
        prev_date, prev_file = oud_files[-2]
        prev_accounts = read_oud_file(os.path.join(OUD_DIR, prev_file))
        prev_accounts = _remap_oud_keys(prev_accounts, name_to_accno)
        result['prev'] = {
            'date': prev_date,
            'accounts': prev_accounts,
            'total_value': sum(a['value'] for a in prev_accounts.values()),
            'total_qty': sum(a['qty'] for a in prev_accounts.values()),
        }
        print(f"   Previous: {prev_file} → {len(prev_accounts)} accounts, Value={result['prev']['total_value']:,.0f}")
        delta_v = result['current']['total_value'] - result['prev']['total_value']
        delta_q = result['current']['total_qty'] - result['prev']['total_qty']
        print(f"   Delta: Value={delta_v:+,.0f}, Qty={delta_q:+,.0f}")

    return result


def _remap_oud_keys(accounts, name_to_accno):
    """Remap account keys: if a key is a customer name and we have a Dealer ACC No mapping, use the ACC No instead."""
    if not name_to_accno:
        return accounts
    remapped = {}
    for key, val in accounts.items():
        # If the key looks like a name (not a numeric ID), try to remap
        if not str(key).isdigit() and key in name_to_accno:
            new_key = name_to_accno[key]
            if new_key in remapped:
                # Merge values if same ACC No
                remapped[new_key]['value'] += val['value']
                remapped[new_key]['qty'] += val['qty']
                for cat, cv in val.get('cats', {}).items():
                    if cat not in remapped[new_key]['cats']:
                        remapped[new_key]['cats'][cat] = {'value': 0, 'qty': 0}
                    remapped[new_key]['cats'][cat]['value'] += cv['value']
                    remapped[new_key]['cats'][cat]['qty'] += cv['qty']
            else:
                remapped[new_key] = val
        else:
            remapped[key] = val
    return remapped


# ============================================================
# STEP 8: Load AR (Overdue) + Collection Data
# ============================================================
# AR paths: try multiple locations (local PC structure vs VM mount)
def _find_ar_dir(subdir, fallback_name):
    """Find AR directory — try user-confirmed path first, then original paths."""
    # Primary path: 00. Raw Data/02. 2026/04. Overdue/...
    user_path = os.path.join(RAW_DATA_BASE, '02. 2026', '04. Overdue', subdir)
    if os.path.isdir(user_path):
        return user_path
    # Fallback: old location under Python Code folder
    local = os.path.join(BASE, '02. 2026', '04. Overdue', subdir)
    if os.path.isdir(local):
        return local
    # VM mount: sibling folder
    vm = os.path.join(os.path.dirname(BASE), fallback_name)
    if os.path.isdir(vm):
        return vm
    return user_path  # return primary as default

AR_OVERDUE_DIR = _find_ar_dir('00. Daily Overdue', '00. Daily Overdue')
AR_COLLECTION_DIR = _find_ar_dir('01. Daily Collection', '01. Daily Collection')

# Aging bucket start days (matches column order [10]~[34])
AGING_BUCKETS = [
    (0, '0 - 30'), (31, '31 - 60'), (61, '61 - 90'), (91, '91 - 120'),
    (121, '121 - 150'), (151, '151 - 180'), (181, '181 - 210'), (211, '211 - 240'),
    (241, '241 - 270'), (271, '271 - 300'), (301, '301 - 330'), (331, '331 - 1 Year'),
    (361, '361 - 450'), (451, '451 - 540'), (541, '541 - 630'), (631, '631 - 2 Year'),
    (730, '2 Year - 810'), (811, '811 - 900'), (901, '901 - 990'), (991, '991 - 3 Year'),
    (1095, '3 Year - 1170'), (1171, '1171 - 1260'), (1261, '1261 - 1350'),
    (1351, '1351 - 4 Year'), (1461, 'Over 4 Years'),
]

def parse_credit_days(terms):
    """Extract credit days from Terms code. e.g. 'Z090-90 Days...' → 90"""
    if not terms:
        return 0  # Cash
    s = str(terms).strip()
    # Try Z### or V### pattern
    m = re.match(r'[A-Z](\d{3})', s)
    if m:
        days = int(m.group(1))
        if days <= 1:  # Z001, Z000 = Cash
            return 0
        return days
    # Try 0001 pattern
    if s.startswith('0001') or s.startswith('0000'):
        return 0
    return 0  # Unknown = Cash


def parse_ar_date(filename):
    """Extract date from filename like '260317_Overdue.xlsx' → '2026-03-17'"""
    m = re.match(r'(\d{6})', filename)
    if not m:
        return None
    d = m.group(1)
    yy, mm, dd = d[:2], d[2:4], d[4:6]
    return f"20{yy}-{mm}-{dd}"


def load_ar_overdue(id_remap=None):
    """Load AR Overdue data from Daily Overdue folder."""
    print("8. Loading AR Overdue Data...")

    if not os.path.isdir(AR_OVERDUE_DIR):
        print(f"   AR Overdue folder not found: {AR_OVERDUE_DIR}")
        return None

    # Find all overdue files
    ar_files = []
    for f in os.listdir(AR_OVERDUE_DIR):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            dt = parse_ar_date(f)
            if dt:
                ar_files.append((dt, f))
    ar_files.sort(key=lambda x: x[0])

    if not ar_files:
        print("   No AR Overdue files found")
        return None

    print(f"   Found {len(ar_files)} files: {[f[1] for f in ar_files]}")

    def read_overdue_file(filepath):
        """Read a single overdue file → {account_id: {...}}"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        headers = rows[0]

        # Build dynamic column map (supports both old and new file formats)
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[str(h).strip()] = i

        # Map required columns (fallback to old fixed indices if header missing)
        col_cid = col_map.get('Customer NO', col_map.get('Customer No', 0))
        col_name = col_map.get('Customer Name', col_map.get('Name', 1))
        col_region = col_map.get('Region Desc', col_map.get('Region', 4))
        col_terms = col_map.get('Terms', 7)
        col_credit_limit = col_map.get('Credit Limit', 8)
        col_balance = col_map.get('Balance', 9)

        # Find aging bucket column indices by header name (handles any format)
        aging_col_map = []  # [(start_day, col_idx), ...]
        for start_day, bucket_name in AGING_BUCKETS:
            idx = col_map.get(bucket_name)
            if idx is not None:
                aging_col_map.append((start_day, idx))
        # Fallback: use fixed range 10~34 if no headers matched
        if not aging_col_map:
            aging_col_map = [(AGING_BUCKETS[i][0], 10 + i) for i in range(len(AGING_BUCKETS))]

        accounts = {}

        for r in rows[1:]:
            if col_cid >= len(r) or not r[col_cid]:
                continue
            cid = normalize_id(r[col_cid])
            if not cid:
                continue
            if id_remap and cid in id_remap:
                cid = id_remap[cid]

            name = str(r[col_name]).strip() if col_name < len(r) and r[col_name] else ''
            terms = str(r[col_terms]).strip() if col_terms < len(r) and r[col_terms] else ''
            credit_days = parse_credit_days(terms)
            credit_limit = float(r[col_credit_limit]) if col_credit_limit < len(r) and r[col_credit_limit] else 0
            balance = float(r[col_balance]) if col_balance < len(r) and r[col_balance] else 0
            region = str(r[col_region]).strip() if col_region < len(r) and r[col_region] else ''

            # Calculate overdue based on credit days (dynamic aging columns)
            overdue = 0
            ovd_buckets = []  # list of overdue amounts per bucket
            for start_day, col_idx in aging_col_map:
                amt = float(r[col_idx]) if col_idx < len(r) and r[col_idx] else 0
                if start_day > credit_days:
                    overdue += amt
                    ovd_buckets.append(amt)

            # Split overdue into 3 sub-buckets: 0-30, 30-60, 60+
            ovd_0_30 = ovd_buckets[0] if len(ovd_buckets) > 0 else 0
            ovd_30_60 = ovd_buckets[1] if len(ovd_buckets) > 1 else 0
            ovd_60_plus = sum(ovd_buckets[2:]) if len(ovd_buckets) > 2 else 0

            # Aggregate if same cid exists (due to remap)
            if cid in accounts:
                a = accounts[cid]
                a['balance'] += balance
                a['overdue'] += overdue
                a['ovd_0_30'] += ovd_0_30
                a['ovd_30_60'] += ovd_30_60
                a['ovd_60_plus'] += ovd_60_plus
                a['credit_limit'] += credit_limit
            else:
                accounts[cid] = {
                    'name': name, 'region': region, 'terms': terms,
                    'credit_days': credit_days, 'credit_limit': credit_limit,
                    'balance': balance, 'overdue': overdue,
                    'ovd_0_30': ovd_0_30, 'ovd_30_60': ovd_30_60,
                    'ovd_60_plus': ovd_60_plus,
                }

        # Merge duplicate Account Names within AR (same logic as Sell-Thru)
        name_ids = {}
        for cid, info in accounts.items():
            nm = info['name']
            if nm not in name_ids:
                name_ids[nm] = {}
            name_ids[nm][cid] = abs(info['balance'])

        ar_name_remap = {}
        for nm, ids_vals in name_ids.items():
            if len(ids_vals) <= 1:
                continue
            rep_id = max(ids_vals, key=lambda x: ids_vals[x])
            for aid in ids_vals:
                if aid != rep_id:
                    ar_name_remap[aid] = rep_id

        if ar_name_remap:
            merged = {}
            for cid, info in accounts.items():
                target = ar_name_remap.get(cid, cid)
                if target in merged:
                    m = merged[target]
                    m['balance'] += info['balance']
                    m['overdue'] += info['overdue']
                    m['ovd_0_30'] += info['ovd_0_30']
                    m['ovd_30_60'] += info['ovd_30_60']
                    m['ovd_60_plus'] += info['ovd_60_plus']
                    m['credit_limit'] += info['credit_limit']
                else:
                    merged[target] = dict(info)
            accounts = merged
            print(f"   AR name merge: {len(ar_name_remap)} IDs remapped → {len(accounts)} accounts")

        return accounts, ar_name_remap

    result = {}
    all_ar_remap = {}  # collect remap from all files for Collection use

    # Current = latest file
    cur_date, cur_file = ar_files[-1]
    cur_accounts, cur_remap = read_overdue_file(os.path.join(AR_OVERDUE_DIR, cur_file))
    all_ar_remap.update(cur_remap)
    total_bal = sum(a['balance'] for a in cur_accounts.values())
    total_ovd = sum(a['overdue'] for a in cur_accounts.values())
    result['current'] = {
        'date': cur_date,
        'accounts': cur_accounts,
        'total_balance': total_bal,
        'total_overdue': total_ovd,
    }
    result['_ar_remap'] = all_ar_remap  # pass to Collection loader
    print(f"   Current: {cur_file} → {len(cur_accounts)} accounts, Balance={total_bal:,.0f}, Overdue={total_ovd:,.0f}")

    # Previous
    if len(ar_files) >= 2:
        prev_date, prev_file = ar_files[-2]
        prev_accounts, prev_remap = read_overdue_file(os.path.join(AR_OVERDUE_DIR, prev_file))
        prev_bal = sum(a['balance'] for a in prev_accounts.values())
        prev_ovd = sum(a['overdue'] for a in prev_accounts.values())
        result['prev'] = {
            'date': prev_date,
            'accounts': prev_accounts,
            'total_balance': prev_bal,
            'total_overdue': prev_ovd,
        }
        print(f"   Previous: {prev_file} → {len(prev_accounts)} accounts")
        print(f"   Delta: Balance={total_bal - prev_bal:+,.0f}, Overdue={total_ovd - prev_ovd:+,.0f}")

    return result


def load_pgi_remain_open(id_remap=None):
    """
    Load PGI, Remaining, and Open data from daily folders.
    Loads latest (current) + previous file for WoW comparison.
    Returns: {
        'pgi':    {'current': {date, accounts, total_value, total_qty}, 'prev': {...} or None},
        'remain': {'current': {...}, 'prev': {...} or None},
        'open':   {'current': {...}, 'prev': {...} or None},
    }
    """
    print("10. Loading PGI/Remaining/Open Data...")

    result = {'pgi': None, 'remain': None, 'open': None}

    # Define data sources
    data_sources = {
        'pgi': {
            'dir': os.path.join(RAW_DATA_BASE, '02. 2026', '01. Daily PGI'),
            'pattern': 'PGI_',
            # PGI Spreadsheet format (Delivery Number col[0], 83 cols):
            # Customer[3], Quantity[39], Value[40], Category[62]
            'cols': {'cid': 3, 'val': 40, 'qty': 39, 'cat': 62}
        },
        'remain': {
            'dir': os.path.join(RAW_DATA_BASE, '02. 2026', '02. Daily Remaining'),
            'pattern': 'Remaining_',
            # Remaining Spreadsheet format (Sales Order Number col[0], 105 cols):
            # Customer[3], Delivery Quantity[6], Delivery Value[7], Category[84]
            'cols': {'cid': 3, 'val': 7, 'qty': 6, 'cat': 84}
        },
        'open': {
            'dir': os.path.join(RAW_DATA_BASE, '02. 2026', '03. Daily Open'),
            'pattern': 'Open_',
            # Open Spreadsheet format (Sales Order Number col[0], 90 cols):
            # Customer[3], Order Quantity[75], Billed List Price[74], Category[69]
            'cols': {'cid': 3, 'val': 74, 'qty': 75, 'cat': 69}
        }
    }

    def _read_pro_file(filepath, cols, id_remap):
        """Read a single PGI/Remain/Open file → {account_id: {value, qty}}"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Auto-detect header row: find row containing known header keywords in first column
        HEADER_FIRST_COLS = {'Sales Order', 'Delivery Number', 'Sales Order Number', 'Product Hierarchy'}
        header_idx = 0
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip() in HEADER_FIRST_COLS:
                header_idx = i
                header_row = row
                break

        # Auto-detect format based on header signature
        if header_row and len(header_row) > 3:
            h0 = str(header_row[0] or '').strip()
            h1 = str(header_row[1] or '').strip()
            ncols = len(header_row)

            if h1 == 'Order Created By':
                # DOM-scroll extracted files (misaligned columns)
                h9 = str(header_row[9] or '').strip()
                if h9 == 'Quantity':
                    # PGI DOM-scroll: cid=3, val=40, qty=39
                    cols = {'cid': 3, 'val': 40, 'qty': 39, 'cat': -1}
                elif h9 == 'Remaining Quantity':
                    # Remaining DOM-scroll: cid=3, val=7, qty=6
                    cols = {'cid': 3, 'val': 7, 'qty': 6, 'cat': -1}
                # Open DOM-scroll: no override → uses data_sources default

            elif h0 == 'Delivery Number' and h1 == 'Delivery Date':
                # PGI Spreadsheet export (old format, 83 cols):
                # Customer[3], Quantity[39], Value[40], Category[62]
                cols = {'cid': 3, 'val': 40, 'qty': 39, 'cat': 62}

            elif h0 == 'Sales Order' and h1 == 'Delivery Date':
                # PGI New Spreadsheet format (77 cols, from 2026-04-23):
                # Customer[35], Quantity[5], Value[6], Category[56]
                cols = {'cid': 35, 'val': 6, 'qty': 5, 'cat': 56}

            elif h0 == 'Sales Order Number' and h1 == 'Order Date':
                if ncols >= 100:
                    # Remaining Spreadsheet (105 cols):
                    # Customer[3], Delivery Quantity[6], Delivery Value[7], Category[84]
                    cols = {'cid': 3, 'val': 7, 'qty': 6, 'cat': 84}
                else:
                    # Open Spreadsheet (90 cols):
                    # Customer[3], Order Quantity[75], Billed List Price[74], Category[69]
                    cols = {'cid': 3, 'val': 74, 'qty': 75, 'cat': 69}

            elif h0 == 'Product Hierarchy' and h1 == 'Group':
                # ZDSAL_OPEN Layout format (96 cols):
                # Customer[8], Open Quantity[10], VALUE[88], Category[56]
                cols = {'cid': 8, 'val': 88, 'qty': 10, 'cat': 56}

            elif h1 == 'Order Date' and ncols >= 88:
                # Old Spreadsheet export with extended columns (legacy format)
                h87 = str(header_row[87] or '').strip() if ncols > 87 else ''
                if h87 == 'VALUE':
                    # Extended Spreadsheet format (97 cols):
                    # VALUE[87] = open order value after discount
                    # Open Quantity[88] = qty still pending delivery
                    cols = {'cid': 33, 'val': 87, 'qty': 88, 'cat': 53}

        accounts = {}
        total_value = 0
        total_qty = 0

        # Determine minimum required columns (cat may be beyond file width)
        min_required = max(cols['cid'], cols['val'], cols['qty'])

        for row in rows[header_idx + 1:]:  # Skip header (auto-detected)
            if not row or len(row) <= min_required:
                continue
            cid = normalize_id(row[cols['cid']])
            if not cid:
                continue
            if id_remap and cid in id_remap:
                cid = id_remap[cid]
            try:
                value = float(row[cols['val']]) if row[cols['val']] else 0
            except (ValueError, TypeError):
                value = 0
            try:
                qty = float(row[cols['qty']]) if row[cols['qty']] else 0
            except (ValueError, TypeError):
                qty = 0
            # Category filtering: skip if cat column specified and category is excluded
            mapped_cat = None
            if cols['cat'] >= 0 and len(row) > cols['cat']:
                raw_cat = row[cols['cat']]
                mapped_cat = map_category(raw_cat)
                if mapped_cat is None:
                    continue
                if mapped_cat in HALF_QTY_CATS:
                    qty = qty / 2
                if mapped_cat in ZERO_QTY_CATS:
                    qty = 0
            if cid not in accounts:
                accounts[cid] = {'value': 0, 'qty': 0, 'cats': {}}
            accounts[cid]['value'] += value
            accounts[cid]['qty'] += qty
            # Store per-category breakdown for dashboard category filter support
            if mapped_cat:
                if mapped_cat not in accounts[cid]['cats']:
                    accounts[cid]['cats'][mapped_cat] = {'v': 0, 'q': 0}
                accounts[cid]['cats'][mapped_cat]['v'] += value
                accounts[cid]['cats'][mapped_cat]['q'] += qty
            total_value += value
            total_qty += qty

        return accounts, total_value, total_qty

    def _parse_date(datestr):
        """YYMMDD → YYYY-MM-DD"""
        yy, mm, dd = int(datestr[0:2]), int(datestr[2:4]), int(datestr[4:6])
        yyyy = 2000 + yy if yy < 70 else 1900 + yy
        return f"{yyyy:04d}-{mm:02d}-{dd:02d}"

    for data_type, source in data_sources.items():
        dir_path = source['dir']

        if not os.path.isdir(dir_path):
            print(f"   {data_type.upper()} folder not found: {dir_path}")
            continue

        # Find all files matching pattern
        files = []
        for f in os.listdir(dir_path):
            if f.startswith(source['pattern']) and f.endswith('.xlsx') and not f.startswith('~$'):
                match = re.search(r'_(\d{6})\.xlsx$', f)
                if match:
                    files.append((match.group(1), f))

        if not files:
            print(f"   No {data_type.upper()} files found")
            continue

        files.sort(key=lambda x: x[0])  # Ascending by date

        # Current = latest file
        cur_date, cur_file = files[-1]
        print(f"   {data_type.upper()} Current: {cur_file}")
        try:
            cur_accts, cur_tv, cur_tq = _read_pro_file(
                os.path.join(dir_path, cur_file), source['cols'], id_remap)
            current = {
                'date': _parse_date(cur_date),
                'accounts': cur_accts,
                'total_value': cur_tv,
                'total_qty': cur_tq,
            }
            print(f"     → {len(cur_accts)} accounts, Value={cur_tv:,.0f}, Qty={cur_tq:,.0f}")
        except Exception as e:
            print(f"   ERROR reading {data_type.upper()} current: {e}")
            continue

        # Previous = file closest to 7 days before current
        prev = None
        if len(files) >= 2:
            from datetime import datetime as _dt, timedelta as _td
            cur_dt = _dt.strptime(cur_date, '%y%m%d')
            target_dt = cur_dt - _td(days=7)
            # Find file closest to target_dt (but not the current file)
            best_file = None
            best_diff = None
            for fdate, fname in files[:-1]:  # exclude current
                fdt = _dt.strptime(fdate, '%y%m%d')
                diff = abs((fdt - target_dt).days)
                if best_diff is None or diff < best_diff:
                    best_diff = diff
                    best_file = (fdate, fname)
            if best_file:
                prev_date, prev_file = best_file
                print(f"   {data_type.upper()} WoW Compare: {prev_file} ({best_diff}d from -7d target)")
                try:
                    prev_accts, prev_tv, prev_tq = _read_pro_file(
                        os.path.join(dir_path, prev_file), source['cols'], id_remap)
                    prev = {
                        'date': _parse_date(prev_date),
                        'accounts': prev_accts,
                        'total_value': prev_tv,
                        'total_qty': prev_tq,
                    }
                    print(f"     → {len(prev_accts)} accounts, Value={prev_tv:,.0f}, Qty={prev_tq:,.0f}")
                except Exception as e:
                    print(f"   ERROR reading {data_type.upper()} prev: {e}")

        # Build all snapshots for date-filter support
        snapshots = {}
        for fdate, fname in files:
            try:
                s_accts, s_tv, s_tq = _read_pro_file(
                    os.path.join(dir_path, fname), source['cols'], id_remap)
                mm_dd = f"{int(fdate[2:4]):02d}-{int(fdate[4:6]):02d}"
                snapshots[mm_dd] = {
                    'date': _parse_date(fdate),
                    'accounts': s_accts,
                    'total_value': s_tv,
                    'total_qty': s_tq,
                }
            except Exception:
                pass

        result[data_type] = {'current': current, 'prev': prev, 'snapshots': snapshots}

    return result


def load_collection(id_remap=None):
    """Load Collection data from Daily Collection folder."""
    print("9. Loading Collection Data...")

    if not os.path.isdir(AR_COLLECTION_DIR):
        print(f"   Collection folder not found: {AR_COLLECTION_DIR}")
        return None

    # Find all collection files
    col_files = []
    for f in os.listdir(AR_COLLECTION_DIR):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            dt = parse_ar_date(f)
            if dt:
                col_files.append((dt, f))
    col_files.sort(key=lambda x: x[0])

    if not col_files:
        print("   No Collection files found")
        return None

    print(f"   Found {len(col_files)} files: {[f[1] for f in col_files]}")

    def read_collection_file(filepath):
        """Read collection file → {account_id: {mtd, ytd, ...}}"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        headers = rows[0]
        # Find column indices dynamically
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[str(h).strip()] = i

        col_cid = col_map.get('Customer No', col_map.get('Customer NO'))
        col_name = col_map.get('Name', col_map.get('Customer Name'))
        col_mtd = col_map.get('MTD')
        col_ytd = col_map.get('YTD')
        col_payment = col_map.get('Payment Amount')
        col_balance = col_map.get('Customer Balance')
        col_over90 = col_map.get('Over 90')

        accounts = {}
        for r in rows[1:]:
            if not r[col_cid]:
                continue
            cid = normalize_id(r[col_cid])
            if not cid:
                continue
            if id_remap and cid in id_remap:
                cid = id_remap[cid]

            name = str(r[col_name]).strip() if col_name is not None and r[col_name] else ''
            mtd = float(r[col_mtd]) if col_mtd is not None and r[col_mtd] else 0
            ytd = float(r[col_ytd]) if col_ytd is not None and r[col_ytd] else 0
            payment = float(r[col_payment]) if col_payment is not None and r[col_payment] else 0
            balance = float(r[col_balance]) if col_balance is not None and r[col_balance] else 0
            over90 = float(r[col_over90]) if col_over90 is not None and r[col_over90] else 0

            # Aggregate if same cid (due to remap)
            if cid in accounts:
                a = accounts[cid]
                a['mtd'] += mtd
                a['ytd'] += ytd
                a['payment'] += payment
            else:
                accounts[cid] = {
                    'name': name, 'mtd': mtd, 'ytd': ytd,
                    'payment': payment, 'balance': balance, 'over90': over90,
                }

        return accounts

    # Current = latest file
    cur_date, cur_file = col_files[-1]
    cur_accounts = read_collection_file(os.path.join(AR_COLLECTION_DIR, cur_file))
    total_mtd = sum(a['mtd'] for a in cur_accounts.values())
    total_ytd = sum(a['ytd'] for a in cur_accounts.values())
    result = {
        'date': cur_date,
        'accounts': cur_accounts,
        'total_mtd': total_mtd,
        'total_ytd': total_ytd,
    }
    print(f"   Current: {cur_file} → {len(cur_accounts)} accounts, MTD={total_mtd:,.0f}, YTD={total_ytd:,.0f}")

    return result


# ============================================================
# ============================================================
# RAW DATA QUALITY VALIDATOR
# ============================================================
_REGION_NAMES = {'East', 'Central', 'West', 'North', 'South'}

def _validate_new_rows(rows: list) -> None:
    """
    신규 append 예정 행의 품질을 검사한다.
    이상 탐지 시 경고를 출력하고 RuntimeError로 파이프라인을 중단한다.
    (Raw 파일이 오염되기 전에 차단하는 것이 목적)

    검증 항목
    ---------
    1. 79-컬럼 오매핑 재발 감지
       col[8]에 Region명이 들어오면서 col[77]=None → 컬럼 매핑이 다시 틀어진 신호.
    2. SAR→Qty 오염 재발 감지
       qty == val AND qty > 1000 → 서비스 요금 SAR 금액이 Qty 컬럼에 들어온 패턴.
    3. Category 누락 비율 > 30% → 새로운 포맷 변경 조기 경보.
    """
    total = len(rows)
    if total == 0:
        return

    region_mismatch = 0   # 체크 1
    sar_qty_poison  = 0   # 체크 2
    cat_missing     = 0   # 체크 3

    for r in rows:
        col8  = r[8]  if len(r) > 8  else None
        col77 = r[77] if len(r) > 77 else None
        col5  = r[5]  if len(r) > 5  else None   # qty
        col6  = r[6]  if len(r) > 6  else None   # val

        # 체크 1: col[8]에 Region명이 들어오면서 Category(col[77])가 None
        if str(col8 or '').strip() in _REGION_NAMES and col77 is None:
            region_mismatch += 1

        # 체크 2: qty == val AND qty > 1000 (SAR 금액이 Qty로 오염)
        # ZERO_QTY_MATERIALS(INSTALLATION CHRGS, ENERGY_SERVICE)는 SAP 원래부터 qty=val 구조이므로 제외
        col70 = str(r[70] if len(r) > 70 else '').strip().upper()
        is_known_sar_mat = col70 in {m.upper() for m in ZERO_QTY_MATERIALS}
        try:
            qty = float(col5 or 0)
            val = float(col6 or 0)
            if qty == val and qty > 1000 and not is_known_sar_mat:
                sar_qty_poison += 1
        except (TypeError, ValueError):
            pass

        # 체크 3: Category(col[77]) 누락
        if col77 is None or str(col77).strip() == '':
            cat_missing += 1

    cat_missing_pct = cat_missing / total * 100
    issues = []

    if region_mismatch > 0:
        issues.append(
            f"⚠️  [체크1] col[8]에 Region명 + Category None 행 {region_mismatch}건 "
            f"→ 79-컬럼 오매핑 재발 의심. Raw 오염 차단."
        )

    if sar_qty_poison > 0:
        issues.append(
            f"⚠️  [체크2] qty==val AND qty>1000 행 {sar_qty_poison}건 "
            f"→ SAR→Qty 오염 재발 의심. Raw 오염 차단."
        )

    if cat_missing_pct > 30:
        issues.append(
            f"⚠️  [체크3] Category 누락 {cat_missing}/{total}건 ({cat_missing_pct:.1f}%) "
            f"→ Daily 파일 포맷 변경 의심. 매핑 점검 필요."
        )

    if issues:
        print("\n" + "=" * 60)
        print("  RAW DATA 안전장치 — 파이프라인 중단")
        print("=" * 60)
        for msg in issues:
            print(" ", msg)
        print("  ▶ Raw 파일은 수정되지 않았습니다.")
        print("  ▶ refresh_dashboard.py의 update_raw_from_daily()를 점검하세요.")
        print("=" * 60 + "\n")
        raise RuntimeError("Raw data quality check failed — pipeline halted to prevent corruption.")

    print(f"   ✓ 안전장치 통과: 신규 {total}행 이상 없음")


# ============================================================
# STEP 0: Auto-update Raw from Daily Sell Thru
# ============================================================
def update_raw_from_daily():
    """
    00. Daily Sell Thru 폴더에서 모든 월별 최신 파일을 읽어
    2026 Sell thru Raw data.xlsx에 신규 날짜 데이터를 자동 추가.
    - 모든 월별로 최신 파일 확인 (이전 월 누락 데이터도 처리)
    - 같은 월 데이터끼리만 비교 (월 단위 마감)
    - 이미 있는 날짜는 건너뜀
    """
    print("0. Updating Raw from Daily Sell Thru...")

    if not os.path.isdir(DAILY_DIR):
        print(f"   Daily folder not found: {DAILY_DIR}")
        return False

    # 1) Find ALL daily files, group by month → pick latest per month
    all_daily = []
    for f in os.listdir(DAILY_DIR):
        if f.endswith('.xlsx') and not f.startswith('~$') and 'Sell thru' in f:
            m = re.search(r'(\d{6})', f)
            if m:
                all_daily.append((m.group(1), f))
    all_daily.sort(key=lambda x: x[0])

    if not all_daily:
        print("   No daily files found")
        return False

    # Group by month (YYMM) → latest file per month
    month_latest = {}  # 'YYMM' → (code, filename)
    for code, fname in all_daily:
        yymm = code[:4]  # first 4 chars = YYMM
        month_latest[yymm] = (code, fname)

    print(f"   Found {len(month_latest)} months: {sorted(month_latest.keys())}")

    def _parse_inv_date(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, str) and val:
            try:
                return datetime.strptime(val, '%d.%m.%Y')
            except ValueError:
                pass
        return None

    # 2) Load lookup tables once (shared across months)
    raw_path = FILES['2026_raw']
    raw_wb2 = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)

    div_map = {}
    for row in raw_wb2['Division'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        code = normalize_id(r[0])
        if code:
            classification = r[2] if len(r) > 2 else None
            new_exist = r[3] if len(r) > 3 else None
            div2 = r[5] if len(r) > 5 else None
            div_map[code] = (classification, new_exist, div2)

    emp_map = {}
    for row in raw_wb2['Employee'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        for i in range(0, len(r) - 1, 2):
            code = normalize_id(r[i])
            if code and r[i + 1]:
                emp_map[code] = str(r[i + 1]).strip()

    raw_wb2.close()
    print(f"   Lookup tables: Division={len(div_map)}, Employee={len(emp_map)}")

    # 3) Read existing dates from Raw (all months)
    raw_wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    raw_ws = raw_wb['Raw']
    existing_by_month = {}  # month_int → set of dates
    for row in raw_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        inv_date = row[1]
        if isinstance(inv_date, datetime):
            mo = inv_date.month
            existing_by_month.setdefault(mo, set()).add(inv_date.date())
    raw_wb.close()

    # 4) Process each month's latest file
    all_new_rows = []
    any_updated = False

    for yymm, (code, fname) in sorted(month_latest.items()):
        daily_month = int(code[2:4])
        daily_path = os.path.join(DAILY_DIR, fname)
        print(f"   Processing month {daily_month} (file: {fname})...")

        daily_wb = openpyxl.load_workbook(daily_path, read_only=True, data_only=True)
        daily_ws = daily_wb[daily_wb.sheetnames[0]]
        daily_rows_all = list(daily_ws.iter_rows(values_only=True))
        daily_wb.close()

        if len(daily_rows_all) < 2:
            continue

        # Find header row
        header_row_idx = 0
        for i, r in enumerate(daily_rows_all):
            if r[1] == 'Invoice Date':
                header_row_idx = i
                break
        daily_data = daily_rows_all[header_row_idx + 1:]

        # Detect format
        # - New (98-col): col[38] header = 'Customer ID (Payer)', col[77] = 'Category'
        # - Old (79-col): col[3] = Customer ID, col[40] = Qty, col[58] = Category
        #   col[8] = Region Descr (NOT Division), col[9] = Division code
        hdr = daily_rows_all[header_row_idx]
        new_format = (len(hdr) >= 39 and str(hdr[38] or '').strip() == 'Customer ID (Payer)')

        # Collect dates in this file (for this month only)
        daily_dates = set()
        for r in daily_data:
            inv_date = _parse_inv_date(r[1])
            if inv_date and inv_date.month == daily_month:
                daily_dates.add(inv_date.date())

        existing_dates = existing_by_month.get(daily_month, set())
        new_dates = daily_dates - existing_dates

        if not new_dates:
            print(f"     Month {daily_month}: no new dates")
            continue

        print(f"     Month {daily_month}: adding dates {sorted(new_dates)}")

        # Build rows for new dates
        for r in daily_data:
            inv_date = _parse_inv_date(r[1])
            if not inv_date:
                continue
            if inv_date.month != daily_month:
                continue
            if inv_date.date() not in new_dates:
                continue

            emp_txn = normalize_emp(r[27])  # Sales Employee Number (same position in both formats)

            if new_format:
                # New format (98-col): columns already match Raw sheet structure
                payer_id = normalize_id(r[38]) if len(r) > 38 else None
                row_list = list(r)
                row_list[1] = inv_date
            else:
                # Old format (79-col): completely different column layout → build fresh 101-col row
                # 79-col key positions:
                #   col[3]=Customer ID, col[4]=Material, col[9]=Division code
                #   col[27]=Sales Emp Txn, col[40]=Qty, col[41]=Value
                #   col[43]=Customer Name, col[58]=Category
                payer_id = normalize_id(r[3]) if len(r) > 3 else None
                row_list = [None] * 101
                row_list[1]  = inv_date
                row_list[5]  = int(r[40])   if len(r) > 40 and isinstance(r[40], (int, float)) else 0
                row_list[6]  = float(r[41]) if len(r) > 41 and isinstance(r[41], (int, float)) else 0.0
                row_list[8]  = r[9]  if len(r) > 9  else None  # Division code
                row_list[27] = r[27] if len(r) > 27 else None  # Sales Emp Txn
                row_list[38] = payer_id
                row_list[39] = r[43] if len(r) > 43 else (r[15] if len(r) > 15 else None)
                row_list[70] = r[4]  if len(r) > 4  else None  # Material
                row_list[77] = r[58] if len(r) > 58 else None  # Category

            div_info = div_map.get(payer_id, (None, None, None))
            classification = div_info[0]
            new_exist = div_info[1]

            division2 = emp_map.get(emp_txn)
            if not division2 and div_info[2]:
                division2 = div_info[2]

            if new_format:
                row_list.extend([classification, new_exist, division2])
            else:
                row_list[98] = classification
                row_list[99] = new_exist
                row_list[100] = division2

            all_new_rows.append(row_list)

        # Update existing_by_month cache to avoid double-adding
        existing_by_month.setdefault(daily_month, set()).update(new_dates)
        any_updated = True

    if not all_new_rows:
        print("   No new dates to add — Raw is up to date")
        return False

    # ── 안전장치: append 전 신규 행 품질 검증 ──────────────────
    _validate_new_rows(all_new_rows)
    # ────────────────────────────────────────────────────────────

    print(f"   Adding {len(all_new_rows)} new rows to Raw sheet...")

    # 5) Append all new rows to Raw sheet
    raw_wb3 = openpyxl.load_workbook(raw_path)
    raw_ws3 = raw_wb3['Raw']

    for row_data in all_new_rows:
        raw_ws3.append(row_data)

    raw_wb3.save(raw_path)
    raw_wb3.close()

    print(f"   Raw data updated: +{len(all_new_rows)} rows total")
    return True


# ============================================================
# AUTO-DEPLOY TO GITHUB PAGES
# ============================================================
def deploy_to_github():
    """Auto commit & push dashboard files to GitHub Pages and sync Shaker-MD-App."""
    import subprocess
    import shutil
    deploy_dir = DEPLOY_DIR
    if not os.path.isdir(os.path.join(deploy_dir, '.git')):
        print("\n11. GitHub Deploy: .git not found — skipping (run git init first)")
        return

    # ── 1. Sell-Thru-Progress-Dashboard 레포 배포 ──
    print("\n11. Deploying to GitHub Pages (Sell-Thru-Progress-Dashboard)...")
    try:
        today = datetime.now().strftime('%Y-%m-%d %H:%M')
        # data.json 포함 모든 대시보드 파일 커밋 (index.html만 하면 data.json 누락됨!)
        deploy_files = ['index.html', 'data.json', 'memo_data.json', 'rsm_fcst_data.json']
        for f in deploy_files:
            if os.path.exists(os.path.join(deploy_dir, f)):
                subprocess.run(['git', 'add', f], cwd=deploy_dir, check=True, capture_output=True)
        # 변경사항 확인
        result = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=deploy_dir, capture_output=True)
        if result.returncode == 0:
            print("   No changes to deploy (Sell-Thru-Progress-Dashboard)")
        else:
            subprocess.run(['git', 'commit', '-m', f'Dashboard update {today}'], cwd=deploy_dir, check=True, capture_output=True)
            subprocess.run(['git', 'push', 'origin', 'main'], cwd=deploy_dir, check=True, capture_output=True)
            print(f"   Pushed to Sell-Thru-Progress-Dashboard ({today})")
    except subprocess.CalledProcessError as e:
        print(f"   GitHub deploy failed: {e}")
        if e.stderr:
            print(f"   {e.stderr.decode('utf-8', errors='ignore')[:200]}")
    except Exception as e:
        print(f"   GitHub deploy error: {e}")

    # ── 2. Shaker-MD-App 레포 자동 동기화 ──
    # Windows 경로와 Linux 서버 경로 모두 지원
    _win_path = r'C:\Users\J_park\Shaker-MD-App'
    _linux_path = os.path.expanduser('~/Shaker-MD-App')
    shaker_root = _linux_path if os.path.isdir(_linux_path) else _win_path
    shaker_sell_thru_dir = os.path.join(shaker_root, 'docs', 'dashboards', 'sell-thru-progress')
    if not os.path.isdir(os.path.join(shaker_root, '.git')):
        print("\n12. Shaker-MD-App: .git not found — skipping sync")
        return
    print("\n12. Syncing to Shaker-MD-App (sell-thru-progress)...")
    try:
        sync_files = ['data.json', 'index.html', 'memo_data.json', 'rsm_fcst_data.json']
        copied = []
        for f in sync_files:
            src = os.path.join(deploy_dir, f)
            dst = os.path.join(shaker_sell_thru_dir, f)
            if os.path.exists(src) and os.path.isdir(shaker_sell_thru_dir):
                shutil.copy2(src, dst)
                copied.append(f)
        if not copied:
            print("   Nothing to sync (no matching files)")
            return
        # git add (Shaker-MD-App 루트 기준 상대 경로)
        for f in copied:
            rel = os.path.join('docs', 'dashboards', 'sell-thru-progress', f)
            subprocess.run(['git', 'add', rel], cwd=shaker_root, check=True, capture_output=True)
        result = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=shaker_root, capture_output=True)
        if result.returncode == 0:
            print("   Shaker-MD-App: No changes to sync")
        else:
            subprocess.run(['git', 'commit', '-m', f'Sync sell-thru-progress dashboard {today}'], cwd=shaker_root, check=True, capture_output=True)
            subprocess.run(['git', 'push', 'origin', 'main'], cwd=shaker_root, check=True, capture_output=True)
            print(f"   Synced to Shaker-MD-App: {', '.join(copied)}")
    except subprocess.CalledProcessError as e:
        print(f"   Shaker-MD-App sync failed: {e}")
        if e.stderr:
            print(f"   {e.stderr.decode('utf-8', errors='ignore')[:200]}")
    except Exception as e:
        print(f"   Shaker-MD-App sync error: {e}")


# ============================================================
# MAIN
# ============================================================
def main():
    t_start = time.time()
    print("=" * 50)
    print("Sell-Thru Dashboard Refresh")
    print("=" * 50)

    # --- Auto-update Raw from Daily Sell Thru ---
    try:
        updated = update_raw_from_daily()
        if updated:
            # Invalidate 2026 cache so it reloads with new data
            _save_cache('raw_2026', 'INVALIDATED', None)
            print("   2026 cache invalidated for reload")
    except Exception as e:
        print(f"   [WARN] update_raw_from_daily() skipped due to error: {e}")
        updated = False

    # --- Classification (needed for all years) ---
    cls_sig = _file_signature(FILES['classification'])
    cached_cls_sig, cached_cls = _load_cache('classification')
    if cached_cls is not None and (cached_cls_sig == cls_sig or cls_sig is None):
        c24, c25 = cached_cls
        print(f"1. Classification → CACHED ({len(c24)}/{len(c25)} accounts)")
    else:
        c24, c25 = load_classification()
        _save_cache('classification', cls_sig, (c24, c25))

    # --- 2024 Data (cache by source file + classification) ---
    sig_24 = f"{_file_signature(FILES['2024_raw'])}|{cls_sig}"
    cached_sig_24, cached_r24 = _load_cache('raw_2024')
    if cached_r24 is not None and (cached_sig_24 == sig_24 or _file_signature(FILES['2024_raw']) is None):
        r24 = cached_r24
        print(f"2. 2024 Raw → CACHED ({len(r24)} daily rows)")
    else:
        r24 = load_2024(c24, c25)
        _save_cache('raw_2024', sig_24, r24)

    # Override hash (cache bust when TEAM_OVERRIDE or SME_EMPLOYEES or qty rules change)
    ovr_hash = (str(sorted(TEAM_OVERRIDE.items())) + str(sorted(SME_EMPLOYEES))
                + str(sorted(STATUS_OVERRIDE.items()))
                + str(sorted(ZERO_QTY_CATS)) + str(sorted(ZERO_QTY_MATERIALS))
                + str(sorted(HALF_QTY_CATS)))

    # --- 2025 Data ---
    sig_25 = f"{_file_signature(FILES['2025_raw'])}|{cls_sig}|{ovr_hash}"
    cached_sig_25, cached_r25 = _load_cache('raw_2025')
    if cached_r25 is not None and (cached_sig_25 == sig_25 or _file_signature(FILES['2025_raw']) is None):
        r25 = cached_r25
        print(f"3. 2025 Raw → CACHED ({len(r25)} rows)")
    else:
        r25 = load_raw_2025_2026(FILES['2025_raw'], 'RAW', c24, c25, '2025')
        _save_cache('raw_2025', sig_25, r25)

    # --- Dealer Mapping 2026 (team 폴백용) ---
    dealer_map_2026 = _load_dealer_map_2026()

    # --- 2026 Data (always check, most likely to change) ---
    sig_26 = f"{_file_signature(FILES['2026_raw'])}|{cls_sig}|{ovr_hash}|{_file_signature(DEALER_MAPPING_2026)}"
    cached_sig_26, cached_r26 = _load_cache('raw_2026')
    if cached_r26 is not None and (cached_sig_26 == sig_26 or _file_signature(FILES['2026_raw']) is None):
        r26 = cached_r26
        print(f"3. 2026 Raw → CACHED ({len(r26)} rows)")
    else:
        r26 = load_raw_2025_2026(FILES['2026_raw'], 'Raw', c24, c25, '2026', dealer_map=dealer_map_2026)
        _save_cache('raw_2026', sig_26, r26)

    df = pd.DataFrame(r24 + r25 + r26)

    # ---- Merge duplicate Account Names → single representative ID ----
    df = merge_duplicate_accounts(df)

    df_agg = df.groupby(['Year', 'Month', 'Day', 'Account_ID', 'Account_Name', 'Team', 'Category']).agg(
        Value=('Value', 'sum'), Quantity=('Quantity', 'sum')).reset_index()

    status, t24, t25, t26, all_accts = classify_accounts(df_agg, c24)
    df_final, dm = save_excel(df_agg, status, t24, t25, t26, all_accts, c24, c25)

    # Build id_remap for OUD (same logic as merge_duplicate_accounts)
    name_ids = {}
    for _, r in df.iterrows():
        nm = str(r['Account_Name']).strip() if r['Account_Name'] else ''
        if not nm: continue
        aid = r['Account_ID']
        val = float(r['Value']) if r['Value'] else 0
        if nm not in name_ids: name_ids[nm] = {}
        name_ids[nm][aid] = name_ids[nm].get(aid, 0) + val
    id_remap = {}
    for nm, ids_vals in name_ids.items():
        if len(ids_vals) <= 1: continue
        rep_id = max(ids_vals, key=lambda x: ids_vals[x])
        for aid in ids_vals:
            if aid != rep_id: id_remap[aid] = rep_id

    oud_data = load_oud(id_remap)
    ar_data = load_ar_overdue(id_remap)
    # Merge AR name remap into id_remap for Collection
    col_remap = dict(id_remap)
    if ar_data and '_ar_remap' in ar_data:
        col_remap.update(ar_data['_ar_remap'])
    col_data = load_collection(col_remap)
    pgi_data = load_pgi_remain_open(id_remap)
    pgi_result = pgi_data['pgi'] if pgi_data else None
    remain_result = pgi_data['remain'] if pgi_data else None
    open_result = pgi_data['open'] if pgi_data else None
    # Get latest daily file date for YTD filter (filename-based, not transaction-based)
    daily_file_date = None
    if os.path.isdir(DAILY_DIR):
        daily_codes = []
        for f in os.listdir(DAILY_DIR):
            if f.endswith('.xlsx') and not f.startswith('~$') and 'Sell thru' in f:
                m_code = re.search(r'(\d{6})', f)
                if m_code:
                    daily_codes.append(m_code.group(1))
        if daily_codes:
            daily_codes.sort()
            lc = daily_codes[-1]
            daily_file_date = f"{lc[2:4]}-{lc[4:6]}"  # MM-DD
            print(f"   Sell Thru YTD date (from filename): {daily_file_date}")

    ir_target = load_ir_target()
    # RSM FCST: 월별 Excel → rsm_fcst_data.json 자동 생성
    rsm_result = load_rsm_fcst()
    # OR RSM FCST: 00. OR 폴더 → 리테일러별 데이터 로딩 후 병합
    or_rsm_result = load_or_rsm_fcst()
    if rsm_result and or_rsm_result:
        # value, qty, models, names 병합 (OR 계정은 IR 계정과 겹치지 않음)
        for k in ('value', 'qty', 'names'):
            rsm_result.setdefault(k, {}).update(or_rsm_result.get(k, {}))
        for model_key, model_val in or_rsm_result.get('models', {}).items():
            if model_key in rsm_result.get('models', {}):
                rsm_result['models'][model_key]['qty'] += model_val['qty']
                rsm_result['models'][model_key]['val'] += model_val['val']
            else:
                rsm_result.setdefault('models', {})[model_key] = model_val
        print(f"   RSM FCST merged: IR+OR total {len(rsm_result.get('qty', {}))} accounts")
    elif or_rsm_result and not rsm_result:
        rsm_result = or_rsm_result
    # rsm_fcst_data.json 저장 (IR+OR 병합 후)
    if rsm_result:
        rsm_json_path = os.path.join(DEPLOY_DIR, 'rsm_fcst_data.json')
        with open(rsm_json_path, 'w', encoding='utf-8') as f:
            json.dump(rsm_result, f, ensure_ascii=False, separators=(',', ':'))
        print(f"   Saved rsm_fcst_data.json (merged): {len(rsm_result.get('qty', {}))} accounts")
    # RSM FCST 계정 중 마스터(dm)에 없는 계정 자동 추가
    if rsm_result and not dm.empty:
        existing_ids = set(str(int(v)) if isinstance(v, float) else str(v)
                          for v in dm['Account_ID'].dropna())
        rsm_missing = []
        for acc_id, qty_v in rsm_result.get('qty', {}).items():
            if str(acc_id) not in existing_ids and qty_v > 0:
                acc_name_v = rsm_result.get('names', {}).get(acc_id, f'Account {acc_id}')
                # OR 계정 여부 판단 (OR RSM FCST에서 온 계정은 Team=OR)
                _or_acc_ids = set(_OR_RETAILER_ACC.values())
                team_v = 'OR' if str(acc_id) in _or_acc_ids else 'IR'
                rsm_missing.append({
                    'Account_ID': float(acc_id) if str(acc_id).isdigit() else acc_id,
                    'Account_Name': acc_name_v,
                    'Team': team_v,     # IR or OR 자동 분류
                    'Account_Status': 'Active',
                    'Value_2024': 0, 'Value_2025': 0, 'Value_2026': 0,
                    'Growth_24_25': None, 'Growth_25_26': None,
                })
        if rsm_missing:
            print(f"   RSM FCST: {len(rsm_missing)}개 누락 계정을 마스터에 추가")
            for r in rsm_missing:
                print(f"     + {r['Account_ID']} / {r['Account_Name']}")
            dm = pd.concat([dm, pd.DataFrame(rsm_missing)], ignore_index=True)
    save_html(df_final, dm, oud_data, ar_data, col_data, pgi_result, remain_result, open_result, sell_thru_date=daily_file_date, ir_target=ir_target, rsm_fcst=rsm_result)

    elapsed = time.time() - t_start
    print("\n" + "=" * 50)
    print(f"DONE! ({elapsed:.1f}s)")
    for yr in [2024, 2025, 2026]:
        sub = df_final[df_final['Year'] == yr]
        print(f"  {yr}: {sub['Account_ID'].nunique()} accounts / Value={sub['Value'].sum():,.0f} / Qty={sub['Quantity'].sum():,.0f}")

    # SME check
    sme = df_final[df_final['Team'] == 'SME']
    print(f"\n  SME transactions: {len(sme)} rows, {sme['Account_ID'].nunique()} accounts")
    print("=" * 50)

    # Auto-deploy to GitHub Pages
    deploy_to_github()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n!! ERROR: {e}")
        import traceback
        traceback.print_exc()
    try:
        input("\nPress Enter to close...")
    except EOFError:
        pass  # 자동화 모드에서는 입력 없이 종료
