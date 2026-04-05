#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SWSG AC 스크래퍼 v12.0
- 대대적 리팩토링
- Category / Sub-category / BTU / Ton / Compressor / Mode 파싱 추가
- amlabel 기반 Free Installation / Cashback 정확 파싱
- Rotary = On-Off 통일
- BTU -> Ton 변환표 적용
"""

import asyncio
import re
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from playwright.async_api import async_playwright

# Windows terminal UTF-8 force setting
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass


# ── BTU → Ton 변환표 ─────────────────────────────────────────────────────────
BTU_TON_TABLE = [
    ( 9000,  15000, 1.0),
    (15000,  21000, 1.5),
    (21000,  27000, 2.0),
    (27000,  33000, 2.5),
    (33000,  39000, 3.0),
    (39000,  45000, 3.5),
    (45000,  51000, 4.0),
    (51000,  57000, 4.5),
    (57000,  63000, 5.0),
]

def btu_to_ton(btu):
    try:
        btu = int(btu)
    except (TypeError, ValueError):
        return ''
    for low, high, ton in BTU_TON_TABLE:
        if low <= btu < high:
            return ton
    return ''


# ── 브랜드 목록 (긴 이름 먼저) ─────────────────────────────────────────────
BRANDS = [
    'GENERAL ELECTRIC', 'GENERAL DAN', 'SUPER GENERAL', 'TECHNO BEST',
    'NIKAI', 'DANSAT', 'GREE', 'HAIER', 'HISENSE', 'ARISTON',
    'PANASONIC', 'DAIKIN', 'SAMSUNG', 'CARRIER', 'FISHER',
    'LG', 'TCL', 'MIDEA', 'MANDO', 'ARROW', 'BASIC', 'PLATINUM',
    'WANSA', 'SANCO', 'TECHNO',
]


class SWSGACScraper:

    def __init__(self):
        self.products = []
        self.ac_types = {'Split': [], 'Window': [], 'Freestanding': []}
        self.log_lines = []
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        self.urls = {
            'Split':       'https://swsg.co/en/air-conditions/split-acs.html?product_list_limit=320',
            'Window':      'https://swsg.co/en/air-conditions/window-acs.html?product_list_limit=320',
            'Freestanding':'https://swsg.co/en/air-conditions.html?a_ac_type=745&product_list_limit=320',
        }

    # ── 로그 ─────────────────────────────────────────────────────────────────
    def log(self, msg):
        print(msg)
        self.log_lines.append(msg)

    def save_log(self):
        filename = Path(__file__).parent / f"swsg_debug_{self.timestamp}.txt"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_lines))
        print(f"\n[LOG] Log saved: {filename.name}")

    # ── 제품명 파싱 ───────────────────────────────────────────────────────────
    def _parse_name(self, name: str) -> dict:
        result = {'brand': 'N/A', 'btu': '', 'ton': '', 'compressor': '', 'mode': ''}
        upper = name.upper()

        # Brand
        for b in BRANDS:
            if b in upper:
                result['brand'] = b.title()
                break

        # BTU  (e.g. "12,000 BTU" / "18000BTU" / "18,300 BTU")
        m = re.search(r'(\d{1,2}[,.]?\d{3})\s*BTU', name, re.IGNORECASE)
        if m:
            btu_int = int(m.group(1).replace(',', '').replace('.', ''))
            result['btu'] = btu_int
            result['ton'] = btu_to_ton(btu_int)

        # Compressor Type  (Rotary = On-Off)
        if re.search(r'\bINVERTER\b', upper):
            result['compressor'] = 'Inverter'
        else:
            result['compressor'] = 'On-Off'   # covers Rotary as well

        # Cooling / Heating Mode
        hc_patterns = [
            r'HOT\s*[&/]\s*C(?:OLD|OOL)',
            r'HOT\s+AND\s+COLD',
            r'HOT\s*/\s*COLD',
            r'COOLING\s+AND\s+HEATING',
            r'H(?:EAT|OT)\s*[&/]\s*C(?:OLD|OOL)',
        ]
        if any(re.search(p, upper) for p in hc_patterns):
            result['mode'] = 'H&C'
        else:
            result['mode'] = 'CO'

        return result

    # ── Sub-category (Split 전용) ──────────────────────────────────────────
    def _subcategory(self, name: str, ac_type: str) -> str:
        if ac_type != 'Split':
            return ac_type
        upper = name.upper()
        if any(k in upper for k in ('FLOOR', 'STANDING', 'FLOOR-STANDING')):
            return 'Floor Standing'
        if 'CASSETTE' in upper:
            return 'Cassette Split'
        return 'Wall Mount Split'

    # ── 메인 스크래핑 ─────────────────────────────────────────────────────
    async def scrape_all(self):
        self.log('\n' + '='*70)
        self.log('[START] SWSG Air Conditioner Scraper v12.0')
        self.log(f'Start: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        self.log('='*70)

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=['--disable-blink-features=AutomationControlled',
                      '--disable-dev-shm-usage', '--no-sandbox',
                      '--disable-gpu', '--no-zygote']
            )
            context = await browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                locale='en-US',
                extra_http_headers={
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.9',
                }
            )
            await context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'plugins',  { get: () => [1,2,3,4,5] });
                window.chrome = { runtime: {} };
            """)
            page = await context.new_page()

            try:
                # 메인 페이지 — 쿠키 획득용, 타임아웃 관대하게 처리
                self.log('\n[BROWSER] Accessing main page (cookie warm-up)...')
                try:
                    resp = await page.goto(
                        'https://swsg.co/en/',
                        wait_until='domcontentloaded',
                        timeout=90000
                    )
                    self.log(f'   [OK] Status: {resp.status if resp else "None"}')
                except Exception as warm_err:
                    self.log(f'   [WARN] Main page timeout (continuing anyway): {warm_err}')

                for _ in range(3):
                    await page.evaluate('window.scrollBy(0, 500)')
                    await asyncio.sleep(0.5)
                await page.evaluate('window.scrollTo(0, 0)')
                await asyncio.sleep(2)

                for ac_type in ['Split', 'Window', 'Freestanding']:
                    await self._scrape_type(page, ac_type)

            except Exception as e:
                self.log(f'\n[ERROR] {e}')
                import traceback
                self.log(traceback.format_exc())
            finally:
                await browser.close()

        self._print_summary()
        self.save_log()

    # ── 타입별 스크래핑 ───────────────────────────────────────────────────
    async def _scrape_type(self, page, ac_type: str):
        self.log(f'\n{"="*70}')
        self.log(f'[{ac_type}] Scraping...')
        self.log(f'   URL: {self.urls[ac_type]}')
        self.log(f'{"="*70}')

        try:
            await page.set_extra_http_headers({'Referer': page.url})
            resp = await page.goto(self.urls[ac_type], wait_until='domcontentloaded', timeout=90000)
            # 동적 요소 추가 로딩 대기
            try:
                await page.wait_for_selector('.item-product', timeout=30000)
            except Exception:
                self.log('   [WARN] .item-product selector timeout — proceeding anyway')
            self.log(f'   [OK] Status: {resp.status if resp else "None"}')

            if 'search' in page.url:
                self.log('   [SKIP] Redirected to search page — skipping')
                return

            await asyncio.sleep(5)

            # lazy-load 트리거용 스크롤
            for _ in range(25):
                await page.evaluate('window.scrollBy(0, 500)')
                await asyncio.sleep(0.3)
            await page.evaluate('window.scrollTo(0, 0)')
            await asyncio.sleep(2)

            # ── JS 추출 ──────────────────────────────────────────────────
            raw_list = await page.evaluate("""
                () => {
                    const results = [];
                    for (const item of document.querySelectorAll('.item-product')) {

                        // 제품명 + URL
                        const nameEl = item.querySelector('a.product-item-link');
                        const name = nameEl ? nameEl.textContent.trim() : '';
                        if (!name || name.length < 5) continue;
                        const productUrl = nameEl ? nameEl.href : '';

                        // Product ID
                        const pbEl = item.querySelector('[data-product-id]');
                        const productId = pbEl ? pbEl.dataset.productId : '';

                        // 현재가 / 정상가
                        // 할인 제품: .special-price (할인가) + .old-price (정상가)
                        // 정가 제품: .special-price 없음 → data-price-type="finalPrice" 사용
                        const spEl  = item.querySelector('.special-price .price');
                        const regEl = item.querySelector('.old-price .price, .regular-price .price');
                        let specialPrice = spEl ? spEl.textContent.replace(/[^0-9]/g, '') : '';
                        let regularPrice = regEl ? regEl.textContent.replace(/[^0-9]/g, '') : '';
                        if (!specialPrice) {
                            // 정가 제품: data-price-amount에서 finalPrice 직접 추출
                            const fpEl = item.querySelector('[data-price-type="finalPrice"][data-price-amount]');
                            if (fpEl) specialPrice = fpEl.dataset.priceAmount;
                        }

                        // 할인율: .hot-onsale 텍스트에서 파싱 ("Discount -42.08%")
                        const hotSaleEl = item.querySelector('.hot-onsale');
                        const hotSaleText = hotSaleEl ? hotSaleEl.textContent.trim().replace(/\\s+/g, ' ') : '';
                        const discMatch = hotSaleText.match(/(-\\d+\\.?\\d*%)/);
                        const discount = discMatch ? discMatch[1] : '';

                        // amlabel 기반 프로모션 파싱
                        // 형태1: "Cash Back 100 SAR 2,549"       (고정금액, 별도 wrapper)
                        // 형태2: "Free Installation + 10% Cashback 2,699" (하나의 wrapper에 합산)
                        let freeInstall   = 'No';
                        let cashbackText  = '';

                        const wrappers = item.querySelectorAll('[class*="amlabel-position"][class*="-prod"]');
                        for (const w of wrappers) {
                            const t = w.textContent.trim().replace(/\\s+/g, ' ');
                            // Free Install 감지 (else if 제거 → 독립 체크)
                            if (/free installation/i.test(t)) {
                                freeInstall = /within riyadh/i.test(t) ? 'Within Riyadh' : 'Yes';
                            }
                            // Cashback 감지 (Free Install과 동시에 있을 수 있음)
                            if (/cash\\s*back/i.test(t)) {
                                cashbackText = t;
                            }
                        }

                        // Cashback 금액 & 적용가 파싱
                        let cashbackAmount = '';
                        let cashbackPrice  = '';
                        if (cashbackText) {
                            // 형태1: "Cash Back 100 SAR ..."  → 고정금액
                            const fixedM = cashbackText.match(/cash\\s*back\\s+([\\d,]+)\\s*SAR/i);
                            // 형태2: "10% Cashback ..."         → 비율
                            const pctM   = cashbackText.match(/(\\d+)\\s*%\\s*cash\\s*back/i);
                            if (fixedM)      cashbackAmount = fixedM[1].replace(/,/g, '') + ' SAR';
                            else if (pctM)   cashbackAmount = pctM[1] + '%';
                            // 적용가: 텍스트 맨 끝 숫자
                            const prM = cashbackText.match(/([\\d,]+)\\s*$/);
                            if (prM) cashbackPrice = prM[1].replace(/,/g, '');
                        }

                        // 재고
                        const stockEl = item.querySelector('.stock, [class*="availability"]');
                        let stock = 'In Stock';
                        if (stockEl) {
                            const t = stockEl.textContent.toLowerCase();
                            if (t.includes('out') || t.includes('notify')) stock = 'Out of Stock';
                        }

                        results.push({
                            name, productId, productUrl,
                            specialPrice, regularPrice, discount,
                            freeInstall, cashbackAmount, cashbackPrice,
                            stock,
                        });
                    }
                    return results;
                }
            """)

            self.log(f'   [OK] Cards found: {len(raw_list)}')

            for raw in raw_list:
                parsed  = self._parse_name(raw['name'])
                subcat  = self._subcategory(raw['name'], ac_type)

                # Final Price 계산:
                # - Cashback 제품: amlabel에서 직접 추출한 cashbackPrice 사용
                # - 일반 제품:     현재 판매가(specialPrice) 그대로 사용
                final_price = raw['cashbackPrice'] if raw['cashbackPrice'] else raw['specialPrice']

                product = {
                    'number':         len(self.products) + 1,
                    'type':           ac_type,
                    'subcategory':    subcat,
                    'product_id':     raw['productId'],
                    'product_url':    raw['productUrl'],
                    'name':           raw['name'],
                    'brand':          parsed['brand'],
                    'btu':            parsed['btu'],
                    'ton':            parsed['ton'],
                    'compressor':     parsed['compressor'],
                    'mode':           parsed['mode'],
                    'price':          raw['specialPrice'],
                    'original_price': raw['regularPrice'],
                    'discount':       raw['discount'],
                    'free_install':   raw['freeInstall'],
                    'cashback_amount':raw['cashbackAmount'],
                    'final_price':    final_price,
                    'stock':          raw['stock'],
                }
                self.ac_types[ac_type].append(product)
                self.products.append(product)

            in_stock = sum(1 for p in self.ac_types[ac_type] if p['stock'] == 'In Stock')
            self.log(f'   [DONE] {ac_type}: total={len(self.ac_types[ac_type])}, in-stock={in_stock}')
            await asyncio.sleep(3)

        except Exception as e:
            self.log(f'   [ERROR] {e}')
            import traceback
            self.log(traceback.format_exc())

    # ── 요약 출력 ─────────────────────────────────────────────────────────
    def _print_summary(self):
        self.log(f'\n{"="*70}')
        self.log('[SUMMARY] Collection Results')
        self.log(f'{"="*70}')
        self.log(f'   Total products : {len(self.products)}')
        for ac_type in ['Split', 'Window', 'Freestanding']:
            prods    = self.ac_types[ac_type]
            in_stock = sum(1 for p in prods if p['stock'] == 'In Stock')
            cashback = sum(1 for p in prods if p['cashback_amount'])
            free_ins = sum(1 for p in prods if p['free_install'] != 'No')
            self.log(f'   {ac_type:12s}: {len(prods):3d}  (In-Stock: {in_stock}, Free Install: {free_ins}, Cashback: {cashback})')

    # ── Excel 저장 (누적 Master 방식) ─────────────────────────────────────
    def save_to_excel(self):
        if not self.products:
            self.log('\n[ERROR] No products to save.')
            return None

        self.log(f'\n{"="*70}')
        self.log('[EXCEL] Saving to Master Excel...')
        self.log(f'{"="*70}')

        MASTER_FILE = Path(__file__).parent / 'SWS_AC_Price_Tracking_Master.xlsx'

        HEADERS = [
            'Timestamp', 'Type', 'Sub-Category', 'Product ID', 'Product Name',
            'Brand', 'Capacity (BTU)', 'Capacity (Ton)',
            'Compressor', 'Mode',
            'Price (SAR)', 'Original Price (SAR)', 'Discount',
            'Free Install', 'Cashback', 'Final Price (SAR)',
            'Stock', 'Product_URL',
        ]
        COL_WIDTHS = {
            'A': 20, 'B': 12, 'C': 18, 'D': 12, 'E': 75,
            'F': 18, 'G': 16, 'H': 16,
            'I': 12, 'J': 8,
            'K': 14, 'L': 18, 'M': 10,
            'N': 22, 'O': 14, 'P': 18,
            'Q': 12, 'R': 60,
        }

        # 스타일
        hdr_fill    = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        hdr_font    = Font(bold=True, color='FFFFFF')
        green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        teal_fill   = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        COL = {h: i+1 for i, h in enumerate(HEADERS)}

        # ── 파일 존재 여부에 따라 로드 또는 신규 생성 ──────────────────
        if MASTER_FILE.exists():
            wb = load_workbook(MASTER_FILE)
            if 'Products_DB' in wb.sheetnames:
                ws = wb['Products_DB']
                self.log(f'   [APPEND] Existing master loaded — last row: {ws.max_row}')
            else:
                ws = wb.create_sheet('Products_DB')
                ws.append(HEADERS)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                for letter, width in COL_WIDTHS.items():
                    ws.column_dimensions[letter].width = width
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products_DB'
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
            for letter, width in COL_WIDTHS.items():
                ws.column_dimensions[letter].width = width
            self.log('   [NEW] Master file created.')

        # ── 데이터 행 추가 ──────────────────────────────────────────────
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for p in self.products:
            price_val = int(p['price'])          if p['price']          else ''
            orig_val  = int(p['original_price']) if p['original_price'] else ''
            final_val = int(p['final_price'])    if p['final_price']    else ''
            ws.append([
                ts, p['type'], p['subcategory'], p['product_id'],
                p['name'], p['brand'], p['btu'], p['ton'],
                p['compressor'], p['mode'],
                price_val, orig_val, p['discount'],
                p['free_install'], p['cashback_amount'], final_val,
                p['stock'], p.get('product_url', ''),
            ])
            row = ws.max_row
            ws.cell(row, COL['Stock']).fill = green_fill if p['stock'] == 'In Stock' else red_fill
            if p['free_install'] != 'No':
                ws.cell(row, COL['Free Install']).fill = teal_fill
            if p['cashback_amount']:
                ws.cell(row, COL['Cashback']).fill          = yellow_fill
                ws.cell(row, COL['Final Price (SAR)']).fill = yellow_fill

        try:
            wb.save(MASTER_FILE)
            self.log(f'\n[SAVED] {MASTER_FILE.name}')
            self.log(f'   Products added this run : {len(self.products)}')
            self.log(f'   Total rows in sheet     : {ws.max_row - 1}')
            return str(MASTER_FILE)
        except Exception as e:
            self.log(f'[ERROR] Save failed: {e}')
            return None


# ── 진입점 ────────────────────────────────────────────────────────────────────
async def main():
    scraper = SWSGACScraper()
    try:
        await scraper.scrape_all()
        scraper.save_to_excel()
    except Exception as e:
        import traceback
        print(f'\n[FATAL ERROR] {e}')
        print(traceback.format_exc())
    finally:
        try:
            input('\nPress Enter to exit...')
        except (EOFError, OSError):
            pass

if __name__ == '__main__':
    asyncio.run(main())
