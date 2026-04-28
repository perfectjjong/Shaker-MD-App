"""
BH AC Price Dashboard — Dual Channel (Wholesale + Retail)
=========================================================
Input:  BH_Subdealer_AC_Master.xlsx
        - Total_Model Info   : 모델 메타데이터
        - Whole selling Price_Master : 도매가 가격 추적
        - Weekly_Price_DB     : 소매가 (BH Store API)
Output: bh_ac_dashboard_v2.html  (self-contained)
"""

import os, json, sys, shutil, subprocess
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ── BH Store API (for product ID mapping) ─────────────────────────
API_URL = "https://api.bhstore.com.sa/commerce/products/"
API_KEY = "2853152294a192f18c3da51ae965f"
API_HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "accept": "application/json",
    "origin": "https://bhstore.com.sa",
    "referer": "https://bhstore.com.sa/sa-en",
    "interface": "web",
    "language": "en",
    "x-api-key": API_KEY,
}
AC_CATEGORIES = {"Split": 12, "Window": 7, "Floor Standing": 21, "Concealed": 15, "Cassette": 18}

def fetch_product_id_map():
    """Fetch model_code → product_id mapping from BH Store API."""
    if not HAS_REQUESTS:
        return {}
    id_map = {}
    for cat_name, cat_id in AC_CATEGORIES.items():
        try:
            r = requests.get(f"{API_URL}?limit=300&category_id={cat_id}",
                             headers=API_HEADERS, timeout=30)
            data = r.json()
            for p in data.get("data", []):
                code = (p.get("code") or "").strip()
                pid = p.get("id")
                if code and pid:
                    id_map[code] = pid
        except Exception:
            pass
    return id_map

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "BH_Subdealer_AC_Master.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "bh_ac_dashboard_v2.html")
SHAKER_DIR = os.path.join(os.path.expanduser("~"), "Shaker-MD-App")
CLOUDFLARE_DEST = os.path.join(SHAKER_DIR, "docs", "dashboards", "bh-price")

# ── Brand colors ─────────────────────────────────────────────────
BRAND_COLORS = {
    "LG": "#A50034", "Samsung": "#1428A0", "SAMSUNG": "#1428A0",
    "Carrier": "#E4002B", "Gree": "#00843D", "Midea": "#0073CF",
    "Haier": "#E30613", "Hisense": "#00A5E0", "TCL": "#2D6DB5",
    "Craft": "#FF8C00", "Dura": "#6B4C9A", "Falcon": "#D4AF37",
    "Gibson": "#2E4057", "Zamil": "#006838", "Zamil Classic": "#006838",
    "Zamil Comfort": "#228B22",
    "Kolin": "#E57300", "Kelvinator": "#4169E1", "Mieling": "#8B4513",
    "York": "#B22222", "AUX": "#FF4500", "Admiral": "#483D8B",
    "White Westinghouse": "#4682B4", "Aston": "#708090",
    "Pan Cool": "#20B2AA", "Daya": "#DC143C", "Daewoo": "#CD853F",
    "Skyworth": "#2F4F4F", "Regent": "#8B008B", "General": "#556B2F",
    "General Max": "#6B8E23", "Crony": "#BC8F8F", "Cool Line": "#5F9EA0",
    "Cool Air": "#87CEEB", "Milano": "#C71585", "Al Jazeera": "#FFD700",
    "Al Ghadeer": "#228B22", "Al Kawthar": "#B8860B", "Haas": "#CD5C5C",
    "Uni Hamburg": "#4B0082", "Symphony": "#DA70D6",
}

# ── Brand normalization ──────────────────────────────────────────
BRAND_NORMALIZE = {
    "Kraft": "Craft", "SAMSUNG": "Samsung", "samsung": "Samsung",
    "Zamil Classic": "Zamil", "Zamil Comfort": "Zamil",
    "Dora": "Dura", "Dora Elegant": "Dura",
}

# ── Category normalization (Retail short → full name) ────────────
CATEGORY_NORMALIZE = {
    "Split": "Split AC", "Window": "Window AC",
    "Floor Standing": "Floor Standing AC",
    "Concealed": "Concealed AC", "Cassette": "Cassette AC",
}

def safe(v):
    """Return None for NaN/None, else the value."""
    if v is None:
        return None
    if isinstance(v, float) and (v != v):  # NaN check
        return None
    return v

def norm_brand(b):
    if not b:
        return b
    b = str(b).strip()
    return BRAND_NORMALIZE.get(b, b)

def norm_category(c):
    if not c:
        return c
    c = str(c).strip()
    return CATEGORY_NORMALIZE.get(c, c)

def load_model_info(wb):
    """Load Total_Model Info sheet → dict keyed by Model Code."""
    mi = {}
    if "Total_Model Info" not in wb.sheetnames:
        return mi
    ws = wb["Total_Model Info"]
    headers = [str(c.value or '').strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        mc = str(rd.get("Model Code", "") or "").strip()
        if not mc:
            continue
        mi[mc] = {
            "brand": norm_brand(rd.get("Brand")),
            "category": str(rd.get("Category", "") or "").strip(),
            "btu": safe(rd.get("BTU")),
            "ton": safe(rd.get("Ton")),
            "type": str(rd.get("Type", "") or "").strip(),  # CO / H&C
            "desc": str(rd.get("Product Description", "") or "").strip(),
        }
    return mi

def load_wholesale(wb, model_info):
    """Load Whole selling Price_Master → list of records."""
    records = []
    if "Whole selling Price_Master" not in wb.sheetnames:
        return records, [], set(), set(), set(), set(), set()
    ws = wb["Whole selling Price_Master"]
    headers = [str(c.value or '').strip() for c in ws[1]]
    dates_set = set()
    brands_set = set()
    cats_set = set()
    types_set = set()
    tons_set = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        mc = str(rd.get("Model Code", "") or "").strip()
        if not mc:
            continue
        brand = norm_brand(rd.get("Brand"))
        cat = str(rd.get("Category", "") or "").strip()
        price = safe(rd.get("Price"))
        status = str(rd.get("Status", "") or "").strip()
        update_date = rd.get("Update Date")
        btu = safe(rd.get("BTU"))
        ton = safe(rd.get("Ton"))
        typ = str(rd.get("Type", "") or "").strip()
        desc = str(rd.get("Product Description", "") or "").strip()

        # Enrich from model_info
        mi = model_info.get(mc, {})
        if not brand and mi.get("brand"):
            brand = mi["brand"]
        if not cat and mi.get("category"):
            cat = mi["category"]
        if not desc and mi.get("desc"):
            desc = mi["desc"]
        if btu is None and mi.get("btu"):
            btu = mi["btu"]
        if ton is None and mi.get("ton"):
            ton = mi["ton"]
        if not typ and mi.get("type"):
            typ = mi["type"]

        # Date formatting
        if update_date:
            if hasattr(update_date, 'strftime'):
                d_str = update_date.strftime('%Y-%m-%d')
            else:
                d_str = str(update_date)[:10]
        else:
            continue

        if brand:
            brands_set.add(brand)
        if cat:
            cats_set.add(cat)
        if typ:
            types_set.add(typ)
        if ton is not None:
            try:
                tons_set.add(float(ton))
            except (ValueError, TypeError):
                pass
        dates_set.add(d_str)

        try:
            price_val = float(price) if price is not None else None
        except (ValueError, TypeError):
            price_val = None
        try:
            ton_val = float(ton) if ton is not None else None
        except (ValueError, TypeError):
            ton_val = None
        try:
            btu_val = int(float(btu)) if btu is not None else None
        except (ValueError, TypeError):
            btu_val = None

        records.append({
            'd': d_str,
            'b': brand or '',
            'm': mc,
            'n': (desc or '')[:70],
            'c': cat or '',
            't': ton_val,
            'btu': btu_val,
            'tp': typ or '',
            'p': price_val,
            'st': status,
        })

    dates_list = sorted(dates_set)
    return records, dates_list, brands_set, cats_set, types_set, tons_set

def load_retail(wb, model_info):
    """Load Weekly_Price_DB → list of records.
    Uses Run Timestamp (col R) to extract date (YYYY-MM-DD) as the time axis.
    """
    records = []
    if "Weekly_Price_DB" not in wb.sheetnames:
        return records, [], set(), set(), set(), set()
    ws = wb["Weekly_Price_DB"]
    headers = [str(c.value or '').strip() for c in ws[1]]
    dates_set = set()
    brands_set = set()
    cats_set = set()
    types_set = set()
    tons_set = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        mc = str(rd.get("Model Code", "") or rd.get("Model_Code", "") or "").strip()
        if not mc:
            continue
        brand = norm_brand(rd.get("Brand"))

        # Extract date from Run Timestamp (e.g. "2026-03-15 08:01:39" → "2026-03-15")
        run_ts = rd.get("Run Timestamp") or rd.get("Run_Timestamp") or ""
        date_str = str(run_ts).strip()[:10]  # "YYYY-MM-DD"
        if len(date_str) != 10 or date_str[4] != '-':
            # fallback to Week column
            date_str = str(rd.get("Week", "") or "").strip()
        if not date_str:
            continue

        cat = str(rd.get("Type", "") or "").strip()  # Type column = category in retail
        cat = norm_category(cat)
        co_hc = str(rd.get("CO/C&H", "") or rd.get("CO_CH", "") or "").strip()
        product_name = str(rd.get("Product Name", "") or rd.get("Product_Name", "") or "").strip()
        regular_price = safe(rd.get("Regular Price") or rd.get("Regular_Price"))
        current_price = safe(rd.get("Current Price") or rd.get("Current_Price"))
        discount_pct = safe(rd.get("Discount %") or rd.get("Discount_Pct"))
        discount_sar = safe(rd.get("Discount SAR") or rd.get("Discount_SAR"))
        stock = safe(rd.get("Stock"))
        btu_raw = safe(rd.get("BTU"))
        url = str(rd.get("Product URL") or rd.get("URL") or "").strip()

        # Enrich from model_info
        mi = model_info.get(mc, {})
        if not brand and mi.get("brand"):
            brand = mi["brand"]
        if not cat and mi.get("category"):
            cat = mi["category"]

        ton_val = None
        btu_val = None
        if mi.get("ton") is not None:
            try:
                ton_val = float(mi["ton"])
            except (ValueError, TypeError):
                pass
        if btu_raw is not None:
            try:
                btu_val = int(float(btu_raw))
            except (ValueError, TypeError):
                pass
        elif mi.get("btu") is not None:
            try:
                btu_val = int(float(mi["btu"]))
            except (ValueError, TypeError):
                pass

        typ = co_hc  # CO / C&H / H&C
        if not typ and mi.get("type"):
            typ = mi["type"]
        # Normalize type
        if typ in ("C&H",):
            typ = "H&C"

        if brand:
            brands_set.add(brand)
        if cat:
            cats_set.add(cat)
        if typ:
            types_set.add(typ)
        if ton_val is not None:
            tons_set.add(ton_val)
        dates_set.add(date_str)

        try:
            rp = float(regular_price) if regular_price is not None else None
        except (ValueError, TypeError):
            rp = None
        try:
            cp = float(current_price) if current_price is not None else None
        except (ValueError, TypeError):
            cp = None
        try:
            dp = float(discount_pct) if discount_pct is not None else None
        except (ValueError, TypeError):
            dp = None
        try:
            ds = float(discount_sar) if discount_sar is not None else None
        except (ValueError, TypeError):
            ds = None

        records.append({
            'd': date_str,    # date from Run Timestamp (YYYY-MM-DD)
            'b': brand or '',
            'm': mc,
            'n': (product_name or '')[:70],
            'c': cat or '',
            'tp': typ or '',
            't': ton_val,
            'btu': btu_val,
            'rp': rp,        # regular price
            'cp': cp,        # current price (discounted)
            'dr': dp,        # discount rate (0~1)
            'ds': ds,        # discount SAR
            'url': url,
        })

    dates_list = sorted(dates_set)
    return records, dates_list, brands_set, cats_set, types_set, tons_set


def main():
    print("[1/4] Loading Excel...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    model_info = load_model_info(wb)
    print(f"  → Total_Model Info: {len(model_info)} models")

    ws_data, ws_dates, ws_brands, ws_cats, ws_types, ws_tons = load_wholesale(wb, model_info)
    print(f"  → Wholesale: {len(ws_data)} records, {len(ws_dates)} dates")

    rt_data, rt_dates, rt_brands, rt_cats, rt_types, rt_tons = load_retail(wb, model_info)
    print(f"  → Retail: {len(rt_data)} records, {len(rt_dates)} dates")
    wb.close()

    # Fetch product ID map from BH Store API (for product page links)
    print("  → Fetching BH Store product IDs...")
    product_id_map = fetch_product_id_map()
    print(f"  → Product ID map: {len(product_id_map)} models")

    # Merge dimensions
    all_brands = sorted(ws_brands | rt_brands)
    all_cats = sorted(ws_cats | rt_cats)
    all_types = sorted(ws_types | rt_types)
    all_tons = sorted(ws_tons | rt_tons)

    # Model overlap stats
    ws_models = set(r['m'] for r in ws_data)
    rt_models = set(r['m'] for r in rt_data)
    matched = ws_models & rt_models
    ws_only = ws_models - rt_models
    rt_only = rt_models - ws_models

    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')

    print("[2/4] Building HTML...")

    # ── HTML HEAD ──────────────────────────────────────────────────
    HTML_HEAD = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>BH AC Price Tracker — Dual Channel</title>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<script>
tailwind.config={theme:{extend:{fontFamily:{sans:['Inter','system-ui','sans-serif']},colors:{navy:{50:'#E8F4FD',100:'#D1E9FB',200:'#A3D3F7',500:'#2E75B6',700:'#1E3A5F',800:'#1F4E79',900:'#0F2A42'}}}}}
</script>
<style>
body{font-family:'Inter','system-ui',sans-serif}
.sort-asc::after{content:' \\25B2';font-size:9px}.sort-desc::after{content:' \\25BC';font-size:9px}
.tbl-wrap{overflow-x:auto}.tbl-wrap table{width:100%;border-collapse:collapse;font-size:12px}
.tbl-wrap th{background:#1F4E79;color:#fff;padding:7px 10px;text-align:left;position:sticky;top:0;z-index:2;white-space:nowrap;cursor:pointer;user-select:none;font-weight:600;font-size:11px}
.tbl-wrap th:hover{background:#2E75B6}
.tbl-wrap td{padding:5px 10px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
.tbl-wrap tr:nth-child(even) td{background:#f8fafc}.tbl-wrap tr:hover td{background:#e8f0fe}
.up-cell{color:#dc2626;font-weight:600}.dn-cell{color:#16a34a;font-weight:600}
.level-cat td{background:#dbeafe!important;font-weight:700;color:#1e3a8a;border-left:4px solid #2563eb}
.level-comp td{background:#eff6ff!important;font-weight:600;border-left:4px solid #60a5fa}
.level-hc td{background:#fef9c3!important;font-weight:500;border-left:4px solid #facc15}
.level-ton td{background:#fff!important;color:#475569;border-left:4px solid #d1d5db}
.type-badge{display:inline-block;padding:1px 6px;border-radius:4px;font-size:9px;font-weight:700;letter-spacing:.3px}
.type-cat{background:#2563eb;color:#fff}.type-comp{background:#60a5fa;color:#fff}.type-hc{background:#facc15;color:#713f12}.type-ton{background:#e5e7eb;color:#374151}
.ms-wrap{position:relative;display:inline-block}
.ms-btn{display:flex;align-items:center;gap:4px;padding:3px 10px;border:1px solid #d1d5db;border-radius:6px;background:#fff;font-size:11px;cursor:pointer;white-space:nowrap;transition:all .15s}
.ms-btn:hover{border-color:#2E75B6;background:#f0f7ff}
.ms-menu{display:none;position:absolute;top:100%;left:0;z-index:50;min-width:180px;max-height:280px;background:#fff;border:1px solid #e2e8f0;border-radius:8px;box-shadow:0 8px 25px rgba(0,0,0,.15);margin-top:2px}
.ms-menu.open{display:block}
.ms-menu .ms-actions{display:flex;gap:6px;padding:6px 10px;border-bottom:1px solid #f0f0f0;font-size:10px}
.ms-menu .ms-actions button{color:#2E75B6;font-weight:600;cursor:pointer;background:none;border:none}
.ms-menu .ms-actions button:hover{text-decoration:underline}
.ms-menu .ms-actions button.ms-none{color:#dc2626}
.ms-menu .ms-list{max-height:230px;overflow-y:auto;padding:4px 0}
.ms-menu label{display:flex;align-items:center;gap:6px;padding:3px 10px;cursor:pointer;font-size:11px;transition:background .1s}
.ms-menu label:hover{background:#f1f5f9}
.ms-menu input[type=checkbox]{width:14px;height:14px;border-radius:3px}
.ms-menu label.ms-disabled{opacity:.35;pointer-events:none}
.ms-menu label.ms-disabled span{text-decoration:line-through}
.pt-sel{font-size:11px;border:1px solid #d1d5db;border-radius:6px;padding:3px 8px;background:#fff;cursor:pointer}
.sec-search{font-size:11px;border:1px solid #d1d5db;border-radius:6px;padding:3px 8px;background:#fff;width:160px}
.sec-search:focus{outline:none;border-color:#2E75B6;box-shadow:0 0 0 2px rgba(46,117,182,.15)}
.tab-btn{padding:8px 24px;font-size:13px;font-weight:600;border-radius:8px 8px 0 0;cursor:pointer;transition:all .2s;border:1px solid #d1d5db;border-bottom:none;background:#f3f4f6;color:#6b7280}
.tab-btn.active{background:#1F4E79;color:#fff;border-color:#1F4E79}
.tab-btn:hover:not(.active){background:#e8f0fe;color:#1F4E79}
.tab-panel{display:none}.tab-panel.active{display:block}
@media print{.no-print{display:none!important}}
/* === DARK MODE (sell-thru-progress unified) === */
body{background:#0f172a!important;color:#e2e8f0!important}
header{background:linear-gradient(135deg,#1e293b,#334155)!important;border-bottom:2px solid #3b82f6}
.bg-white,.bg-gray-50{background:#1e293b!important}
.bg-white\/95{background:rgba(30,41,59,.95)!important}
section{background:#1e293b!important;border-color:#334155!important}
.border-gray-100,.border-gray-200,.border-gray-300{border-color:#334155!important}
.text-gray-800,.text-gray-700,.text-gray-600,.text-gray-500{color:#e2e8f0!important}
.text-gray-400{color:#94a3b8!important}
.text-navy-800{color:#60a5fa!important}
.text-navy-700{color:#93c5fd!important}
.bg-navy-800{background:#3b82f6!important}
.bg-navy-50{background:#1e3a5f!important}
.tbl-wrap td{border-bottom-color:#334155!important;color:#e2e8f0!important}
.tbl-wrap tr:nth-child(even) td{background:#1e293b!important}
.tbl-wrap tr:nth-child(odd) td{background:#0f172a!important}
.tbl-wrap tr:hover td{background:#334155!important}
.tbl-wrap th{background:#334155!important;color:#ffffff!important}
.tbl-wrap th:hover{background:#475569!important}
.tbl-wrap a{color:#60a5fa!important}
a.text-blue-600{color:#60a5fa!important}
.level-cat td{background:#1e3a5f!important;color:#60a5fa!important;border-left-color:#3b82f6!important}
.level-comp td{background:#172554!important;color:#93c5fd!important;border-left-color:#60a5fa!important}
.level-hc td{background:#422006!important;color:#fbbf24!important;border-left-color:#f59e0b!important}
.level-ton td{background:#0f172a!important;color:#94a3b8!important;border-left-color:#475569!important}
/* Filter readability - dark buttons */
.ms-btn{background:#0f172a!important;color:#f1f5f9!important;border-color:#64748b!important;font-size:12px!important;font-weight:500!important}
.ms-btn:hover{border-color:#60a5fa!important;background:#1e293b!important}
.ms-btn b{color:#93c5fd!important}
.ms-menu{background:#1e293b!important;border-color:#475569!important;box-shadow:0 8px 25px rgba(0,0,0,.4)!important}
.ms-menu .ms-actions{border-bottom-color:#334155!important}
.ms-menu .ms-actions button{color:#60a5fa!important}
.ms-menu .ms-actions button.ms-none{color:#f87171!important}
.ms-menu label{color:#e2e8f0!important}
.ms-menu label:hover{background:#334155!important}
.pt-sel,.sec-search{background:#0f172a!important;color:#e2e8f0!important;border-color:#475569!important}
.sec-search:focus{border-color:#3b82f6!important;box-shadow:0 0 0 2px rgba(59,130,246,.15)!important}
.shadow-sm{box-shadow:0 2px 8px rgba(0,0,0,.3)!important}
a[class*="bg-white"]{background:#1e293b!important;color:#94a3b8!important;border-color:#334155!important}
a[class*="bg-white"]:hover{background:#334155!important}
.bg-gray-100{background:#334155!important;color:#e2e8f0!important}
.bg-gray-100:hover,.bg-gray-200{background:#475569!important}
.bg-gray-700{background:#334155!important}
.bg-green-600{background:#10b981!important}
.text-green-700,.text-green-600{color:#34d399!important}
.text-red-700,.text-red-600{color:#f87171!important}
.text-purple-600{color:#a78bfa!important}
nav .flex .px-3.py-1{border-color:#334155!important}
.w-px{background:#334155!important}
input[type=checkbox]{accent-color:#3b82f6}
select{background:#0f172a!important;color:#e2e8f0!important;border-color:#475569!important}
/* KPI card backgrounds */
.bg-amber-50{background:#2d1f05!important}
.bg-red-50{background:#2d0f0f!important}
.bg-green-50{background:#0a2618!important}
.bg-blue-50{background:#0f1d3d!important}
.bg-orange-50{background:#2d1507!important}
.bg-teal-50{background:#0a2625!important}
.bg-purple-50{background:#1f0a3d!important}
.bg-cyan-50{background:#0a2833!important}
.bg-indigo-50{background:#1a1840!important}
.bg-pink-50{background:#2d0a1a!important}
/* KPI card text colors */
.text-amber-700,.text-amber-600{color:#fbbf24!important}
.text-blue-600{color:#60a5fa!important}
.text-orange-600,.text-orange-700{color:#fb923c!important}
.text-teal-600,.text-teal-700{color:#2dd4bf!important}
.text-purple-700{color:#c4b5fd!important}
.text-cyan-600{color:#22d3ee!important}
/* Card borders */
.border-green-200{border-color:#166534!important}
.border-red-200{border-color:#991b1b!important}
.border-amber-200{border-color:#92400e!important}
/* Stock badges */
.stk-ok{background:#052e16!important;color:#4ade80!important}
.stk-high{background:#172554!important;color:#60a5fa!important}
.stk-low{background:#422006!important;color:#fbbf24!important}
.stk-critical{background:#450a0a!important;color:#fca5a5!important}
.stk-out{background:#3b0a0a!important;color:#fca5a5!important}
/* Cashback/Install badges */
.cb-yes{background:#052e16!important;color:#4ade80!important}
.cb-no{background:#334155!important;color:#94a3b8!important}
.fi-yes{background:#172554!important;color:#60a5fa!important}
.fi-riyadh{background:#422006!important;color:#fbbf24!important}
.fi-no{background:#334155!important;color:#94a3b8!important}
/* Up/Down cell contrast */
.up-cell{color:#f87171!important;font-weight:700}
.dn-cell{color:#4ade80!important;font-weight:700}
/* Nav pills */
nav a.rounded-full{color:#94a3b8!important;border-color:#475569!important}
nav a.rounded-full:hover{background:#334155!important;color:#e2e8f0!important}
nav a.bg-navy-800.rounded-full{color:#fff!important}
/* Section filter bars */
.text-\[10px\].font-bold.text-gray-400.uppercase{color:#64748b!important}
/* New/Disc card text */
.bg-green-50 .text-gray-700,.bg-red-50 .text-gray-700{color:#f1f5f9!important}
.bg-green-50 .text-gray-400,.bg-red-50 .text-gray-400{color:#94a3b8!important}
.bg-green-50,.bg-red-50{color:#e2e8f0!important}
.bg-green-50 a,.bg-red-50 a{color:#93c5fd!important}
#newCards span[style*="color"],#discCards span[style*="color"]{color:#e2e8f0!important;text-shadow:none}
/* Section heading */
.border-navy-800{border-color:#3b82f6!important}
.text-sm.font-bold.text-gray-700{color:#f1f5f9!important}
/* Mode buttons */
.dir-btn.active,.s5agg-btn.active,.agg-btn.active{background:#3b82f6!important;color:#fff!important}
.dir-btn:not(.active),.s5agg-btn:not(.active),.agg-btn:not(.active){background:#1e293b!important;color:#94a3b8!important;border-color:#475569!important}
/* Type badges */
.type-cat{background:#3b82f6!important}.type-comp{background:#60a5fa!important}.type-hc{background:#f59e0b!important;color:#422006!important}.type-ton{background:#475569!important;color:#e2e8f0!important}
/* Sticky filter bar */
.sticky.top-0{background:rgba(15,23,42,.95)!important;border-bottom-color:#334155!important}
/* Promo badge */
.promo-badge{background:#422006!important;color:#fbbf24!important}
</style></head>
"""

    # ── HTML BODY ──────────────────────────────────────────────────
    HTML_BODY = """<body class="bg-gray-50 font-sans text-gray-800 text-sm">
<!-- Header -->
<div class="bg-gradient-to-r from-navy-900 via-navy-800 to-navy-700 text-white py-4 px-6 shadow-lg no-print">
  <div class="max-w-[1600px] mx-auto flex items-center justify-between flex-wrap gap-3">
    <div>
      <h1 class="text-xl font-bold tracking-tight">BH AC Price Tracker</h1>
      <p class="text-navy-200 text-xs mt-1">Wholesale + Retail Dual-Channel Dashboard</p>
    </div>
    <div class="flex gap-4 text-xs text-navy-200">
      <span>Generated: <b class="text-white" id="metaDate">GENERATED_AT</b></span>
      <span>WS Models: <b class="text-white" id="metaWS">0</b></span>
      <span>RT Models: <b class="text-white" id="metaRT">0</b></span>
      <span>Matched: <b class="text-white" id="metaMatch">0</b></span>
    </div>
  </div>
</div>

<!-- Tab Navigation -->
<div class="sticky top-0 z-40 bg-white border-b shadow-sm no-print">
  <div class="max-w-[1600px] mx-auto px-4 pt-2 flex gap-1">
    <button class="tab-btn active" data-tab="integrated" onclick="switchTab('integrated')">Integrated</button>
    <button class="tab-btn" data-tab="wholesale" onclick="switchTab('wholesale')">Wholesale</button>
    <button class="tab-btn" data-tab="retail" onclick="switchTab('retail')">Retail</button>
  </div>
  <!-- Global Filter Bar -->
  <div class="max-w-[1600px] mx-auto px-4 py-2 flex items-center gap-3 flex-wrap">
    <span class="text-[10px] text-gray-400 font-semibold uppercase tracking-wider">Filter:</span>
    <div id="gf_cat"></div>
    <div id="gf_type"></div>
    <div id="gf_ton"></div>
    <div id="gf_brand"></div>
    <button onclick="resetGlobal()" class="text-[10px] text-red-500 font-bold hover:underline ml-2">Reset</button>
  </div>
</div>

<div class="max-w-[1600px] mx-auto px-4 py-4">

<!-- ═══════════════════════════════════════════════════════════════
     TAB 1: INTEGRATED
     ═══════════════════════════════════════════════════════════════ -->
<div id="tab_integrated" class="tab-panel active">

  <!-- I-S1: Integrated KPI Cards -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Integrated Overview</h2>
    <div class="flex items-center gap-3 mb-3 text-xs">
      <span class="font-semibold text-gray-500">WS Date:</span>
      <select id="int_ws_date" class="pt-sel"></select>
      <span class="font-semibold text-gray-500 ml-3">Retail Date:</span>
      <select id="int_rt_week" class="pt-sel"></select>
    </div>
    <div id="intKpiGrid" class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-6 gap-3"></div>
  </div>

  <!-- I-S2: Price Gap Analysis Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Price Gap Analysis (Wholesale vs Retail)</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="int_search" class="sec-search" placeholder="Search model..."/>
        <button class="dir-int-btn text-[10px] px-2 py-1 rounded bg-gray-700 text-white active" data-dir="all" onclick="setIntDir(this)">All</button>
        <button class="dir-int-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="high" onclick="setIntDir(this)">High Markup</button>
        <button class="dir-int-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="low" onclick="setIntDir(this)">Low Markup</button>
        <button class="dir-int-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="neg" onclick="setIntDir(this)">Negative</button>
        <span id="int_gap_count" class="text-[10px] text-gray-400 ml-1"></span>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:500px;overflow-y:auto">
      <table id="tblIntGap"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- I-S3: Markup Distribution Charts -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Markup Distribution</h2>
    <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
      <div style="height:400px"><canvas id="intMarkupBar"></canvas></div>
      <div style="height:400px"><canvas id="intMarkupDonut"></canvas></div>
    </div>
  </div>

  <!-- I-S4: Channel Coverage Matrix -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Channel Coverage Matrix (Brand × Category)</h2>
    <div class="tbl-wrap" style="max-height:500px;overflow-y:auto">
      <table id="tblCoverage"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- I-S5: Matched Models Full Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Matched Models — Wholesale vs Retail Detail</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="int_s5_search" class="sec-search" placeholder="Search model..." oninput="ST.intS5q=this.value;renderIntMatched()"/>
        <button onclick="downloadIntMatchedExcel()" class="text-[10px] px-3 py-1 rounded bg-green-600 text-white font-semibold hover:bg-green-700">Excel</button>
        <span id="int_s5_count" class="text-[10px] text-gray-400 ml-1"></span>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:600px;overflow-y:auto">
      <table id="tblIntMatched"><thead></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════
     TAB 2: WHOLESALE
     ═══════════════════════════════════════════════════════════════ -->
<div id="tab_wholesale" class="tab-panel">

  <!-- WS-S1: KPI Cards -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Wholesale Overview</h2>
    <div class="flex items-center gap-3 mb-3 text-xs">
      <span class="font-semibold text-gray-500">Current Date:</span>
      <select id="ws_cur_date" class="pt-sel" onchange="renderWS()"></select>
      <span class="font-semibold text-gray-500 ml-3">Compare To:</span>
      <select id="ws_prev_date" class="pt-sel" onchange="renderWS()"></select>
      <span id="ws_compare" class="text-[10px] text-gray-400 ml-2"></span>
    </div>
    <div id="wsKpiGrid" class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-6 gap-3"></div>
  </div>

  <!-- WS-S2: Price Alerts -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Wholesale Price Alerts</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="ws_alert_search" class="sec-search" placeholder="Search..."/>
        <button class="ws-dir-btn text-[10px] px-2 py-1 rounded bg-gray-700 text-white active" data-dir="all" onclick="setWsDir(this)">All</button>
        <button class="ws-dir-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="up" onclick="setWsDir(this)">Up</button>
        <button class="ws-dir-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="down" onclick="setWsDir(this)">Down</button>
        <span id="ws_alert_count" class="text-[10px] text-gray-400 ml-1"></span>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:500px;overflow-y:auto">
      <table id="tblWsAlert"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- WS-S3: New & Discontinued -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">New & Discontinued Models</h2>
    <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
      <div>
        <h3 class="text-sm font-semibold text-green-700 mb-2">New Models <span id="ws_new_count" class="text-gray-400"></span></h3>
        <div id="wsNewGrid" class="grid grid-cols-1 gap-2 max-h-[400px] overflow-y-auto"></div>
      </div>
      <div>
        <h3 class="text-sm font-semibold text-red-700 mb-2">Discontinued <span id="ws_disc_count" class="text-gray-400"></span></h3>
        <div id="wsDiscGrid" class="grid grid-cols-1 gap-2 max-h-[400px] overflow-y-auto"></div>
      </div>
    </div>
  </div>

  <!-- WS-S4: Category KPI Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Wholesale Category KPI</h2>
      <input type="text" id="ws_s4_search" class="sec-search" placeholder="Search..."/>
    </div>
    <div class="tbl-wrap" style="max-height:600px;overflow-y:auto">
      <table id="tblWsCatKpi"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- WS-S5: Brand Comparison Chart -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Wholesale Brand Price Comparison</h2>
    <div style="height:450px"><canvas id="wsBrandBar"></canvas></div>
  </div>

  <!-- WS-S6: Price Trend -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Wholesale Price Trend</h2>
      <input type="text" id="ws_s6_search" class="sec-search" placeholder="Search model..."/>
    </div>
    <div style="height:400px"><canvas id="wsTrendChart"></canvas></div>
  </div>

  <!-- WS-S7: Full Model Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Wholesale Full Model Table</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="ws_s7_search" class="sec-search" placeholder="Search..."/>
        <button onclick="downloadWsExcel()" class="text-[10px] px-3 py-1 rounded bg-green-600 text-white font-semibold hover:bg-green-700">Excel</button>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:600px;overflow-y:auto">
      <table id="tblWsFull"><thead></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════
     TAB 3: RETAIL
     ═══════════════════════════════════════════════════════════════ -->
<div id="tab_retail" class="tab-panel">

  <!-- RT-S1: KPI Cards -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Retail Overview</h2>
    <div class="flex items-center gap-3 mb-3 text-xs">
      <span class="font-semibold text-gray-500">Current Date:</span>
      <select id="rt_cur_week" class="pt-sel" onchange="renderRT()"></select>
      <span class="font-semibold text-gray-500 ml-3">Compare To:</span>
      <select id="rt_prev_week" class="pt-sel" onchange="renderRT()"></select>
      <span id="rt_compare" class="text-[10px] text-gray-400 ml-2"></span>
    </div>
    <div id="rtKpiGrid" class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-6 gap-3"></div>
  </div>

  <!-- RT-S2: Price Alerts -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Retail Price Alerts</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="rt_alert_search" class="sec-search" placeholder="Search..."/>
        <select id="rt_alert_pt" class="pt-sel" onchange="renderRtAlerts()">
          <option value="cp">Current Price</option>
          <option value="rp">Regular Price</option>
        </select>
        <button class="rt-dir-btn text-[10px] px-2 py-1 rounded bg-gray-700 text-white active" data-dir="all" onclick="setRtDir(this)">All</button>
        <button class="rt-dir-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="up" onclick="setRtDir(this)">Up</button>
        <button class="rt-dir-btn text-[10px] px-2 py-1 rounded bg-gray-50" data-dir="down" onclick="setRtDir(this)">Down</button>
        <span id="rt_alert_count" class="text-[10px] text-gray-400 ml-1"></span>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:500px;overflow-y:auto">
      <table id="tblRtAlert"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- RT-S3: New & Discontinued -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">New & Discontinued SKUs</h2>
    <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
      <div>
        <h3 class="text-sm font-semibold text-green-700 mb-2">New SKUs <span id="rt_new_count" class="text-gray-400"></span></h3>
        <div id="rtNewGrid" class="grid grid-cols-1 gap-2 max-h-[400px] overflow-y-auto"></div>
      </div>
      <div>
        <h3 class="text-sm font-semibold text-red-700 mb-2">Removed SKUs <span id="rt_disc_count" class="text-gray-400"></span></h3>
        <div id="rtDiscGrid" class="grid grid-cols-1 gap-2 max-h-[400px] overflow-y-auto"></div>
      </div>
    </div>
  </div>

  <!-- RT-S4: Category KPI Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Retail Category KPI</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="rt_s4_search" class="sec-search" placeholder="Search..."/>
        <select id="rt_s4_pt" class="pt-sel" onchange="renderRtCatKpi()">
          <option value="cp">Current Price</option>
          <option value="rp">Regular Price</option>
        </select>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:600px;overflow-y:auto">
      <table id="tblRtCatKpi"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <!-- RT-S5: Brand Comparison -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <h2 class="text-base font-bold text-navy-800 mb-3">Retail Brand Price Comparison</h2>
    <div class="flex items-center gap-2 mb-3">
      <select id="rt_s5_pt" class="pt-sel" onchange="renderRtBrandBar()">
        <option value="cp">Current Price</option>
        <option value="rp">Regular Price</option>
      </select>
    </div>
    <div style="height:450px"><canvas id="rtBrandBar"></canvas></div>
  </div>

  <!-- RT-S6: Price Trend -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Retail Price Trend</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="rt_s6_search" class="sec-search" placeholder="Search model..."/>
        <select id="rt_s6_pt" class="pt-sel" onchange="renderRtTrend()">
          <option value="cp">Current Price</option>
          <option value="rp">Regular Price</option>
        </select>
      </div>
    </div>
    <div style="height:400px"><canvas id="rtTrendChart"></canvas></div>
  </div>

  <!-- RT-S7: Full SKU Table -->
  <div class="bg-white rounded-xl border p-4 mb-4 shadow-sm">
    <div class="flex items-center justify-between mb-3 flex-wrap gap-2">
      <h2 class="text-base font-bold text-navy-800">Retail Full SKU Table</h2>
      <div class="flex items-center gap-2">
        <input type="text" id="rt_s7_search" class="sec-search" placeholder="Search..."/>
        <button onclick="downloadRtExcel()" class="text-[10px] px-3 py-1 rounded bg-green-600 text-white font-semibold hover:bg-green-700">Excel</button>
      </div>
    </div>
    <div class="tbl-wrap" style="max-height:600px;overflow-y:auto">
      <table id="tblRtFull"><thead></thead><tbody></tbody></table>
    </div>
  </div>
</div>

</div><!-- max-w container -->

<!-- Footer -->
<div class="text-center text-gray-400 text-[10px] py-4 no-print">
  BH AC Price Tracker — Generated GENERATED_AT
</div>
"""

    # ── HTML DATA ──────────────────────────────────────────────────
    HTML_DATA = f"""<script>
const WS={json.dumps(ws_data,ensure_ascii=False)};
const RT={json.dumps(rt_data,ensure_ascii=False)};
const WS_DATES={json.dumps(ws_dates)};
const RT_DATES={json.dumps(rt_dates)};
const BRANDS={json.dumps(all_brands)};
const CATEGORIES={json.dumps(all_cats)};
const TYPES={json.dumps(all_types)};
const TONS={json.dumps([float(t) for t in sorted(all_tons)])};
const BRAND_COLORS={json.dumps(BRAND_COLORS)};
const PID={json.dumps(product_id_map)};
function bhLink(model,label){{const id=PID[model];return id?`<a href="https://bhstore.com.sa/sa-en/details/${{id}}" target="_blank" class="text-blue-600 hover:underline">${{label}}</a>`:label;}}
const MODEL_OVERLAP={{matched:{len(matched)},wsOnly:{len(ws_only)},rtOnly:{len(rt_only)},total:{len(ws_models|rt_models)}}};
const GENERATED_AT='{generated_at}';
</script>
"""

    # ── HTML LOGIC ─────────────────────────────────────────────────
    HTML_LOGIC = """<script>
Chart.register(ChartDataLabels);
Chart.defaults.plugins.datalabels={display:false};

// ── Utilities ──
const fmtSAR=v=>v==null?'-':Number(v).toLocaleString('en-SA',{maximumFractionDigits:0});
const fmtPct=v=>v==null?'-':(v*100).toFixed(1)+'%';
const fmtPctR=v=>v==null?'-':v.toFixed(1)+'%';
const fmtChg=v=>{if(v==null)return'-';return(v>0?'+':'')+Number(v).toLocaleString('en-SA',{maximumFractionDigits:0});};
const colorOf=br=>BRAND_COLORS[br]||'#6b7280';
const alphaC=(hex,a)=>{const r=parseInt(hex.slice(1,3),16),g=parseInt(hex.slice(3,5),16),b=parseInt(hex.slice(5,7),16);return`rgba(${r},${g},${b},${a})`;};
const mean=arr=>arr.length?arr.reduce((s,v)=>s+v,0)/arr.length:null;
const searchMatch=(r,q)=>!q||((r.b||'')+(r.m||'')+(r.n||'')+(r.c||'')+(r.tp||'')).toLowerCase().includes(q);

// ── MultiSelect Class ──
class MS{
  constructor(el,opts,label,cb,colorFn){
    this.el=el;this.opts=opts;this.label=label;this.cb=cb;this.colorFn=colorFn||null;
    this.sel=new Set(opts.map(String));this._build();
  }
  _build(){
    const w=document.createElement('div');w.className='ms-wrap';
    const btn=document.createElement('button');btn.type='button';btn.className='ms-btn';
    this.btnEl=btn;w.appendChild(btn);
    const menu=document.createElement('div');menu.className='ms-menu';menu.onclick=function(e){e.stopPropagation()};
    const acts=document.createElement('div');acts.className='ms-actions';
    const aAll=document.createElement('button');aAll.textContent='All';aAll.onclick=()=>this.selectAll();
    const aNone=document.createElement('button');aNone.textContent='None';aNone.className='ms-none';aNone.onclick=()=>this.selectNone();
    acts.appendChild(aAll);acts.appendChild(aNone);menu.appendChild(acts);
    const list=document.createElement('div');list.className='ms-list';
    this.opts.forEach(o=>{
      const lbl=document.createElement('label');
      const cb=document.createElement('input');cb.type='checkbox';cb.checked=true;cb.value=String(o);
      cb.addEventListener('change',()=>{if(cb.checked)this.sel.add(String(o));else this.sel.delete(String(o));this._upd();this.cb();});
      lbl.appendChild(cb);
      const sp=document.createElement('span');
      const txt=this.label==='Ton'?parseFloat(o).toFixed(1)+'T':String(o);
      sp.textContent=txt;
      if(this.colorFn)sp.style.cssText=`color:${this.colorFn(o)};font-weight:600`;
      lbl.appendChild(sp);list.appendChild(lbl);
    });
    menu.appendChild(list);w.appendChild(menu);this.menuEl=menu;this.listEl=list;
    btn.addEventListener('click',e=>{e.stopPropagation();document.querySelectorAll('.ms-menu.open').forEach(m=>{if(m!==menu)m.classList.remove('open')});menu.classList.toggle('open');});
    this.el.appendChild(w);this._upd();
  }
  _upd(){const av=this._availCount();this.btnEl.innerHTML=`${this.label} <b class="text-navy-700">${this.sel.size}/${av}</b> <span class="text-gray-400 text-[9px]">&#9662;</span>`;}
  _availCount(){let n=0;this.listEl.querySelectorAll('label').forEach(l=>{if(!l.classList.contains('ms-disabled'))n++;});return n||this.opts.length;}
  selectAll(){this.sel=new Set();this.listEl.querySelectorAll('input').forEach(c=>{if(!c.disabled){this.sel.add(c.value);c.checked=true;}else{c.checked=false;}});this._upd();this.cb();}
  selectNone(){this.sel.clear();this.listEl.querySelectorAll('input').forEach(c=>c.checked=false);this._upd();this.cb();}
  reset(){this.selectAll();}
  getSelected(){return this.sel;}
  setSelected(vals){this.sel=new Set(vals.map(String));this.listEl.querySelectorAll('input').forEach(cb=>{cb.checked=this.sel.has(cb.value);});this._upd();}
  updateAvailable(availSet){this.listEl.querySelectorAll('label').forEach(lbl=>{const cb=lbl.querySelector('input');const avail=availSet.has(cb.value);lbl.classList.toggle('ms-disabled',!avail);cb.disabled=!avail;});this._upd();}
}
document.addEventListener('click',()=>document.querySelectorAll('.ms-menu.open').forEach(m=>m.classList.remove('open')));

// ── Global Filter ──
let GF;
function applyGF(rows){
  return rows.filter(r=>{
    if(r.c&&!GF.cat.getSelected().has(r.c))return false;
    if(r.tp&&!GF.type.getSelected().has(r.tp))return false;
    if(r.t!=null&&!GF.ton.getSelected().has(String(r.t)))return false;
    if(r.b&&!GF.brand.getSelected().has(r.b))return false;
    return true;
  });
}
function resetGlobal(){GF.cat.reset();GF.type.reset();GF.ton.reset();GF.brand.reset();refreshAll();}

// ── Table Sort ──
const _ss={};
function sortTbl(id,col){
  const tbl=document.getElementById(id),tb=tbl.querySelector('tbody'),rows=Array.from(tb.querySelectorAll('tr')),ths=tbl.querySelectorAll('th');
  const p=_ss[id]||{col:-1,asc:true};const asc=p.col===col?!p.asc:true;_ss[id]={col,asc};
  ths.forEach((th,i)=>{th.classList.remove('sort-asc','sort-desc');if(i===col)th.classList.add(asc?'sort-asc':'sort-desc');});
  rows.sort((a,b)=>{const ta=a.cells[col]?.textContent.trim()||'',tb2=b.cells[col]?.textContent.trim()||'';
    const na=parseFloat(ta.replace(/[^0-9.\\-+]/g,'')),nb=parseFloat(tb2.replace(/[^0-9.\\-+]/g,''));
    if(!isNaN(na)&&!isNaN(nb))return asc?na-nb:nb-na;return asc?ta.localeCompare(tb2):tb2.localeCompare(ta);});
  rows.forEach(r=>tb.appendChild(r));
}

// ── endLabels Plugin ──
const endLabelsPlugin={id:'endLabels',afterDatasetsDraw(chart){const ctx=chart.ctx;chart.data.datasets.forEach((ds,i)=>{const meta=chart.getDatasetMeta(i);if(!meta.visible)return;for(let j=meta.data.length-1;j>=0;j--){const pt=meta.data[j];if(pt&&ds.data[j]!=null){ctx.save();ctx.font='bold 10px Inter,sans-serif';ctx.fillStyle=ds.borderColor||'#333';ctx.textBaseline='middle';const lbl=ds.label.length>18?ds.label.substring(0,18)+'..':ds.label;ctx.fillText(lbl,pt.x+6,pt.y);ctx.restore();break;}}});}};

// ── KPI Card HTML ──
function kpiCard(v,l,borderColor,textColor){
  return `<div class="rounded-lg border-l-4 ${borderColor} p-3"><div class="text-lg font-bold ${textColor}">${v}</div><div class="text-[10px] text-gray-500 mt-1 font-medium">${l}</div></div>`;
}

// ── Tab Switch ──
let activeTab='integrated';
function switchTab(tab){
  activeTab=tab;
  document.querySelectorAll('.tab-btn').forEach(b=>{b.classList.toggle('active',b.dataset.tab===tab);});
  document.querySelectorAll('.tab-panel').forEach(p=>{p.classList.toggle('active',p.id==='tab_'+tab);});
  refreshAll();
}

// ── State ──
const ST={intDir:'all',intQ:'',intS5q:'',wsDir:'all',wsQ:'',wsS4q:'',wsS6q:'',wsS7q:'',rtDir:'all',rtQ:'',rtS4q:'',rtS6q:'',rtS7q:''};

function setIntDir(btn){document.querySelectorAll('.dir-int-btn').forEach(b=>{b.classList.remove('active');b.className=b.className.replace(/bg-gray-700 text-white/g,'bg-gray-50')});btn.classList.add('active');btn.className=btn.className.replace(/bg-gray-50(?!\\s)/g,'bg-gray-700 text-white');ST.intDir=btn.dataset.dir;renderIntGap();}
function setWsDir(btn){document.querySelectorAll('.ws-dir-btn').forEach(b=>{b.classList.remove('active');b.className=b.className.replace(/bg-gray-700 text-white/g,'bg-gray-50')});btn.classList.add('active');btn.className=btn.className.replace(/bg-gray-50(?!\\s)/g,'bg-gray-700 text-white');ST.wsDir=btn.dataset.dir;renderWsAlerts();}
function setRtDir(btn){document.querySelectorAll('.rt-dir-btn').forEach(b=>{b.classList.remove('active');b.className=b.className.replace(/bg-gray-700 text-white/g,'bg-gray-50')});btn.classList.add('active');btn.className=btn.className.replace(/bg-gray-50(?!\\s)/g,'bg-gray-700 text-white');ST.rtDir=btn.dataset.dir;renderRtAlerts();}

// ══════════════════════════════════════════════════════════════════
// TAB 1: INTEGRATED
// ══════════════════════════════════════════════════════════════════
let intMarkupBarObj=null,intMarkupDonutObj=null;

function renderIntKpis(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtWeek=document.getElementById('int_rt_week').value;
  const wsD=applyGF(WS.filter(r=>r.d===wsDate));
  const rtD=applyGF(RT.filter(r=>r.d===rtWeek));
  const wsModels=new Set(wsD.map(r=>r.m));
  const rtModels=new Set(rtD.map(r=>r.m));
  const matched=new Set([...wsModels].filter(m=>rtModels.has(m)));
  const wsOnly=new Set([...wsModels].filter(m=>!rtModels.has(m)));
  const rtOnly=new Set([...rtModels].filter(m=>!wsModels.has(m)));
  const totalModels=new Set([...wsModels,...rtModels]).size;

  // Avg prices for matched models
  const wsMap={};wsD.forEach(r=>{if(r.p!=null)wsMap[r.m]=r.p;});
  const rtMap={};rtD.forEach(r=>{if(r.cp!=null)rtMap[r.m]=r.cp;});
  let markups=[];
  matched.forEach(m=>{
    if(wsMap[m]!=null&&rtMap[m]!=null&&wsMap[m]>0){
      markups.push((rtMap[m]-wsMap[m])/wsMap[m]*100);
    }
  });
  const avgMarkup=markups.length?Math.round(mean(markups)*10)/10:null;
  const wsAvg=wsD.filter(r=>r.p!=null).length?Math.round(mean(wsD.filter(r=>r.p!=null).map(r=>r.p))):null;
  const rtAvg=rtD.filter(r=>r.cp!=null).length?Math.round(mean(rtD.filter(r=>r.cp!=null).map(r=>r.cp))):null;

  document.getElementById('intKpiGrid').innerHTML=[
    kpiCard(totalModels,'Total Models','border-l-navy-800 bg-navy-50','text-navy-800'),
    kpiCard(matched.size,'Matched Models','border-l-green-500 bg-green-50','text-green-600'),
    kpiCard(wsOnly.size,'WS Only','border-l-amber-500 bg-amber-50','text-amber-700'),
    kpiCard(rtOnly.size,'Retail Only','border-l-blue-500 bg-blue-50','text-blue-600'),
    kpiCard(wsAvg!=null?'SAR '+fmtSAR(wsAvg):'-','Avg WS Price','border-l-purple-500 bg-purple-50','text-purple-700'),
    kpiCard(avgMarkup!=null?avgMarkup.toFixed(1)+'%':'-','Avg Markup %','border-l-red-500 bg-red-50','text-red-600'),
  ].join('');
}

function renderIntGap(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtWeek=document.getElementById('int_rt_week').value;
  const wsD=applyGF(WS.filter(r=>r.d===wsDate));
  const rtD=applyGF(RT.filter(r=>r.d===rtWeek));
  const q=ST.intQ;
  const wsMap={};wsD.forEach(r=>{wsMap[r.m]=r;});
  const rtMap={};rtD.forEach(r=>{rtMap[r.m]=r;});
  const allModels=new Set([...Object.keys(wsMap),...Object.keys(rtMap)]);
  let rows=[];
  allModels.forEach(m=>{
    const ws=wsMap[m],rt=rtMap[m];
    const ref=ws||rt;
    if(q&&!searchMatch(ref,q))return;
    const wsP=ws?ws.p:null;
    const rtReg=rt?rt.rp:null;
    const rtCur=rt?rt.cp:null;
    const markup=(wsP!=null&&rtCur!=null&&wsP>0)?((rtCur-wsP)/wsP*100):null;
    const markupSAR=(wsP!=null&&rtCur!=null)?(rtCur-wsP):null;
    rows.push({b:ref.b,m,c:ref.c,t:ref.t,tp:ref.tp,n:ref.n,wsP,rtReg,rtCur,markup,markupSAR,hasWS:!!ws,hasRT:!!rt});
  });
  // Direction filter
  if(ST.intDir==='high'){const avg=mean(rows.filter(r=>r.markup!=null).map(r=>r.markup))||0;rows=rows.filter(r=>r.markup!=null&&r.markup>avg);}
  if(ST.intDir==='low'){const avg=mean(rows.filter(r=>r.markup!=null).map(r=>r.markup))||0;rows=rows.filter(r=>r.markup!=null&&r.markup<=avg&&r.markup>=0);}
  if(ST.intDir==='neg')rows=rows.filter(r=>r.markup!=null&&r.markup<0);
  rows.sort((a,b)=>(b.markup||0)-(a.markup||0));
  document.getElementById('int_gap_count').textContent=rows.length+' models';

  const cols=[{k:'b',l:'Brand'},{k:'m',l:'Model'},{k:'c',l:'Category'},{k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'wsP',l:'WS Price'},{k:'rtReg',l:'Retail Regular'},{k:'rtCur',l:'Retail Current'},{k:'markupSAR',l:'Markup SAR'},{k:'markup',l:'Markup %'}];
  const tbl=document.getElementById('tblIntGap');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblIntGap',${i})">${c.l}</th>`).join('')+'</tr>';
  tbl.querySelector('tbody').innerHTML=rows.length?rows.map(r=>'<tr>'+cols.map(c=>{
    let v;
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='t')v=r.t!=null?r.t.toFixed(1)+'T':'-';
    else if(c.k==='wsP'||c.k==='rtReg'||c.k==='rtCur')v=r[c.k]!=null?fmtSAR(r[c.k]):(r.hasWS||c.k==='wsP'?'-':'<span class="text-gray-300">N/A</span>');
    else if(c.k==='markupSAR'){v=r.markupSAR!=null?fmtChg(r.markupSAR):'-';return`<td class="${r.markupSAR>0?'up-cell':r.markupSAR<0?'dn-cell':''}">${v}</td>`;}
    else if(c.k==='markup'){v=r.markup!=null?r.markup.toFixed(1)+'%':'-';return`<td class="${r.markup!=null&&r.markup>50?'up-cell':r.markup!=null&&r.markup<10?'dn-cell':''}">${v}</td>`;}
    else v=r[c.k]||'-';
    return`<td>${v}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="10" class="text-center text-gray-400 py-6">No data</td></tr>';
}

function renderIntCharts(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtWeek=document.getElementById('int_rt_week').value;
  const wsD=applyGF(WS.filter(r=>r.d===wsDate));
  const rtD=applyGF(RT.filter(r=>r.d===rtWeek));
  const wsMap={};wsD.forEach(r=>{if(r.p!=null)wsMap[r.m]=r;});
  const rtMap={};rtD.forEach(r=>{if(r.cp!=null)rtMap[r.m]=r;});

  // Brand avg markup
  const brandMarkup={};
  Object.keys(wsMap).forEach(m=>{
    if(!rtMap[m])return;
    const ws=wsMap[m],rt=rtMap[m];
    if(ws.p>0){
      const br=ws.b||rt.b;
      if(!brandMarkup[br])brandMarkup[br]=[];
      brandMarkup[br].push((rt.cp-ws.p)/ws.p*100);
    }
  });
  let items=Object.entries(brandMarkup).map(([b,arr])=>({b,avg:Math.round(mean(arr)*10)/10})).sort((a,b)=>a.avg-b.avg);
  const ctx1=document.getElementById('intMarkupBar').getContext('2d');
  if(intMarkupBarObj)intMarkupBarObj.destroy();
  intMarkupBarObj=new Chart(ctx1,{type:'bar',data:{labels:items.map(x=>x.b),datasets:[{label:'Avg Markup %',data:items.map(x=>x.avg),backgroundColor:items.map(x=>x.avg>=0?alphaC(colorOf(x.b),.7):alphaC('#dc2626',.5)),borderColor:items.map(x=>colorOf(x.b)),borderWidth:1}]},options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},title:{display:true,text:'Avg Markup % by Brand (Retail vs Wholesale)',font:{size:12}},datalabels:{display:true,anchor:'end',align:'right',color:'#1e3a5f',font:{size:10,weight:'bold'},formatter:v=>v.toFixed(1)+'%',clip:false}},scales:{x:{grid:{color:'#f0f0f0'},ticks:{callback:v=>v+'%'}},y:{ticks:{font:{size:10}}}}}});

  // Category donut
  const catMarkup={};
  Object.keys(wsMap).forEach(m=>{
    if(!rtMap[m])return;
    const ws=wsMap[m],rt=rtMap[m];
    if(ws.p>0){
      const cat=ws.c||rt.c||'Other';
      if(!catMarkup[cat])catMarkup[cat]=[];
      catMarkup[cat].push((rt.cp-ws.p)/ws.p*100);
    }
  });
  const catItems=Object.entries(catMarkup).map(([c,arr])=>({c,avg:Math.round(mean(arr)*10)/10}));
  const catCols=['#1F4E79','#ED7D31','#A9D18E','#FF0000','#70AD47','#FFC000','#4472C4','#9E480E'];
  const ctx2=document.getElementById('intMarkupDonut').getContext('2d');
  if(intMarkupDonutObj)intMarkupDonutObj.destroy();
  intMarkupDonutObj=new Chart(ctx2,{type:'doughnut',data:{labels:catItems.map(x=>x.c),datasets:[{data:catItems.map(x=>x.avg),backgroundColor:catCols.slice(0,catItems.length)}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{font:{size:10}}},title:{display:true,text:'Avg Markup % by Category',font:{size:12}},datalabels:{display:true,color:'#fff',font:{size:12,weight:'bold'},formatter:v=>v.toFixed(1)+'%'}}}});
}

function renderCoverage(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtWeek=document.getElementById('int_rt_week').value;
  const wsD=applyGF(WS.filter(r=>r.d===wsDate));
  const rtD=applyGF(RT.filter(r=>r.d===rtWeek));

  // Brand × Category matrix
  const brands=[...GF.brand.getSelected()].sort();
  const cats=[...GF.cat.getSelected()].sort();
  const wsBC={};wsD.forEach(r=>{const k=r.b+'|'+r.c;wsBC[k]=(wsBC[k]||new Set()).add(r.m);});
  const rtBC={};rtD.forEach(r=>{const k=r.b+'|'+r.c;rtBC[k]=(rtBC[k]||new Set()).add(r.m);});

  const tbl=document.getElementById('tblCoverage');
  tbl.querySelector('thead').innerHTML='<tr><th>Brand</th>'+cats.map(c=>`<th>${c}</th>`).join('')+'<th>Total</th></tr>';
  let tbody='';
  brands.forEach(br=>{
    let totalWs=0,totalRt=0;
    let cells=cats.map(cat=>{
      const k=br+'|'+cat;
      const wc=(wsBC[k]||new Set()).size;
      const rc=(rtBC[k]||new Set()).size;
      totalWs+=wc;totalRt+=rc;
      const bg=wc>0&&rc>0?'bg-green-50':wc>0||rc>0?'bg-amber-50':'bg-gray-50';
      return`<td class="${bg} text-center text-xs">${wc>0||rc>0?wc+' / '+rc:'-'}</td>`;
    }).join('');
    tbody+=`<tr><td><span style="color:${colorOf(br)};font-weight:600">${br}</span></td>${cells}<td class="text-center font-semibold text-xs">${totalWs} / ${totalRt}</td></tr>`;
  });
  tbl.querySelector('tbody').innerHTML=tbody||'<tr><td colspan="'+(cats.length+2)+'" class="text-center text-gray-400 py-4">No data</td></tr>';
}

// ── I-S5: Matched Models Full Table ──
let _intMatchedRows=[];
function renderIntMatched(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtDate=document.getElementById('int_rt_week').value;
  const wsD=applyGF(WS.filter(r=>r.d===wsDate));
  const rtD=applyGF(RT.filter(r=>r.d===rtDate));
  const q=ST.intS5q;
  const wsMap={};wsD.forEach(r=>{wsMap[r.m]=r;});
  const rtMap={};rtD.forEach(r=>{rtMap[r.m]=r;});
  // Only matched models (exist in both)
  const matched=Object.keys(wsMap).filter(m=>rtMap[m]);
  let rows=[];
  matched.forEach(m=>{
    const ws=wsMap[m],rt=rtMap[m];
    if(q&&!searchMatch(ws,q)&&!searchMatch(rt,q))return;
    const wsP=ws.p;
    const rtReg=rt.rp;
    const rtCur=rt.cp;
    const markup=(wsP!=null&&rtCur!=null&&wsP>0)?((rtCur-wsP)/wsP*100):null;
    const markupSAR=(wsP!=null&&rtCur!=null)?(rtCur-wsP):null;
    const dr=rt.dr;
    rows.push({b:ws.b||rt.b,m,n:rt.n||ws.n||'',c:ws.c||rt.c||'',tp:ws.tp||rt.tp||'',t:ws.t||rt.t,btu:ws.btu||rt.btu,wsP,rtReg,rtCur,dr,markup,markupSAR,url:rt.url||''});
  });
  rows.sort((a,b)=>(a.b||'').localeCompare(b.b||'')||(a.m||'').localeCompare(b.m||''));
  _intMatchedRows=rows;
  document.getElementById('int_s5_count').textContent=rows.length+' matched models';

  const cols=[
    {k:'b',l:'Brand'},{k:'m',l:'Model Code'},{k:'n',l:'Product Name'},{k:'c',l:'Category'},
    {k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'btu',l:'BTU'},
    {k:'wsP',l:'WS Price'},{k:'rtReg',l:'Retail Regular'},{k:'rtCur',l:'Retail Current'},
    {k:'dr',l:'Discount %'},{k:'markupSAR',l:'Markup SAR'},{k:'markup',l:'Markup %'}
  ];
  const tbl=document.getElementById('tblIntMatched');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblIntMatched',${i})">${c.l}</th>`).join('')+'</tr>';
  tbl.querySelector('tbody').innerHTML=rows.length?rows.map(r=>'<tr>'+cols.map(c=>{
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='n')return`<td>${bhLink(r.m,r.n||'-')}</td>`;
    if(c.k==='t')return`<td>${r.t!=null?r.t.toFixed(1)+'T':'-'}</td>`;
    if(c.k==='btu')return`<td>${r.btu!=null?r.btu.toLocaleString():'-'}</td>`;
    if(c.k==='wsP'||c.k==='rtReg'||c.k==='rtCur')return`<td>${r[c.k]!=null?fmtSAR(r[c.k]):'-'}</td>`;
    if(c.k==='dr')return`<td>${r.dr!=null?fmtPct(r.dr):'-'}</td>`;
    if(c.k==='markupSAR'){const v=r.markupSAR;return`<td class="${v!=null&&v>0?'up-cell':v!=null&&v<0?'dn-cell':''}">${v!=null?fmtChg(v):'-'}</td>`;}
    if(c.k==='markup'){const v=r.markup;return`<td class="${v!=null&&v>50?'up-cell':v!=null&&v<10?'dn-cell':''}">${v!=null?v.toFixed(1)+'%':'-'}</td>`;}
    return`<td>${r[c.k]||'-'}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="13" class="text-center text-gray-400 py-6">No matched models</td></tr>';
}

function downloadIntMatchedExcel(){
  const wsDate=document.getElementById('int_ws_date').value;
  const rtDate=document.getElementById('int_rt_week').value;
  const hdr=['Brand','Model Code','Product Name','Category','Type','Ton','BTU','WS Price','Retail Regular','Retail Current','Discount %','Markup SAR','Markup %'];
  const data=_intMatchedRows.map(r=>[r.b,r.m,r.n,r.c,r.tp,r.t,r.btu,r.wsP,r.rtReg,r.rtCur,r.dr!=null?r.dr:null,r.markupSAR,r.markup!=null?r.markup/100:null]);
  const ws2=XLSX.utils.aoa_to_sheet([hdr,...data]);
  // Format Discount % and Markup % columns
  const range=XLSX.utils.decode_range(ws2['!ref']);
  for(let R=1;R<=range.e.r;R++){
    const dCell=XLSX.utils.encode_cell({r:R,c:10});if(ws2[dCell]&&ws2[dCell].v!=null)ws2[dCell].z='0.0%';
    const mCell=XLSX.utils.encode_cell({r:R,c:12});if(ws2[mCell]&&ws2[mCell].v!=null)ws2[mCell].z='0.0%';
  }
  // Auto column widths
  ws2['!cols']=hdr.map((_,i)=>({wch:i===2?40:i===1?18:i===3?20:12}));
  const wbx=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wbx,ws2,'Matched Models');
  XLSX.writeFile(wbx,`BH_Matched_Models_WS${wsDate}_RT${rtDate}.xlsx`);
}

function renderIntegrated(){renderIntKpis();renderIntGap();renderIntCharts();renderCoverage();renderIntMatched();}

// ══════════════════════════════════════════════════════════════════
// TAB 2: WHOLESALE
// ══════════════════════════════════════════════════════════════════
let wsBrandBarObj=null,wsTrendObj=null;

function getWsDates(){
  return{cur:document.getElementById('ws_cur_date').value,prev:document.getElementById('ws_prev_date').value};
}

function renderWsKpis(){
  const{cur,prev}=getWsDates();
  const curD=applyGF(WS.filter(r=>r.d===cur));
  const prevD=prev?applyGF(WS.filter(r=>r.d===prev)):[];
  document.getElementById('ws_compare').textContent='Viewing: '+cur+(prev?' vs '+prev:'');

  const curModels=new Set(curD.map(r=>r.m));
  const prevModels=new Set(prevD.map(r=>r.m));
  const newM=new Set([...curModels].filter(m=>!prevModels.has(m)));
  const discM=new Set([...prevModels].filter(m=>!curModels.has(m)));

  // Price changes
  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});
  let upCnt=0,dnCnt=0;
  curD.forEach(r=>{
    const pr=prevMap[r.m];
    if(pr&&r.p!=null&&pr.p!=null){
      if(r.p>pr.p)upCnt++;
      if(r.p<pr.p)dnCnt++;
    }
  });

  const avgP=curD.filter(r=>r.p!=null).length?Math.round(mean(curD.filter(r=>r.p!=null).map(r=>r.p))):null;
  const topBrand={};curD.forEach(r=>{if(r.b)topBrand[r.b]=(topBrand[r.b]||0)+1;});
  const topB=Object.entries(topBrand).sort((a,b)=>b[1]-a[1]);

  document.getElementById('wsKpiGrid').innerHTML=[
    kpiCard(curModels.size,'Total Models','border-l-navy-800 bg-navy-50','text-navy-800'),
    kpiCard(avgP!=null?'SAR '+fmtSAR(avgP):'-','Avg Price','border-l-blue-500 bg-blue-50','text-blue-600'),
    kpiCard(upCnt,'Price Up','border-l-red-500 bg-red-50','text-red-600'),
    kpiCard(dnCnt,'Price Down','border-l-green-500 bg-green-50','text-green-600'),
    kpiCard(newM.size,'New Models','border-l-emerald-500 bg-emerald-50','text-emerald-600'),
    kpiCard(discM.size,'Discontinued','border-l-orange-500 bg-orange-50','text-orange-600'),
  ].join('');
}

function renderWsAlerts(){
  const{cur,prev}=getWsDates();
  const curD=applyGF(WS.filter(r=>r.d===cur));
  const prevD=applyGF(WS.filter(r=>r.d===prev));
  const q=ST.wsQ;
  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});
  let rows=[];
  curD.forEach(r=>{
    const pr=prevMap[r.m];
    if(!pr||r.p==null||pr.p==null)return;
    const chg=r.p-pr.p;
    if(chg===0)return;
    if(q&&!searchMatch(r,q))return;
    const chgPct=pr.p!==0?((chg/pr.p)*100).toFixed(1)+'%':'';
    rows.push({b:r.b,m:r.m,n:r.n,c:r.c,t:r.t,tp:r.tp,prev:pr.p,curr:r.p,chg,chgPct,dir:chg>0?'up':'down'});
  });
  if(ST.wsDir==='up')rows=rows.filter(r=>r.dir==='up');
  if(ST.wsDir==='down')rows=rows.filter(r=>r.dir==='down');
  rows.sort((a,b)=>Math.abs(b.chg)-Math.abs(a.chg));
  document.getElementById('ws_alert_count').textContent=rows.length+' changes';

  const cols=[{k:'b',l:'Brand'},{k:'m',l:'Model'},{k:'n',l:'Product'},{k:'c',l:'Category'},{k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'prev',l:'Prev Price'},{k:'curr',l:'Curr Price'},{k:'chg',l:'Change'},{k:'chgPct',l:'Change %'}];
  const tbl=document.getElementById('tblWsAlert');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblWsAlert',${i})">${c.l}</th>`).join('')+'</tr>';
  tbl.querySelector('tbody').innerHTML=rows.length?rows.map(r=>'<tr>'+cols.map(c=>{
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='t')return`<td>${r.t!=null?r.t.toFixed(1)+'T':'-'}</td>`;
    if(c.k==='prev'||c.k==='curr')return`<td>${fmtSAR(r[c.k])}</td>`;
    if(c.k==='chg')return`<td class="${r.chg>0?'up-cell':'dn-cell'}">${fmtChg(r.chg)}</td>`;
    if(c.k==='chgPct')return`<td class="${r.dir==='up'?'up-cell':'dn-cell'}">${r.chgPct}</td>`;
    return`<td>${r[c.k]||'-'}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="10" class="text-center text-gray-400 py-6">No price changes</td></tr>';
}

function renderWsNewDisc(){
  const{cur,prev}=getWsDates();
  const curD=applyGF(WS.filter(r=>r.d===cur));
  const prevD=applyGF(WS.filter(r=>r.d===prev));
  const curModels=new Set(curD.map(r=>r.m));
  const prevModels=new Set(prevD.map(r=>r.m));

  const newModels=curD.filter(r=>!prevModels.has(r.m));
  const discModels=prevD.filter(r=>!curModels.has(r.m));
  const card=(r,type)=>`<div class="border rounded-lg p-2 ${type==='new'?'border-green-200 bg-green-50':'border-red-200 bg-red-50'}"><div class="flex items-center gap-2"><span style="color:${colorOf(r.b)};font-weight:700" class="text-xs">${r.b}</span><span class="text-[10px] text-gray-500">${r.m}</span></div><div class="text-[10px] text-gray-600 mt-1 truncate">${r.n||'-'}</div><div class="flex gap-2 mt-1 text-[10px]"><span class="text-gray-500">${r.c||'-'}</span><span class="font-semibold">${r.p!=null?'SAR '+fmtSAR(r.p):'-'}</span></div></div>`;
  document.getElementById('wsNewGrid').innerHTML=newModels.length?newModels.map(r=>card(r,'new')).join(''):'<div class="text-gray-400 text-xs py-4 text-center">No new models</div>';
  document.getElementById('wsDiscGrid').innerHTML=discModels.length?discModels.map(r=>card(r,'disc')).join(''):'<div class="text-gray-400 text-xs py-4 text-center">No discontinued models</div>';
  document.getElementById('ws_new_count').textContent='('+newModels.length+')';
  document.getElementById('ws_disc_count').textContent='('+discModels.length+')';
}

function renderWsCatKpi(){
  const{cur}=getWsDates();
  let data=applyGF(WS.filter(r=>r.d===cur));
  if(ST.wsS4q)data=data.filter(r=>searchMatch(r,ST.wsS4q));

  function segKpi(d){
    const cnt=new Set(d.map(r=>r.m)).size;
    const ps=d.filter(r=>r.p!=null).map(r=>r.p);
    const avg=ps.length?Math.round(mean(ps)):null;
    const mn=ps.length?Math.min(...ps):null;
    const mx=ps.length?Math.max(...ps):null;
    // LG comparison
    const lgD=d.filter(r=>r.b==='LG'&&r.p!=null);const lgAvg=lgD.length?Math.round(mean(lgD.map(r=>r.p))):null;
    const lgGap=(lgAvg!=null&&avg!=null)?lgAvg-avg:null;
    return{cnt,avg,mn,mx,lgAvg,lgGap};
  }

  const cats=[...new Set(data.map(r=>r.c).filter(Boolean))].sort();
  const thCols=['Segment','Models','Avg Price','Min','Max','LG Avg','LG Gap'];
  const tbl=document.getElementById('tblWsCatKpi');
  tbl.querySelector('thead').innerHTML='<tr>'+thCols.map((c,i)=>`<th onclick="sortTbl('tblWsCatKpi',${i})">${c}</th>`).join('')+'</tr>';
  let rows='';
  cats.forEach(cat=>{
    const cd=data.filter(r=>r.c===cat);
    const ck=segKpi(cd);
    rows+=`<tr class="level-cat"><td><span class="type-badge type-cat">${cat}</span></td><td>${ck.cnt}</td><td>${fmtSAR(ck.avg)}</td><td>${fmtSAR(ck.mn)}</td><td>${fmtSAR(ck.mx)}</td><td>${fmtSAR(ck.lgAvg)}</td><td class="${ck.lgGap>0?'up-cell':ck.lgGap<0?'dn-cell':''}">${fmtChg(ck.lgGap)}</td></tr>`;
    // Type sub-level
    const types=[...new Set(cd.map(r=>r.tp).filter(Boolean))].sort();
    types.forEach(tp=>{
      const td=cd.filter(r=>r.tp===tp);
      const tk=segKpi(td);
      rows+=`<tr class="level-comp"><td>&nbsp;&nbsp;<span class="type-badge type-comp">${tp}</span></td><td>${tk.cnt}</td><td>${fmtSAR(tk.avg)}</td><td>${fmtSAR(tk.mn)}</td><td>${fmtSAR(tk.mx)}</td><td>${fmtSAR(tk.lgAvg)}</td><td class="${tk.lgGap>0?'up-cell':tk.lgGap<0?'dn-cell':''}">${fmtChg(tk.lgGap)}</td></tr>`;
      // Ton sub-level
      const tons=[...new Set(td.map(r=>r.t).filter(v=>v!=null))].sort((a,b)=>a-b);
      tons.forEach(ton=>{
        const tond=td.filter(r=>r.t===ton);
        const tonk=segKpi(tond);
        rows+=`<tr class="level-ton"><td>&nbsp;&nbsp;&nbsp;&nbsp;<span class="type-badge type-ton">${ton.toFixed(1)}T</span></td><td>${tonk.cnt}</td><td>${fmtSAR(tonk.avg)}</td><td>${fmtSAR(tonk.mn)}</td><td>${fmtSAR(tonk.mx)}</td><td>${fmtSAR(tonk.lgAvg)}</td><td class="${tonk.lgGap>0?'up-cell':tonk.lgGap<0?'dn-cell':''}">${fmtChg(tonk.lgGap)}</td></tr>`;
      });
    });
  });
  // Total row
  const tk=segKpi(data);
  rows+=`<tr class="level-cat" style="border-top:2px solid #1F4E79"><td><span class="type-badge type-cat">TOTAL</span></td><td>${tk.cnt}</td><td>${fmtSAR(tk.avg)}</td><td>${fmtSAR(tk.mn)}</td><td>${fmtSAR(tk.mx)}</td><td>${fmtSAR(tk.lgAvg)}</td><td class="${tk.lgGap>0?'up-cell':tk.lgGap<0?'dn-cell':''}">${fmtChg(tk.lgGap)}</td></tr>`;
  tbl.querySelector('tbody').innerHTML=rows;
}

function renderWsBrandBar(){
  const{cur}=getWsDates();
  const data=applyGF(WS.filter(r=>r.d===cur));
  const bAvg={};data.forEach(r=>{if(r.p==null)return;if(!bAvg[r.b])bAvg[r.b]=[];bAvg[r.b].push(r.p);});
  let items=Object.entries(bAvg).map(([b,v])=>({b,avg:Math.round(mean(v))})).sort((a,b)=>a.avg-b.avg);
  const ctx=document.getElementById('wsBrandBar').getContext('2d');
  if(wsBrandBarObj)wsBrandBarObj.destroy();
  wsBrandBarObj=new Chart(ctx,{type:'bar',data:{labels:items.map(x=>x.b),datasets:[{label:'Avg Price',data:items.map(x=>x.avg),backgroundColor:items.map(x=>colorOf(x.b)),borderWidth:1}]},options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},title:{display:true,text:'Wholesale Avg Price by Brand',font:{size:12}},datalabels:{display:true,anchor:'end',align:'right',color:'#1e3a5f',font:{size:10,weight:'bold'},formatter:v=>fmtSAR(v),clip:false}},scales:{x:{ticks:{callback:v=>fmtSAR(v)},grid:{color:'#f0f0f0'},suggestedMax:items.length?Math.max(...items.map(x=>x.avg))*1.15:undefined},y:{ticks:{font:{size:10}}}}}});
}

function renderWsTrend(){
  const data=applyGF(WS);
  const q=ST.wsS6q;
  let datasets=[];
  const mkt=WS_DATES.map(d=>{let dd=applyGF(WS.filter(r=>r.d===d));if(q)dd=dd.filter(r=>searchMatch(r,q));dd=dd.filter(r=>r.p!=null);return dd.length?Math.round(mean(dd.map(r=>r.p))):null;});
  datasets.push({label:'Market Avg',data:mkt,borderColor:'#6b7280',borderWidth:2,borderDash:[6,3],pointRadius:2,tension:.3,fill:false});
  const bSet=GF.brand.getSelected();
  [...bSet].forEach(br=>{
    const vals=WS_DATES.map(d=>{let dd=applyGF(WS.filter(r=>r.d===d&&r.b===br));if(q)dd=dd.filter(r=>searchMatch(r,q));dd=dd.filter(r=>r.p!=null);return dd.length?Math.round(mean(dd.map(r=>r.p))):null;});
    if(vals.every(v=>v===null))return;const c=colorOf(br);
    datasets.push({label:br,data:vals,borderColor:c,backgroundColor:alphaC(c,.08),borderWidth:1.5,pointRadius:2,tension:.3,fill:false,spanGaps:true});
  });
  const ctx=document.getElementById('wsTrendChart').getContext('2d');
  if(wsTrendObj)wsTrendObj.destroy();
  wsTrendObj=new Chart(ctx,{type:'line',data:{labels:WS_DATES,datasets},options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},layout:{padding:{right:90}},plugins:{legend:{position:'bottom',labels:{boxWidth:10,font:{size:10}}},tooltip:{callbacks:{label:c=>c.dataset.label+': SAR '+fmtSAR(c.raw)}},datalabels:{display:false}},scales:{x:{ticks:{font:{size:10},maxRotation:45},grid:{color:'#f5f5f5'}},y:{ticks:{callback:v=>fmtSAR(v)},grid:{color:'#f5f5f5'}}}},plugins:[endLabelsPlugin]});
}

function renderWsFull(){
  const{cur,prev}=getWsDates();
  let curD=applyGF(WS.filter(r=>r.d===cur));
  const prevD=applyGF(WS.filter(r=>r.d===prev));
  if(ST.wsS7q)curD=curD.filter(r=>searchMatch(r,ST.wsS7q));
  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});

  const cols=[{k:'b',l:'Brand'},{k:'m',l:'Model'},{k:'n',l:'Product'},{k:'c',l:'Category'},{k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'btu',l:'BTU'},{k:'p',l:'Price'},{k:'prev',l:'Prev Price'},{k:'chg',l:'Change'},{k:'chgPct',l:'Chg%'},{k:'st',l:'Status'}];
  const tbl=document.getElementById('tblWsFull');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblWsFull',${i})">${c.l}</th>`).join('')+'</tr>';

  _wsFullData=curD.map(r=>{
    const pr=prevMap[r.m];
    const prev=pr?pr.p:null;
    const chg=(r.p!=null&&prev!=null)?r.p-prev:null;
    const chgPct=(chg!=null&&prev)?((chg/prev)*100).toFixed(1)+'%':null;
    return{...r,prev,chg,chgPct};
  });

  tbl.querySelector('tbody').innerHTML=_wsFullData.length?_wsFullData.map(r=>'<tr>'+cols.map(c=>{
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='t')return`<td>${r.t!=null?r.t.toFixed(1)+'T':'-'}</td>`;
    if(c.k==='btu')return`<td>${r.btu||'-'}</td>`;
    if(c.k==='p'||c.k==='prev')return`<td>${fmtSAR(r[c.k])}</td>`;
    if(c.k==='chg')return`<td class="${r.chg>0?'up-cell':r.chg<0?'dn-cell':''}">${fmtChg(r.chg)}</td>`;
    if(c.k==='chgPct')return`<td class="${r.chg>0?'up-cell':r.chg<0?'dn-cell':''}">${r.chgPct||'-'}</td>`;
    if(c.k==='st'){const sc=r.st==='New'?'text-green-600':r.st==='Discontinue'?'text-red-600':r.st==='Re-Active'?'text-blue-600':'text-gray-600';return`<td class="${sc} font-semibold text-xs">${r.st||'-'}</td>`;}
    return`<td>${r[c.k]||'-'}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="12" class="text-center text-gray-400 py-6">No data</td></tr>';
}
let _wsFullData=[];

function downloadWsExcel(){
  if(!_wsFullData.length){alert('No data');return;}
  const{cur}=getWsDates();
  const xl=_wsFullData.map(r=>({Brand:r.b,Model:r.m,Product:r.n,Category:r.c,Type:r.tp,Ton:r.t,BTU:r.btu,Price:r.p,'Prev Price':r.prev,Change:r.chg,'Change %':r.chgPct,Status:r.st}));
  const ws=XLSX.utils.json_to_sheet(xl);const wb=XLSX.utils.book_new();XLSX.utils.book_append_sheet(wb,ws,'WS Data');XLSX.writeFile(wb,'BH_Wholesale_'+cur+'.xlsx');
}

function renderWS(){renderWsKpis();renderWsAlerts();renderWsNewDisc();renderWsCatKpi();renderWsBrandBar();renderWsTrend();renderWsFull();}

// ══════════════════════════════════════════════════════════════════
// TAB 3: RETAIL
// ══════════════════════════════════════════════════════════════════
let rtBrandBarObj=null,rtTrendObj=null;

function getRtWeeks(){
  return{cur:document.getElementById('rt_cur_week').value,prev:document.getElementById('rt_prev_week').value};
}

function renderRtKpis(){
  const{cur,prev}=getRtWeeks();
  const curD=applyGF(RT.filter(r=>r.d===cur));
  const prevD=prev?applyGF(RT.filter(r=>r.d===prev)):[];
  document.getElementById('rt_compare').textContent='Viewing: '+cur+(prev?' vs '+prev:'');

  const curModels=new Set(curD.map(r=>r.m));
  const prevModels=new Set(prevD.map(r=>r.m));
  const newM=[...curModels].filter(m=>!prevModels.has(m)).length;
  const discM=[...prevModels].filter(m=>!curModels.has(m)).length;

  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});
  let upCnt=0,dnCnt=0;
  curD.forEach(r=>{
    const pr=prevMap[r.m];
    if(pr&&r.cp!=null&&pr.cp!=null){
      if(r.cp>pr.cp)upCnt++;
      if(r.cp<pr.cp)dnCnt++;
    }
  });

  const avgCp=curD.filter(r=>r.cp!=null).length?Math.round(mean(curD.filter(r=>r.cp!=null).map(r=>r.cp))):null;
  const drs=curD.filter(r=>r.dr!=null).map(r=>r.dr*100);
  const avgDisc=drs.length?Math.round(mean(drs)*10)/10:null;

  document.getElementById('rtKpiGrid').innerHTML=[
    kpiCard(curModels.size,'Total SKUs','border-l-navy-800 bg-navy-50','text-navy-800'),
    kpiCard(avgCp!=null?'SAR '+fmtSAR(avgCp):'-','Avg Current Price','border-l-blue-500 bg-blue-50','text-blue-600'),
    kpiCard(avgDisc!=null?avgDisc.toFixed(1)+'%':'-','Avg Discount','border-l-purple-500 bg-purple-50','text-purple-700'),
    kpiCard(upCnt,'Price Up','border-l-red-500 bg-red-50','text-red-600'),
    kpiCard(dnCnt,'Price Down','border-l-green-500 bg-green-50','text-green-600'),
    kpiCard(newM+' / '+discM,'New / Removed','border-l-amber-500 bg-amber-50','text-amber-700'),
  ].join('');
}

function renderRtAlerts(){
  const{cur,prev}=getRtWeeks();
  const pk=document.getElementById('rt_alert_pt').value;
  const curD=applyGF(RT.filter(r=>r.d===cur));
  const prevD=applyGF(RT.filter(r=>r.d===prev));
  const q=ST.rtQ;
  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});
  let rows=[];
  curD.forEach(r=>{
    const pr=prevMap[r.m];
    if(!pr||r[pk]==null||pr[pk]==null)return;
    const chg=r[pk]-pr[pk];
    if(chg===0)return;
    if(q&&!searchMatch(r,q))return;
    const chgPct=pr[pk]!==0?((chg/pr[pk])*100).toFixed(1)+'%':'';
    rows.push({b:r.b,m:r.m,n:r.n,c:r.c,tp:r.tp,t:r.t,prev:pr[pk],curr:r[pk],chg,chgPct,dir:chg>0?'up':'down'});
  });
  if(ST.rtDir==='up')rows=rows.filter(r=>r.dir==='up');
  if(ST.rtDir==='down')rows=rows.filter(r=>r.dir==='down');
  rows.sort((a,b)=>Math.abs(b.chg)-Math.abs(a.chg));
  document.getElementById('rt_alert_count').textContent=rows.length+' changes';

  const cols=[{k:'b',l:'Brand'},{k:'m',l:'Model'},{k:'n',l:'Product'},{k:'c',l:'Category'},{k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'prev',l:'Prev'},{k:'curr',l:'Current'},{k:'chg',l:'Change'},{k:'chgPct',l:'Chg%'}];
  const tbl=document.getElementById('tblRtAlert');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblRtAlert',${i})">${c.l}</th>`).join('')+'</tr>';
  tbl.querySelector('tbody').innerHTML=rows.length?rows.map(r=>'<tr>'+cols.map(c=>{
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='n')return`<td>${bhLink(r.m,r.n||'-')}</td>`;
    if(c.k==='t')return`<td>${r.t!=null?r.t.toFixed(1)+'T':'-'}</td>`;
    if(c.k==='prev'||c.k==='curr')return`<td>${fmtSAR(r[c.k])}</td>`;
    if(c.k==='chg')return`<td class="${r.chg>0?'up-cell':'dn-cell'}">${fmtChg(r.chg)}</td>`;
    if(c.k==='chgPct')return`<td class="${r.dir==='up'?'up-cell':'dn-cell'}">${r.chgPct}</td>`;
    return`<td>${r[c.k]||'-'}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="10" class="text-center text-gray-400 py-6">No price changes</td></tr>';
}

function renderRtNewDisc(){
  const{cur,prev}=getRtWeeks();
  const curD=applyGF(RT.filter(r=>r.d===cur));
  const prevD=applyGF(RT.filter(r=>r.d===prev));
  const curModels=new Set(curD.map(r=>r.m));
  const prevModels=new Set(prevD.map(r=>r.m));
  const newModels=curD.filter(r=>!prevModels.has(r.m));
  const discModels=prevD.filter(r=>!curModels.has(r.m));

  const card=(r,type)=>`<div class="border rounded-lg p-2 ${type==='new'?'border-green-200 bg-green-50':'border-red-200 bg-red-50'}"><div class="flex items-center gap-2"><span style="color:${colorOf(r.b)};font-weight:700" class="text-xs">${r.b}</span><span class="text-[10px] text-gray-500">${r.m}</span></div><div class="text-[10px] text-gray-600 mt-1 truncate">${r.n||'-'}</div><div class="flex gap-2 mt-1 text-[10px]"><span class="text-gray-500">${r.c||'-'}</span><span class="font-semibold">${r.cp!=null?'SAR '+fmtSAR(r.cp):'-'}</span></div></div>`;
  document.getElementById('rtNewGrid').innerHTML=newModels.length?newModels.map(r=>card(r,'new')).join(''):'<div class="text-gray-400 text-xs py-4 text-center">No new SKUs</div>';
  document.getElementById('rtDiscGrid').innerHTML=discModels.length?discModels.map(r=>card(r,'disc')).join(''):'<div class="text-gray-400 text-xs py-4 text-center">No removed SKUs</div>';
  document.getElementById('rt_new_count').textContent='('+newModels.length+')';
  document.getElementById('rt_disc_count').textContent='('+discModels.length+')';
}

function renderRtCatKpi(){
  const{cur}=getRtWeeks();
  const pk=document.getElementById('rt_s4_pt').value;
  let data=applyGF(RT.filter(r=>r.d===cur));
  if(ST.rtS4q)data=data.filter(r=>searchMatch(r,ST.rtS4q));

  function segKpi(d){
    const cnt=new Set(d.map(r=>r.m)).size;
    const ps=d.filter(r=>r[pk]!=null).map(r=>r[pk]);
    const avg=ps.length?Math.round(mean(ps)):null;
    const rps=d.filter(r=>r.rp!=null).map(r=>r.rp);
    const avgReg=rps.length?Math.round(mean(rps)):null;
    const drs=d.filter(r=>r.dr!=null).map(r=>r.dr*100);
    const avgDisc=drs.length?Math.round(mean(drs)*10)/10:null;
    const lgD=d.filter(r=>r.b==='LG'&&r[pk]!=null);const lgAvg=lgD.length?Math.round(mean(lgD.map(r=>r[pk]))):null;
    const lgGap=(lgAvg!=null&&avg!=null)?lgAvg-avg:null;
    return{cnt,avg,avgReg,avgDisc,lgAvg,lgGap};
  }

  const cats=[...new Set(data.map(r=>r.c).filter(Boolean))].sort();
  const thCols=['Segment','SKUs','Avg Price','Avg Regular','Avg Disc%','LG Avg','LG Gap'];
  const tbl=document.getElementById('tblRtCatKpi');
  tbl.querySelector('thead').innerHTML='<tr>'+thCols.map((c,i)=>`<th onclick="sortTbl('tblRtCatKpi',${i})">${c}</th>`).join('')+'</tr>';
  let rows='';
  cats.forEach(cat=>{
    const cd=data.filter(r=>r.c===cat);
    const ck=segKpi(cd);
    rows+=`<tr class="level-cat"><td><span class="type-badge type-cat">${cat}</span></td><td>${ck.cnt}</td><td>${fmtSAR(ck.avg)}</td><td>${fmtSAR(ck.avgReg)}</td><td>${ck.avgDisc!=null?ck.avgDisc.toFixed(1)+'%':'-'}</td><td>${fmtSAR(ck.lgAvg)}</td><td class="${ck.lgGap>0?'up-cell':ck.lgGap<0?'dn-cell':''}">${fmtChg(ck.lgGap)}</td></tr>`;
    const types=[...new Set(cd.map(r=>r.tp).filter(Boolean))].sort();
    types.forEach(tp=>{
      const td=cd.filter(r=>r.tp===tp);
      const tk=segKpi(td);
      rows+=`<tr class="level-comp"><td>&nbsp;&nbsp;<span class="type-badge type-comp">${tp}</span></td><td>${tk.cnt}</td><td>${fmtSAR(tk.avg)}</td><td>${fmtSAR(tk.avgReg)}</td><td>${tk.avgDisc!=null?tk.avgDisc.toFixed(1)+'%':'-'}</td><td>${fmtSAR(tk.lgAvg)}</td><td class="${tk.lgGap>0?'up-cell':tk.lgGap<0?'dn-cell':''}">${fmtChg(tk.lgGap)}</td></tr>`;
      const tons=[...new Set(td.map(r=>r.t).filter(v=>v!=null))].sort((a,b)=>a-b);
      tons.forEach(ton=>{
        const tond=td.filter(r=>r.t===ton);
        const tonk=segKpi(tond);
        rows+=`<tr class="level-ton"><td>&nbsp;&nbsp;&nbsp;&nbsp;<span class="type-badge type-ton">${ton.toFixed(1)}T</span></td><td>${tonk.cnt}</td><td>${fmtSAR(tonk.avg)}</td><td>${fmtSAR(tonk.avgReg)}</td><td>${tonk.avgDisc!=null?tonk.avgDisc.toFixed(1)+'%':'-'}</td><td>${fmtSAR(tonk.lgAvg)}</td><td class="${tonk.lgGap>0?'up-cell':tonk.lgGap<0?'dn-cell':''}">${fmtChg(tonk.lgGap)}</td></tr>`;
      });
    });
  });
  const tk=segKpi(data);
  rows+=`<tr class="level-cat" style="border-top:2px solid #1F4E79"><td><span class="type-badge type-cat">TOTAL</span></td><td>${tk.cnt}</td><td>${fmtSAR(tk.avg)}</td><td>${fmtSAR(tk.avgReg)}</td><td>${tk.avgDisc!=null?tk.avgDisc.toFixed(1)+'%':'-'}</td><td>${fmtSAR(tk.lgAvg)}</td><td class="${tk.lgGap>0?'up-cell':tk.lgGap<0?'dn-cell':''}">${fmtChg(tk.lgGap)}</td></tr>`;
  tbl.querySelector('tbody').innerHTML=rows;
}

function renderRtBrandBar(){
  const{cur}=getRtWeeks();
  const pk=document.getElementById('rt_s5_pt').value;
  const data=applyGF(RT.filter(r=>r.d===cur));
  const bAvg={};data.forEach(r=>{if(r[pk]==null)return;if(!bAvg[r.b])bAvg[r.b]=[];bAvg[r.b].push(r[pk]);});
  let items=Object.entries(bAvg).map(([b,v])=>({b,avg:Math.round(mean(v))})).sort((a,b)=>a.avg-b.avg);
  const ctx=document.getElementById('rtBrandBar').getContext('2d');
  if(rtBrandBarObj)rtBrandBarObj.destroy();
  rtBrandBarObj=new Chart(ctx,{type:'bar',data:{labels:items.map(x=>x.b),datasets:[{label:'Avg Price',data:items.map(x=>x.avg),backgroundColor:items.map(x=>colorOf(x.b)),borderWidth:1}]},options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},title:{display:true,text:'Retail Avg Price by Brand ('+pk.toUpperCase()+')',font:{size:12}},datalabels:{display:true,anchor:'end',align:'right',color:'#1e3a5f',font:{size:10,weight:'bold'},formatter:v=>fmtSAR(v),clip:false}},scales:{x:{ticks:{callback:v=>fmtSAR(v)},grid:{color:'#f0f0f0'},suggestedMax:items.length?Math.max(...items.map(x=>x.avg))*1.15:undefined},y:{ticks:{font:{size:10}}}}}});
}

function renderRtTrend(){
  const pk=document.getElementById('rt_s6_pt').value;
  const q=ST.rtS6q;
  let datasets=[];
  const mkt=RT_DATES.map(w=>{let dd=applyGF(RT.filter(r=>r.d===w));if(q)dd=dd.filter(r=>searchMatch(r,q));dd=dd.filter(r=>r[pk]!=null);return dd.length?Math.round(mean(dd.map(r=>r[pk]))):null;});
  datasets.push({label:'Market Avg',data:mkt,borderColor:'#6b7280',borderWidth:2,borderDash:[6,3],pointRadius:2,tension:.3,fill:false});
  const bSet=GF.brand.getSelected();
  [...bSet].forEach(br=>{
    const vals=RT_DATES.map(w=>{let dd=applyGF(RT.filter(r=>r.d===w&&r.b===br));if(q)dd=dd.filter(r=>searchMatch(r,q));dd=dd.filter(r=>r[pk]!=null);return dd.length?Math.round(mean(dd.map(r=>r[pk]))):null;});
    if(vals.every(v=>v===null))return;const c=colorOf(br);
    datasets.push({label:br,data:vals,borderColor:c,backgroundColor:alphaC(c,.08),borderWidth:1.5,pointRadius:2,tension:.3,fill:false,spanGaps:true});
  });
  const ctx=document.getElementById('rtTrendChart').getContext('2d');
  if(rtTrendObj)rtTrendObj.destroy();
  rtTrendObj=new Chart(ctx,{type:'line',data:{labels:RT_DATES,datasets},options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},layout:{padding:{right:90}},plugins:{legend:{position:'bottom',labels:{boxWidth:10,font:{size:10}}},tooltip:{callbacks:{label:c=>c.dataset.label+': SAR '+fmtSAR(c.raw)}},datalabels:{display:false}},scales:{x:{ticks:{font:{size:10}},grid:{color:'#f5f5f5'}},y:{ticks:{callback:v=>fmtSAR(v)},grid:{color:'#f5f5f5'}}}},plugins:[endLabelsPlugin]});
}

function renderRtFull(){
  const{cur,prev}=getRtWeeks();
  let curD=applyGF(RT.filter(r=>r.d===cur));
  const prevD=applyGF(RT.filter(r=>r.d===prev));
  if(ST.rtS7q)curD=curD.filter(r=>searchMatch(r,ST.rtS7q));
  const prevMap={};prevD.forEach(r=>{prevMap[r.m]=r;});

  const cols=[{k:'b',l:'Brand'},{k:'m',l:'Model'},{k:'n',l:'Product'},{k:'c',l:'Category'},{k:'tp',l:'Type'},{k:'t',l:'Ton'},{k:'rp',l:'Regular'},{k:'cp',l:'Current'},{k:'dr',l:'Disc%'},{k:'ds',l:'Disc SAR'},{k:'prev',l:'Prev'},{k:'chg',l:'Change'},{k:'chgPct',l:'Chg%'}];
  const tbl=document.getElementById('tblRtFull');
  tbl.querySelector('thead').innerHTML='<tr>'+cols.map((c,i)=>`<th onclick="sortTbl('tblRtFull',${i})">${c.l}</th>`).join('')+'</tr>';

  _rtFullData=curD.map(r=>{
    const pr=prevMap[r.m];
    const prev=pr?pr.cp:null;
    const chg=(r.cp!=null&&prev!=null)?r.cp-prev:null;
    const chgPct=(chg!=null&&prev)?((chg/prev)*100).toFixed(1)+'%':null;
    return{...r,prev,chg,chgPct};
  });

  tbl.querySelector('tbody').innerHTML=_rtFullData.length?_rtFullData.map(r=>'<tr>'+cols.map(c=>{
    if(c.k==='b')return`<td><span style="color:${colorOf(r.b)};font-weight:600">${r.b}</span></td>`;
    if(c.k==='n')return`<td>${bhLink(r.m,r.n||'-')}</td>`;
    if(c.k==='t')return`<td>${r.t!=null?r.t.toFixed(1)+'T':'-'}</td>`;
    if(c.k==='rp'||c.k==='cp'||c.k==='prev'||c.k==='ds')return`<td>${fmtSAR(r[c.k])}</td>`;
    if(c.k==='dr')return`<td>${r.dr!=null?(r.dr*100).toFixed(1)+'%':'-'}</td>`;
    if(c.k==='chg')return`<td class="${r.chg>0?'up-cell':r.chg<0?'dn-cell':''}">${fmtChg(r.chg)}</td>`;
    if(c.k==='chgPct')return`<td class="${r.chg>0?'up-cell':r.chg<0?'dn-cell':''}">${r.chgPct||'-'}</td>`;
    return`<td>${r[c.k]||'-'}</td>`;
  }).join('')+'</tr>').join(''):'<tr><td colspan="13" class="text-center text-gray-400 py-6">No data</td></tr>';
}
let _rtFullData=[];

function downloadRtExcel(){
  if(!_rtFullData.length){alert('No data');return;}
  const{cur}=getRtWeeks();
  const xl=_rtFullData.map(r=>({Brand:r.b,Model:r.m,'Product Name':r.n,Category:r.c,Type:r.tp,Ton:r.t,BTU:r.btu,'Regular Price':r.rp,'Current Price':r.cp,'Discount %':r.dr!=null?Math.round(r.dr*1000)/10:null,'Discount SAR':r.ds,'Prev Price':r.prev,Change:r.chg,'Change %':r.chgPct}));
  const ws=XLSX.utils.json_to_sheet(xl);const wb=XLSX.utils.book_new();XLSX.utils.book_append_sheet(wb,ws,'RT Data');XLSX.writeFile(wb,'BH_Retail_'+cur+'.xlsx');
}

function renderRT(){renderRtKpis();renderRtAlerts();renderRtNewDisc();renderRtCatKpi();renderRtBrandBar();renderRtTrend();renderRtFull();}

// ══════════════════════════════════════════════════════════════════
// INIT
// ══════════════════════════════════════════════════════════════════
function refreshAll(){
  if(activeTab==='integrated')renderIntegrated();
  else if(activeTab==='wholesale')renderWS();
  else if(activeTab==='retail')renderRT();
}

function init(){
  // Global Filters
  GF={
    cat: new MS(document.getElementById('gf_cat'),CATEGORIES,'Category',refreshAll),
    type: new MS(document.getElementById('gf_type'),TYPES,'Type',refreshAll),
    ton: new MS(document.getElementById('gf_ton'),TONS.map(String),'Ton',refreshAll),
    brand: new MS(document.getElementById('gf_brand'),BRANDS,'Brand',refreshAll,colorOf),
  };

  // Integrated date selectors
  const intWs=document.getElementById('int_ws_date');
  const intRt=document.getElementById('int_rt_week');
  WS_DATES.forEach(d=>intWs.add(new Option(d,d)));
  RT_DATES.forEach(w=>intRt.add(new Option(w,w)));
  if(WS_DATES.length)intWs.value=WS_DATES[WS_DATES.length-1];
  if(RT_DATES.length)intRt.value=RT_DATES[RT_DATES.length-1];
  intWs.addEventListener('change',renderIntegrated);
  intRt.addEventListener('change',renderIntegrated);

  // Wholesale date selectors
  const wsCur=document.getElementById('ws_cur_date');
  const wsPrev=document.getElementById('ws_prev_date');
  WS_DATES.forEach(d=>{wsCur.add(new Option(d,d));wsPrev.add(new Option(d,d));});
  if(WS_DATES.length>=2){wsCur.value=WS_DATES[WS_DATES.length-1];wsPrev.value=WS_DATES[WS_DATES.length-2];}
  else if(WS_DATES.length===1){wsCur.value=WS_DATES[0];wsPrev.value=WS_DATES[0];}

  // Retail date selectors
  const rtCur=document.getElementById('rt_cur_week');
  const rtPrev=document.getElementById('rt_prev_week');
  RT_DATES.forEach(w=>{rtCur.add(new Option(w,w));rtPrev.add(new Option(w,w));});
  if(RT_DATES.length>=2){rtCur.value=RT_DATES[RT_DATES.length-1];rtPrev.value=RT_DATES[RT_DATES.length-2];}
  else if(RT_DATES.length===1){rtCur.value=RT_DATES[0];rtPrev.value=RT_DATES[0];}

  // Search handlers
  const searches=[
    ['int_search','intQ',renderIntGap],
    ['ws_alert_search','wsQ',renderWsAlerts],
    ['ws_s4_search','wsS4q',renderWsCatKpi],
    ['ws_s6_search','wsS6q',renderWsTrend],
    ['ws_s7_search','wsS7q',renderWsFull],
    ['rt_alert_search','rtQ',renderRtAlerts],
    ['rt_s4_search','rtS4q',renderRtCatKpi],
    ['rt_s6_search','rtS6q',renderRtTrend],
    ['rt_s7_search','rtS7q',renderRtFull],
  ];
  searches.forEach(([id,key,fn])=>{
    let to=null;const el=document.getElementById(id);
    if(el)el.addEventListener('input',e=>{clearTimeout(to);to=setTimeout(()=>{ST[key]=e.target.value.toLowerCase().trim();fn();},150);});
  });

  // Meta
  document.getElementById('metaWS').textContent=new Set(WS.map(r=>r.m)).size;
  document.getElementById('metaRT').textContent=new Set(RT.map(r=>r.m)).size;
  document.getElementById('metaMatch').textContent=MODEL_OVERLAP.matched;

  refreshAll();
}

document.addEventListener('DOMContentLoaded',init);
</script>
"""

    # ── Assemble HTML ──────────────────────────────────────────────
    html = HTML_HEAD + HTML_BODY.replace('GENERATED_AT', generated_at) + HTML_DATA + HTML_LOGIC + "\n</body>\n</html>"

    print("[3/4] Writing HTML...")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(html)
    fsize = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"  → {OUTPUT_FILE} ({fsize:.0f} KB)")

    # ── Deploy to GitHub Pages ─────────────────────────────────────
    print("[4/5] Deploying to GitHub Pages...")
    DEPLOY_DIR = os.path.join(os.path.expanduser("~"), "tmp_deploy_bh")
    if os.path.isdir(os.path.join(DEPLOY_DIR, ".git")):
        try:
            shutil.copy2(OUTPUT_FILE, os.path.join(DEPLOY_DIR, "index.html"))
            subprocess.run(["git", "add", "index.html"], cwd=DEPLOY_DIR, check=True, capture_output=True)
            subprocess.run(["git", "commit", "-m", f"Update BH dashboard {generated_at}"], cwd=DEPLOY_DIR, check=True, capture_output=True)
            try:
                subprocess.run(["git", "push"], cwd=DEPLOY_DIR, check=True, capture_output=True)
            except subprocess.CalledProcessError:
                subprocess.run(["git", "pull", "--rebase"], cwd=DEPLOY_DIR, capture_output=True)
                subprocess.run(["git", "push"], cwd=DEPLOY_DIR, check=True, capture_output=True)
            print("  → Deployed to GitHub Pages")
        except subprocess.CalledProcessError as e:
            print(f"  → Deploy skipped: {e}")
    else:
        print("  → No deploy repo found (~/tmp_deploy_bh). Skipping.")

    # ── Deploy to Cloudflare (Shaker-MD-App) ──────────────────────
    print("[5/5] Deploying to Cloudflare...")
    if os.path.exists(os.path.join(SHAKER_DIR, ".git")):
        os.makedirs(CLOUDFLARE_DEST, exist_ok=True)
        dest = os.path.join(CLOUDFLARE_DEST, "index.html")
        shutil.copy2(OUTPUT_FILE, dest)
        print(f"  📋 Copied to {dest}")
        try:
            subprocess.run(["git", "add", "docs/dashboards/bh-price/index.html"],
                           cwd=SHAKER_DIR, check=True, capture_output=True)
            result = subprocess.run(
                ["git", "commit", "-m",
                 f"Update BH Price dashboard ({generated_at})"],
                cwd=SHAKER_DIR, capture_output=True, text=True
            )
            if result.returncode == 0:
                push = subprocess.run(["git", "push"], cwd=SHAKER_DIR,
                                      capture_output=True, text=True)
                if push.returncode != 0:
                    subprocess.run(["git", "pull", "--rebase"], cwd=SHAKER_DIR,
                                   capture_output=True)
                    subprocess.run(["git", "push"], cwd=SHAKER_DIR,
                                   capture_output=True)
                print("  🚀 Pushed to Shaker-MD-App (Cloudflare auto-deploy)")
            else:
                if "nothing to commit" in (result.stdout or "") + (result.stderr or ""):
                    print("  ℹ️ No changes to deploy (dashboard unchanged)")
                else:
                    print(f"  ⚠️ Git commit failed: {result.stderr}")
        except Exception as e:
            print(f"  ⚠️ Cloudflare deploy error: {e}")
    else:
        print(f"  ⚠️ Shaker-MD-App not found at {SHAKER_DIR}")

    print("\nDone!")


if __name__ == "__main__":
    main()
