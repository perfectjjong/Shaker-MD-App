"""
BH AC Price Tracker — Unified Pipeline
=======================================
1. Retail: BH Store API 스크래핑 -> Weekly_Price_DB 시트 업데이트 (매일)
2. Wholesale: BH_Subdealer_Pricelist_*.xlsx -> Whole selling Price_Master (신규 파일 있을 때만)
3. Model Info: 3단계 우선순위 (사용자 편집 > Retail API > Wholesale 자동추출)

Output: BH_Subdealer_AC_Master.xlsx
  - "Total_Model Info": 모델 메타데이터 (수동 편집 보존)
  - "Whole selling Price_Master": Wholesale 가격 추적
  - "Weekly_Price_DB": Retail 가격 추적 (API 스크래핑)

실행: python consolidate_ac.py
"""

import sys
import os, glob, re, unicodedata
from datetime import datetime

# Windows 콘솔 인코딩 문제 방지
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ── BH Store Retail API 설정 ─────────────────────────────────────
API_URL = "https://api.bhstore.com.sa/commerce/products/"
API_KEY = "2853152294a192f18c3da51ae965f"

API_HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "accept": "application/json",
    "origin": "https://bhstore.com.sa",
    "referer": "https://bhstore.com.sa/sa-en",
    "interface": "web",
    "language": "en",
    "x-api-key": API_KEY,
}

# BH Store category IDs
RETAIL_CATEGORIES = {
    "Split": 12,
    "Window": 7,
    "Floor Standing": 21,
    "Concealed": 15,
    "Cassette": 18,
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(SCRIPT_DIR, "01. Sub Dealer Price List", "01. Excel")
INPUT_PATTERN = os.path.join(INPUT_DIR, "BH_Subdealer_Pricelist_*.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "BH_Subdealer_AC_Master.xlsx")

# ── Retail data source (Weekly_Price_DB sheet in the same output file) ──
RETAIL_SHEET = "Weekly_Price_DB"
# Weekly_Price_DB columns: Week(0), Retailer(1), Brand(2), Model Code(3),
#   Product Name(4), Type(5), BTU(6), CO/C&H(7), ...

# ── Category name normalization (Retail -> Wholesale naming) ────────
CATEGORY_NORMALIZE = {
    "Split": "Split AC",
    "Window": "Window AC",
    "Floor Standing": "Floor Standing AC",
    "Concealed": "Concealed AC",
    "Cassette": "Cassette AC",
}

# ── Type normalization (Retail CO/C&H -> Wholesale CO/H&C) ─────────
TYPE_NORMALIZE = {
    "CO": "CO",
    "C&H": "H&C",
    "H&C": "H&C",
}

# ── Categories to EXCLUDE ──────────────────────────────────────────
EXCLUDED_CATEGORIES = {"Portable AC", "Desert Cooler", "Cassette AC Face Panel"}

# ── BTU -> Ton mapping table ────────────────────────────────────────
# Each tuple: (btu_value, lower_bound_inclusive, upper_bound_exclusive, ton)
BTU_TO_TON = [
    (12000,  9000, 15000, 1.0),
    (18000, 15000, 21000, 1.5),
    (24000, 21000, 27000, 2.0),
    (30000, 27000, 33000, 2.5),
    (36000, 33000, 39000, 3.0),
    (42000, 39000, 45000, 3.5),
    (48000, 45000, 51000, 4.0),
    (54000, 51000, 57000, 4.5),
    (60000, 57000, 63000, 5.0),
]

def btu_to_ton(btu):
    """Convert BTU value to Ton (Capacity) using range mapping."""
    if btu is None:
        return None
    try:
        btu = float(btu)
    except (ValueError, TypeError):
        return None
    for _, lo, hi, ton in BTU_TO_TON:
        if lo <= btu < hi:
            return ton
    return None

# ── Brand normalization (fix past mistranslations) ─────────────────
BRAND_NORMALIZE = {
    "Kraft": "Craft",
}

# ── Arabic -> English brand mapping ──────────────────────────────────
BRAND_MAP_AR = {
    "استون": "Aston",
    "بان كول": "Pan Cool",
    "بان كوول": "Pan Cool",
    "كارير": "Carrier",
    "كاريير": "Carrier",
    "تي سي ال": "TCL",
    "تى سى ال": "TCL",
    "كرافت": "Craft",
    "دورا": "Dura",
    "فالكون": "Falcon",
    "فالكون انسيابي": "Falcon",
    "جيبسون": "Gibson",
    "الزامل": "Zamil",
    "الزامل كلاسيك": "Zamil",
    "الزامل كمفورت": "Zamil",
    "هاير": "Haier",
    "هايسنس": "Hisense",
    "هاس": "Haas",
    "كولين": "Kolin",
    "كول _": "Kolin",
    "كول ᣌᡧ": "Kolin",
    "كلفنيتور": "Kelvinator",
    "ميلنج": "Mieling",
    "مايديا": "Midea",
    "ميديا": "Midea",
    "ماييديا": "Midea",
    "مايدييا": "Midea",
    "وايت وستنجهاوس": "White Westinghouse",
    "ادميرال": "Admiral",
    "سامسونج": "Samsung",
    "اوكس": "AUX",
    "جري": "Gree",
    "كول لاين": "Cool Line",
    "ال جي": "LG",
    "ال جي اسلامي": "LG",
    "ال جي تيتان": "LG",
    "ال جي سمارت": "LG",
    "ال جي سمارت انفرتر": "LG",
    "ال جي فريش": "LG",
    "ال جي اولترا": "LG",
    "دايا": "Daya",
    "دايو": "Daewoo",
    "سكاي وورث": "Skyworth",
    "يورك": "York",
    "يوني هامبورج": "Uni Hamburg",
    "جنرال": "General",
    "جنرال ماكس": "General Max",
    "كروني": "Crony",
    "ريجنت": "Regent",
    "رجنت": "Regent",
    "كول اير": "Cool Air",
    "ميلانو": "Milano",
    "الجزيرة": "Al Jazeera",
    "الغدير": "Al Ghadeer",
    "هاس": "Haas",
    "TCL": "TCL",
    "الكوثر": "Al Kawthar",
    "سيمفوني": "Symphony",
    "نقال سيمفوني": "Symphony",
    "هايسينس": "Hisense",
    "هايسينس اسلامي": "Hisense",
    "ميلنج": "Mieling",
    "مᘭلنج": "Mieling",
}

# ── Arabic -> English category mapping ──────────────────────────────
CATEGORY_MAP = {
    "مكيف شباك": "Window AC",
    "مكيف اسبليت سقفي ارضي": "Ceiling Floor Split AC",
    "مكيف اسبليت": "Split AC",
    "مكيف كونسيلد": "Concealed AC",
    "مكيف كاسيت": "Cassette AC",
    "مكيف دولابي": "Floor Standing AC",
    "مكيف دوبي": "Floor Standing AC",
    "مكيف صحراوي": "Desert Cooler",
    "مكيف متنقل": "Portable AC",
    "مكيف باكج مركزي": "Central Package AC",
    "وش مكيف كاسيت": "Cassette AC Face Panel",
    "مكيف مراوح": "Fan AC",
}

# ── Type mapping from header ────────────────────────────────────────
TYPE_KEYWORDS = {
    "بارد": "CO",
    "ح/ب": "H&C",
    "قش": "CO",
    "مياه": "CO",
    "فريون": "CO",
    "دولابي": "CO",
    "كرتون": "CO",
    "غاز": "CO",
    "كهرباء": "CO",
}


def arabic_to_western(text):
    """Convert Arabic-Indic digits (٠١٢٣٤٥٦٧٨٩) to Western digits."""
    if not text:
        return text
    arabic_digits = "٠١٢٣٤٥٦٧٨٩"
    result = list(text)
    for i, ch in enumerate(result):
        if ch in arabic_digits:
            result[i] = str(arabic_digits.index(ch))
    return "".join(result)


def extract_btu(product_name):
    """Extract BTU value from product description."""
    if not product_name:
        return None
    name = arabic_to_western(str(product_name))
    # Replace Arabic decimal separator (U+066B) and other separators
    name = name.replace("٫", ".").replace(",", "").replace("،", "")

    # Pattern: number before وحدة (units) or after سعة التبريد/تبريد
    # Look for numbers like 18000, 18.000, 21800, 21.800, etc.
    patterns = [
        r'(\d{2}[\.,]?\d{3})\s*وحد',        # 18000 وحدة or 18.000 وحدة
        r'(\d{2}[\.,]?\d{3})\s*تدف',         # before تدفئة
        r'تبريد\s*(\d{2}[\.,]?\d{3})',        # after تبريد
        r'سعة\s*(?:ال)?ت[بـ]ريد\s*(\d{2}[\.,]?\d{3})',  # after سعة التبريد
        r'(\d{2}[\.,]?\d{3})',                 # fallback: any 5-digit-ish number
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if m:
            val = m.group(1).replace(".", "").replace(",", "")
            try:
                btu = int(val)
                if 5000 <= btu <= 80000:
                    return btu
            except ValueError:
                continue
    return None


def normalize_garbled(text):
    """Normalize garbled Arabic text by removing non-standard chars."""
    if not text:
        return text
    # Common garbled char replacements
    replacements = {
        "ᘭ": "ي",
        "ᘘ": "ب",
        "ᗫ": "ب",
        "ᗷ": "ب",
        "ᣆ": "ض",
        "ᡧ": "ن",
        "ᣎ": "ط",
        "\u173b": "ك",
        "\u16ff": "ك",
        "ᚏ": "ي",
        "ᛞ": "ب",
        "ᛒ": "ي",
        "ᢔ": "ي",
        "ᣂ": "ر",
        "ᙬ": "ت",
        "ᙫ": "ب",
        "ᚱ": "ر",
        "ᖁ": "ي",
        "ᣐ": "ج",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


# Additional garbled category patterns
CATEGORY_MAP_GARBLED = {
    "مكيف شباك": "Window AC",
    "مكᘭف شᘘاك": "Window AC",
    "مكيف اسبليت سقفي ارضي": "Ceiling Floor Split AC",
    "مكيف اسبليت": "Split AC",
    "مكᘭف اسᘘلᘭت": "Split AC",
    "مكᘭف اسᘭᙫلت": "Split AC",
    "مكيف كونسيلد": "Concealed AC",
    "مكᘭف كونسᘭلد": "Concealed AC",
    "مكيف كاسيت": "Cassette AC",
    "مكᘭف كاسᘭت": "Cassette AC",
    "مكيف دولابي": "Floor Standing AC",
    "مكيف دوبي": "Floor Standing AC",
    "مكᘭف دولاᗫ": "Floor Standing AC",
    "مكيف صحراوي": "Desert Cooler",
    "مكᘭف صحراو": "Desert Cooler",
    "مكيف متنقل": "Portable AC",
    "مكᘭف متنقل": "Portable AC",
    "مكيف باكج مركزي": "Central Package AC",
    "مكᘭف باكج": "Central Package AC",
    "وش مكيف كاسيت": "Cassette AC Face Panel",
    "وش مكᘭف": "Cassette AC Face Panel",
    "مكيف مراوح": "Fan AC",
}


def parse_header(header_text):
    """Parse section header to extract Category (EN) and Type (CO/H&C)."""
    if not header_text:
        return None, None

    text = header_text.replace("--", "").strip()

    # Normalize garbled text for better matching
    norm = normalize_garbled(text)

    # Determine type from header
    ac_type = None
    for ar_key, en_type in TYPE_KEYWORDS.items():
        if ar_key in text or ar_key in norm:
            ac_type = en_type
            break
    # Also check garbled ح/ب
    if ac_type is None:
        if "ح/ب" in text or "ح/ᗷ" in text:
            ac_type = "H&C"
        elif "ᗷارد" in text or "بارد" in text or "ارد" in text:
            ac_type = "CO"

    # Determine category - try both clean and garbled maps
    category = None
    all_cats = {**CATEGORY_MAP, **CATEGORY_MAP_GARBLED}

    # Try on original text first
    for ar_cat in sorted(all_cats.keys(), key=len, reverse=True):
        if ar_cat in text:
            category = all_cats[ar_cat]
            break

    # Try on normalized text
    if category is None:
        for ar_cat in sorted(CATEGORY_MAP.keys(), key=len, reverse=True):
            if ar_cat in norm:
                category = CATEGORY_MAP[ar_cat]
                break

    # Fallback: detect AC by checking if text contains garbled "مك" + AC type keywords
    if category is None:
        # Check for garbled AC keywords via model code patterns or partial text matches
        text_lower = text + " " + norm
        if any(kw in text_lower for kw in ["مك", "مكيف", "مكᘭف"]):
            # Try to determine category from partial matches
            if any(kw in text_lower for kw in ["سقف", "سقᘭ", "ارض", "ار "]):
                category = "Ceiling Floor Split AC"
            elif any(kw in text_lower for kw in ["اسبل", "اسᘘل", "اسᘭ"]):
                category = "Split AC"
            elif any(kw in text_lower for kw in ["شباك", "شᘘاك"]):
                category = "Window AC"
            elif any(kw in text_lower for kw in ["كونس", "كوᙏس"]):
                category = "Concealed AC"
            elif any(kw in text_lower for kw in ["كاسي", "كاسᘭ", "ᛳاس", "اس᛿"]):
                category = "Cassette AC"
            elif any(kw in text_lower for kw in ["دولاب", "دوﻻ"]):
                category = "Floor Standing AC"
            elif any(kw in text_lower for kw in ["صحرا", "صحراو"]):
                category = "Desert Cooler"
            elif any(kw in text_lower for kw in ["متنق"]):
                category = "Portable AC"
            elif any(kw in text_lower for kw in ["باكج", "مركز"]):
                category = "Central Package AC"
            elif any(kw in text_lower for kw in ["وش"]):
                category = "Cassette AC Face Panel"

    return category, ac_type


def extract_brand(product_name):
    """Extract brand name from Arabic product description and translate to English."""
    if not product_name:
        return "Unknown"
    name = str(product_name).strip()

    # Remove category prefixes
    prefixes = [
        "مكيف شباك انفرتر", "مكيف شباك روتاري",
        "مكيف شباك", "مكيف اسبليت انفرتر", "مكيف اسبليت",
        "مكيف كونسيلد", "مكيف كاسيت", "مكيف دولابي", "مكيف دوبي",
        "مكيف صحراوي", "مكيف متنقل", "مكيف باكج",
        "مكيف سقفي ارضي",
        # Garbled versions
        "مكᘭف شᘘاك انفرتر", "مكᘭف شᘘاك",
        "مكᘭف اسᘘلᘭت انفرتر", "مكᘭف اسᘘلᘭت", "مكᘭف اسᘭᙫلت",
        "مكᘭف كونسᘭلد", "مكᘭف كاسᘭت", "مكᘭف دولاᗫ",
        "مكيف صحراوى", "مكيف نقال",
    ]
    rest = name
    for p in sorted(prefixes, key=len, reverse=True):
        if name.startswith(p):
            rest = name[len(p):].strip()
            break

    # Try matching brand from the rest (longest match first)
    for ar_brand in sorted(BRAND_MAP_AR.keys(), key=len, reverse=True):
        if rest.startswith(ar_brand):
            return BRAND_MAP_AR[ar_brand]

    # Fallback: try to find brand anywhere in the product name
    for ar_brand in sorted(BRAND_MAP_AR.keys(), key=len, reverse=True):
        if ar_brand in name:
            return BRAND_MAP_AR[ar_brand]

    # Check for English brand names in the text
    english_brands = ["TCL", "SAMSUNG", "LG", "GREE", "AUX", "MIDEA", "HAIER",
                      "CARRIER", "DAIKIN", "YORK", "SUPER COOL", "ELEGANT",
                      "OASIS", "VICTORY", "FROSTY", "TURBO COOL",
                      "MISSION XTREME"]
    # Map English brand names to canonical names
    english_brand_map = {
        "MISSION XTREME": "Midea",
    }
    name_upper = name.upper()
    for eb in english_brands:
        if eb in name_upper:
            return english_brand_map.get(eb, eb.title())

    return "Unknown"


def detect_columns(rows):
    """Detect column indices for Price, Product Name, Model Code based on file format."""
    if not rows:
        return None, None, None

    num_cols = len(rows[0])

    if num_cols >= 14:
        # 20260125 format: price=0, name=2, code=10
        return 0, 2, 10
    elif num_cols == 4:
        # 20260113/20260217 format: price=0, name=1, code=2
        return 0, 1, 2
    elif num_cols == 3:
        # 20260130/20260308/20260311 format: price=0, name=1, code=2
        return 0, 1, 2
    else:
        return 0, 1, 2


def is_section_header(row, col_price, col_name, col_code, num_cols):
    """Check if a row is a section header (category separator)."""
    # Headers typically have '--' and null product/code columns
    for c in range(num_cols):
        val = row[c] if c < len(row) else None
        if val is not None and "--" in str(val):
            return True, str(val)

    # Also check: price column has text (not number), other cols are None
    price_val = row[col_price] if col_price < len(row) else None
    name_val = row[col_name] if col_name < len(row) else None
    code_val = row[col_code] if col_code < len(row) else None

    if isinstance(price_val, str) and ("مكيف" in price_val or "مكᘭف" in price_val):
        if name_val is None and code_val is None:
            return True, price_val

    return False, None


def parse_file(filepath):
    """Parse a single Excel file and return list of AC product dicts."""
    basename = os.path.basename(filepath)
    # Extract date from filename: BH_Subdealer_Pricelist_YYYYMMDD.xlsx
    m = re.search(r'(\d{8})', basename)
    if not m:
        return []
    date_str = m.group(1)
    update_date = datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    col_price, col_name, col_code = detect_columns(rows)
    num_cols = len(rows[0])

    products = []
    current_category = None
    current_type = None
    is_ac_section = False

    for i, row in enumerate(rows):
        # Skip header row (row 0)
        if i == 0:
            continue

        # Check for section header
        is_header, header_text = is_section_header(row, col_price, col_name, col_code, num_cols)
        if is_header and header_text:
            cat, typ = parse_header(header_text)
            if cat is not None:
                current_category = cat
                current_type = typ
                is_ac_section = True
            else:
                # Non-AC header encountered
                is_ac_section = False
            continue

        if not is_ac_section:
            continue

        # Skip excluded categories
        if current_category in EXCLUDED_CATEGORIES:
            continue

        # Extract product data
        price_val = row[col_price] if col_price < len(row) else None
        name_val = row[col_name] if col_name < len(row) else None
        code_val = row[col_code] if col_code < len(row) else None

        # Skip rows without valid price or code
        if not isinstance(price_val, (int, float)) or price_val <= 0:
            continue
        if code_val is None or str(code_val).strip() == "":
            continue
        if name_val is None:
            continue

        model_code = str(code_val).strip().replace("\n", "")
        product_name = str(name_val).strip().replace("\n", " ")
        price = float(price_val)

        # Extract brand
        brand = extract_brand(product_name)

        # Extract BTU
        btu = extract_btu(product_name)

        products.append({
            "Brand": brand,
            "Category": current_category,
            "Model_Code": model_code,
            "BTU": btu,
            "Ton": btu_to_ton(btu),
            "Type": current_type,
            "Product_Description": product_name,
            "Price": price,
            "Update_Date": update_date,
        })

    return products


def translate_product_desc(text):
    """Translate key Arabic terms in product description to English."""
    if not text:
        return text

    text = arabic_to_western(text)
    text = text.replace("٫", ".")

    translations = {
        # Category
        "مكيف شباك": "Window AC",
        "مكيف اسبليت": "Split AC",
        "مكيف كونسيلد": "Concealed AC",
        "مكيف كاسيت": "Cassette AC",
        "مكيف دولابي": "Floor Standing AC",
    "مكيف دوبي": "Floor Standing AC",
        "مكيف صحراوي": "Desert Cooler",
        "مكيف متنقل": "Portable AC",
        "مكيف سقفي ارضي": "Ceiling Floor Split AC",
        "مكيف باكج": "Package AC",
        # Garbled category
        "مكᘭف شᘘاك": "Window AC",
        "مكᘭف اسᘘلᘭت": "Split AC",
        "مكᘭف اسᘭᙫلت": "Split AC",
        "مكᘭف كونسᘭلد": "Concealed AC",
        # Common terms
        "سعة التبريد": "Cooling Capacity",
        "سعة تبريد": "Cooling Capacity",
        "سعة الت": "Cooling Capacity",
        "وحدة": "BTU",
        "وحده": "BTU",
        "انفرتر": "Inverter",
        "روتاري": "Rotary",
        "روتارى": "Rotary",
        "واي فاي": "WiFi",
        "تدفئة": "Heating",
        "كيلوواط": "kW",
        "كيلو واط": "kW",
        "واط": "W",
        "وطني": "National",
        "بني": "Brown",
        "اسود": "Black",
        "ابيض": "White",
        "بدون مواسير": "Without Pipes",
        "تنظيف ذاتي": "Self-Clean",
        "ريموت": "Remote",
        "اتجاهات": "Directions",
        "شاشة ديجيتال": "Digital Display",
        "تحكم": "Control",
        "سرعات": "Speeds",
        "حصان": "HP",
    }

    for ar, en in sorted(translations.items(), key=lambda x: len(x[0]), reverse=True):
        text = text.replace(ar, en)

    # Also translate brand names
    for ar, en in sorted(BRAND_MAP_AR.items(), key=lambda x: len(x[0]), reverse=True):
        text = text.replace(ar, en)

    return text.strip()


## ═══════════════════════════════════════════════════════════════════
## BH Store Retail API — 스크래핑 함수
## ═══════════════════════════════════════════════════════════════════

def extract_cold_hot(name):
    """제품명에서 Cold Only / Cold & Hot 구분 (Retail API 용)."""
    name_lower = name.lower()
    if 'hot' in name_lower and 'cold' in name_lower:
        return 'C&H'
    elif 'heat' in name_lower and 'cold' in name_lower:
        return 'C&H'
    elif 'heating' in name_lower:
        return 'C&H'
    elif 'hot' in name_lower:
        return 'C&H'
    elif 'cold only' in name_lower:
        return 'CO'
    elif 'cold' in name_lower:
        return 'CO'
    elif 'cool' in name_lower:
        return 'CO'
    return ''


def extract_btu_retail(name):
    """제품명에서 BTU 추출 (Retail API 영문 제품명용)."""
    match = re.search(r'(\d{1,2}[,.]?\d{3})\s*(?:BTU|Btu|btu)?', name)
    if match:
        btu = match.group(1).replace(',', '').replace('.', '')
        return int(btu)
    return None


def fetch_products_by_category(category_name, category_id):
    """카테고리 ID로 BH Store 제품 수집."""
    print(f"  [*] {category_name} (category_id={category_id}) 수집 중...")
    url = f"{API_URL}?limit=300&category_id={category_id}"
    try:
        r = requests.get(url, headers=API_HEADERS, timeout=30)
        if r.status_code != 200:
            print(f"    [!] API 에러: {r.status_code}")
            return []
        data = r.json()
        products = data.get("data", [])
        print(f"    [+] {len(products)}개 발견")
        return products
    except Exception as e:
        print(f"    [!] 오류: {e}")
        return []


def fetch_all_ac_products():
    """모든 카테고리에서 BH Store AC 제품 수집."""
    all_products = {}
    for ac_type, category_id in RETAIL_CATEGORIES.items():
        products = fetch_products_by_category(ac_type, category_id)
        for p in products:
            code = p.get("code", "").strip()
            if not code or code in all_products:
                continue
            name = p.get("name", "")
            all_products[code] = {
                "code": code,
                "brand": p.get("brand", {}).get("name", ""),
                "name": name,
                "ac_type": ac_type,
                "cold_hot": extract_cold_hot(name),
                "btu": extract_btu_retail(name),
                "price": p.get("price"),
                "current_price": p.get("current_price"),
                "saved_percent": p.get("saved_percent"),
                "saved_amount": p.get("saved_amount"),
                "available_stock": p.get("available_stock"),
                "special_from": p.get("special_from"),
                "special_to": p.get("special_to"),
                "updated_at": p.get("updated_at"),
                "slug": p.get("slug", ""),
            }
    print(f"  [+] 총 {len(all_products)}개 AC 제품 수집 완료")
    return all_products


def get_existing_run_dates(output_path):
    """Weekly_Price_DB에 이미 기록된 Run_Timestamp의 날짜 목록 반환 (중복 방지용)."""
    if not os.path.exists(output_path):
        return set()
    try:
        wb = openpyxl.load_workbook(output_path, read_only=True)
    except Exception:
        return set()
    if RETAIL_SHEET not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[RETAIL_SHEET]
    dates = set()
    # Run_Timestamp = R열 (18번째 컬럼)
    for row in ws.iter_rows(min_row=2, max_col=18, values_only=True):
        if row and len(row) >= 18 and row[17]:
            ts = str(row[17]).strip()
            # "2026-03-14 10:30:00" -> "2026-03-14"
            date_part = ts[:10]
            dates.add(date_part)
    wb.close()
    return dates


def update_weekly_price_db(api_products, output_path):
    """
    Weekly_Price_DB 시트에 오늘의 스크래핑 데이터 추가.
    같은 날짜에 이미 데이터가 있으면 스킵 (중복 방지).
    """
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today = datetime.now().strftime("%Y-%m-%d")
    week = datetime.now().strftime("%Y-W%U")

    # 중복 체크 (날짜 기준)
    existing_dates = get_existing_run_dates(output_path)
    if today in existing_dates:
        print(f"  [!] {today} 데이터가 이미 존재합니다. 스킵합니다.")
        return 0

    # 엑셀 파일 로드 (없으면 새로 생성)
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()

    # Weekly_Price_DB 시트가 없으면 생성
    if RETAIL_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(RETAIL_SHEET)
        # 헤더 작성
        headers = [
            "Week", "Retailer", "Brand", "Model_Code", "Product Name",
            "Type", "BTU", "CO/C&H", "Regular Price", "Current Price",
            "Discount_%", "Discount SAR", "Stock", "Promo From", "Promo To",
            "Last Updated", "URL", "Run_Timestamp",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
    else:
        ws = wb[RETAIL_SHEET]

    # 첫 번째 빈 행 찾기 (A열 기준)
    current_row = 2
    for row_num in range(2, ws.max_row + 2):
        if ws.cell(row=row_num, column=1).value is None:
            current_row = row_num
            break
        current_row = row_num + 1

    count = 0
    for code, product in api_products.items():
        discount_ratio = product["saved_percent"] / 100 if product["saved_percent"] else None
        slug = product.get("slug", "")
        url = f"https://bhstore.com.sa/sa-en/{slug}" if slug else ""

        data = [
            week,                          # A: Week
            "BH Store",                    # B: Retailer
            product["brand"],              # C: Brand
            product["code"],               # D: Model_Code
            product["name"],               # E: Product Name
            product["ac_type"],            # F: Type
            product["btu"],                # G: BTU
            product["cold_hot"],           # H: CO/C&H
            product["price"],              # I: Regular Price
            product["current_price"],      # J: Current Price
            discount_ratio,                # K: Discount_%
            product["saved_amount"],       # L: Discount SAR
            product["available_stock"],    # M: Stock
            product["special_from"],       # N: Promo From
            product["special_to"],         # O: Promo To
            product["updated_at"],         # P: Last Updated
            url,                           # Q: URL
            run_ts,                        # R: Run_Timestamp
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=current_row, column=col, value=value)

        # Discount % 서식
        ws.cell(row=current_row, column=11).number_format = '0%'

        current_row += 1
        count += 1

    wb.save(output_path)
    print(f"  [+] Weekly_Price_DB: {count}개 제품 추가 (Week: {week})")
    return count


def compute_status(all_data_by_date):
    """
    Compute Status for each model at each date:
    - First file (earliest date): all models are "New"
    - Subsequent files:
      - Model not seen before: "New"
      - Model was in previous file: "Active" (existing)
      - Model was in previous file but not in current: "Discontinue"
      - Model was discontinued but reappears: "Re-Active"
    """
    dates = sorted(all_data_by_date.keys())
    if not dates:
        return []

    # Track model status history
    model_status = {}  # model_code -> last known status
    model_ever_seen = set()
    results = []

    for d_idx, date in enumerate(dates):
        current_models = set()
        date_products = all_data_by_date[date]

        for prod in date_products:
            mc = prod["Model_Code"]
            current_models.add(mc)

            if d_idx == 0:
                prod["Status"] = "New"
                model_status[mc] = "New"
            else:
                if mc not in model_ever_seen:
                    prod["Status"] = "New"
                    model_status[mc] = "New"
                elif model_status.get(mc) == "Discontinue":
                    prod["Status"] = "Re-Active"
                    model_status[mc] = "Re-Active"
                else:
                    prod["Status"] = "Active"
                    model_status[mc] = "Active"

            results.append(prod)

        model_ever_seen.update(current_models)

        # Mark discontinued models (were active/new but not in current file)
        if d_idx > 0:
            prev_active = {mc for mc, st in model_status.items()
                          if st in ("New", "Active", "Re-Active") and mc in model_ever_seen}
            discontinued = prev_active - current_models
            for mc in discontinued:
                model_status[mc] = "Discontinue"
                # Find last known product info for this model
                last_info = None
                for prev_date in reversed(dates[:d_idx]):
                    for p in all_data_by_date[prev_date]:
                        if p["Model_Code"] == mc:
                            last_info = p
                            break
                    if last_info:
                        break
                if last_info:
                    results.append({
                        "Brand": last_info["Brand"],
                        "Category": last_info["Category"],
                        "Model_Code": mc,
                        "BTU": last_info["BTU"],
                        "Ton": last_info.get("Ton"),
                        "Type": last_info["Type"],
                        "Product_Description": last_info["Product_Description"],
                        "Status": "Discontinue",
                        "Price": None,
                        "Update_Date": date,
                    })

    return results


def enrich_from_auto(results):
    """
    Use clean-file data to fill in 'Unknown' brands, BTU, and descriptions
    using model_code lookup across all files (auto-extracted cross-reference).
    """
    model_brand = {}
    model_desc = {}
    model_btu = {}
    for r in results:
        mc = r["Model_Code"]
        brand = r["Brand"]
        desc = r.get("Product_Description", "")
        btu = r.get("BTU")
        if brand != "Unknown":
            model_brand[mc] = brand
        if desc and "ᘭ" not in desc and "ᗫ" not in desc:
            model_desc[mc] = desc
        if btu is not None:
            model_btu[mc] = btu

    for r in results:
        mc = r["Model_Code"]
        if r["Brand"] == "Unknown" and mc in model_brand:
            r["Brand"] = model_brand[mc]
        if mc in model_desc:
            r["Product_Description"] = model_desc[mc]
        if r.get("BTU") is None and mc in model_btu:
            r["BTU"] = model_btu[mc]
        # Recalculate Ton from BTU if missing
        if r.get("Ton") is None and r.get("BTU") is not None:
            r["Ton"] = btu_to_ton(r["BTU"])


def load_existing_model_info(output_path):
    """
    Load existing 'Model Info' sheet from the output file.
    Returns dict: model_code -> {Brand, Category, Model_Code, BTU, Type, Product_Description}
    User edits in this sheet take priority.
    """
    if not os.path.exists(output_path):
        return {}

    try:
        wb = openpyxl.load_workbook(output_path, read_only=True)
    except Exception:
        return {}

    if "Total_Model Info" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Total_Model Info"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return {}

    # Expected columns: Brand, Category, Model Code, BTU, Ton, Type, Product Description
    model_info = {}
    for row in rows[1:]:
        if len(row) < 6:
            continue
        # Support both old (6-col) and new (7-col with Ton) formats
        if len(row) >= 7:
            mc = str(row[2]).strip() if row[2] else None
            if not mc:
                continue
            model_info[mc] = {
                "Brand": row[0] if row[0] else None,
                "Category": row[1] if row[1] else None,
                "Model_Code": mc,
                "BTU": row[3] if row[3] else None,
                "Ton": row[4] if row[4] is not None else None,
                "Type": row[5] if row[5] else None,
                "Product_Description": row[6] if row[6] else None,
            }
        else:
            mc = str(row[2]).strip() if row[2] else None
            if not mc:
                continue
            model_info[mc] = {
                "Brand": row[0] if row[0] else None,
                "Category": row[1] if row[1] else None,
                "Model_Code": mc,
                "BTU": row[3] if row[3] else None,
                "Ton": None,
                "Type": row[4] if row[4] else None,
                "Product_Description": row[5] if row[5] else None,
            }

    return model_info


def load_retail_model_info(output_path):
    """
    Load model info from BH Store Retail API data (Weekly_Price_DB sheet in output file).
    Returns dict: model_code -> {Brand, Category, Model_Code, BTU, Ton, Type, Product_Description}
    Uses the most recent entry per model for best data quality.
    """
    if not os.path.exists(output_path):
        return {}

    try:
        wb = openpyxl.load_workbook(output_path, read_only=True)
    except Exception:
        return {}

    if RETAIL_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[RETAIL_SHEET]
    retail_info = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        mc = str(row[3]).strip() if row[3] else None
        if not mc:
            continue

        brand = str(row[2]).strip() if row[2] else None
        category = str(row[5]).strip() if row[5] else None
        btu = row[6] if row[6] is not None else None
        cold_hot = str(row[7]).strip() if row[7] else None
        product_name = str(row[4]).strip() if row[4] else None

        # Normalize category and type to match Wholesale naming
        if category:
            category = CATEGORY_NORMALIZE.get(category, category)
        if cold_hot:
            cold_hot = TYPE_NORMALIZE.get(cold_hot, cold_hot)

        # Calculate Ton from BTU
        ton = btu_to_ton(btu) if btu is not None else None

        # Later rows override earlier (more recent data wins)
        retail_info[mc] = {
            "Brand": brand,
            "Category": category,
            "Model_Code": mc,
            "BTU": btu,
            "Ton": ton,
            "Type": cold_hot,
            "Product_Description": product_name,
        }

    wb.close()
    return retail_info


def build_model_info(results, existing_model_info, retail_model_info=None):
    """
    Build the Total_Model Info dictionary from Retail + Wholesale sources.
    Priority: Retail API (영문, 정확) > Wholesale auto-extract (아랍어 번역)
    사용자 수동 편집은 이 빌드 이후에 직접 엑셀에서 진행.
    Returns dict: model_code -> info dict
    """
    # Wholesale auto-extracted data
    auto_info = {}
    for r in results:
        mc = r["Model_Code"]
        if mc not in auto_info:
            auto_info[mc] = {
                "Brand": r.get("Brand", ""),
                "Category": r.get("Category", ""),
                "Model_Code": mc,
                "BTU": r.get("BTU"),
                "Ton": r.get("Ton"),
                "Type": r.get("Type", ""),
                "Product_Description": r.get("Product_Description", ""),
            }

    # Merge priority: Retail API > Wholesale auto-extract
    if retail_model_info is None:
        retail_model_info = {}

    merged = {}
    all_models = set(auto_info.keys()) | set(retail_model_info.keys())

    for mc in all_models:
        auto = auto_info.get(mc, {})            # tier 2: Wholesale auto-extract
        retail = retail_model_info.get(mc, {})   # tier 1: Retail API (highest)

        # Each field: retail > auto
        btu = (retail.get("BTU") if retail.get("BTU") is not None
               else auto.get("BTU"))
        ton = (retail.get("Ton") if retail.get("Ton") is not None
               else auto.get("Ton"))
        if ton is None and btu is not None:
            ton = btu_to_ton(btu)

        brand = retail.get("Brand") or auto.get("Brand", "")
        brand = BRAND_NORMALIZE.get(brand, brand)

        merged[mc] = {
            "Brand": brand,
            "Category": retail.get("Category") or auto.get("Category", ""),
            "Model_Code": mc,
            "BTU": btu,
            "Ton": ton,
            "Type": retail.get("Type") or auto.get("Type", ""),
            "Product_Description": retail.get("Product_Description") or auto.get("Product_Description", ""),
        }

    return merged


def apply_model_info(results, model_info):
    """
    Apply Model Info sheet data to results.
    Model Info values override auto-extracted values.
    """
    for r in results:
        mc = r["Model_Code"]
        if mc in model_info:
            info = model_info[mc]
            if info.get("Brand"):
                r["Brand"] = info["Brand"]
            if info.get("Category"):
                r["Category"] = info["Category"]
            if info.get("BTU") is not None:
                r["BTU"] = info["BTU"]
            if info.get("Ton") is not None:
                r["Ton"] = info["Ton"]
            if info.get("Type"):
                r["Type"] = info["Type"]
            if info.get("Product_Description"):
                r["Product_Description"] = info["Product_Description"]


def write_excel(results, model_info, output_path):
    """Write results to Excel file, preserving extra sheets (e.g. Weekly_Price_DB)."""
    # Load existing workbook to preserve extra sheets, or create new
    preserve_sheets = set()
    if os.path.exists(output_path):
        try:
            wb = openpyxl.load_workbook(output_path)
            # Remember which sheets to preserve (not ours)
            our_sheets = {"Total_Model Info", "Whole selling_Model Info", "Whole selling Price_Master"}
            for sn in wb.sheetnames:
                if sn in our_sheets:
                    wb.remove(wb[sn])
                else:
                    preserve_sheets.add(sn)
        except Exception:
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()

    # If workbook has only the default "Sheet", remove it after we create ours
    has_default = "Sheet" in wb.sheetnames and "Sheet" not in preserve_sheets

    # ── Shared styles ────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)

    # ════════════════════════════════════════════════════════════════
    # Sheet 1: Model Info
    # ════════════════════════════════════════════════════════════════
    ws_info = wb.create_sheet("Total_Model Info", 0)

    info_columns = [
        ("Brand", 18),
        ("Category", 25),
        ("Model Code", 28),
        ("BTU", 12),
        ("Ton", 8),
        ("Type", 8),
        ("Product Description", 65),
    ]

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(info_columns, 1):
        cell = ws_info.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws_info.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Sort models by Category, Brand, Model Code
    sorted_models = sorted(
        model_info.values(),
        key=lambda m: (m.get("Category", ""), m.get("Brand", ""), m.get("Model_Code", ""))
    )

    for row_idx, info in enumerate(sorted_models, 2):
        desc = translate_product_desc(info.get("Product_Description", ""))
        values = [
            info.get("Brand", ""),
            info.get("Category", ""),
            info.get("Model_Code", ""),
            info.get("BTU"),
            info.get("Ton"),
            info.get("Type", ""),
            desc,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if col_idx == 4 and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
            if col_idx == 5 and isinstance(val, (int, float)):
                cell.number_format = '0.0'

    ws_info.freeze_panes = "A2"
    ws_info.auto_filter.ref = f"A1:{get_column_letter(len(info_columns))}{len(sorted_models) + 1}"

    # ════════════════════════════════════════════════════════════════
    # Sheet 2: AC Master (price tracking)
    # ════════════════════════════════════════════════════════════════
    ws_master = wb.create_sheet("Whole selling Price_Master", 1)

    master_columns = [
        ("Brand", 18),
        ("Category", 25),
        ("Model Code", 28),
        ("BTU", 12),
        ("Ton", 8),
        ("Type", 8),
        ("Product Description", 60),
        ("Status", 14),
        ("Price", 12),
        ("Update Date", 14),
    ]

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(master_columns, 1):
        cell = ws_master.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws_master.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Status colors
    status_fills = {
        "New": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Active": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "Discontinue": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Re-Active": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    }
    status_fonts = {
        "New": Font(name="Calibri", size=10, color="006100"),
        "Active": Font(name="Calibri", size=10),
        "Discontinue": Font(name="Calibri", size=10, color="9C0006"),
        "Re-Active": Font(name="Calibri", size=10, color="1F4E79"),
    }

    # Sort results by Update_Date, Category, Brand, Model_Code
    results.sort(key=lambda r: (r["Update_Date"], r.get("Category", ""), r.get("Brand", ""), r.get("Model_Code", "")))

    price_fmt = '#,##0'

    for row_idx, rec in enumerate(results, 2):
        desc = translate_product_desc(rec.get("Product_Description", ""))
        status = rec.get("Status", "")

        values = [
            rec.get("Brand", ""),
            rec.get("Category", ""),
            rec.get("Model_Code", ""),
            rec.get("BTU"),
            rec.get("Ton"),
            rec.get("Type", ""),
            desc,
            status,
            rec.get("Price"),
            rec.get("Update_Date", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_master.cell(row=row_idx, column=col_idx, value=val)
            cell.font = status_fonts.get(status, data_font)
            cell.alignment = data_align
            cell.border = thin_border

            # Status column coloring
            if col_idx == 8:
                cell.fill = status_fills.get(status, PatternFill())
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Price formatting
            if col_idx == 9 and isinstance(val, (int, float)):
                cell.number_format = price_fmt

            # BTU formatting
            if col_idx == 4 and isinstance(val, (int, float)):
                cell.number_format = '#,##0'

            # Ton formatting
            if col_idx == 5 and isinstance(val, (int, float)):
                cell.number_format = '0.0'

    ws_master.freeze_panes = "A2"
    ws_master.auto_filter.ref = f"A1:{get_column_letter(len(master_columns))}{len(results) + 1}"

    # Remove default "Sheet" if it was auto-created and not a preserved sheet
    if has_default and "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"  Model Info: {len(sorted_models)} unique models")
    print(f"  AC Master:  {len(results)} records")
    if preserve_sheets:
        print(f"  Preserved:  {', '.join(sorted(preserve_sheets))}")


def get_processed_wholesale_dates(output_path):
    """
    이미 처리된 Wholesale 파일의 날짜 목록을 반환.
    Whole selling Price_Master 시트의 Update Date 컬럼에서 추출.
    """
    if not os.path.exists(output_path):
        return set()
    try:
        wb = openpyxl.load_workbook(output_path, read_only=True)
    except Exception:
        return set()
    if "Whole selling Price_Master" not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb["Whole selling Price_Master"]
    dates = set()
    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        if row and len(row) >= 10 and row[9]:
            dates.add(str(row[9]).strip())
    wb.close()
    return dates


def main():
    print("=" * 60)
    print("BH AC Price Tracker — Unified Pipeline")
    print("=" * 60)

    # ══════════════════════════════════════════════════════════════
    # Phase 1: Retail API -> Weekly_Price_DB (매일 실행)
    # ══════════════════════════════════════════════════════════════
    print("\n[Phase 1] BH Store Retail API -> Weekly_Price_DB")
    print("-" * 50)

    retail_count = 0
    api_products = {}

    if not HAS_REQUESTS:
        print("  [!] requests 라이브러리 미설치. Retail 스크래핑 스킵.")
        print("      설치: pip install requests")
    else:
        api_products = fetch_all_ac_products()
        if api_products:
            retail_count = update_weekly_price_db(api_products, OUTPUT_FILE)
        else:
            print("  [!] API에서 수집된 제품이 없습니다.")

    # ══════════════════════════════════════════════════════════════
    # Phase 2: Wholesale 엑셀 -> Whole selling Price_Master (신규만)
    # ══════════════════════════════════════════════════════════════
    print(f"\n[Phase 2] Wholesale 엑셀 -> Whole selling Price_Master")
    print("-" * 50)

    files = sorted(glob.glob(INPUT_PATTERN))
    if not files:
        print("  [!] BH_Subdealer_Pricelist_*.xlsx 파일 없음.")
        if retail_count > 0:
            print(f"\n  완료: Retail {retail_count}개 업데이트됨, Wholesale 스킵")
        return

    # 이미 처리된 날짜 확인
    processed_dates = get_processed_wholesale_dates(OUTPUT_FILE)

    # 신규 파일만 필터링
    new_files = []
    for filepath in files:
        basename = os.path.basename(filepath)
        if "AC_Master" in basename:
            continue
        m = re.search(r'(\d{8})', basename)
        if m:
            date_str = datetime.strptime(m.group(1), "%Y%m%d").strftime("%Y-%m-%d")
            if date_str not in processed_dates:
                new_files.append(filepath)

    if not new_files and processed_dates:
        print(f"  [*] 전체 {len(files)}개 파일 중 신규 없음 (기존 {len(processed_dates)}개 날짜 처리됨)")

        # Model Info는 Retail 데이터 반영을 위해 항상 재빌드
        if retail_count > 0 or api_products:
            print("\n[Phase 3] Model Info 재빌드 (Retail 정보 반영)")
            print("-" * 50)
            _rebuild_model_info_only(files, OUTPUT_FILE)

        print(f"\n{'=' * 60}")
        print(f"완료!")
        print(f"  Retail: {retail_count}개 가격 업데이트")
        print(f"  Wholesale: 신규 파일 없음 (스킵)")
        print(f"{'=' * 60}")
        return

    # 신규 파일이 있으면 전체 Wholesale 재처리
    print(f"  [*] 전체 {len(files)}개 파일, 신규 {len(new_files)}개:")
    for f in new_files:
        print(f"    NEW -> {os.path.basename(f)}")
    print()

    # ── Step 1: Load Retail model info (Weekly_Price_DB in same file) ──
    retail_model_info = load_retail_model_info(OUTPUT_FILE)
    if retail_model_info:
        print(f"  Loaded {len(retail_model_info)} retail model entries (Weekly_Price_DB)")
    print()

    # ── Step 2: Parse ALL wholesale files (전체 재처리 - Status 계산 필요) ──
    all_data_by_date = {}
    for filepath in files:
        basename = os.path.basename(filepath)
        if "AC_Master" in basename:
            continue
        products = parse_file(filepath)
        m = re.search(r'(\d{8})', basename)
        if m:
            date_str = datetime.strptime(m.group(1), "%Y%m%d").strftime("%Y-%m-%d")
            all_data_by_date[date_str] = products
            print(f"  {basename}: {len(products)} AC products")

    print()

    # ── Step 3: Compute status ──────────────────────────────────────
    results = compute_status(all_data_by_date)

    # ── Step 4: Auto-enrich from cross-file data ────────────────────
    enrich_from_auto(results)

    # ── Step 5: Build Total_Model Info (Retail > Wholesale) ─────────
    model_info = build_model_info(results, {}, retail_model_info)

    # ── Step 6: Apply Model Info back to results ────────────────────
    apply_model_info(results, model_info)

    # ── Step 7: Write output ────────────────────────────────────────
    write_excel(results, model_info, OUTPUT_FILE)

    # Summary
    print(f"\n{'=' * 60}")
    print("완료!")
    print(f"  Retail: {retail_count}개 가격 업데이트 (Weekly_Price_DB)")
    dates = sorted(all_data_by_date.keys())
    for d in dates:
        status_counts = {}
        for r in results:
            if r["Update_Date"] == d:
                s = r["Status"]
                status_counts[s] = status_counts.get(s, 0) + 1
        print(f"  Wholesale {d}: {status_counts}")
    print(f"{'=' * 60}")


def _rebuild_model_info_only(files, output_path):
    """
    Wholesale 신규 파일이 없어도, Retail API 정보로 Total_Model Info 재빌드.
    Wholesale Price_Master는 건드리지 않음.
    """
    retail_model_info = load_retail_model_info(output_path)

    # 기존 Wholesale 결과 파싱 (Model Info 빌드용)
    all_data_by_date = {}
    for filepath in files:
        basename = os.path.basename(filepath)
        if "AC_Master" in basename:
            continue
        products = parse_file(filepath)
        m = re.search(r'(\d{8})', basename)
        if m:
            date_str = datetime.strptime(m.group(1), "%Y%m%d").strftime("%Y-%m-%d")
            all_data_by_date[date_str] = products

    results = compute_status(all_data_by_date)
    enrich_from_auto(results)

    model_info = build_model_info(results, {}, retail_model_info)
    apply_model_info(results, model_info)

    # Total_Model Info 시트 재작성 (Wholesale Price_Master + Weekly_Price_DB 보존)
    write_excel(results, model_info, output_path)
    print(f"  Total_Model Info 재빌드 완료: {len(model_info)}개 모델")


if __name__ == "__main__":
    main()
