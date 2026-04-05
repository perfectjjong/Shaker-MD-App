#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bin Momen AC Dashboard Builder
================================
Source  : Binmomen_AC_Data.xlsx  (cumulative scrape data)
Output  : Binmomen_AC_Prices_Tracking_Master.xlsx

Sheets:
  1. Dashboard_Summary       - KPI / Category dist / Brand share / vs Previous
  2. Price_Change_Alert      - Sale price changes (latest vs previous)
  3. New_Discontinued_SKUs   - New / discontinued models
  4. Brand_Price_Compare     - Brand avg price by Category x Ton, LG gap
  5. Full_Price_Tracking     - Full latest snapshot + prev price delta
"""

import os, sys

if sys.stdout and sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[ERROR] Missing package: {e}")
    print("  >> pip install pandas openpyxl numpy")
    sys.exit(1)

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
DATA_FILE  = os.path.join(SCRIPT_DIR, "Binmomen_AC_Data.xlsx")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Binmomen_AC_Prices_Tracking_Master.xlsx")

# ── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=13, color='1F4E79')
SUB_FONT   = Font(name='Arial', bold=True, size=10, color='666666')
SEC_FONT   = Font(name='Arial', bold=True, size=11, color='1F4E79')
DATA_FONT  = Font(name='Arial', size=10)
BOLD_FONT  = Font(name='Arial', bold=True, size=10)
UP_FONT    = Font(name='Arial', size=10, color='C00000', bold=True)
DOWN_FONT  = Font(name='Arial', size=10, color='375623', bold=True)
NEW_FONT   = Font(name='Arial', size=10, color='0070C0', bold=True)
DISC_FONT  = Font(name='Arial', size=10, color='999999')
TK_BORDER  = Border(bottom=Side(style='medium', color='1F4E79'))
THIN_BORDER= Border(bottom=Side(style='thin', color='D9D9D9'))
CTR        = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left', vertical='center')
WRAP       = Alignment(horizontal='left', vertical='center', wrap_text=True)
LIGHT_FILL = PatternFill('solid', fgColor='F2F7FB')
CAT_FILL   = PatternFill('solid', fgColor='D6E4F0')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
RED_FILL   = PatternFill('solid', fgColor='FCE4EC')

CAT_ORDER  = ['Split AC', 'Window AC', 'Floor Standing AC']


# ── Utils ────────────────────────────────────────────────────────────────────
def _hdr(ws, row, n_col):
    for c in range(1, n_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CTR
        cell.border = TK_BORDER


def _auto(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx = 0
        ltr = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[ltr].width = min(max(mx + 3, min_w), max_w)


def _del(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def _fmt_ton(v):
    try:
        f = float(v)
        return f"{f:.1f}T" if f % 1 else f"{int(f)}T"
    except Exception:
        return str(v) if v else '-'


def _pct(val, denom):
    if not denom or denom == 0:
        return 0
    return val / denom


# ── Load Data ────────────────────────────────────────────────────────────────
def load_data():
    df = pd.read_excel(DATA_FILE)
    df['Scrape_Date'] = pd.to_datetime(df['Scrape_Date'], errors='coerce')
    df['scrape_day'] = df['Scrape_Date'].dt.normalize()

    for col in ['Original_Price', 'Sale_Price', 'BTU', 'Tonnage', 'Stock_Qty']:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df['Discount_Pct'] = df['Discount'].astype(str).str.replace('%', '').str.strip()
    df['Discount_Pct'] = pd.to_numeric(df['Discount_Pct'], errors='coerce').astype(float)
    df.loc[df['Discount_Pct'] >= 1, 'Discount_Pct'] = df.loc[df['Discount_Pct'] >= 1, 'Discount_Pct'] / 100

    df['In_Stock_Bool'] = df['In_Stock'].astype(str).str.lower().isin(['yes', 'true', '1'])

    for col in ['Brand', 'Category', 'Cooling_Type', 'Inverter', 'Compressor', 'SKU']:
        df[col] = df[col].fillna('').astype(str).str.strip()

    # Normalize category
    def _norm_cat(c):
        c = str(c).strip()
        for std in CAT_ORDER:
            if std.lower().replace(' ', '') in c.lower().replace(' ', ''):
                return std
        return c if c else 'Other'
    df['Category'] = df['Category'].apply(_norm_cat)

    return df


# ═════════════════════════════════════════════════════════════════════════════
# 1. Dashboard_Summary
# ═════════════════════════════════════════════════════════════════════════════
def build_summary(wb, df):
    sn = 'Dashboard_Summary'
    _del(wb, sn)
    ws = wb.create_sheet(sn, 0)

    dates = sorted(df['scrape_day'].dropna().unique())
    latest = dates[-1]
    prev = dates[-2] if len(dates) >= 2 else None
    dl = df[df['scrape_day'] == latest]
    dp = df[df['scrape_day'] == prev] if prev else pd.DataFrame()

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Bin Momen — AC Price Tracking Dashboard'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = (
        f"Latest: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
        f"Total SKUs: {len(dl)}  |  "
        f"Brands: {dl['Brand'].replace('', pd.NA).dropna().nunique()}  |  "
        f"In Stock: {dl['In_Stock_Bool'].sum()} / {len(dl)}"
    )
    ws['A2'].font = SUB_FONT

    # ── Section 1: KPI by Category / Cooling / Ton ───────────────────────
    r = 4
    ws.merge_cells(f'A{r}:K{r}')
    ws[f'A{r}'] = '▶  Price KPI by Category / Cooling Type / Capacity'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    kpi_h = ['Category', 'Cooling Type', 'Inverter', 'Capacity',
             'SKUs', 'Avg Orig (SAR)', 'Avg Sale (SAR)', 'Avg Disc %',
             'LG SKUs', 'LG Avg Sale', 'LG vs Mkt %']
    for ci, h in enumerate(kpi_h, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(kpi_h))
    r += 1

    for cat in CAT_ORDER:
        dc = dl[dl['Category'] == cat]
        if dc.empty:
            continue

        # Category subtotal row
        mkt_avg = round(dc['Sale_Price'].mean()) if dc['Sale_Price'].notna().any() else 0
        lg_dc = dc[dc['Brand'].str.upper() == 'LG']
        lg_avg = round(lg_dc['Sale_Price'].mean()) if not lg_dc.empty else None
        gap_pct = _pct((lg_avg - mkt_avg), mkt_avg) if lg_avg else None

        for ci in range(1, len(kpi_h) + 1):
            ws.cell(row=r, column=ci).fill = CAT_FILL
        ws.cell(r, 1, cat).font = BOLD_FONT
        ws.cell(r, 5, len(dc)).font = BOLD_FONT
        ws.cell(r, 5).alignment = CTR
        if dc['Original_Price'].notna().any():
            c = ws.cell(r, 6, round(dc['Original_Price'].mean()))
            c.number_format = '#,##0'; c.font = BOLD_FONT
        if mkt_avg:
            c = ws.cell(r, 7, mkt_avg)
            c.number_format = '#,##0'; c.font = BOLD_FONT
        if dc['Discount_Pct'].notna().any():
            c = ws.cell(r, 8, dc['Discount_Pct'].mean())
            c.number_format = '0.0%'; c.font = BOLD_FONT; c.alignment = CTR
        if not lg_dc.empty:
            ws.cell(r, 9, len(lg_dc)).font = BOLD_FONT
            ws.cell(r, 9).alignment = CTR
            if lg_avg:
                c = ws.cell(r, 10, lg_avg)
                c.number_format = '#,##0'; c.font = BOLD_FONT
        if gap_pct is not None:
            c = ws.cell(r, 11, gap_pct)
            c.number_format = '0.0%'
            c.font = UP_FONT if gap_pct > 0 else DOWN_FONT
        r += 1

        # Detail rows by Cooling x Inverter x Tonnage
        for _, grp in dc.groupby(['Cooling_Type', 'Inverter']):
            cool = grp['Cooling_Type'].iloc[0]
            inv = grp['Inverter'].iloc[0]
            for ton, tg in grp.groupby('Tonnage'):
                fill = LIGHT_FILL if (r % 2 == 0) else None
                ws.cell(r, 2, cool).font = DATA_FONT
                ws.cell(r, 3, inv).font = DATA_FONT
                ws.cell(r, 4, _fmt_ton(ton)).font = DATA_FONT
                ws.cell(r, 5, len(tg)).font = DATA_FONT; ws.cell(r, 5).alignment = CTR
                if tg['Original_Price'].notna().any():
                    c = ws.cell(r, 6, round(tg['Original_Price'].mean()))
                    c.number_format = '#,##0'; c.font = DATA_FONT
                if tg['Sale_Price'].notna().any():
                    c = ws.cell(r, 7, round(tg['Sale_Price'].mean()))
                    c.number_format = '#,##0'; c.font = DATA_FONT
                if tg['Discount_Pct'].notna().any():
                    c = ws.cell(r, 8, tg['Discount_Pct'].mean())
                    c.number_format = '0.0%'; c.font = DATA_FONT; c.alignment = CTR

                lg_t = tg[tg['Brand'].str.upper() == 'LG']
                if not lg_t.empty:
                    ws.cell(r, 9, len(lg_t)).font = DATA_FONT; ws.cell(r, 9).alignment = CTR
                    lg_a = round(lg_t['Sale_Price'].mean())
                    ws.cell(r, 10, lg_a).number_format = '#,##0'
                    mkt_a = round(tg['Sale_Price'].mean())
                    if mkt_a:
                        gp = _pct(lg_a - mkt_a, mkt_a)
                        c = ws.cell(r, 11, gp)
                        c.number_format = '0.0%'
                        c.font = UP_FONT if gp > 0 else DOWN_FONT

                if fill:
                    for ci in range(1, len(kpi_h) + 1):
                        ws.cell(r, ci).fill = fill
                r += 1

    # ── Section 2: Brand Distribution ─────────────────────────────────────
    r += 2
    ws.merge_cells(f'A{r}:K{r}')
    ws[f'A{r}'] = '▶  Brand Distribution (Latest Scrape)'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    brand_h = ['Brand', 'SKUs', 'Share %', 'In Stock', 'Stock Rate',
               'Avg Orig (SAR)', 'Avg Sale (SAR)', 'Min Sale', 'Max Sale']
    for ci, h in enumerate(brand_h, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(brand_h))
    r += 1

    brand_g = dl.groupby('Brand')
    brand_stats = []
    for brand, bg in brand_g:
        if not brand:
            continue
        brand_stats.append({
            'Brand': brand, 'SKUs': len(bg),
            'Share': len(bg) / len(dl) if len(dl) else 0,
            'InStock': bg['In_Stock_Bool'].sum(),
            'StockRate': bg['In_Stock_Bool'].mean(),
            'AvgOrig': round(bg['Original_Price'].mean()) if bg['Original_Price'].notna().any() else 0,
            'AvgSale': round(bg['Sale_Price'].mean()) if bg['Sale_Price'].notna().any() else 0,
            'MinSale': bg['Sale_Price'].min() if bg['Sale_Price'].notna().any() else 0,
            'MaxSale': bg['Sale_Price'].max() if bg['Sale_Price'].notna().any() else 0,
        })
    brand_stats.sort(key=lambda x: x['SKUs'], reverse=True)

    for bs in brand_stats:
        fill = LIGHT_FILL if (r % 2 == 0) else None
        ws.cell(r, 1, bs['Brand']).font = BOLD_FONT
        ws.cell(r, 2, bs['SKUs']).font = DATA_FONT; ws.cell(r, 2).alignment = CTR
        ws.cell(r, 3, bs['Share']).number_format = '0.0%'; ws.cell(r, 3).alignment = CTR
        ws.cell(r, 4, bs['InStock']).font = DATA_FONT; ws.cell(r, 4).alignment = CTR
        ws.cell(r, 5, bs['StockRate']).number_format = '0.0%'; ws.cell(r, 5).alignment = CTR
        ws.cell(r, 6, bs['AvgOrig']).number_format = '#,##0'
        ws.cell(r, 7, bs['AvgSale']).number_format = '#,##0'
        ws.cell(r, 8, bs['MinSale']).number_format = '#,##0'
        ws.cell(r, 9, bs['MaxSale']).number_format = '#,##0'
        if fill:
            for ci in range(1, len(brand_h) + 1):
                ws.cell(r, ci).fill = fill
        r += 1

    # ── Section 3: Changes vs Previous ────────────────────────────────────
    if not dp.empty:
        r += 2
        ws.merge_cells(f'A{r}:K{r}')
        prev_str = pd.Timestamp(prev).strftime('%Y-%m-%d')
        lat_str = pd.Timestamp(latest).strftime('%Y-%m-%d')
        ws[f'A{r}'] = f'▶  Changes: {prev_str} → {lat_str}'
        ws[f'A{r}'].font = SEC_FONT
        r += 2

        prev_skus = set(dp['SKU'].unique())
        lat_skus = set(dl['SKU'].unique())
        new_skus = lat_skus - prev_skus
        disc_skus = prev_skus - lat_skus

        # Price changes
        merged = dl[['SKU', 'Sale_Price']].merge(
            dp[['SKU', 'Sale_Price']].rename(columns={'Sale_Price': 'Prev_Price'}),
            on='SKU', how='inner'
        )
        changed = merged[merged['Sale_Price'] != merged['Prev_Price']]
        up_cnt = (changed['Sale_Price'] > changed['Prev_Price']).sum()
        dn_cnt = (changed['Sale_Price'] < changed['Prev_Price']).sum()

        chg_h = ['Metric', 'Count']
        for ci, h in enumerate(chg_h, 1):
            ws.cell(row=r, column=ci, value=h)
        _hdr(ws, r, 2)
        r += 1

        metrics = [
            ('New SKUs', len(new_skus)),
            ('Discontinued SKUs', len(disc_skus)),
            ('Price Increased ▲', up_cnt),
            ('Price Decreased ▼', dn_cnt),
            ('Price Unchanged', len(merged) - len(changed)),
        ]
        for label, val in metrics:
            ws.cell(r, 1, label).font = DATA_FONT
            ws.cell(r, 2, val).font = BOLD_FONT
            ws.cell(r, 2).alignment = CTR
            r += 1

    _auto(ws)
    ws.freeze_panes = 'A4'
    print(f"  [OK] {sn}")


# ═════════════════════════════════════════════════════════════════════════════
# 2. Price_Change_Alert
# ═════════════════════════════════════════════════════════════════════════════
def build_price_alert(wb, df):
    sn = 'Price_Change_Alert'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['scrape_day'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Not enough scrape dates for comparison.'
        print(f"  [SKIP] {sn} - need >=2 dates")
        return

    latest, prev = dates[-1], dates[-2]
    dl = df[df['scrape_day'] == latest]
    dp = df[df['scrape_day'] == prev]

    merged = dl[['SKU', 'Brand', 'Category', 'Product_Name_EN', 'Tonnage', 'Sale_Price', 'In_Stock']].merge(
        dp[['SKU', 'Sale_Price']].rename(columns={'Sale_Price': 'Prev_Sale'}),
        on='SKU', how='inner'
    )
    changed = merged[merged['Sale_Price'] != merged['Prev_Sale']].copy()

    if changed.empty:
        ws['A1'] = f'No price changes between {pd.Timestamp(prev).strftime("%Y-%m-%d")} and {pd.Timestamp(latest).strftime("%Y-%m-%d")}.'
        print(f"  [OK] {sn} - no changes")
        return

    changed['Delta'] = changed['Sale_Price'] - changed['Prev_Sale']
    changed['Delta_Pct'] = changed['Delta'] / changed['Prev_Sale']
    changed = changed.sort_values('Delta_Pct')

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f'Price Changes: {pd.Timestamp(prev).strftime("%Y-%m-%d")} → {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A1'].font = TITLE_FONT

    r = 3
    hdrs = ['Brand', 'Category', 'Product', 'SKU', 'Ton',
            'Prev Price', 'New Price', 'Change (SAR)', 'Change %']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(hdrs))
    r += 1

    for _, row in changed.iterrows():
        ws.cell(r, 1, row['Brand']).font = BOLD_FONT
        ws.cell(r, 2, row['Category']).font = DATA_FONT
        ws.cell(r, 3, row['Product_Name_EN']).font = DATA_FONT
        ws.cell(r, 4, row['SKU']).font = DATA_FONT
        ws.cell(r, 5, _fmt_ton(row['Tonnage'])).font = DATA_FONT; ws.cell(r, 5).alignment = CTR
        ws.cell(r, 6, row['Prev_Sale']).number_format = '#,##0'
        ws.cell(r, 7, row['Sale_Price']).number_format = '#,##0'

        delta = row['Delta']
        delta_pct = row['Delta_Pct']
        font = UP_FONT if delta > 0 else DOWN_FONT
        prefix = '▲ ' if delta > 0 else '▼ '

        c = ws.cell(r, 8, abs(delta))
        c.number_format = '#,##0'; c.font = font
        c.value = f"{prefix}{abs(int(delta))}"

        c = ws.cell(r, 9, abs(delta_pct))
        c.number_format = '0.0%'; c.font = font

        if r % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    _auto(ws)
    ws.freeze_panes = 'A4'
    print(f"  [OK] {sn} ({len(changed)} changes)")


# ═════════════════════════════════════════════════════════════════════════════
# 3. New_Discontinued_SKUs
# ═════════════════════════════════════════════════════════════════════════════
def build_new_disc(wb, df):
    sn = 'New_Discontinued_SKUs'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['scrape_day'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Not enough scrape dates for comparison.'
        print(f"  [SKIP] {sn}")
        return

    latest, prev = dates[-1], dates[-2]
    dl = df[df['scrape_day'] == latest]
    dp = df[df['scrape_day'] == prev]

    new_skus = set(dl['SKU'].unique()) - set(dp['SKU'].unique())
    disc_skus = set(dp['SKU'].unique()) - set(dl['SKU'].unique())

    r = 1
    ws.merge_cells('A1:G1')
    ws['A1'] = f'New & Discontinued: {pd.Timestamp(prev).strftime("%Y-%m-%d")} → {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A1'].font = TITLE_FONT

    # New SKUs
    r = 3
    ws[f'A{r}'] = f'▶ New SKUs ({len(new_skus)})'
    ws[f'A{r}'].font = SEC_FONT
    r += 1

    hdrs = ['Brand', 'Category', 'Product', 'SKU', 'Ton', 'Sale Price', 'In Stock']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(hdrs))
    r += 1

    if new_skus:
        new_df = dl[dl['SKU'].isin(new_skus)].sort_values(['Brand', 'Category'])
        for _, row in new_df.iterrows():
            ws.cell(r, 1, row['Brand']).font = NEW_FONT
            ws.cell(r, 2, row['Category']).font = DATA_FONT
            ws.cell(r, 3, row['Product_Name_EN']).font = DATA_FONT
            ws.cell(r, 4, row['SKU']).font = DATA_FONT
            ws.cell(r, 5, _fmt_ton(row['Tonnage'])).font = DATA_FONT; ws.cell(r, 5).alignment = CTR
            ws.cell(r, 6, row['Sale_Price']).number_format = '#,##0'
            ws.cell(r, 7, row['In_Stock']).font = DATA_FONT
            r += 1
    else:
        ws.cell(r, 1, 'No new SKUs').font = DATA_FONT
        r += 1

    # Discontinued SKUs
    r += 2
    ws[f'A{r}'] = f'▶ Discontinued SKUs ({len(disc_skus)})'
    ws[f'A{r}'].font = SEC_FONT
    r += 1

    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(hdrs))
    r += 1

    if disc_skus:
        disc_df = dp[dp['SKU'].isin(disc_skus)].sort_values(['Brand', 'Category'])
        for _, row in disc_df.iterrows():
            ws.cell(r, 1, row['Brand']).font = DISC_FONT
            ws.cell(r, 2, row['Category']).font = DISC_FONT
            ws.cell(r, 3, row['Product_Name_EN']).font = DISC_FONT
            ws.cell(r, 4, row['SKU']).font = DISC_FONT
            ws.cell(r, 5, _fmt_ton(row['Tonnage'])).font = DISC_FONT; ws.cell(r, 5).alignment = CTR
            ws.cell(r, 6, row['Sale_Price']).number_format = '#,##0'
            ws.cell(r, 7, row['In_Stock']).font = DISC_FONT
            r += 1
    else:
        ws.cell(r, 1, 'No discontinued SKUs').font = DATA_FONT
        r += 1

    _auto(ws)
    print(f"  [OK] {sn} (new={len(new_skus)}, disc={len(disc_skus)})")


# ═════════════════════════════════════════════════════════════════════════════
# 4. Brand_Price_Compare
# ═════════════════════════════════════════════════════════════════════════════
def build_brand_compare(wb, df):
    sn = 'Brand_Price_Compare'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['scrape_day'].dropna().unique())
    latest = dates[-1]
    dl = df[df['scrape_day'] == latest]

    ws.merge_cells('A1:I1')
    ws['A1'] = 'Brand Price Comparison by Category × Capacity'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Based on: {pd.Timestamp(latest).strftime("%Y-%m-%d")}  |  LG vs Market Gap highlighted'
    ws['A2'].font = SUB_FONT

    r = 4
    for cat in CAT_ORDER:
        dc = dl[dl['Category'] == cat]
        if dc.empty:
            continue

        ws.merge_cells(f'A{r}:I{r}')
        ws[f'A{r}'] = f'▶ {cat}'
        ws[f'A{r}'].font = SEC_FONT
        r += 1

        tons = sorted(dc['Tonnage'].dropna().unique())
        brands = sorted(dc['Brand'].dropna().unique())

        # Header: Brand, then each Ton
        hdrs = ['Brand'] + [_fmt_ton(t) for t in tons] + ['Overall Avg']
        for ci, h in enumerate(hdrs, 1):
            ws.cell(row=r, column=ci, value=h)
        _hdr(ws, r, len(hdrs))
        r += 1

        # Market average row
        for ci in range(1, len(hdrs) + 1):
            ws.cell(r, ci).fill = CAT_FILL
        ws.cell(r, 1, 'Market Avg').font = BOLD_FONT
        for ti, ton in enumerate(tons):
            tg = dc[dc['Tonnage'] == ton]
            if tg['Sale_Price'].notna().any():
                c = ws.cell(r, 2 + ti, round(tg['Sale_Price'].mean()))
                c.number_format = '#,##0'; c.font = BOLD_FONT
        if dc['Sale_Price'].notna().any():
            c = ws.cell(r, len(hdrs), round(dc['Sale_Price'].mean()))
            c.number_format = '#,##0'; c.font = BOLD_FONT
        r += 1

        # Per brand rows
        for brand in brands:
            bg = dc[dc['Brand'] == brand]
            is_lg = brand.upper() == 'LG'
            font = BOLD_FONT if is_lg else DATA_FONT

            ws.cell(r, 1, brand).font = font
            for ti, ton in enumerate(tons):
                tg = bg[bg['Tonnage'] == ton]
                if tg['Sale_Price'].notna().any():
                    val = round(tg['Sale_Price'].mean())
                    c = ws.cell(r, 2 + ti, val)
                    c.number_format = '#,##0'; c.font = font

                    # LG vs Market gap coloring
                    if is_lg:
                        mkt = dc[(dc['Tonnage'] == ton)]['Sale_Price'].mean()
                        if mkt and mkt > 0:
                            gap = (val - mkt) / mkt
                            if gap > 0.05:
                                c.fill = RED_FILL
                            elif gap < -0.05:
                                c.fill = GREEN_FILL

            if bg['Sale_Price'].notna().any():
                c = ws.cell(r, len(hdrs), round(bg['Sale_Price'].mean()))
                c.number_format = '#,##0'; c.font = font

            if r % 2 == 0:
                for ci in range(1, len(hdrs) + 1):
                    if not ws.cell(r, ci).fill or ws.cell(r, ci).fill.fgColor.rgb in ('00000000', None):
                        ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

        r += 1  # gap between categories

    _auto(ws)
    ws.freeze_panes = 'B5'
    print(f"  [OK] {sn}")


# ═════════════════════════════════════════════════════════════════════════════
# 5. Full_Price_Tracking
# ═════════════════════════════════════════════════════════════════════════════
def build_full_tracking(wb, df):
    sn = 'Full_Price_Tracking'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['scrape_day'].dropna().unique())
    latest = dates[-1]
    prev = dates[-2] if len(dates) >= 2 else None
    dl = df[df['scrape_day'] == latest].copy()
    dp = df[df['scrape_day'] == prev] if prev else pd.DataFrame()

    if not dp.empty:
        prev_prices = dp.groupby('SKU')['Sale_Price'].first().to_dict()
        dl['Prev_Sale'] = dl['SKU'].map(prev_prices)
        dl['Price_Delta'] = dl['Sale_Price'] - dl['Prev_Sale']
    else:
        dl['Prev_Sale'] = np.nan
        dl['Price_Delta'] = np.nan

    # Sort by Category order, then Tonnage, then Brand
    dl['_cat_sort'] = dl['Category'].map({c: i for i, c in enumerate(CAT_ORDER)}).fillna(99)
    dl = dl.sort_values(['_cat_sort', 'Tonnage', 'Brand']).drop(columns='_cat_sort')

    ws.merge_cells('A1:N1')
    ws['A1'] = f'Full Product List — {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A1'].font = TITLE_FONT

    r = 3
    hdrs = ['Brand', 'Category', 'Product Name', 'SKU', 'Ton', 'Cooling', 'Inverter',
            'Compressor', 'Orig Price', 'Sale Price', 'Discount', 'Prev Sale',
            'Price Δ', 'In Stock', 'Stock Qty']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(hdrs))
    r += 1

    for _, row in dl.iterrows():
        ws.cell(r, 1, row['Brand']).font = BOLD_FONT
        ws.cell(r, 2, row['Category']).font = DATA_FONT
        ws.cell(r, 3, row['Product_Name_EN']).font = DATA_FONT
        ws.cell(r, 4, row['SKU']).font = DATA_FONT
        ws.cell(r, 5, _fmt_ton(row['Tonnage'])).font = DATA_FONT; ws.cell(r, 5).alignment = CTR
        ws.cell(r, 6, row['Cooling_Type']).font = DATA_FONT
        ws.cell(r, 7, row['Inverter']).font = DATA_FONT; ws.cell(r, 7).alignment = CTR
        ws.cell(r, 8, row['Compressor']).font = DATA_FONT
        ws.cell(r, 9, row['Original_Price']).number_format = '#,##0'
        ws.cell(r, 10, row['Sale_Price']).number_format = '#,##0'
        ws.cell(r, 11, row['Discount']).font = DATA_FONT; ws.cell(r, 11).alignment = CTR

        if pd.notna(row.get('Prev_Sale')):
            ws.cell(r, 12, row['Prev_Sale']).number_format = '#,##0'
        if pd.notna(row.get('Price_Delta')) and row['Price_Delta'] != 0:
            delta = row['Price_Delta']
            prefix = '▲' if delta > 0 else '▼'
            c = ws.cell(r, 13, f"{prefix} {abs(int(delta))}")
            c.font = UP_FONT if delta > 0 else DOWN_FONT

        ws.cell(r, 14, row['In_Stock']).font = DATA_FONT; ws.cell(r, 14).alignment = CTR
        if pd.notna(row.get('Stock_Qty')):
            ws.cell(r, 15, int(row['Stock_Qty'])).font = DATA_FONT

        if r % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    _auto(ws)
    ws.freeze_panes = 'A4'
    print(f"  [OK] {sn} ({len(dl)} rows)")


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("Bin Momen AC Dashboard Builder")
    print("=" * 60)

    if not os.path.exists(DATA_FILE):
        print(f"[ERROR] Data file not found: {DATA_FILE}")
        sys.exit(1)

    print(f"[1/6] Loading data from {os.path.basename(DATA_FILE)}...")
    df = load_data()
    dates = sorted(df['scrape_day'].dropna().unique())
    print(f"       Rows: {len(df)}, Dates: {len(dates)}, "
          f"Latest: {pd.Timestamp(dates[-1]).strftime('%Y-%m-%d') if dates else 'N/A'}")

    # Create or load workbook
    if os.path.exists(OUT_FILE):
        wb = load_workbook(OUT_FILE)
        print(f"[2/6] Loaded existing {os.path.basename(OUT_FILE)}")
    else:
        wb = Workbook()
        # Copy raw data as Prices_DB sheet
        ws_db = wb.active
        ws_db.title = 'Prices_DB'
        raw_df = pd.read_excel(DATA_FILE)
        for ci, col in enumerate(raw_df.columns, 1):
            ws_db.cell(1, ci, col)
        for ri, row in raw_df.iterrows():
            for ci, val in enumerate(row, 1):
                ws_db.cell(ri + 2, ci, val)
        print(f"[2/6] Created new {os.path.basename(OUT_FILE)} with Prices_DB sheet")

    print("[3/6] Building Dashboard_Summary...")
    build_summary(wb, df)

    print("[4/6] Building Price_Change_Alert & New_Discontinued...")
    build_price_alert(wb, df)
    build_new_disc(wb, df)

    print("[5/6] Building Brand_Price_Compare & Full_Price_Tracking...")
    build_brand_compare(wb, df)
    build_full_tracking(wb, df)

    # Reorder sheets
    desired = ['Dashboard_Summary', 'Price_Change_Alert', 'New_Discontinued_SKUs',
               'Brand_Price_Compare', 'Full_Price_Tracking', 'Prices_DB']
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != idx:
                wb.move_sheet(name, offset=idx - cur)

    print(f"[6/6] Saving to {os.path.basename(OUT_FILE)}...")
    try:
        wb.save(OUT_FILE)
        print(f"\n✅ Done! → {OUT_FILE}")
    except PermissionError:
        alt = OUT_FILE.replace('.xlsx', '_new.xlsx')
        wb.save(alt)
        print(f"\n⚠️ File locked, saved to: {alt}")

    print(f"Sheet order: {wb.sheetnames}")


if __name__ == '__main__':
    main()
