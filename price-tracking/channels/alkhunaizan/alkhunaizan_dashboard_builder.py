#!/usr/bin/env python3
"""
Al-Khunaizan AC - Dashboard Auto Builder v2
- 세분화: Category > Compressor > Type > Ton
- Brand Price Comparison: 평균가, Ton 기준
- Full_Price_Tracking: Ton 기준 정렬
- Stock_Status_Analysis (AlKhunaizan 전용)
- Gree_vs_LG_Analysis (경쟁사 분석)
Author: 핍쫑이
"""

import os, sys, re

try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
except ImportError as e:
    print(f"❌ 필수 패키지 설치 필요: {e}")
    print("py -m pip install pandas openpyxl numpy --break-system-packages")
    sys.exit(1)

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__)) if os.path.dirname(os.path.abspath(__file__)) else os.getcwd()
INPUT_FILE = os.path.join(CURRENT_DIR, "AlKhunaizan_AC_Prices Tracking_Master.xlsx")

# ── 스타일 (eXtra 동일) ──────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT   = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT= Font(name='Arial', bold=True, size=10, color='666666')
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E79')
DATA_FONT    = Font(name='Arial', size=10)
NUM_FONT     = Font(name='Arial', size=10)
BOLD_FONT    = Font(name='Arial', bold=True, size=10)
UP_FONT      = Font(name='Arial', size=10, color='FF0000', bold=True)
DOWN_FONT    = Font(name='Arial', size=10, color='008000', bold=True)
NEW_FONT     = Font(name='Arial', size=10, color='0070C0', bold=True)
DISC_FONT    = Font(name='Arial', size=10, color='999999')
THICK_BORDER = Border(bottom=Side(style='medium', color='1F4E79'))
CENTER = Alignment(horizontal='center', vertical='center')
WRAP   = Alignment(horizontal='left', vertical='center', wrap_text=True)
LIGHT_FILL = PatternFill('solid', fgColor='F2F7FB')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
RED_FILL   = PatternFill('solid', fgColor='FCE4EC')
CAT_FILL   = PatternFill('solid', fgColor='D6E4F0')

# AlKhunaizan 카테고리 순서
CAT_ORDER_PREF = ['Window AC', 'Split AC', 'Cassette AC', 'Floor Standing AC', 'Portable AC']


# ── 공통 헬퍼 ────────────────────────────────────────────────
def safe_round(val, default=0):
    """NaN-safe round: NaN이면 default 반환"""
    try:
        return round(val) if pd.notna(val) else default
    except (TypeError, ValueError):
        return default


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THICK_BORDER


def auto_width(ws, min_width=10, max_width=35):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def delete_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def get_cat_order(dl):
    cats = dl['Category'].dropna().unique().tolist()
    ordered = [c for c in CAT_ORDER_PREF if c in cats]
    ordered += [c for c in cats if c not in ordered]
    return ordered


# ── 데이터 로드 & 전처리 ─────────────────────────────────────
def load_data(filepath):
    df = pd.read_excel(filepath, sheet_name='Products_DB', engine='openpyxl')
    df['Scraped_At'] = pd.to_datetime(df['Scraped_At'])

    for col in ['Promotion Price (SAR)', 'Original Price (SAR)', 'Only Pay Price (SAR)', 'Save amount (SAR)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Effective Price: Only Pay 우선, 없으면 Promotion Price
    only_pay = df['Only Pay Price (SAR)'] if 'Only Pay Price (SAR)' in df.columns else pd.Series(np.nan, index=df.index)
    promo    = df['Promotion Price (SAR)'] if 'Promotion Price (SAR)' in df.columns else pd.Series(np.nan, index=df.index)
    df['Effective_Price'] = np.where(only_pay.notna() & (only_pay > 0), only_pay, promo)

    # Discount Rate
    if 'Discount Rate (%)' in df.columns:
        dr = df['Discount Rate (%)'].astype(str).str.replace('%', '', regex=False).str.strip()
        df['Discount_Numeric'] = pd.to_numeric(dr, errors='coerce')   # 12.1 형태
        df['Discount_Rate']    = df['Discount_Numeric'] / 100          # 0.121 형태
    else:
        df['Discount_Numeric'] = 0.0
        df['Discount_Rate']    = 0.0

    # Nominal Capacity: 이중 공백 정규화 + BTU 기반 오류 보정
    # (스펙 테이블 실패 시 구 BTU 범위로 잘못 저장된 값을 수정, e.g. 3/9 데이터)
    if 'Nominal Capacity' in df.columns:
        df['Nominal Capacity'] = (
            df['Nominal Capacity'].astype(str)
            .str.replace(r'\s+', ' ', regex=True).str.strip()
            .replace('nan', np.nan)
        )

    if 'Capacity' in df.columns and 'Nominal Capacity' in df.columns:
        _TON_MAP = [
            (7000,  11999, '0.75 Ton'),
            (12000, 15999, '1 Ton'),
            (16000, 20999, '1.5 Ton'),   # 18000 BTU = 1.5 Ton (구 코드 오류 수정)
            (21000, 24999, '2 Ton'),
            (25000, 32999, '2.5 Ton'),
            (33000, 38999, '3 Ton'),
            (39000, 44999, '3.5 Ton'),   # 42000 BTU = 3.5 Ton
            (45000, 55999, '4 Ton'),
            (56000, 75000, '5 Ton'),
        ]
        def _btu_to_nom(cap_str):
            if pd.isna(cap_str):
                return ''
            m = re.search(r'(\d{4,6})', str(cap_str))
            if not m:
                return ''
            btu = int(m.group(1))
            for lo, hi, label in _TON_MAP:
                if lo <= btu <= hi:
                    return label
            return ''

        btu_derived = df['Capacity'].apply(_btu_to_nom)
        has_btu     = btu_derived.notna() & (btu_derived != '')
        # BTU 도출 값이 있으면 덮어쓰기 (공백·잘못된 구 BTU 범위 값 보정)
        df['Nominal Capacity'] = np.where(has_btu, btu_derived, df['Nominal Capacity'])

    # Cooling_Capacity_Ton  (예: "1.5 Ton" → 1.5)
    if 'Nominal Capacity' in df.columns:
        df['Cooling_Capacity_Ton'] = (
            df['Nominal Capacity'].astype(str)
            .str.extract(r'(\d+\.?\d*)')[0]
            .astype(float)
        )
    else:
        df['Cooling_Capacity_Ton'] = np.nan

    # Compressor / Type / Stock
    df['Compressor_Type'] = (df['Compressor Type'].fillna('N/A').astype(str).str.strip()
                             if 'Compressor Type' in df.columns
                             else pd.Series(['N/A'] * len(df)))
    df['Compressor_Type'] = df['Compressor_Type'].replace('', 'N/A')

    df['Cold_or_HC'] = (df['Type'].fillna('N/A').astype(str).str.strip()
                        if 'Type' in df.columns
                        else pd.Series(['N/A'] * len(df)))

    df['Stock_Status'] = (df['Stock Status'].fillna('Unknown').astype(str).str.strip()
                          if 'Stock Status' in df.columns
                          else pd.Series(['Unknown'] * len(df)))

    return df


# ============================================================
# 1. Dashboard_Summary
# ============================================================
def build_dashboard_summary(wb, df):
    sn = 'Dashboard_Summary'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn, 0)

    dates = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    prev   = dates[-2] if len(dates) >= 2 else None
    dl = df[df['Scraped_At'] == latest]
    dp = df[df['Scraped_At'] == prev] if prev is not None else pd.DataFrame()

    ws.merge_cells('A1:L1')
    ws['A1'] = '🏪 Al-Khunaizan AC Price Tracking Dashboard'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Last Updated: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Total SKUs: {len(dl)}'
    ws['A2'].font = SUBTITLE_FONT

    # ── KPI 세분화 테이블 ──
    r = 4
    ws.merge_cells(f'A{r}:L{r}')
    ws[f'A{r}'] = '📊 Key Performance Indicators (Detailed)'
    ws[f'A{r}'].font = SECTION_FONT

    r = 6
    headers = ['Category', 'Compressor', 'Type', 'Ton',
               'SKU Count', 'Avg Orig Price', 'Avg Promo Price', 'Avg Eff Price',
               'Avg Discount', 'LG Avg', 'LG Gap', 'LG Gap %']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    cat_order  = get_cat_order(dl)
    row_idx    = r + 1
    grand_total = 0

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue

        # Category 소계 행
        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=row_idx, column=5).alignment = CENTER
        ws.cell(row=row_idx, column=6, value=safe_round(dc['Original Price (SAR)'].mean())).font = BOLD_FONT
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        ws.cell(row=row_idx, column=7, value=safe_round(dc['Promotion Price (SAR)'].mean())).font = BOLD_FONT
        ws.cell(row=row_idx, column=7).number_format = '#,##0'
        mkt_avg_cat = safe_round(dc['Effective_Price'].mean())
        ws.cell(row=row_idx, column=8, value=mkt_avg_cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=8).number_format = '#,##0'
        disc_avg = dc['Discount_Rate'].mean()
        disc_str = f'{disc_avg*100:.1f}%' if pd.notna(disc_avg) else '-'
        ws.cell(row=row_idx, column=9, value=disc_str).font = BOLD_FONT
        ws.cell(row=row_idx, column=9).alignment = CENTER
        # LG vs Market
        lg_cat = dc[dc['Brand'] == 'LG']
        if len(lg_cat) > 0:
            lg_avg_cat = safe_round(lg_cat['Effective_Price'].mean())
            gap_sar = lg_avg_cat - mkt_avg_cat
            gap_pct = gap_sar / mkt_avg_cat if mkt_avg_cat else 0
            ws.cell(row=row_idx, column=10, value=lg_avg_cat).font = BOLD_FONT
            ws.cell(row=row_idx, column=10).number_format = '#,##0'
            cg = ws.cell(row=row_idx, column=11, value=gap_sar)
            cg.number_format = '#,##0'
            cg.font = UP_FONT if gap_sar > 0 else DOWN_FONT if gap_sar < 0 else BOLD_FONT
            cp = ws.cell(row=row_idx, column=12, value=gap_pct)
            cp.number_format = '0%'
            cp.font = UP_FONT if gap_sar > 0 else DOWN_FONT if gap_sar < 0 else BOLD_FONT
        else:
            for ci in [10, 11, 12]:
                ws.cell(row=row_idx, column=ci, value='-').font = BOLD_FONT
        for ci in range(10, 13):
            ws.cell(row=row_idx, column=ci).alignment = CENTER
        for c in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1
        grand_total += len(dc)

        # Sub-rows: Compressor > Type > Ton
        groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
        for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2] if pd.notna(x[0][2]) else 0)):
            ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=f'{ton}T' if pd.notna(ton) else 'N/A').font = DATA_FONT
            ws.cell(row=row_idx, column=4).alignment = CENTER
            ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
            ws.cell(row=row_idx, column=5).alignment = CENTER
            ws.cell(row=row_idx, column=6, value=safe_round(grp['Original Price (SAR)'].mean())).font = NUM_FONT
            ws.cell(row=row_idx, column=6).number_format = '#,##0'
            ws.cell(row=row_idx, column=7, value=safe_round(grp['Promotion Price (SAR)'].mean())).font = NUM_FONT
            ws.cell(row=row_idx, column=7).number_format = '#,##0'
            mkt_avg = safe_round(grp['Effective_Price'].mean())
            ws.cell(row=row_idx, column=8, value=mkt_avg).font = NUM_FONT
            ws.cell(row=row_idx, column=8).number_format = '#,##0'
            d = grp['Discount_Rate'].mean()
            ws.cell(row=row_idx, column=9, value=f'{d*100:.1f}%' if pd.notna(d) else '-').font = NUM_FONT
            ws.cell(row=row_idx, column=9).alignment = CENTER
            lg_grp = grp[grp['Brand'] == 'LG']
            if len(lg_grp) > 0:
                lg_avg = safe_round(lg_grp['Effective_Price'].mean())
                gap_s  = lg_avg - mkt_avg
                gap_p  = gap_s / mkt_avg if mkt_avg else 0
                ws.cell(row=row_idx, column=10, value=lg_avg).font = NUM_FONT
                ws.cell(row=row_idx, column=10).number_format = '#,##0'
                cg = ws.cell(row=row_idx, column=11, value=gap_s)
                cg.number_format = '#,##0'
                cg.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
                cp = ws.cell(row=row_idx, column=12, value=gap_p)
                cp.number_format = '0%'
                cp.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
            else:
                for ci in [10, 11, 12]:
                    ws.cell(row=row_idx, column=ci, value='-').font = NUM_FONT
            for ci in range(10, 13):
                ws.cell(row=row_idx, column=ci).alignment = CENTER
            if row_idx % 2 == 0:
                for c in range(1, len(headers)+1):
                    ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
            row_idx += 1

    ws.cell(row=row_idx, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=row_idx, column=5, value=grand_total).font = BOLD_FONT
    ws.cell(row=row_idx, column=5).alignment = CENTER
    for c in range(1, len(headers)+1):
        ws.cell(row=row_idx, column=c).fill = CAT_FILL
    row_idx += 1

    # ── Category Distribution by Brand ──
    row_idx += 1
    ws.merge_cells(f'A{row_idx}:L{row_idx}')
    ws[f'A{row_idx}'] = '📦 Category Distribution by Brand'
    ws[f'A{row_idx}'].font = SECTION_FONT
    row_idx += 2

    brands_top = dl['Brand'].value_counts().head(8).index.tolist()
    if 'LG' not in brands_top:
        brands_top.append('LG')
    dist_headers = ['Category', 'Compressor', 'Type', 'Ton', 'Total'] + brands_top
    for i, h in enumerate(dist_headers):
        ws.cell(row=row_idx, column=i+1, value=h)
    style_header_row(ws, row_idx, len(dist_headers))
    row_idx += 1

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue
        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=row_idx, column=5).alignment = CENTER
        for bi, brand in enumerate(brands_top):
            cnt = len(dc[dc['Brand'] == brand])
            ws.cell(row=row_idx, column=6+bi, value=cnt if cnt > 0 else '-').font = BOLD_FONT
            ws.cell(row=row_idx, column=6+bi).alignment = CENTER
        for c in range(1, len(dist_headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1

        groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
        for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2] if pd.notna(x[0][2]) else 0)):
            ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=f'{ton}T' if pd.notna(ton) else 'N/A').font = DATA_FONT
            ws.cell(row=row_idx, column=4).alignment = CENTER
            ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
            ws.cell(row=row_idx, column=5).alignment = CENTER
            for bi, brand in enumerate(brands_top):
                cnt = len(grp[grp['Brand'] == brand])
                ws.cell(row=row_idx, column=6+bi, value=cnt if cnt > 0 else '-').font = NUM_FONT
                ws.cell(row=row_idx, column=6+bi).alignment = CENTER
            if row_idx % 2 == 0:
                for c in range(1, len(dist_headers)+1):
                    ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
            row_idx += 1

    ws.cell(row=row_idx, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=row_idx, column=5, value=len(dl)).font = BOLD_FONT
    ws.cell(row=row_idx, column=5).alignment = CENTER
    for bi, brand in enumerate(brands_top):
        ws.cell(row=row_idx, column=6+bi, value=len(dl[dl['Brand'] == brand])).font = BOLD_FONT
        ws.cell(row=row_idx, column=6+bi).alignment = CENTER
    for c in range(1, len(dist_headers)+1):
        ws.cell(row=row_idx, column=c).fill = CAT_FILL
    row_idx += 1

    # ── Changes vs Previous ──
    if not dp.empty:
        row_idx += 1
        ws.merge_cells(f'A{row_idx}:J{row_idx}')
        ws[f'A{row_idx}'] = f'🔄 Changes vs Previous ({pd.Timestamp(prev).strftime("%Y-%m-%d")})'
        ws[f'A{row_idx}'].font = SECTION_FONT
        row_idx += 2

        prev_skus = set(dp['SKU'].astype(str))
        curr_skus = set(dl['SKU'].astype(str))
        merged = pd.merge(
            dp[['SKU', 'Effective_Price']].astype({'SKU': str}),
            dl[['SKU', 'Effective_Price']].astype({'SKU': str}),
            on='SKU', suffixes=('_prev', '_curr')
        )
        merged['Change'] = merged['Effective_Price_curr'] - merged['Effective_Price_prev']

        items = [
            ('New SKUs',           len(curr_skus - prev_skus)),
            ('Removed SKUs',       len(prev_skus - curr_skus)),
            ('Price Increased ▲', len(merged[merged['Change'] > 0])),
            ('Price Decreased ▼', len(merged[merged['Change'] < 0])),
            ('No Change',          len(merged[merged['Change'] == 0])),
        ]
        ws.cell(row=row_idx, column=1, value='Metric')
        ws.cell(row=row_idx, column=2, value='Count')
        style_header_row(ws, row_idx, 2)
        for i, (m, cnt) in enumerate(items):
            rr = row_idx + 1 + i
            ws.cell(row=rr, column=1, value=m).font = DATA_FONT
            cell = ws.cell(row=rr, column=2, value=cnt)
            cell.alignment = CENTER
            if '▲' in m and cnt > 0:   cell.font = UP_FONT
            elif '▼' in m and cnt > 0: cell.font = DOWN_FONT
            else:                       cell.font = NUM_FONT

    auto_width(ws)
    ws.sheet_properties.tabColor = '1F4E79'


# ============================================================
# 2. Price_Change_Alert
# ============================================================
def build_price_change_alert(wb, df):
    sn = 'Price_Change_Alert'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    if len(dates) < 2:
        ws['A1'] = '⚠️ 최소 2회 이상 스크래핑 데이터가 필요합니다.'
        ws['A1'].font = SUBTITLE_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev].copy()
    dc = df[df['Scraped_At'] == latest].copy()

    cols = ['SKU', 'Brand', 'Product Name', 'Category',
            'Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton', 'Effective_Price']
    merged = pd.merge(
        dp[cols].astype({'SKU': str}),
        dc[cols].astype({'SKU': str}),
        on='SKU', suffixes=('_prev', '_curr')
    )
    merged['Change_SAR'] = merged['Effective_Price_curr'] - merged['Effective_Price_prev']
    merged['Change_Pct'] = merged['Change_SAR'] / merged['Effective_Price_prev']
    changed = merged[merged['Change_SAR'] != 0].sort_values('Change_Pct')

    ws.merge_cells('A1:K1')
    ws['A1'] = '🔔 Price Change Alert (Effective Price)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = (f'Price changes: {pd.Timestamp(prev).strftime("%Y-%m-%d")} → '
                f'{pd.Timestamp(latest).strftime("%Y-%m-%d")} ({len(changed)} models)')
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Brand', 'SKU', 'Product Name', 'Category', 'Compressor',
               'Type', 'Ton', 'Prev Price', 'Curr Price', 'Change (SAR)', 'Change %']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    for idx, (_, rd) in enumerate(changed.iterrows()):
        row = r + 1 + idx
        ws.cell(row=row, column=1, value=rd['Brand_curr']).font = DATA_FONT
        ws.cell(row=row, column=2, value=rd['SKU']).font = DATA_FONT
        ws.cell(row=row, column=3, value=rd['Product Name_curr']).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=4, value=rd['Category_curr']).font = DATA_FONT
        ws.cell(row=row, column=5, value=rd['Compressor_Type_curr']).font = DATA_FONT
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=6, value=rd['Cold_or_HC_curr']).font = DATA_FONT
        ton = rd['Cooling_Capacity_Ton_curr']
        ws.cell(row=row, column=7, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=8, value=rd['Effective_Price_prev']).font = NUM_FONT
        ws.cell(row=row, column=8).number_format = '#,##0'
        ws.cell(row=row, column=9, value=rd['Effective_Price_curr']).font = NUM_FONT
        ws.cell(row=row, column=9).number_format = '#,##0'
        chg = rd['Change_SAR']
        c = ws.cell(row=row, column=10, value=chg)
        c.number_format = '#,##0'
        c.font = DOWN_FONT if chg < 0 else UP_FONT
        p = ws.cell(row=row, column=11, value=rd['Change_Pct'])
        p.number_format = '0%'
        p.font = DOWN_FONT if chg < 0 else UP_FONT
        if row % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=row, column=c).fill = LIGHT_FILL

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = 'FF6600'


# ============================================================
# 3. New_Discontinued_SKUs
# ============================================================
def build_new_discontinued(wb, df):
    sn = 'New_Discontinued_SKUs'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    if len(dates) < 2:
        ws['A1'] = '⚠️ 최소 2회 이상 스크래핑 데이터가 필요합니다.'
        ws['A1'].font = SUBTITLE_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev]
    dc = df[df['Scraped_At'] == latest]
    prev_skus = set(dp['SKU'].astype(str))
    curr_skus = set(dc['SKU'].astype(str))
    new_skus  = curr_skus - prev_skus
    disc_skus = prev_skus - curr_skus

    ws.merge_cells('A1:I1')
    ws['A1'] = '🆕 New & Discontinued SKUs'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:I2')
    ws['A2'] = f'{pd.Timestamp(prev).strftime("%Y-%m-%d")} → {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A2'].font = SUBTITLE_FONT

    headers = ['Brand', 'SKU', 'Product Name', 'Category',
               'Compressor', 'Type', 'Ton', 'Eff Price', 'Stock']

    def write_section(title, title_font, skus, source_df, fill, font, start_row):
        ws.merge_cells(f'A{start_row}:I{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = title_font
        hr = start_row + 1
        for i, h in enumerate(headers):
            ws.cell(row=hr, column=i+1, value=h)
        style_header_row(ws, hr, len(headers))
        df_sub = source_df[source_df['SKU'].astype(str).isin(skus)]
        for idx, (_, rd) in enumerate(df_sub.iterrows()):
            row = hr + 1 + idx
            ws.cell(row=row, column=1, value=rd['Brand']).font = font
            ws.cell(row=row, column=2, value=str(rd['SKU'])).font = font
            ws.cell(row=row, column=3, value=rd['Product Name']).font = DATA_FONT
            ws.cell(row=row, column=3).alignment = WRAP
            ws.cell(row=row, column=4, value=rd['Category']).font = DATA_FONT
            ws.cell(row=row, column=5, value=rd['Compressor_Type']).font = DATA_FONT
            ws.cell(row=row, column=5).alignment = CENTER
            ws.cell(row=row, column=6, value=rd['Cold_or_HC']).font = DATA_FONT
            ton = rd['Cooling_Capacity_Ton']
            ws.cell(row=row, column=7, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
            ws.cell(row=row, column=7).alignment = CENTER
            c = ws.cell(row=row, column=8, value=rd.get('Effective_Price', ''))
            c.font = NUM_FONT
            c.number_format = '#,##0'
            ws.cell(row=row, column=9, value=rd.get('Stock_Status', '')).font = DATA_FONT
            ws.cell(row=row, column=9).alignment = CENTER
            for ci in range(1, len(headers)+1):
                ws.cell(row=row, column=ci).fill = fill
        return hr + 1 + len(df_sub)

    r = 4
    end = write_section(f'🟢 New SKUs ({len(new_skus)})',
                        Font(name='Arial', bold=True, size=12, color='008000'),
                        new_skus, dc, GREEN_FILL, NEW_FONT, r)
    r = end + 2
    write_section(f'🔴 Discontinued SKUs ({len(disc_skus)})',
                  Font(name='Arial', bold=True, size=12, color='CC0000'),
                  disc_skus, dp, RED_FILL, DISC_FONT, r)

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = '00B050'


# ============================================================
# 4. Brand_Avg_Price_Comparison
# ============================================================
def build_brand_price_comparison(wb, df):
    sn = 'Brand_Avg_Price_Comparison'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates    = sorted(df['Scraped_At'].unique())
    latest   = dates[-1]
    dl       = df[df['Scraped_At'] == latest].copy()
    price_col = 'Effective_Price'
    brands_top = dl['Brand'].value_counts().head(7).index.tolist()
    if 'LG' not in brands_top:
        brands_top.append('LG')

    ws.merge_cells('A1:L1')
    ws['A1'] = '💰 Brand Avg Price Comparison by Category & Ton'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Price: Effective Price (Avg)'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Category', 'Compressor', 'Type', 'Ton'] + brands_top + ['Market Avg', 'LG Min']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    cat_order = get_cat_order(dl)
    row_idx   = r + 1

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue
        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        for c in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1

        groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
        for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2] if pd.notna(x[0][2]) else 0)):
            ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=f'{ton}T' if pd.notna(ton) else 'N/A').font = DATA_FONT
            ws.cell(row=row_idx, column=4).alignment = CENTER
            mkt_avg = grp[price_col].mean()
            for bi, brand in enumerate(brands_top):
                bg = grp[grp['Brand'] == brand]
                if len(bg) > 0:
                    c = ws.cell(row=row_idx, column=5+bi, value=round(bg[price_col].mean()))
                    c.number_format = '#,##0'
                    c.font = NUM_FONT
                else:
                    ws.cell(row=row_idx, column=5+bi, value='-').font = NUM_FONT
                ws.cell(row=row_idx, column=5+bi).alignment = CENTER
            mc = ws.cell(row=row_idx, column=5+len(brands_top), value=round(mkt_avg))
            mc.number_format = '#,##0'
            mc.font = BOLD_FONT
            mc.alignment = CENTER
            # LG Min
            lg_min_col = 6 + len(brands_top)
            lg_grp = grp[grp['Brand'] == 'LG']
            if len(lg_grp) > 0:
                lm = ws.cell(row=row_idx, column=lg_min_col, value=round(lg_grp[price_col].min()))
                lm.number_format = '#,##0'
                lm.font = Font(name='Arial', size=10, color='0070C0', bold=True)
            else:
                ws.cell(row=row_idx, column=lg_min_col, value='-').font = NUM_FONT
            ws.cell(row=row_idx, column=lg_min_col).alignment = CENTER
            if row_idx % 2 == 0:
                for c in range(1, len(headers)+1):
                    ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
            row_idx += 1

    auto_width(ws)
    ws.sheet_properties.tabColor = 'FFC000'


# ============================================================
# 5. Full_Price_Tracking
# ============================================================
def build_full_tracking(wb, df):
    sn = 'Full_Price_Tracking'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates    = sorted(df['Scraped_At'].unique())
    latest   = dates[-1]
    prev     = dates[-2] if len(dates) >= 2 else None
    dc       = df[df['Scraped_At'] == latest].copy()
    price_col = 'Effective_Price'

    prev_map = {}
    if prev is not None:
        dp = df[df['Scraped_At'] == prev]
        prev_map = dp.set_index(dp['SKU'].astype(str))[price_col].to_dict()

    ws.merge_cells('A1:P1')
    ws['A1'] = '📋 Full Price Tracking (Latest Snapshot)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:P2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Total: {len(dc)} SKUs'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Brand', 'SKU', 'Product Name', 'Category', 'Compressor', 'Type', 'Ton',
               'Orig Price', 'Promo Price', 'Only Pay', 'Eff Price', 'Discount %',
               'Prev Price', 'Change (SAR)', 'Stock', 'Energy']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    dc_sorted = dc.sort_values(['Category', 'Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton', 'Brand'])

    for idx, (_, rd) in enumerate(dc_sorted.iterrows()):
        row     = r + 1 + idx
        sku_str = str(rd['SKU'])
        ws.cell(row=row, column=1, value=rd['Brand']).font = DATA_FONT
        ws.cell(row=row, column=2, value=sku_str).font = DATA_FONT
        ws.cell(row=row, column=3, value=rd['Product Name']).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=4, value=rd['Category']).font = DATA_FONT
        ws.cell(row=row, column=5, value=rd['Compressor_Type']).font = DATA_FONT
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=6, value=rd['Cold_or_HC']).font = DATA_FONT
        ton = rd['Cooling_Capacity_Ton']
        ws.cell(row=row, column=7, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
        ws.cell(row=row, column=7).alignment = CENTER

        for ci, col_name in enumerate(['Original Price (SAR)', 'Promotion Price (SAR)',
                                        'Only Pay Price (SAR)', price_col], start=8):
            val = rd.get(col_name)
            c   = ws.cell(row=row, column=ci, value=val if pd.notna(val) else '')
            if pd.notna(val):
                c.number_format = '#,##0'
            c.font = NUM_FONT

        disc = rd.get('Discount_Numeric')
        if pd.notna(disc):
            ws.cell(row=row, column=12, value=f'{disc:.1f}%').font = NUM_FONT
        ws.cell(row=row, column=12).alignment = CENTER

        prev_price = prev_map.get(sku_str)
        curr_price = rd.get(price_col)
        if prev_price and pd.notna(prev_price):
            ws.cell(row=row, column=13, value=prev_price).font = NUM_FONT
            ws.cell(row=row, column=13).number_format = '#,##0'
            if curr_price and pd.notna(curr_price):
                change = curr_price - prev_price
                chg = ws.cell(row=row, column=14, value=change)
                chg.number_format = '#,##0'
                chg.font = DOWN_FONT if change < 0 else UP_FONT if change > 0 else NUM_FONT
        else:
            ws.cell(row=row, column=13, value='NEW').font = NEW_FONT
            ws.cell(row=row, column=13).alignment = CENTER

        ws.cell(row=row, column=15, value=rd.get('Stock_Status', '')).font = DATA_FONT
        ws.cell(row=row, column=15).alignment = CENTER
        ws.cell(row=row, column=16, value=rd.get('Energy Grade', '')).font = DATA_FONT
        ws.cell(row=row, column=16).alignment = CENTER

        if row % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=row, column=c).fill = LIGHT_FILL

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = '4472C4'


# ============================================================
# 6. Stock_Status_Analysis  (AlKhunaizan 전용)
# ============================================================
def build_stock_analysis(wb, df):
    sn = 'Stock_Status_Analysis'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    dl     = df[df['Scraped_At'] == latest]

    ws.merge_cells('A1:F1')
    ws['A1'] = '📦 Stock Status Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Total: {len(dl)} SKUs'
    ws['A2'].font = SUBTITLE_FONT

    total    = len(dl)
    in_stock = len(dl[dl['Stock_Status'] == 'In Stock'])
    out_stock = total - in_stock

    # ── Overall ──
    r = 4
    ws[f'A{r}'] = '📌 Overall Stock Summary'
    ws[f'A{r}'].font = SECTION_FONT
    r += 2

    hdr = ['Status', 'Count', 'Rate']
    for i, h in enumerate(hdr):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, 3)
    r += 1

    rows_ovr = [
        ('In Stock',     in_stock,  f'{in_stock/total*100:.1f}%'  if total else '-', '008000'),
        ('Out of Stock', out_stock, f'{out_stock/total*100:.1f}%' if total else '-', 'FF0000'),
        ('Total',        total,     '',                                               '1F4E79'),
    ]
    for label, cnt, rate, color in rows_ovr:
        ws.cell(row=r, column=1, value=label).font = Font(name='Arial', size=10, color=color, bold=True)
        ws.cell(row=r, column=2, value=cnt).font = Font(name='Arial', size=10, color=color, bold=True)
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3, value=rate).font = DATA_FONT
        ws.cell(row=r, column=3).alignment = CENTER
        if label == 'Total':
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = CAT_FILL
        r += 1

    # ── By Brand ──
    r += 2
    ws[f'A{r}'] = '🏷️ Stock Status by Brand'
    ws[f'A{r}'].font = SECTION_FONT
    r += 2

    brand_hdr = ['Brand', 'Total', 'In Stock', 'Out of Stock', 'In Stock %']
    for i, h in enumerate(brand_hdr):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, 5)
    r += 1

    brand_stock = (dl.groupby('Brand')
                   .apply(lambda x: pd.Series({'Total': len(x),
                                               'In_Stock': len(x[x['Stock_Status'] == 'In Stock'])}))
                   .reset_index())
    brand_stock['Out_Stock'] = brand_stock['Total'] - brand_stock['In_Stock']
    brand_stock['Rate']      = brand_stock['In_Stock'] / brand_stock['Total']
    brand_stock = brand_stock.sort_values('Total', ascending=False)

    for _, row_data in brand_stock.iterrows():
        rate = row_data['Rate']
        rate_color = '008000' if rate >= 0.8 else 'FF6600' if rate >= 0.5 else 'FF0000'
        ws.cell(row=r, column=1, value=row_data['Brand']).font = DATA_FONT
        ws.cell(row=r, column=2, value=int(row_data['Total'])).font = NUM_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3, value=int(row_data['In_Stock'])).font = Font(name='Arial', size=10, color='008000')
        ws.cell(row=r, column=3).alignment = CENTER
        oos_val = int(row_data['Out_Stock'])
        ws.cell(row=r, column=4, value=oos_val).font = (Font(name='Arial', size=10, color='FF0000')
                                                         if oos_val > 0 else NUM_FONT)
        ws.cell(row=r, column=4).alignment = CENTER
        rc = ws.cell(row=r, column=5, value=rate)
        rc.number_format = '0%'
        rc.font = Font(name='Arial', size=10, color=rate_color, bold=True)
        rc.alignment = CENTER
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = LIGHT_FILL
        r += 1

    # ── By Category ──
    r += 2
    ws[f'A{r}'] = '📂 Stock Status by Category'
    ws[f'A{r}'].font = SECTION_FONT
    r += 2

    for i, h in enumerate(brand_hdr):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, 5)
    r += 1

    cat_stock = (dl.groupby('Category')
                 .apply(lambda x: pd.Series({'Total': len(x),
                                             'In_Stock': len(x[x['Stock_Status'] == 'In Stock'])}))
                 .reset_index())
    cat_stock['Out_Stock'] = cat_stock['Total'] - cat_stock['In_Stock']
    cat_stock['Rate']      = cat_stock['In_Stock'] / cat_stock['Total']
    cat_stock = cat_stock.sort_values('Total', ascending=False)

    for _, row_data in cat_stock.iterrows():
        rate = row_data['Rate']
        rate_color = '008000' if rate >= 0.8 else 'FF6600' if rate >= 0.5 else 'FF0000'
        ws.cell(row=r, column=1, value=row_data['Category']).font = DATA_FONT
        ws.cell(row=r, column=2, value=int(row_data['Total'])).font = NUM_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3, value=int(row_data['In_Stock'])).font = Font(name='Arial', size=10, color='008000')
        ws.cell(row=r, column=3).alignment = CENTER
        oos_val = int(row_data['Out_Stock'])
        ws.cell(row=r, column=4, value=oos_val).font = (Font(name='Arial', size=10, color='FF0000')
                                                         if oos_val > 0 else NUM_FONT)
        ws.cell(row=r, column=4).alignment = CENTER
        rc = ws.cell(row=r, column=5, value=rate)
        rc.number_format = '0%'
        rc.font = Font(name='Arial', size=10, color=rate_color, bold=True)
        rc.alignment = CENTER
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = LIGHT_FILL
        r += 1

    # ── Stock Change vs Previous ──
    if len(dates) >= 2:
        prev = dates[-2]
        dp   = df[df['Scraped_At'] == prev]
        r += 2
        ws[f'A{r}'] = f'🔄 Stock Changes vs Previous ({pd.Timestamp(prev).strftime("%Y-%m-%d")})'
        ws[f'A{r}'].font = SECTION_FONT
        r += 2

        prev_in  = set(dp[dp['Stock_Status'] == 'In Stock']['SKU'].astype(str))
        curr_in  = set(dl[dl['Stock_Status'] == 'In Stock']['SKU'].astype(str))
        back_in  = curr_in - prev_in
        went_oos = prev_in - curr_in

        ws.cell(row=r, column=1, value='Change')
        ws.cell(row=r, column=2, value='Count')
        style_header_row(ws, r, 2)
        r += 1
        for label, cnt, color in [('Back In Stock', len(back_in), '008000'),
                                   ('Went Out of Stock', len(went_oos), 'FF0000')]:
            ws.cell(row=r, column=1, value=label).font = DATA_FONT
            ws.cell(row=r, column=2, value=cnt).font = Font(name='Arial', size=10, color=color, bold=True)
            ws.cell(row=r, column=2).alignment = CENTER
            r += 1

    auto_width(ws)
    ws.sheet_properties.tabColor = 'FF9900'


# ============================================================
# 7. Gree_vs_LG_Analysis  (AlKhunaizan 전용)
# ============================================================
def build_gree_vs_lg(wb, df):
    sn = 'Gree_vs_LG_Analysis'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    dl     = df[df['Scraped_At'] == latest]
    gree_df = dl[dl['Brand'] == 'GREE']
    lg_df   = dl[dl['Brand'] == 'LG']

    ws.merge_cells('A1:L1')
    ws['A1'] = '⚔️ GREE vs LG Competitive Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = (f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | '
                f'GREE: {len(gree_df)} SKUs | LG: {len(lg_df)} SKUs')
    ws['A2'].font = SUBTITLE_FONT

    GREE_FILL = PatternFill('solid', fgColor='E2EFDA')
    LG_FILL   = PatternFill('solid', fgColor='DDEEFF')

    # ── Overall Metrics ──
    r = 4
    ws[f'A{r}'] = '📊 Overall Comparison'
    ws[f'A{r}'].font = SECTION_FONT
    r += 2

    hdr = ['Metric', 'GREE', 'LG', 'Difference', 'Winner']
    for i, h in enumerate(hdr):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, 5)
    r += 1

    def write_metric(label, gv, lv, lower_better=True):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        ws.cell(row=r, column=2, value=gv).font = DATA_FONT
        ws.cell(row=r, column=3, value=lv).font = DATA_FONT
        try:
            gn = float(str(gv).replace('%', '').replace(',', ''))
            ln = float(str(lv).replace('%', '').replace(',', ''))
            diff = gn - ln
            ws.cell(row=r, column=4, value=f'{diff:+.1f}').font = DATA_FONT
            winner = ('GREE' if (gn < ln if lower_better else gn > ln)
                      else 'LG' if (ln < gn if lower_better else ln > gn)
                      else 'Tie')
            wc = ws.cell(row=r, column=5, value=winner)
            wc.alignment = CENTER
            if winner == 'GREE':
                wc.font = Font(name='Arial', size=10, color='008000', bold=True)
                wc.fill = GREE_FILL
            elif winner == 'LG':
                wc.font = Font(name='Arial', size=10, color='0070C0', bold=True)
                wc.fill = LG_FILL
            else:
                wc.font = BOLD_FONT
        except Exception:
            ws.cell(row=r, column=4, value='-').font = DATA_FONT
            ws.cell(row=r, column=5, value='-').font = DATA_FONT
        r += 1

    if len(gree_df) > 0 and len(lg_df) > 0:
        write_metric('Total SKUs',             len(gree_df),                    len(lg_df),                    lower_better=False)
        write_metric('Avg Orig Price (SAR)',   f"{gree_df['Original Price (SAR)'].mean():.0f}",
                                               f"{lg_df['Original Price (SAR)'].mean():.0f}",   lower_better=True)
        write_metric('Avg Promo Price (SAR)',  f"{gree_df['Promotion Price (SAR)'].mean():.0f}",
                                               f"{lg_df['Promotion Price (SAR)'].mean():.0f}",  lower_better=True)
        write_metric('Avg Discount (%)',       f"{gree_df['Discount_Numeric'].mean():.1f}",
                                               f"{lg_df['Discount_Numeric'].mean():.1f}",       lower_better=False)
        gree_in = len(gree_df[gree_df['Stock_Status'] == 'In Stock'])
        lg_in   = len(lg_df[lg_df['Stock_Status'] == 'In Stock'])
        write_metric('In Stock Count',         gree_in,                          lg_in,                         lower_better=False)
        write_metric('In Stock Rate (%)',      f"{gree_in/len(gree_df)*100:.1f}",
                                               f"{lg_in/len(lg_df)*100:.1f}",                   lower_better=False)
    else:
        ws.cell(row=r, column=1, value='⚠️ GREE 또는 LG 데이터가 없습니다.').font = SUBTITLE_FONT
        r += 1

    # ── Category × Ton 세분화 비교 ──
    r += 2
    ws[f'A{r}'] = '📦 SKU Count & Price by Category & Ton'
    ws[f'A{r}'].font = SECTION_FONT
    r += 2

    cat_hdr = ['Category', 'Compressor', 'Type', 'Ton',
               'GREE SKUs', 'LG SKUs',
               'GREE Avg (SAR)', 'LG Avg (SAR)',
               'LG vs GREE Gap (SAR)', 'Gap %']
    for i, h in enumerate(cat_hdr):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(cat_hdr))
    r += 1

    cat_order = get_cat_order(dl)

    for cat in cat_order:
        dc   = dl[dl['Category'] == cat]
        gdc  = dc[dc['Brand'] == 'GREE']
        ldc  = dc[dc['Brand'] == 'LG']
        if len(gdc) == 0 and len(ldc) == 0:
            continue

        # 카테고리 소계
        ws.cell(row=r, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=r, column=5, value=len(gdc)).font = BOLD_FONT
        ws.cell(row=r, column=5).alignment = CENTER
        ws.cell(row=r, column=6, value=len(ldc)).font = BOLD_FONT
        ws.cell(row=r, column=6).alignment = CENTER
        if len(gdc) > 0:
            ws.cell(row=r, column=7, value=round(gdc['Effective_Price'].mean())).font = BOLD_FONT
            ws.cell(row=r, column=7).number_format = '#,##0'
        if len(ldc) > 0:
            ws.cell(row=r, column=8, value=round(ldc['Effective_Price'].mean())).font = BOLD_FONT
            ws.cell(row=r, column=8).number_format = '#,##0'
        if len(gdc) > 0 and len(ldc) > 0:
            gap     = round(ldc['Effective_Price'].mean() - gdc['Effective_Price'].mean())
            gap_pct = gap / gdc['Effective_Price'].mean() if gdc['Effective_Price'].mean() > 0 else 0
            cg = ws.cell(row=r, column=9, value=gap)
            cg.number_format = '#,##0'
            cg.font = UP_FONT if gap > 0 else DOWN_FONT if gap < 0 else BOLD_FONT
            cp = ws.cell(row=r, column=10, value=gap_pct)
            cp.number_format = '0%'
            cp.font = UP_FONT if gap > 0 else DOWN_FONT if gap < 0 else BOLD_FONT
        for c in range(1, len(cat_hdr)+1):
            ws.cell(row=r, column=c).fill = CAT_FILL
        r += 1

        groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
        for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2] if pd.notna(x[0][2]) else 0)):
            g_grp = grp[grp['Brand'] == 'GREE']
            l_grp = grp[grp['Brand'] == 'LG']
            if len(g_grp) == 0 and len(l_grp) == 0:
                continue

            ws.cell(row=r, column=2, value=comp).font = DATA_FONT
            ws.cell(row=r, column=3, value=hc).font = DATA_FONT
            ws.cell(row=r, column=4, value=f'{ton}T' if pd.notna(ton) else 'N/A').font = DATA_FONT
            ws.cell(row=r, column=4).alignment = CENTER
            ws.cell(row=r, column=5, value=len(g_grp) if len(g_grp) > 0 else '-').font = NUM_FONT
            ws.cell(row=r, column=5).alignment = CENTER
            ws.cell(row=r, column=6, value=len(l_grp) if len(l_grp) > 0 else '-').font = NUM_FONT
            ws.cell(row=r, column=6).alignment = CENTER

            g_avg = round(g_grp['Effective_Price'].mean()) if len(g_grp) > 0 else None
            l_avg = round(l_grp['Effective_Price'].mean()) if len(l_grp) > 0 else None

            if g_avg is not None:
                ws.cell(row=r, column=7, value=g_avg).font = NUM_FONT
                ws.cell(row=r, column=7).number_format = '#,##0'
            else:
                ws.cell(row=r, column=7, value='-').font = NUM_FONT
            ws.cell(row=r, column=7).alignment = CENTER

            if l_avg is not None:
                ws.cell(row=r, column=8, value=l_avg).font = NUM_FONT
                ws.cell(row=r, column=8).number_format = '#,##0'
            else:
                ws.cell(row=r, column=8, value='-').font = NUM_FONT
            ws.cell(row=r, column=8).alignment = CENTER

            if g_avg is not None and l_avg is not None:
                gap     = l_avg - g_avg
                gap_pct = gap / g_avg if g_avg > 0 else 0
                cg = ws.cell(row=r, column=9, value=gap)
                cg.number_format = '#,##0'
                cg.font = UP_FONT if gap > 0 else DOWN_FONT if gap < 0 else NUM_FONT
                cp = ws.cell(row=r, column=10, value=gap_pct)
                cp.number_format = '0%'
                cp.font = UP_FONT if gap > 0 else DOWN_FONT if gap < 0 else NUM_FONT
            else:
                ws.cell(row=r, column=9,  value='-').font = NUM_FONT
                ws.cell(row=r, column=10, value='-').font = NUM_FONT
            for ci in [9, 10]:
                ws.cell(row=r, column=ci).alignment = CENTER

            if r % 2 == 0:
                for c in range(1, len(cat_hdr)+1):
                    ws.cell(row=r, column=c).fill = LIGHT_FILL
            r += 1

    auto_width(ws)
    ws.sheet_properties.tabColor = '7030A0'


# ============================================================
# MAIN
# ============================================================
def main():
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    print("=" * 60)
    print("📊 Al-Khunaizan AC Dashboard Builder v2")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ 파일 없음: {INPUT_FILE}")
        sys.exit(1)

    print(f"📂 파일 로드: {INPUT_FILE}")
    df = load_data(INPUT_FILE)

    dates = sorted(df['Scraped_At'].unique())
    print(f"📅 스크래핑 일자: {len(dates)}회")
    for d in dates:
        print(f"   - {pd.Timestamp(d).strftime('%Y-%m-%d')}: {len(df[df['Scraped_At']==d])}개 SKU")

    wb = load_workbook(INPUT_FILE)

    print("\n🔧 대시보드 시트 생성 중...")
    print("  1/7 Dashboard_Summary...")
    build_dashboard_summary(wb, df)
    print("  2/7 Price_Change_Alert...")
    build_price_change_alert(wb, df)
    print("  3/7 New_Discontinued_SKUs...")
    build_new_discontinued(wb, df)
    print("  4/7 Brand_Avg_Price_Comparison...")
    build_brand_price_comparison(wb, df)
    print("  5/7 Full_Price_Tracking...")
    build_full_tracking(wb, df)
    print("  6/7 Stock_Status_Analysis...")
    build_stock_analysis(wb, df)
    print("  7/7 Gree_vs_LG_Analysis...")
    build_gree_vs_lg(wb, df)

    # 시트 순서 정렬
    desired_order = [
        'Dashboard_Summary', 'Price_Change_Alert', 'New_Discontinued_SKUs',
        'Brand_Avg_Price_Comparison', 'Full_Price_Tracking',
        'Stock_Status_Analysis', 'Gree_vs_LG_Analysis', 'Products_DB'
    ]
    current = wb.sheetnames
    for i, name in enumerate(desired_order):
        if name in current:
            wb.move_sheet(name, offset=i - current.index(name))
            current = wb.sheetnames

    # 이전 버전 시트 정리
    for old in ['Executive_Summary', 'Gree_vs_LG', 'Price_Analysis', 'Product_Catalog']:
        if old in wb.sheetnames:
            del wb[old]

    wb.save(INPUT_FILE)
    print(f"\n✅ 대시보드 업데이트 완료!")
    print(f"📁 파일: {INPUT_FILE}")
    print("=" * 60)

    if sys.stdin.isatty():
        input("\n 엔터를 누르면 종료...")


if __name__ == "__main__":
    main()
