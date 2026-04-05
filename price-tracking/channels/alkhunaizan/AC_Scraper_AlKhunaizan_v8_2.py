#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Al-Khunaizan 스크래퍼 v8.2
─────────────────────────────────────────────────────────────────────
[수집 방식]
  0단계) 사이트 필터 URL로 Wifi Available SKU 세트 사전 수집
         → /en/air-conditioner-c-6?pageSize=300&wifi_controle=143
         → 상품명 파싱보다 훨씬 정확 (사이트 내부 DB 기준)

  1단계) 카테고리 목록 페이지(pageSize=300)
         → 가격·재고·Only Pay Price·Energy Grade(배지)·브랜드 수집

  2단계) 각 상품 상세 페이지 (5개 병렬)
         → 스펙 테이블에서 정확한 값 수집
           · Nominal Capacity (1 Ton / 1.5 Ton …)
           · Energy efficiency class (A~G)
           · Compressor type (On/Off / Inverter / Dual Inverter)
           · COOL & HOT OR COOL ONLY → Type (Hot & Cold / Cold Only)
           · Color (White / Black …)

  Wifi  → 0단계 SKU 세트 매핑 (Available / Not Available)
  Color → 스펙 테이블 우선, 없으면 상품명 파싱

[포함 카테고리]  Window / Split / Concealed / Free Stand / Cassette AC
[제외 카테고리]  Portable AC, Desert Cooler, Air Curtains
─────────────────────────────────────────────────────────────────────
"""

import asyncio
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

DETAIL_CONCURRENCY = 5

# Wi-Fi Available 필터 URL (사이트 내부 필터 ID = 143)
WIFI_FILTER_URL = "/en/air-conditioner-c-6?pageSize=300&wifi_controle=143"


class AlKhunaizanScraper:
    def __init__(self):
        self.products = []
        self.seen_skus = {}
        self.wifi_sku_set = set()   # 0단계에서 수집한 Wi-Fi Available SKU 세트

        self.categories = [
            {"name": "Window AC",     "url": "/en/window-ac-c-34?pageSize=300"},
            {"name": "Split AC",      "url": "/en/split-ac-c-56?pageSize=300"},
            {"name": "Concealed AC",  "url": "/en/concealed-ac-c-461?pageSize=300"},
            {"name": "Free Stand AC", "url": "/en/free-stand-ac-c-57?pageSize=300"},
            {"name": "Cassette AC",   "url": "/en/cassette-ac-c-58?pageSize=300"},
        ]

        # 스펙 테이블 키 매핑 (소문자 비교)
        self.SPEC_KEY_MAP = {
            'nominal capacity':        'nominalCap',
            'energy efficiency class': 'energyGrade',
            'compressor type':         'compressor',
            'cool& hot or cool only':  'coolType',
            'color':                   'color',
        }

    # ──────────────────────────────────────────
    # 파싱 헬퍼
    # ──────────────────────────────────────────
    def _parse_ac_type_from_name(self, name):
        u = name.upper()
        is_inv = 'INVERTER' in u
        if 'HOT' in u and 'COLD' in u:
            base = 'Hot & Cold'
        elif 'COLD ONLY' in u or ('COLD' in u and 'HOT' not in u):
            base = 'Cold Only'
        elif 'HOT ONLY' in u:
            base = 'Hot Only'
        else:
            base = 'N/A'
        if is_inv and base != 'N/A':
            return f"{base}, Inverter"
        elif is_inv:
            return 'Inverter'
        return base

    def _parse_ac_type_from_spec(self, cool_type_val):
        mapping = {
            'Hot / Cool': 'Hot & Cold',
            'Hot/Cool':   'Hot & Cold',
            'Cool Only':  'Cold Only',
            'Cold Only':  'Cold Only',
            'Hot Only':   'Hot Only',
        }
        return mapping.get(cool_type_val.strip(), cool_type_val.strip())

    def _parse_capacity_from_name(self, name):
        m = re.search(r'(\d{4,6})\s*(?:BTU|Btu|Units|Unit)', name, re.IGNORECASE)
        if m:
            return m.group(1)
        m = re.search(r'\b(\d{5})\b', name)
        if m:
            return m.group(1)
        return ''

    def _parse_color_from_name(self, name):
        u = name.upper()
        if 'BLACK' in u:
            return 'Black'
        if 'GREY' in u or 'GRAY' in u:
            return 'Grey'
        if 'BROWN' in u:
            return 'Brown'
        return 'White'

    def _parse_compressor_from_name(self, name):
        u = name.upper()
        if 'DUAL INVERTER' in u or 'INVERTER' in u:
            return 'Inverter'
        return 'On/Off'

    def _btu_to_nominal(self, capacity_str):
        if not capacity_str:
            return ''
        m = re.search(r'(\d{4,6})', capacity_str)
        if not m:
            return ''
        btu = int(m.group(1))
        # 실제 AlKhunaizan 제품 BTU 기준으로 보정된 범위
        # 18000 BTU = 1.5 Ton (이전 코드 오류: 18000→'2 Ton' 잘못 매핑)
        # 3.5 Ton 범위 추가 (42000 BTU), 4 Ton / 5 Ton 범위 교정
        ton_map = [
            (7000,  11999,  '0.75 Ton'),
            (12000, 15999,  '1 Ton'),
            (16000, 20999,  '1.5 Ton'),  # 17600, 18000, 18400 BTU → 1.5 Ton
            (21000, 24999,  '2 Ton'),    # 24000 BTU → 2 Ton
            (25000, 32999,  '2.5 Ton'),  # 25600~31400 BTU → 2.5 Ton (AlKhunaizan 실측값)
            (33000, 38999,  '3 Ton'),    # 36000 BTU → 3 Ton
            (39000, 44999,  '3.5 Ton'),  # 42000 BTU → 3.5 Ton (신규 추가)
            (45000, 55999,  '4 Ton'),    # 48000 BTU → 4 Ton
            (56000, 75000,  '5 Ton'),    # 60000 BTU → 5 Ton
        ]
        for lo, hi, label in ton_map:
            if lo <= btu <= hi:
                return label
        return ''

    # ──────────────────────────────────────────
    # 0단계: Wi-Fi Available SKU 사전 수집
    # ──────────────────────────────────────────
    async def _collect_wifi_skus(self, page, base_url):
        """사이트 Wi-Fi 필터 URL → Available SKU 세트 수집"""
        print("\n" + "="*70)
        print("📶 0단계: Wi-Fi Available SKU 수집")
        print(f"   URL: {base_url}{WIFI_FILTER_URL}")
        print("="*70)

        try:
            await page.goto(f"{base_url}{WIFI_FILTER_URL}",
                            wait_until="networkidle", timeout=60000)
            await asyncio.sleep(3)
            await self._scroll_to_load_all(page)

            skus = await page.evaluate('''() => {
                const cards = document.querySelectorAll('.rounded-md.border.bg-white.p-3');
                const skus = [];
                cards.forEach(card => {
                    const link = card.querySelector('a[href*="/en/product/"]');
                    if (link) {
                        const href = link.getAttribute('href');
                        const m = href.match(/p-([^\\/]+)$/);
                        if (m) skus.push(m[1]);
                    }
                });
                return skus;
            }''')

            self.wifi_sku_set = set(skus)
            print(f"   ✅ Wi-Fi Available SKU 수집 완료: {len(self.wifi_sku_set)}개")

        except Exception as e:
            print(f"   ⚠️ Wi-Fi SKU 수집 실패: {e}")
            print("   → 상품명 파싱으로 대체합니다")

    # ──────────────────────────────────────────
    # 스크래핑 메인
    # ──────────────────────────────────────────
    async def scrape_alkhunaizan(self):
        print("\n" + "="*70)
        print("🌐 AL-KHUNAIZAN 스크래퍼 v8.2")
        print("   0단계: Wi-Fi 필터 SKU 사전 수집 (사이트 DB 기준)")
        print("   1단계: 카테고리 목록 수집")
        print("   2단계: 상세 페이지 스펙 수집 (병렬)")
        print("="*70)

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                channel="chrome",
                headless=True,
                args=[
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-gpu',
                ]
            )
            context = await browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                viewport={'width': 1920, 'height': 1080}
            )

            base_url = "https://www.alkhunaizan.sa"
            list_page = await context.new_page()

            # ── 0단계: Wi-Fi SKU 세트 수집 ────────────────────────────
            await self._collect_wifi_skus(list_page, base_url)

            # ── 1단계: 카테고리 목록 수집 ─────────────────────────────
            print("\n" + "="*70)
            print("📋 1단계: 카테고리 목록 수집")
            print("="*70)

            for idx, cat in enumerate(self.categories, 1):
                print(f"\n{'='*50}")
                print(f"📂 [{idx}/{len(self.categories)}] {cat['name']}")
                print(f"{'='*50}")
                cat_url = f"{base_url}{cat['url']}"
                print(f"   URL: {cat_url}")

                max_retries = 3
                for attempt in range(1, max_retries + 1):
                    try:
                        await list_page.goto(cat_url, wait_until="networkidle", timeout=60000)
                        await asyncio.sleep(3)

                        try:
                            body_text = await list_page.evaluate('document.body.innerText')
                            rm = re.search(r'Results\s+(\d+)\s+of\s+(\d+)\s+products', body_text)
                            if rm:
                                print(f"   📊 {rm.group(0)}")
                        except:
                            pass

                        await self._scroll_to_load_all(list_page)
                        count_before = len(self.seen_skus)
                        await self._extract_list_products(list_page, cat['name'])
                        count_after = len(self.seen_skus)
                        new_count = count_after - count_before
                        print(f"   ✅ {new_count}개 추가 (누적 {count_after}개)")

                        # 상품이 0개면 재시도 (사이트 로딩 실패 가능성)
                        if new_count == 0 and attempt < max_retries:
                            print(f"   ⚠️ 0개 수집 — 재시도 {attempt}/{max_retries}")
                            await asyncio.sleep(5)
                            continue
                        break  # 성공 시 루프 탈출

                    except Exception as e:
                        print(f"   ⚠️ 카테고리 수집 실패 (시도 {attempt}/{max_retries}): {e}")
                        if attempt < max_retries:
                            await asyncio.sleep(5)
                            continue
                        break

            await list_page.close()

            # ── 2단계: 상세 페이지 병렬 수집 ──────────────────────────
            all_skus = list(self.seen_skus.keys())
            total = len(all_skus)
            print(f"\n{'='*70}")
            print(f"🔍 2단계: 상세 페이지 스펙 수집 — 총 {total}개")
            print(f"   동시 요청: {DETAIL_CONCURRENCY}개")
            print(f"{'='*70}")

            semaphore   = asyncio.Semaphore(DETAIL_CONCURRENCY)
            detail_pages = [await context.new_page() for _ in range(DETAIL_CONCURRENCY)]
            page_queue  = asyncio.Queue()
            for dp in detail_pages:
                await page_queue.put(dp)

            done_count = [0]

            async def fetch_spec(sku):
                product = self.seen_skus[sku]
                url = product['url']
                page = await page_queue.get()
                try:
                    async with semaphore:
                        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                        await asyncio.sleep(1)
                        spec = await self._extract_spec_table(page)
                        self._apply_spec(sku, spec)
                except Exception:
                    pass
                finally:
                    done_count[0] += 1
                    if done_count[0] % 20 == 0 or done_count[0] == total:
                        print(f"   ⏳ {done_count[0]}/{total} 완료...")
                    await page_queue.put(page)

            await asyncio.gather(*[fetch_spec(sku) for sku in all_skus])

            for dp in detail_pages:
                await dp.close()
            await browser.close()
            print("\n✓ 브라우저 종료")

        self.products = list(self.seen_skus.values())
        print(f"\n📊 총 수집 상품: {len(self.products)}개")

        # Wi-Fi 통계 출력
        wifi_available = sum(1 for p in self.products if p.get('wifi') == 'Available')
        wifi_not = sum(1 for p in self.products if p.get('wifi') == 'Not Available')
        print(f"📶 Wifi Available: {wifi_available}개 / Not Available: {wifi_not}개")

        return self.products

    # ──────────────────────────────────────────
    # 스크롤 로딩
    # ──────────────────────────────────────────
    async def _scroll_to_load_all(self, page):
        last_height = 0
        same_count  = 0
        scroll_count = 0
        while same_count < 3 and scroll_count < 50:
            await page.evaluate('window.scrollBy(0, 1200)')
            await asyncio.sleep(0.4)
            current_height = await page.evaluate('document.documentElement.scrollHeight')
            if current_height == last_height:
                same_count += 1
            else:
                same_count = 0
            last_height = current_height
            scroll_count += 1
        await page.evaluate('window.scrollTo(0, 0)')
        await asyncio.sleep(0.3)

    # ──────────────────────────────────────────
    # 1단계: 카테고리 목록 데이터 추출
    # ──────────────────────────────────────────
    async def _extract_list_products(self, page, category_name):
        products_data = await page.evaluate('''() => {
            const cards = document.querySelectorAll('.rounded-md.border.bg-white.p-3');
            const products = [];
            const seenUrls = new Set();

            const BRANDS = [
                'WHITE WESTINGHOUSE', 'SMART ELECTRIC', 'SUPER GENERAL', 'STAR WAY',
                'NEW HOUSE', 'GREE', 'LG', 'TCL', 'HAIER', 'BASIC', 'HISENSE',
                'BANCOOL', 'CRAFFT', 'UNIX', 'ALJAZERAH', 'SHARP', 'KION', 'FUJI',
                'BREEZAIR', 'SYMPHONY', 'BENCHMARK', 'COOLINE', 'MAAS', 'HAAS',
                'GENERIC', 'PLATINUM'
            ];
            // Arabic brand name → English mapping
            const AR_BRANDS = {
                'تي سي ال': 'TCL',
                'ال جي': 'LG',
                'هايسنس': 'HISENSE',
                'هاير': 'HAIER',
                'جري': 'GREE',
                'شارب': 'SHARP',
                'سيمفوني': 'SYMPHONY',
                'بنكول': 'BANCOOL',
                'كرافت': 'CRAFFT',
                'يونكس': 'UNIX',
                'كيون': 'KION',
                'فوجي': 'FUJI',
                'بلاتينيوم': 'PLATINUM',
                'بيسك': 'BASIC',
                'ستار واي': 'STAR WAY',
                'سوبر جنرال': 'SUPER GENERAL',
                'وستنجهاوس': 'WHITE WESTINGHOUSE',
            };

            cards.forEach(card => {
                const productLink = card.querySelector('a[href*="/en/product/"]');
                if (!productLink) return;
                const href = productLink.getAttribute('href');
                if (!href || seenUrls.has(href)) return;
                seenUrls.add(href);

                const nameEl = card.querySelector('a.line-clamp-3');
                const name = nameEl ? nameEl.textContent.trim() : '';
                if (!name || name.length < 5) return;

                // 브랜드 (영문 먼저, 아랍어 fallback)
                let brand = 'N/A';
                const nameUpper = name.toUpperCase();
                for (const b of BRANDS) {
                    if (nameUpper.includes(b)) { brand = b; break; }
                }
                if (brand === 'N/A') {
                    for (const [arName, enName] of Object.entries(AR_BRANDS)) {
                        if (name.includes(arName)) { brand = enName; break; }
                    }
                }

                // 현재 가격
                let price = 0;
                const priceEl = card.querySelector('span.flex.items-center.gap-1.flex-wrap.text-xl');
                if (priceEl) {
                    const m = priceEl.textContent.match(/[\\d,]+/);
                    if (m) price = parseInt(m[0].replace(/,/g, ''));
                }
                if (price === 0) return;

                // 원가
                let beforePrice = 0;
                const beforeEl = card.querySelector('span.flex.items-center.gap-1.visible.text-xs');
                if (beforeEl) {
                    const m = beforeEl.textContent.match(/Before\\s*([\\d,]+)/);
                    if (m) beforePrice = parseInt(m[1].replace(/,/g, ''));
                }

                // SKU
                const skuMatch = href.match(/p-([^\\/]+)$/);
                const sku = skuMatch ? skuMatch[1] : '';

                // 재고
                const cardText = card.textContent;
                let stock = 'In Stock';
                if (cardText.includes('Last Piece'))          stock = 'Last Piece';
                else if (cardText.includes('Out of Stock'))   stock = 'Out of Stock';
                else if (cardText.includes('Selling Out Fast')) stock = 'Low Stock';

                // Energy Grade (카드 배지 — 우측 하단)
                let energyGrade = '';
                const energyEl = card.querySelector(
                    'div[class*="absolute"][class*="right-0"][class*="bottom-0"]'
                );
                if (energyEl) {
                    const t = energyEl.textContent.trim();
                    if (/^[A-G]$/.test(t)) energyGrade = t;
                }

                // Only Pay Price (Extra Savings 배지 — 중앙 하단)
                let onlyPayPrice = 0;
                const onlyPayEl = card.querySelector(
                    'div[class*="absolute"][class*="left-1\\/2"]'
                );
                if (onlyPayEl) {
                    const m = onlyPayEl.textContent.match(/Only Pay\\s*([\\d,]+)/);
                    if (m) onlyPayPrice = parseInt(m[1].replace(/,/g, ''));
                }

                products.push({
                    name, brand, price, beforePrice, sku, stock,
                    energyGrade, onlyPayPrice,
                    url: 'https://www.alkhunaizan.sa' + href
                });
            });
            return products;
        }''')

        if not products_data:
            return

        for p in products_data:
            sku   = p['sku']
            stock = p['stock']

            # 중복 처리: Last Piece 우선 삭제
            if sku in self.seen_skus:
                existing = self.seen_skus[sku]
                if existing['stock'] == 'Last Piece' and stock != 'Last Piece':
                    pass
                elif stock == 'Last Piece' and existing['stock'] != 'Last Piece':
                    continue
                else:
                    continue

            # 할인율
            discount_rate = 0
            savings       = 0
            if p['beforePrice'] > 0 and p['price'] > 0:
                savings = p['beforePrice'] - p['price']
                if savings > 0:
                    discount_rate = round((savings / p['beforePrice']) * 100, 2)

            # ★ Wifi: 0단계 SKU 세트로 판별 (사이트 DB 기준)
            wifi = 'Available' if sku in self.wifi_sku_set else 'Not Available'

            cap_from_name = self._parse_capacity_from_name(p['name'])

            self.seen_skus[sku] = {
                'category':     category_name,
                'name':         p['name'],
                'brand':        p['brand'],
                'acType':       self._parse_ac_type_from_name(p['name']),   # 2단계에서 정확한 값으로 덮어쓰기
                'capacity':     cap_from_name,
                'compressor':   self._parse_compressor_from_name(p['name']), # 2단계에서 덮어쓰기
                'wifi':         wifi,                                          # ★ 사이트 필터 기준
                'color':        self._parse_color_from_name(p['name']),       # 2단계에서 덮어쓰기
                'nominalCap':   self._btu_to_nominal(cap_from_name),          # 2단계에서 덮어쓰기
                'energyGrade':  p['energyGrade'],                              # 2단계에서 보강
                'price':        p['price'],
                'beforePrice':  p['beforePrice'],
                'discountRate': discount_rate,
                'savings':      savings,
                'onlyPayPrice': p['onlyPayPrice'],
                'sku':          sku,
                'stock':        stock,
                'url':          p['url'],
            }

    # ──────────────────────────────────────────
    # 2단계: 상세 페이지 스펙 테이블 추출
    # ──────────────────────────────────────────
    async def _extract_spec_table(self, page):
        spec = await page.evaluate('''() => {
            const result = {};
            document.querySelectorAll('table tr').forEach(tr => {
                const cells = [...tr.querySelectorAll('td, th')];
                if (cells.length >= 2) {
                    const key = cells[0].textContent.trim();
                    const val = cells[1].textContent.trim();
                    if (key && val) result[key] = val;
                }
            });
            return result;
        }''')
        return spec

    def _apply_spec(self, sku, spec):
        """스펙 테이블 dict → seen_skus[sku] 반영"""
        if not spec or sku not in self.seen_skus:
            return

        product = self.seen_skus[sku]

        for raw_key, raw_val in spec.items():
            key_lower = raw_key.strip().lower().rstrip('"')   # stray " 제거
            val       = ' '.join(raw_val.split())             # 내부 공백 정규화 ('2.5  Ton' → '2.5 Ton')
            if not val:
                continue

            mapped = self.SPEC_KEY_MAP.get(key_lower)

            if mapped == 'nominalCap':
                product['nominalCap'] = val

            elif mapped == 'energyGrade':
                if re.match(r'^[A-G](\+{1,3})?$', val):
                    product['energyGrade'] = val

            elif mapped == 'compressor':
                v = val.lower()
                if 'dual' in v or 'inverter' in v:
                    product['compressor'] = 'Inverter'
                else:
                    product['compressor'] = 'On/Off'

            elif mapped == 'coolType':
                converted = self._parse_ac_type_from_spec(val)
                # Inverter 정보는 compressor 필드에서 이미 관리하므로 Type은 Cold/Hot 만 표기
                product['acType'] = converted

            elif mapped == 'color':
                product['color'] = val

        # Nominal Capacity 보조: 스펙 테이블에도 없으면 BTU→Ton 변환
        if not product.get('nominalCap') and product.get('capacity'):
            product['nominalCap'] = self._btu_to_nominal(product['capacity'])

    # ──────────────────────────────────────────
    # 행 빌더
    # ──────────────────────────────────────────
    def _build_row(self, product, scraped_at):
        """
        컬럼 순서 (A~T, 20열):
        A  Reference
        B  Category
        C  Product Name
        D  Brand
        E  Type  (Hot & Cold / Cold Only …)
        F  Capacity  (BTU/Units — 상품명 기반)
        G  Compressor Type  (스펙 테이블)
        H  Wifi  (사이트 필터 기준 ★)
        I  Color  (스펙 테이블 / 상품명)
        J  Nominal Capacity  (스펙 테이블)
        K  Energy Grade  (스펙 테이블 / 카드 배지)
        L  Promotion Price (SAR)
        M  Original Price (SAR)
        N  Discount Rate (%)
        O  Save amount (SAR)
        P  Only Pay Price (SAR)
        Q  SKU
        R  Stock Status
        S  URL
        T  Scraped_At
        """
        dr  = product.get('discountRate', 0)
        bp  = product.get('beforePrice',  0)
        sv  = product.get('savings',      0)
        opp = product.get('onlyPayPrice', 0)
        return [
            'Al-Khunaizan',
            product.get('category', ''),
            product.get('name', ''),
            product.get('brand', ''),
            product.get('acType', ''),
            product.get('capacity', ''),
            product.get('compressor', ''),
            product.get('wifi', ''),
            product.get('color', ''),
            product.get('nominalCap', ''),
            product.get('energyGrade', ''),
            product.get('price', 0),
            bp  if bp  > 0 else product.get('price', 0),
            f"{dr:.1f}%" if dr > 0 else '0%',
            sv  if sv  > 0 else 0,
            opp if opp > 0 else '',
            product.get('sku', ''),
            product.get('stock', ''),
            product.get('url', ''),
            scraped_at,
        ]

    # ──────────────────────────────────────────
    # Master DB 업데이트
    # ──────────────────────────────────────────
    def update_master_db(self):
        print("\n" + "="*70)
        print("📂 MASTER DB 업데이트")
        print("="*70)

        if not self.products:
            print("❌ 업데이트할 상품이 없습니다")
            return False

        script_dir  = os.path.dirname(os.path.abspath(__file__))
        master_path = os.path.join(script_dir, "AlKhunaizan_AC_Prices Tracking_Master.xlsx")

        if not os.path.exists(master_path):
            print(f"❌ Master 파일 없음: {master_path}")
            return False

        print(f"   파일: {master_path}")

        try:
            wb = load_workbook(master_path)
        except Exception as e:
            print(f"❌ Master 파일 열기 실패: {e}")
            return False

        if 'Products_DB' not in wb.sheetnames:
            print("❌ 'Products_DB' 시트 없음")
            return False

        ws = wb['Products_DB']

        HEADERS = [
            'Reference', 'Category', 'Product Name', 'Brand', 'Type', 'Capacity',
            'Compressor Type', 'Wifi', 'Color', 'Nominal Capacity', 'Energy Grade',
            'Promotion Price (SAR)', 'Original Price (SAR)', 'Discount Rate (%)',
            'Save amount (SAR)', 'Only Pay Price (SAR)',
            'SKU', 'Stock Status', 'URL', 'Scraped_At'
        ]
        header_row   = [cell.value for cell in ws[1]]
        current_cols = len([h for h in header_row if h is not None])

        if current_cols < len(HEADERS):
            print(f"   ⚠️  헤더 확장: {current_cols}열 → {len(HEADERS)}열")
            hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")
            for ci, hdr in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=ci, value=hdr)
                cell.fill  = hfill
                cell.font  = hfont
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 실제 데이터가 있는 마지막 행 탐색
        # (서식만 있는 빈 행을 건너뛰기 위해 max_row 대신 직접 탐색)
        actual_last_row = 1
        for r in range(ws.max_row, 1, -1):
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, 6)):
                actual_last_row = r
                break

        existing_data_rows = actual_last_row - 1  # 헤더 제외
        print(f"   기존 데이터: {existing_data_rows}행 (헤더 제외, 실제 데이터 기준)")

        scraped_at = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        added = 0
        next_row = actual_last_row + 1
        for product in self.products:
            row_data = self._build_row(product, scraped_at)
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col_idx, value=value)
            ws.cell(row=next_row, column=20).number_format = 'YYYY-MM-DD'
            next_row += 1
            added += 1

        try:
            wb.save(master_path)
            print(f"   ✅ {added}개 행 추가 완료")
            print(f"   📊 총 누적 데이터: {existing_data_rows + added}행")
            print(f"   📅 스크래핑 날짜: {scraped_at.strftime('%Y-%m-%d')}")

            from collections import Counter
            print("\n📊 카테고리별:")
            for cat, cnt in Counter(p['category'] for p in self.products).most_common():
                print(f"   • {cat}: {cnt}개")
            print("\n📊 브랜드별 Top10:")
            for b, cnt in Counter(p['brand'] for p in self.products if p['brand'] != 'N/A').most_common(10):
                print(f"   • {b}: {cnt}개")
            print("\n📊 Nominal Capacity별:")
            for nc, cnt in Counter(p['nominalCap'] for p in self.products if p['nominalCap']).most_common():
                print(f"   • {nc}: {cnt}개")
            print("\n📊 Energy Grade별:")
            for g, cnt in Counter(p['energyGrade'] for p in self.products if p['energyGrade']).most_common():
                print(f"   • {g}: {cnt}개")
            print("\n📊 Compressor Type별:")
            for c, cnt in Counter(p['compressor'] for p in self.products if p['compressor']).most_common():
                print(f"   • {c}: {cnt}개")
            print("\n📊 Wifi별:")
            for w, cnt in Counter(p['wifi'] for p in self.products if p['wifi']).most_common():
                print(f"   • {w}: {cnt}개")
            print(f"\n📊 Only Pay (Extra Savings): "
                  f"{sum(1 for p in self.products if p.get('onlyPayPrice', 0) > 0)}개")

            return True
        except Exception as e:
            print(f"❌ Master 파일 저장 실패: {e}")
            print("   (파일이 Excel에서 열려 있으면 닫고 다시 시도하세요)")
            return False


# ──────────────────────────────────────────
# 실행
# ──────────────────────────────────────────
async def main():
    scraper = AlKhunaizanScraper()

    print("\n" + "="*70)
    print("🔥 AL-KHUNAIZAN AC 스크래퍼 v8.2")
    print("   Window / Split / Concealed / Free Stand / Cassette AC")
    print("   Wifi: 사이트 필터 기반 (DB 직접 매핑)")
    print("   Energy Grade · Nominal Capacity · Compressor: 상세 페이지 스펙 테이블")
    print("="*70)

    await scraper.scrape_alkhunaizan()
    scraper.update_master_db()


if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n⛔ 사용자 중단")
    except Exception as e:
        print(f"\n❌ 에러: {e}")
        import traceback
        traceback.print_exc()

    print("\n엔터를 눌러 종료...")
    input()
