#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Almanea AC Scraper v3.4
- Al Manea API 기반 AC 제품 정보 수집
- 아랍어 → 영어 번역 강화 (Country / Offer_Detail 완전 처리)
- Almanea_AC_Price_Tracking_Master.xlsx > Products_DB 시트에 날짜별 누적 저장
- 컬럼 구조 고정: 중복/혼재 없이 단일 표준 스키마 유지
- W열(Final_Promo_Price), X열(AlAhli_Price) 파이썬으로 자동 계산
- Capacity_Ton: BTU 범위 테이블 기반 표준 등급 매핑 (÷12000 단순 나눗셈 제거)
"""

import requests
import time
import re
import os
import sys
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[ERROR] 필수 패키지 없음: {e}")
    print("  >> py -m pip install requests pandas openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "Almanea_AC_Price_Tracking_Master.xlsx")
DB_SHEET    = "Products_DB"

# ─────────────────────────────────────────────
# 표준 컬럼 순서 (Products_DB 고정 스키마)
# ─────────────────────────────────────────────
COLUMNS = [
    "Scraped_At",              # 수집 날짜 (YYYY-MM-DD)
    "SKU",                     # 제품 고유 ID
    "Brand",                   # 브랜드
    "Model",                   # 모델번호
    "Product_Name",            # 제품명 (영문)
    "Category",                # 카테고리 (Split AC / Window AC / Floor Standing / ...)
    "Function",                # Cold Only / Cold & Hot
    "Compressor_Type",         # Inverter / Fixed Speed / Rotary / Dual Inverter
    "Capacity_Ton",            # 냉방 용량 (Ton 단위 숫자, e.g. 1.5)
    "BTU",                     # 냉방 용량 (BTU 숫자, e.g. 18000)
    "Energy_Rating",           # 에너지 등급 (A / B / C / ...)
    "Color",                   # 색상
    "Country",                 # 제조국
    "Warranty_Yr",             # 일반 보증 (숫자, 년)
    "Compressor_Warranty_Yr",  # 컴프레서 보증 (숫자, 년)
    "Original_Price",          # 정가 (SAR, 정수)
    "Promo_Price",             # 프로모션 가격 (SAR, 정수)
    "Discount_Pct",            # 할인율 (소수, e.g. 0.20)
    "Has_Offer",               # 오퍼 여부 (Yes/No)
    "Offer_Detail",            # 오퍼 상세 (영문)
    "Free_Gift",               # 무료 증정품 (영문)
    "Stock",                   # 재고 수량
    "Final_Promo_Price",       # 최종 프로모션 적용가 (Cashback 할인 반영, SAR 정수)
    "AlAhli_Price",            # Al Ahli 은행 카드 추가 10% 할인가 (SAR 정수)
    "URL_Key",                 # 상품 페이지 URL slug (almanea.sa/en/{url_key})
]

# ─────────────────────────────────────────────
# 아랍어 → 영어 번역 테이블  (v3.3 강화)
# ─────────────────────────────────────────────
AR_TO_EN = {
    # ── 컴프레서 타입
    "إنفرتر مزدوج": "Dual Inverter",   # 반드시 단순 Inverter보다 먼저
    "انفرتر مزدوج": "Dual Inverter",
    "إنفرتر": "Inverter",
    "انفرتر": "Inverter",
    "ثابت السرعة": "Fixed Speed",
    "ثابت": "Fixed Speed",
    "دوار": "Rotary",
    "روتاري": "Rotary",
    # ── Cold/Hot 기능 (긴 구문 먼저)
    "تبريد وتدفئة": "Cold & Hot",
    "تدفئة وتبريد": "Cold & Hot",
    "تبريد فقط": "Cold Only",
    "تبريد": "Cold Only",
    "تدفئة": "Heating",
    # ── 에너지 등급
    "ممتاز جداً": "Excellent",
    "ممتاز": "Excellent",
    "جيد جداً": "Very Good",
    "جيد": "Good",
    "مقبول": "Acceptable",
    # ── 색상
    "أبيض": "White",
    "فضي": "Silver",
    "ذهبي": "Gold",
    "أسود": "Black",
    "رمادي": "Grey",
    "بيج": "Beige",
    "أزرق": "Blue",
    "أحمر": "Red",
    # ── 제조국 (긴 구문 먼저 — 부분 치환 오류 방지)
    "المملكة العربية السعودية": "Saudi Arabia",
    "المملكة العربية": "Saudi Arabia",
    "المملكة": "Saudi Arabia",
    "الصين": "China",
    "كوريا الجنوبية": "South Korea",
    "كوريا": "Korea",
    "اليابان": "Japan",
    "تايلاند": "Thailand",
    "تايوان": "Taiwan",
    "السعودية": "Saudi Arabia",
    "مصر": "Egypt",
    "ماليزيا": "Malaysia",
    "الإمارات العربية المتحدة": "UAE",
    "الإمارات": "UAE",
    "تركيا": "Turkey",
    "الهند": "India",
    "إيطاليا": "Italy",
    "ألمانيا": "Germany",
    # ── 오퍼 / 프로모션 문구 (긴 구문 먼저)
    "خصم 15% كاش باك": "15% Cashback Discount",
    "خصم 10% كاش باك": "10% Cashback Discount",
    "كاش باك 15%": "15% Cashback",
    "كاش باك 10%": "10% Cashback",
    "كاش باك": "Cashback",
    "خصم بنك الأهلي 10% بحد أقصى": "Al Ahli Bank 10% Discount with Max Cap",
    "خصم بنك الأهلي": "Al Ahli Bank Discount",
    "بنك الأهلي": "Al Ahli Bank",
    "خدمة التركيب المجانية": "Free Installation Service",
    "خدمة التركيب": "Installation Service",
    "تركيب مجاني": "Free Installation",
    "تركيب": "Installation",
    "خصم إضافي": "Additional Discount",
    "خصم": "Discount",
    "عرض خاص": "Special Offer",
    "عرض محدود": "Limited Offer",
    "عرض": "Offer",
    "هدية مجانية": "Free Gift",
    "هدية": "Gift",
    "مجاناً": "Free",
    "مجاني": "Free",
    "شراء": "Purchase",
    "استبدال": "Replacement",
    "تنزيل": "Discount",
    # ── Air Swing / 방향
    "تلقائي": "Auto",
    "يدوي": "Manual",
    "اتجاهين": "2-Way",
    "أربعة اتجاهات": "4-Way",
    # ── 라벨 / 배너
    "الأكثر مبيعاً": "Best Seller",
    "الأفضل مبيعاً": "Best Seller",
    "جديد": "New",
    "موصى به": "Recommended",
    "حصري": "Exclusive",
    # ── 보증 (숫자 추출용 — 전체 일치 우선 처리)
    "عشر سنوات": "10",
    "10 سنوات": "10",
    "سبع سنوات": "7",
    "7 سنوات": "7",
    "خمس سنوات": "5",
    "5 سنوات": "5",
    "ثلاث سنوات": "3",
    "3 سنوات": "3",
    "سنتان": "2",
    "سنة واحدة": "1",
    "سنة": "1",
    "سنوات": "Years",
    # ── 기타 스펙 단위/용어
    "بي تي يو": "BTU",
    "طن": "Ton",
    "واط": "Watt",
    "كيلو واط": "kW",
    "نعم": "Yes",
    "لا": "No",
}

BRAND_AR_TO_EN = {
    "ال جي": "LG",
    "إل جي": "LG",
    "سامسونج": "Samsung",
    "سامسونغ": "Samsung",
    "جري": "Gree",
    "غري": "Gree",
    "ميديا": "Midea",
    "ميدياS": "Midea",
    "كاريير": "Carrier",
    "كارير": "Carrier",
    "داييكن": "Daikin",
    "دايكن": "Daikin",
    "توشيبا": "Toshiba",
    "هيتاشي": "Hitachi",
    "شارب": "Sharp",
    "باناسونيك": "Panasonic",
    "هاير": "Haier",
    "هايير": "Haier",
    "يورك": "York",
    "زاميل": "Zamil",
    "هام": "Haam",
    "كلفينيتور": "Kelvinator",
    "أريستون": "Ariston",
    "فوجيتسو": "Fujitsu",
    "ميتسوبيشي": "Mitsubishi",
    "هيونداي": "Hyundai",
    "بوش": "Bosch",
    "إليكترولوكس": "Electrolux",
}

# ─────────────────────────────────────────────
# 번역 유틸리티  (v3.3 강화)
# ─────────────────────────────────────────────
# 긴 구문 우선 정렬 캐시 (모듈 로드 시 1회 계산)
_AR_SORTED = sorted(AR_TO_EN.items(),      key=lambda x: len(x[0]), reverse=True)
_BR_SORTED = sorted(BRAND_AR_TO_EN.items(), key=lambda x: len(x[0]), reverse=True)

# 아랍어 유니코드 블록 전체 (아랍어·확장·표현형 모두 포함)
_AR_PATTERN = re.compile(
    r'[\u0600-\u06FF'    # Arabic
    r'\u0750-\u077F'     # Arabic Supplement
    r'\u08A0-\u08FF'     # Arabic Extended-A
    r'\uFB50-\uFDFF'     # Arabic Presentation Forms-A
    r'\uFE70-\uFEFF]+'   # Arabic Presentation Forms-B
)


def _has_arabic(text: str) -> bool:
    if not text:
        return False
    return bool(_AR_PATTERN.search(text))


def translate(text) -> str:
    """아랍어가 포함된 경우 영어로 변환. 영어면 그대로 반환.

    처리 순서:
    1. 전체 문자열 완전 일치 (AR_TO_EN → BRAND_AR_TO_EN)
    2. 긴 구문 → 짧은 구문 순서로 부분 치환
    3. 잔여 아랍어 블록 완전 제거
    4. 공백·구두점 정리
    """
    if not isinstance(text, str):
        return text
    t = text.strip()
    if not t or not _has_arabic(t):
        return t

    # 1. 완전 일치
    if t in AR_TO_EN:
        return AR_TO_EN[t]
    if t in BRAND_AR_TO_EN:
        return BRAND_AR_TO_EN[t]

    # 2. 부분 치환 (긴 구문 우선 — 캐시된 정렬 사용)
    result = t
    for ar, en in _AR_SORTED:
        if ar in result:
            result = result.replace(ar, en)
    for ar, en in _BR_SORTED:
        if ar in result:
            result = result.replace(ar, en)

    # 3. 잔여 아랍어 블록 제거
    result = _AR_PATTERN.sub('', result)

    # 4. 중복 공백·앞뒤 구두점 정리
    result = re.sub(r'\s{2,}', ' ', result).strip(' |,-')

    return result if result else t  # 완전 빈 결과면 원문 반환


# ─────────────────────────────────────────────
# 값 정규화 헬퍼
# ─────────────────────────────────────────────
def _norm_compressor(val: str) -> str:
    """컴프레서 타입 정규화"""
    v = val.strip().upper()
    if not v or v == 'N/A':
        return ''
    if 'DUAL' in v and 'INVERTER' in v:
        return 'Dual Inverter'
    if 'INVERTER' in v or 'ROTARY INVERTER' in v:
        return 'Inverter'
    if 'FIXED' in v or 'ON/OF' in v or 'ONSTANT' in v:
        return 'Fixed Speed'
    if 'ROTARY' in v or 'روتاري' in val:
        return 'Rotary'
    # 알려진 브랜드/컴포넌트 이름이면 그대로
    for kw in ['GMCC', 'HIGHLY', 'LANDA', 'GREE', 'WELLING']:
        if kw in v:
            return val.strip().title()
    return val.strip()


def _norm_function(val: str) -> str:
    """Function 정규화: Cold Only / Cold & Hot"""
    v = translate(val).lower()
    if 'hot' in v or 'heat' in v:
        return 'Cold & Hot'
    if 'cold' in v or 'cool' in v:
        return 'Cold Only'
    return val.strip() if val.strip() else ''


def _extract_number(val) -> float | None:
    """문자열에서 첫 번째 숫자 추출. 숫자면 그대로."""
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except:
        m = re.search(r'[\d,]+\.?\d*', str(val).replace(',', ''))
        if m:
            try:
                return float(m.group())
            except:
                pass
    return None


# ─────────────────────────────────────────────
# BTU → Ton 범위 테이블 기반 매핑  (v3.4)
# ─────────────────────────────────────────────
# Promotion Condition 시트의 G열 테이블과 동일한 규칙 (min <= BTU < max)
# 코드에 내장하여 엑셀 시트 의존성 없이 동일한 결과를 반환
BTU_TON_TABLE = [
    # (min_btu,  max_btu,  ton_grade)
    (  9_000,  15_000, 1.0),
    ( 15_000,  21_000, 1.5),
    ( 21_000,  27_000, 2.0),
    ( 27_000,  33_000, 2.5),
    ( 33_000,  39_000, 3.0),
    ( 39_000,  45_000, 3.5),
    ( 45_000,  51_000, 4.0),
    ( 51_000,  57_000, 4.5),
    ( 57_000,  63_000, 5.0),
]


def map_btu_to_ton(btu_value) -> float | None:
    """BTU 값을 표준 Ton 등급으로 변환.

    규칙: min_btu <= BTU < max_btu 인 구간의 ton_grade 반환
    - BTU 없음(None/빈값) → None
    - 범위 밖(9,000 미만 또는 63,000 이상) → None

    Examples:
        map_btu_to_ton(12000) → 1.0
        map_btu_to_ton(18400) → 1.5
        map_btu_to_ton(15000) → 1.5   # 경계: 15000은 1.5T 구간 시작
        map_btu_to_ton(14999) → 1.0   # 경계: 14999는 1T 구간 끝
        map_btu_to_ton(8999)  → None  # 범위 미달
        map_btu_to_ton(65000) → None  # 범위 초과
        map_btu_to_ton(None)  → None
    """
    if btu_value is None:
        return None
    # 콤마 포함 문자열도 처리 (e.g. "18,400")
    try:
        btu = float(str(btu_value).replace(',', '').strip())
    except (ValueError, TypeError):
        return None

    for min_b, max_b, ton in BTU_TON_TABLE:
        if min_b <= btu < max_b:
            return ton

    return None  # 범위 밖


# ─────────────────────────────────────────────
# W열·X열 가격 계산 로직
# ─────────────────────────────────────────────
def _calc_final_promo(promo_price, offer_detail: str):
    """
    W열: Final_Promo_Price
    Offer_Detail 문자열에서 캐시백/할인율을 읽어 Promo_Price에 적용.

    규칙 (실제 엑셀 수식 역공학 기준):
      - offer에 '15%' 가 있으면 → Promo_Price × 0.85
      - offer에 '10%' 가 있으면 → Promo_Price × 0.90
        (단 'Al Ahli' 전용 10%는 W열에서 제외 — X열에서만 적용)
      - 그 외                   → Promo_Price 그대로
    """
    if promo_price is None:
        return None
    try:
        p = float(promo_price)
    except:
        return promo_price

    offer = (offer_detail or '').lower()

    # 'Al Ahli' 전용 오퍼만 있는 경우 W = Q (Al Ahli 할인은 X열에서 처리)
    # cashback/extra 오퍼 내 15% 감지
    if re.search(r'(cashback|cash back|كاش باك).{0,30}15%', offer) or \
       re.search(r'15%.{0,30}(cashback|cash back|كاش باك)', offer) or \
       '(15% off)' in offer:
        return int(round(p * 0.85))

    # cashback/extra 오퍼 내 10% 감지 — Al Ahli만 있는 경우는 제외
    has_10 = re.search(r'(cashback|cash back|extra|كاش باك).{0,30}10%', offer) or \
             re.search(r'10%.{0,30}(cashback|cash back|extra|كاش باك)', offer) or \
             '(10% off)' in offer
    only_alahli = bool(re.search(r'al ahli', offer)) and not re.search(
        r'(cashback|cash back|extra|كاش باك)', offer)

    if has_10 and not only_alahli:
        return int(round(p * 0.90))

    return int(round(p))


def _calc_alahli_price(final_promo_price, promo_price, offer_detail: str):
    """
    X열: AlAhli_Price
    Al Ahli 은행 카드 10% 추가 할인가.

    규칙 (실제 엑셀 수식 역공학 기준):
      - offer에 'Al Ahli' 관련 10% 가 있으면 → Final_Promo_Price × 0.90
      - 그 외                                 → Final_Promo_Price 그대로
    """
    base = final_promo_price if final_promo_price is not None else promo_price
    if base is None:
        return None
    try:
        b = float(base)
    except:
        return base

    offer = (offer_detail or '').lower()

    has_alahli_10 = bool(re.search(
        r'al ahli.{0,50}10%|10%.{0,50}al ahli'
        r'|al ahli bank 10'
        r'|alahli.{0,30}10',
        offer
    ))

    if has_alahli_10:
        return int(round(b * 0.90))

    return int(round(b))


# ─────────────────────────────────────────────
# 스크래퍼 클래스
# ─────────────────────────────────────────────
class AlmaneaScraper:
    BASE_URL  = "https://api-preprod.dev-almanea.com/api/v1/facets/categoryV2/535"
    SITE_URL  = "https://www.almanea.sa/en"
    # 만료 시 자동 갱신됨 — 하드코딩값은 fallback용
    TOKEN     = ("Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
                 ".eyJzdG9yZSI6ImVuIiwic2Vzc2lvbklEIjoiaEZCZGRlLXhL"
                 "RnBISkFlcS14akctNEFSWFY3bjk4Y2EiLCJpYXQiOjE3NzIy"
                 "OTg5NDMsImV4cCI6MTc3NDg5MDk0M30"
                 ".M4LoHZotd3KWLAGo6219sWiWc8_XB1VE06cqWA3X2dI")

    # ── 사이트에서 JWT 토큰 자동 갱신
    @classmethod
    def _fetch_token(cls) -> str:
        """almanea.sa 페이지에서 최신 Bearer 토큰 자동 추출.
        서버 다운 시 fallback 토큰 사용 + 잔여 유효기간 표시."""
        import re as _re, base64 as _b64, json as _json

        # ── 0) 현재 fallback 토큰 유효기간 확인 (디버그용)
        def _token_info(tok: str) -> str:
            try:
                parts = tok.replace('Bearer ', '').split('.')
                pad = parts[1] + '=' * (4 - len(parts[1]) % 4)
                p = _json.loads(_b64.b64decode(pad))
                exp_dt = datetime.fromtimestamp(p['exp'])
                remain = exp_dt - datetime.now()
                if remain.total_seconds() <= 0:
                    return f"EXPIRED ({exp_dt.strftime('%Y-%m-%d')})"
                return f"valid until {exp_dt.strftime('%Y-%m-%d')} ({remain.days}d remaining)"
            except Exception:
                return "unknown expiry"

        print(f"  [TOKEN] fallback token: {_token_info(cls.TOKEN)}")
        print("  [TOKEN] 사이트에서 최신 토큰 취득 시도...")

        hdrs = {
            'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/122.0.0.0 Safari/537.36'),
            'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
        }

        # ── 1) 웹사이트에서 HTML 가져오기 (재시도 포함)
        urls_to_try = [
            'https://www.almanea.sa/en',
            'https://www.almanea.sa',
            'https://almanea.sa/en',
        ]
        html = ''
        MAX_RETRIES = 2
        for _url in urls_to_try:
            for attempt in range(MAX_RETRIES):
                try:
                    r = requests.get(_url, headers=hdrs, timeout=20)
                    html = r.text
                    if html:
                        print(f"  [TOKEN] {_url} 접속 성공 ({len(html)} bytes)")
                        break
                except requests.exceptions.Timeout:
                    if attempt < MAX_RETRIES - 1:
                        print(f"  [TOKEN] {_url} timeout, 재시도 {attempt+2}/{MAX_RETRIES}...")
                        time.sleep(2)
                    continue
                except Exception:
                    break
            if html:
                break

        if not html:
            print("  [TOKEN] 사이트 접속 불가 (서버 다운 또는 네트워크 문제)")
            print(f"  [TOKEN] fallback 토큰 사용 — {_token_info(cls.TOKEN)}")
            return cls.TOKEN

        # ── 2) __NEXT_DATA__ JSON 안에 JWT 탐색
        try:
            nd = _re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, _re.S)
            if nd:
                jwt = _re.search(
                    r'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9\.[A-Za-z0-9_\-]+\.[A-Za-z0-9_\-]+',
                    nd.group(1)
                )
                if jwt:
                    token = 'Bearer ' + jwt.group()
                    print(f"  [TOKEN] 자동 갱신 성공 (__NEXT_DATA__) — {_token_info(token)}")
                    return token

            # 3) 전체 HTML 에서 JWT 패턴 탐색
            jwt = _re.search(
                r'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9\.[A-Za-z0-9_\-]+\.[A-Za-z0-9_\-]+',
                html
            )
            if jwt:
                token = 'Bearer ' + jwt.group()
                print(f"  [TOKEN] 자동 갱신 성공 (HTML) — {_token_info(token)}")
                return token

            print("  [TOKEN] HTML에서 토큰 미발견 — JS 번들 탐색 중...")

            # 4) 인라인 script src 목록에서 JS 번들 다운로드 후 탐색
            srcs = _re.findall(r'<script[^>]+src=["\']([^"\']+)["\']', html)
            for src in srcs:
                if not src.startswith('http'):
                    src = 'https://www.almanea.sa' + src
                try:
                    jr = requests.get(src, headers=hdrs, timeout=10)
                    jm = _re.search(
                        r'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9\.[A-Za-z0-9_\-]+\.[A-Za-z0-9_\-]+',
                        jr.text
                    )
                    if jm:
                        token = 'Bearer ' + jm.group()
                        print(f"  [TOKEN] 자동 갱신 성공 (JS bundle) — {_token_info(token)}")
                        return token
                except Exception:
                    pass

        except Exception as e:
            print(f"  [TOKEN] 자동 갱신 실패: {e}")

        print(f"  [TOKEN] fallback 토큰 사용 — {_token_info(cls.TOKEN)}")
        return cls.TOKEN

    def __init__(self):
        token = self._fetch_token()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/122.0.0.0 Safari/537.36'),
            'Authorization': token,
        })

    # ── 카테고리 추출 (제품명 기반)
    def _category(self, name: str) -> str:
        u = name.upper()
        if 'WINDOW'                               in u: return 'Window AC'
        if 'FLOOR' in u or 'FREESTANDING'        in u: return 'Floor Standing'
        if 'CASSETTE' in u or 'CEILING' in u or 'CONCEALED' in u: return 'Cassette & Ceiling'
        if 'PORTABLE'                             in u: return 'Portable'
        if 'SPLIT' in u or 'WALL' in u           : return 'Split AC'
        return 'Other'

    # ── product_attributes에서 값 추출 (영문 key=2 우선, 아랍어 key=14 fallback)
    def _attr(self, pa: list, keyword: str) -> str:
        if not pa:
            return ''
        kw = keyword.lower()
        for section in pa:
            if not isinstance(section, dict):
                continue
            for key in ('2', 2, '14', 14):
                attr = section.get(key, {})
                if not isinstance(attr, dict):
                    continue
                label = translate(str(attr.get('label', ''))).strip().lower()
                value = translate(str(attr.get('value', ''))).strip()
                if kw in label and value and value != 'None':
                    return value
        return ''

    # ── Offer 텍스트 조합
    def _offers(self, src: dict) -> tuple[str, str]:
        """(offer_detail, free_gift) 반환 — 모두 영문"""
        parts   = []
        gifts   = []

        # cart_rule
        for rule in (src.get('cart_rule') or []):
            if not isinstance(rule, dict):
                continue
            name = translate(str(rule.get('name', ''))).strip()
            amt  = rule.get('discount_amount', '')
            if name:
                try:
                    a = float(amt)
                    if a == 1.0:
                        parts.append(f"{name} (Free)")
                    elif a < 100:
                        parts.append(f"{name} ({a:.0f}% off)")
                    else:
                        parts.append(f"{name} (SAR {a:.0f} off)")
                except:
                    parts.append(name)

        # free_gifts
        for g in (src.get('free_gifts') or []):
            if not isinstance(g, dict):
                continue
            gn = translate(str(g.get('product_name', ''))).strip()
            if gn:
                try:
                    gp = float(g.get('product_price', 0))
                    gifts.append(f"{gn} (SAR {gp:.0f})")
                except:
                    gifts.append(gn)

        # label
        for lbl in (src.get('label') or []):
            if isinstance(lbl, dict):
                ln = translate(str(lbl.get('name', ''))).strip()
                if ln:
                    parts.append(f"[{ln}]")

        # ac_replacement
        if src.get('ac_replacement'):
            d = src.get('ac_replacement_discount', '')
            if d:
                parts.append(f"AC Replacement: SAR {d} off")

        # 중복 제거
        def dedup(lst):
            seen, out = set(), []
            for x in lst:
                if x.lower() not in seen:
                    seen.add(x.lower())
                    out.append(x)
            return out

        return ' | '.join(dedup(parts)), ', '.join(dedup(gifts))

    # ── 단일 regionId로 전체 페이지 수집 (내부 헬퍼)
    def _fetch_region(self, region_id: int) -> list:
        """지정 regionId로 페이지네이션 수집. 401 시 자동 토큰 재발급, timeout 시 재시도."""
        params   = {'pageSize': 32, 'sortBy': 'position', 'sortDir': 'ASC',
                    'regionId': region_id}
        products = []
        page     = 0
        total    = None
        MAX_RETRIES = 3

        while True:
            params['pageNo'] = page
            r = None
            for attempt in range(MAX_RETRIES):
                try:
                    r = self.session.get(self.BASE_URL, params=params, timeout=20)
                    break  # 성공
                except requests.exceptions.Timeout:
                    if attempt < MAX_RETRIES - 1:
                        wait = 3 * (attempt + 1)
                        print(f"    Page {page+1} timeout, {wait}초 후 재시도 ({attempt+2}/{MAX_RETRIES})...")
                        time.sleep(wait)
                    else:
                        print(f"    Page {page+1} — API 서버 응답 없음 (timeout x{MAX_RETRIES})")
                        return products
                except requests.exceptions.ConnectionError as e:
                    print(f"    연결 오류: {e}")
                    return products
                except Exception as e:
                    print(f"    오류: {e}")
                    return products

            if r is None:
                break

            if r.status_code == 401:
                print(f"    HTTP 401 — 토큰 재발급 시도...")
                new_token = self._fetch_token()
                self.session.headers.update({'Authorization': new_token})
                try:
                    r = self.session.get(self.BASE_URL, params=params, timeout=20)
                except Exception as e:
                    print(f"    재발급 후 요청 실패: {e}")
                    break

            if r.status_code != 200:
                print(f"    HTTP {r.status_code} — 중단")
                break

            try:
                data = r.json()
            except Exception:
                print(f"    JSON 파싱 실패 — 중단")
                break

            batch = data.get('products', [])
            if not batch:
                break
            products.extend(batch)
            if total is None:
                total = data.get('totalProduct') or 0
            if total > 0 and len(products) >= total:
                break
            page += 1
            time.sleep(0.8)

        return products

    # ── 페이지네이션으로 전체 제품 수집 (다중 regionId 병합)
    def fetch_all(self) -> list:
        print("\n" + "═"*65)
        print("  [STEP 1] Al Manea API — 제품 목록 수집")
        print("═"*65)

        # ── Region 전략 ─────────────────────────────────────────────
        # regionId=1114: 전국 최대 커버리지 (205개)
        # regionId=1101: Riyadh 지역 추가 (1114에 없는 2개 보완)
        # → 두 region 병합으로 최대 207개 수집
        # ────────────────────────────────────────────────────────────
        REGION_IDS = [1114, 1101]   # 1114 우선, 1101로 보완
        seen_skus  = {}             # sku → product (1114 우선 저장)

        for rid in REGION_IDS:
            print(f"\n  [Region {rid}] 수집 시작...", flush=True)
            batch_products = self._fetch_region(rid)
            added = 0
            for p in batch_products:
                src = p.get('_source', p)
                sku = src.get('sku', '')
                if sku and sku not in seen_skus:
                    seen_skus[sku] = p
                    added += 1
            print(f"  [Region {rid}] {len(batch_products)}개 수신 / 신규 {added}개 추가")
            time.sleep(1.0)

        products = list(seen_skus.values())
        print(f"\n  총 {len(products)}개 제품 수집 완료 (region 병합)\n")
        return products

    # ── 제품 1개 → 표준 dict 변환
    def _parse(self, idx: int, product: dict) -> dict:
        src = product.get('_source', product)

        def _first(v):
            return (v[0] if v else '') if isinstance(v, list) else v

        # ── 기본 필드
        name  = translate(_first(src.get('name', '')) or '')
        brand = translate(_first(src.get('option_text_brand', '')) or '')
        model = (_first(src.get('model', '')) or '').strip()
        sku   = src.get('sku', '')

        # ── 가격
        prices  = src.get('prices_with_tax', {}) or {}
        orig    = prices.get('original_price')
        promo   = prices.get('discounted_price')
        try:    orig  = int(round(float(orig)))
        except: orig  = None
        try:    promo = int(round(float(promo)))
        except: promo = None
        discount_pct = None
        if orig and promo and orig > 0:
            discount_pct = round((orig - promo) / orig, 4)

        # ── 재고
        stock_info = src.get('stock', {})
        try:
            stock = int(stock_info.get('qty') or 0) if isinstance(stock_info, dict) else 0
        except (TypeError, ValueError):
            stock = 0

        # ── 카테고리 / Function
        category = self._category(name)
        cold_raw = translate(_first(src.get('option_text_cold_or_hot_cold', '')) or '')
        function = _norm_function(cold_raw)

        # ── product_attributes 스펙
        pa = src.get('product_attributes', []) or []

        comp_raw = self._attr(pa, 'compressor type') or self._attr(pa, 'type of compressor')
        comp_type = _norm_compressor(translate(comp_raw))

        # 보증 (숫자만 추출)
        def _warranty_yr(keyword):
            raw = self._attr(pa, keyword)
            n = _extract_number(raw)
            return int(n) if n and n > 0 else None

        warranty_yr   = _warranty_yr('warranty')
        comp_warr_yr  = _warranty_yr('compressor warranty')

        # BTU
        btu_raw = self._attr(pa, 'cooling capacity') or self._attr(pa, 'btu')
        btu_num = _extract_number(btu_raw)
        # product_attributes 없으면 제품명에서 추출
        if not btu_num:
            m = re.search(r'(\d[\d,]+)\s*BTU', name, re.IGNORECASE)
            if m:
                btu_num = _extract_number(m.group(1))
        btu = int(btu_num) if btu_num else None

        # Capacity_Ton — BTU 범위 테이블 기준 표준 등급
        # 1순위: BTU 값으로 테이블 매핑
        # 2순위: product_attributes capacity 값이 Ton 단위 소수라면 테이블 역매핑
        cap_ton = map_btu_to_ton(btu)

        if cap_ton is None:
            # BTU로 매핑 실패 시 capacity 원시값 시도
            cap_raw = self._attr(pa, 'capacity')
            cap_num = _extract_number(cap_raw)
            if cap_num:
                if cap_num > 50:
                    # 큰 숫자 = BTU 단위 값으로 간주 → 테이블 매핑
                    cap_ton = map_btu_to_ton(cap_num)
                else:
                    # 작은 숫자 = 이미 Ton 단위 → 테이블에서 가장 가까운 표준 등급 찾기
                    cap_ton = map_btu_to_ton(cap_num * 12000)

        # Energy Rating
        energy = self._attr(pa, 'energy efficiency') or self._attr(pa, 'energy rating')
        energy = translate(energy)
        # 단일 알파벳 등급만 남김 (A / A+ / B / C ...)
        if energy and len(energy) > 5:
            m2 = re.search(r'\b(A\+{0,3}|B|C|D|E|F)\b', energy, re.IGNORECASE)
            energy = m2.group(1).upper() if m2 else energy[:10]

        # Color / Country
        color   = translate(_first(src.get('option_text_color', '')) or '')
        country_raw = self._attr(pa, 'country') or self._attr(pa, 'origin')
        country = translate(country_raw)
        # 하이브리드 아랍어+영어 잔여 처리
        if country and _has_arabic(country):
            country = translate(country)

        # ── Offer
        offer_detail, free_gift = self._offers(src)

        # ── W열·X열 자동 계산
        final_promo = _calc_final_promo(promo, offer_detail)
        alahli      = _calc_alahli_price(final_promo, promo, offer_detail)

        # ── URL Key (상품 페이지 링크용)
        url_key_raw = _first(src.get('url_key', '')) or src.get('rewrite_url', '') or ''
        url_key = url_key_raw.strip() if isinstance(url_key_raw, str) else ''

        return {
            "Scraped_At":             datetime.now().date(),
            "SKU":                    sku,
            "Brand":                  brand,
            "Model":                  model,
            "Product_Name":           name,
            "Category":               category,
            "Function":               function,
            "Compressor_Type":        comp_type,
            "Capacity_Ton":           cap_ton,
            "BTU":                    btu,
            "Energy_Rating":          energy or '',
            "Color":                  color or '',
            "Country":                country or '',
            "Warranty_Yr":            warranty_yr,
            "Compressor_Warranty_Yr": comp_warr_yr,
            "Original_Price":         orig,
            "Promo_Price":            promo,
            "Discount_Pct":           discount_pct,
            "Has_Offer":              'Yes' if offer_detail else 'No',
            "Offer_Detail":           offer_detail,
            "Free_Gift":              free_gift,
            "Stock":                  stock,
            "Final_Promo_Price":      final_promo,
            "AlAhli_Price":           alahli,
            "URL_Key":                url_key,
        }

    # ── 전체 파싱
    def parse_all(self, products: list) -> list:
        print("═"*65)
        print("  [STEP 2] 제품 정보 파싱 및 정규화")
        print("═"*65)
        results = []
        for i, p in enumerate(products, 1):
            try:
                results.append(self._parse(i, p))
            except Exception as e:
                print(f"  [WARN] #{i} 파싱 실패: {e}")
            if i % 50 == 0:
                print(f"  {i}/{len(products)} 처리 완료")

        # 수집 통계
        total = len(results)
        print(f"\n  파싱 완료: {total}개")
        stats = [
            ("Compressor_Type",        "컴프레서 타입"),
            ("BTU",                    "BTU"),
            ("Capacity_Ton",           "Ton"),
            ("Energy_Rating",          "에너지 등급"),
            ("Country",                "제조국"),
            ("Warranty_Yr",            "보증"),
            ("Compressor_Warranty_Yr", "컴프레서 보증"),
            ("Has_Offer",              "오퍼 보유"),
        ]
        for col, label in stats:
            if col == "Has_Offer":
                cnt = sum(1 for r in results if r[col] == 'Yes')
            else:
                cnt = sum(1 for r in results if r.get(col))
            print(f"    {label:18s}: {cnt:3d} / {total}")
        print()
        return results

    # ── Excel 누적 저장
    def save(self, records: list) -> str:
        print("═"*65)
        print("  [STEP 3] Products_DB 시트 누적 저장")
        print("═"*65)

        df_new = pd.DataFrame(records, columns=COLUMNS)

        # ── 기존 데이터 로드
        if os.path.exists(OUTPUT_FILE):
            try:
                xl = pd.ExcelFile(OUTPUT_FILE, engine='openpyxl')
                if DB_SHEET in xl.sheet_names:
                    df_old = xl.parse(DB_SHEET)
                    # 구버전 컬럼명 → 현재 표준명으로 정규화
                    df_old.rename(columns={
                        'Final Promotion Price':  'Final_Promo_Price',
                        'Al Ahli Bank Promotion': 'AlAhli_Price',
                    }, inplace=True)
                    # 스키마 불일치 처리: 기존 파일의 컬럼이 다를 수 있음
                    df_old = df_old.reindex(columns=COLUMNS)
                    df_all = pd.concat([df_old, df_new], ignore_index=True)
                    print(f"  기존 {len(df_old)}행 + 신규 {len(df_new)}행 = 누적 {len(df_all)}행")
                else:
                    df_all = df_new
                    print(f"  Products_DB 시트 신규 생성: {len(df_all)}행")
                xl.close()
            except Exception as e:
                print(f"  [WARN] 기존 파일 읽기 실패 ({e}) → 신규 생성")
                df_all = df_new
        else:
            df_all = df_new
            print(f"  신규 파일 생성: {len(df_all)}행")

        # ── openpyxl로 시트만 교체 (나머지 시트 보존)
        try:
            if os.path.exists(OUTPUT_FILE):
                wb = load_workbook(OUTPUT_FILE)
                if DB_SHEET in wb.sheetnames:
                    del wb[DB_SHEET]
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            ws = wb.create_sheet(DB_SHEET)

            # 헤더 스타일
            hdr_fill = PatternFill('solid', fgColor='1F4E79')
            hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            ctr      = Alignment(horizontal='center', vertical='center', wrap_text=False)

            for ci, col in enumerate(df_all.columns, 1):
                c = ws.cell(row=1, column=ci, value=col)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = ctr

            # Scraped_At 열 인덱스 (0-based)
            scraped_at_ci = list(df_all.columns).index('Scraped_At') if 'Scraped_At' in df_all.columns else -1

            # 데이터 기록
            for ri, row in enumerate(df_all.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    # NaN → 빈 문자열
                    if isinstance(val, float) and pd.isna(val):
                        val = ''
                    # Scraped_At: Timestamp/datetime/문자열 → date 객체로 통일
                    if ci == scraped_at_ci + 1 and val != '':
                        try:
                            if hasattr(val, 'date'):      # Timestamp or datetime
                                val = val.date()
                            elif isinstance(val, str):    # '2026-02-28' 형식 문자열
                                from datetime import date as _date
                                val = _date.fromisoformat(val[:10])
                        except Exception:
                            pass
                    ws.cell(row=ri, column=ci, value=val)

            # 컬럼 너비
            widths = {
                "Scraped_At": 13, "SKU": 18, "Brand": 12, "Model": 24,
                "Product_Name": 60, "Category": 22, "Function": 12,
                "Compressor_Type": 16, "Capacity_Ton": 12, "BTU": 10,
                "Energy_Rating": 12, "Color": 10, "Country": 14,
                "Warranty_Yr": 12, "Compressor_Warranty_Yr": 20,
                "Original_Price": 14, "Promo_Price": 14, "Discount_Pct": 12,
                "Has_Offer": 10, "Offer_Detail": 55,
                "Free_Gift": 30, "Stock": 8,
                "Final_Promo_Price": 18, "AlAhli_Price": 16,
                "URL_Key": 65,
            }
            for ci, col in enumerate(df_all.columns, 1):
                ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

            # 숫자 포맷 적용
            col_list = list(df_all.columns)

            # Discount_Pct → 0% 포맷
            disc_col = col_list.index("Discount_Pct") + 1
            for ri in range(2, len(df_all) + 2):
                ws.cell(row=ri, column=disc_col).number_format = '0%'

            # Final_Promo_Price, AlAhli_Price → #,##0 포맷
            for col_name in ("Final_Promo_Price", "AlAhli_Price"):
                if col_name in col_list:
                    c_idx = col_list.index(col_name) + 1
                    for ri in range(2, len(df_all) + 2):
                        ws.cell(row=ri, column=c_idx).number_format = '#,##0'

            ws.sheet_properties.tabColor = '1F4E79'
            ws.freeze_panes = 'A2'

            wb.save(OUTPUT_FILE)
            print(f"\n  저장 완료 → {OUTPUT_FILE}")
            print(f"  시트: {DB_SHEET}  |  총 {len(df_all)}행  |  컬럼: {len(COLUMNS)}개")

            # 카테고리/브랜드 요약
            df_today = df_new
            print(f"\n  ── 이번 수집 요약 ({df_today['Scraped_At'].iloc[0]}) ──")
            for cat, grp in df_today.groupby('Category'):
                brands = grp['Brand'].value_counts().head(4)
                brand_str = ', '.join(f"{b}({n})" for b, n in brands.items())
                print(f"    {cat:<28s}: {len(grp):3d}개  [{brand_str}]")
            return OUTPUT_FILE

        except Exception as e:
            import traceback
            print(f"\n  [ERROR] 저장 실패: {e}")
            traceback.print_exc()
            return ''


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    if sys.platform == 'win32':
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    print("\n" + "═"*65)
    print("  Almanea AC Scraper v3.4")
    print("═"*65)

    scraper  = AlmaneaScraper()
    products = scraper.fetch_all()

    if not products:
        print("[ERROR] 수집된 제품이 없습니다. 네트워크/토큰을 확인하세요.")
        input("\n엔터로 종료...")
        return

    records  = scraper.parse_all(products)
    scraper.save(records)

    print("\n" + "═"*65)
    print("  완료!")
    print("═"*65)
    input("\n엔터를 누르면 종료...")


if __name__ == '__main__':
    main()
