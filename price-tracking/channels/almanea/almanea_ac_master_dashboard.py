#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Almanea AC Master Dashboard Builder v2.0
- almanea_ac_dashboard_builder.py + create_stock_dashboard.py 통합본
- 실행: python almanea_ac_master_dashboard.py
- 생성 시트 (총 9개):
    [Price Dashboards]
    1. Dashboard_Summary       — KPI / Category 분포 / 이전 대비 변동 요약
    2. Price_Change_Alert      — 가격 변동 SKU 목록
    3. New_Discontinued_SKUs   — 신규 / 단종 SKU
    4. Brand_Avg_Price_Compare — 브랜드별 평균가 비교
    5. Full_Price_Tracking     — 전체 제품 최신 스냅샷
    6. Offer_Analysis          — 오퍼 / Free Gift 분석
    [Stock Dashboards]
    7. Stock_Dashboard_Summary — 재고 Executive Summary
    8. Stock_Dashboard_Brand   — 카테고리×브랜드 재고 비중
    9. Stock_Dashboard_Detail  — 전체 SKU 재고 상세
"""

import os
import sys

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[ERROR] 패키지 없음: {e}")
    print("  >> py -m pip install pandas openpyxl numpy")
    sys.exit(1)

# ─────────────────────────────────────────────
# 경로
# ─────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
TARGET_FILE = os.path.join(SCRIPT_DIR, "Almanea_AC_Price_Tracking_Master.xlsx")
DB_SHEET    = "Products_DB"

# ─────────────────────────────────────────────
# 스타일 상수
# ─────────────────────────────────────────────
HDR_FILL        = PatternFill('solid', fgColor='1F4E79')
HDR_FONT        = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT      = Font(name='Arial', bold=True, size=13, color='1F4E79')
SUB_FONT        = Font(name='Arial', bold=True, size=10, color='595959')
SEC_FONT        = Font(name='Arial', bold=True, size=11, color='1F4E79')
DATA_FONT       = Font(name='Arial', size=10)
NUM_FONT        = Font(name='Arial', size=10)
BOLD_FONT       = Font(name='Arial', bold=True, size=10)
UP_FONT         = Font(name='Arial', size=10, color='C00000', bold=True)   # 가격 상승
DOWN_FONT       = Font(name='Arial', size=10, color='375623', bold=True)   # 가격 하락
STOCK_UP        = Font(name='Arial', size=10, color='375623', bold=True)   # 재고 증가
STOCK_DOWN      = Font(name='Arial', size=10, color='C00000', bold=True)   # 재고 감소
NEW_FONT        = Font(name='Arial', size=10, color='0070C0', bold=True)
DISC_FONT       = Font(name='Arial', size=10, color='999999')
HIGH_FONT       = Font(name='Arial', size=10, color='1F4E79', bold=True)
TK_BORDER       = Border(bottom=Side(style='medium', color='1F4E79'))
CTR             = Alignment(horizontal='center', vertical='center')
LEFT            = Alignment(horizontal='left',   vertical='center')
WRAP            = Alignment(horizontal='left',   vertical='center', wrap_text=True)
LIGHT_FILL      = PatternFill('solid', fgColor='F2F7FB')
CAT_FILL        = PatternFill('solid', fgColor='D6E4F0')
GREEN_FILL      = PatternFill('solid', fgColor='E2EFDA')
RED_FILL        = PatternFill('solid', fgColor='FCE4EC')
CRITICAL_FILL   = PatternFill('solid', fgColor='FCE4EC')
OUT_OF_STOCK_FILL = PatternFill('solid', fgColor='CCCCCC')
HIGH_FILL       = PatternFill('solid', fgColor='D6E4F0')

CAT_ORDER = ['Split AC', 'Window AC', 'Floor Standing', 'Cassette & Ceiling']

# ─────────────────────────────────────────────
# Products_DB 속성 정규화 설정
# ─────────────────────────────────────────────
FILL_COLS = [
    'Category', 'Function', 'Compressor_Type', 'Capacity_Ton',
    'BTU', 'Energy_Rating', 'Color', 'Country',
    'Warranty_Yr', 'Compressor_Warranty_Yr',
]

NORMALIZE_MAP = {
    'Function': {
        'Cold & Hot': 'Cold/Hot',
        'cold & hot': 'Cold/Hot',
        'cold/hot':   'Cold/Hot',
        'Cold only':  'Cold Only',
        'cold only':  'Cold Only',
    },
}


# ─────────────────────────────────────────────
# Products_DB 속성 백필 / 정규화
# ─────────────────────────────────────────────
def fill_product_attributes():
    """Products_DB F~O열을 과거 날짜 기준으로 덮어쓰고 값 정규화."""
    print("\n[STEP 0] Products_DB 속성 백필 & 정규화 시작")

    df = pd.read_excel(TARGET_FILE, sheet_name=DB_SHEET, engine='openpyxl')
    df['Scraped_At'] = pd.to_datetime(df['Scraped_At'], errors='coerce')
    dates = sorted(df['Scraped_At'].dropna().unique())

    if len(dates) < 2:
        print("  [SKIP] 이전 데이터 없음 - 백필 건너뜀")
        return

    df = df.copy()
    total_changed = 0

    for date_idx in range(1, len(dates)):
        current_date = dates[date_idx]
        past_dates   = dates[:date_idx]
        date_str     = pd.Timestamp(current_date).strftime('%Y-%m-%d')

        cur_mask  = df['Scraped_At'] == current_date
        past_df   = df[df['Scraped_At'].isin(past_dates)].sort_values('Scraped_At', ascending=False)
        sku_ref   = past_df.drop_duplicates('SKU',   keep='first').set_index('SKU')
        model_ref = past_df.drop_duplicates('Model', keep='first').set_index('Model')

        changed = 0
        for col in FILL_COLS:
            if col not in df.columns:
                continue
            for idx in df.index[cur_mask]:
                sku   = df.at[idx, 'SKU']
                model = df.at[idx, 'Model']
                ref   = None

                if sku in sku_ref.index:
                    v = sku_ref.at[sku, col]
                    if pd.notna(v) and str(v).strip() != '':
                        ref = v

                if ref is None and pd.notna(model) and str(model).strip() != '':
                    if model in model_ref.index:
                        v = model_ref.at[model, col]
                        if pd.notna(v) and str(v).strip() != '':
                            ref = v

                if ref is not None:
                    old = df.at[idx, col]
                    if str(old) != str(ref):
                        changed += 1
                    df.at[idx, col] = ref

        print(f"  [{date_str}] {changed}개 셀 변경")
        total_changed += changed

    # 값 정규화
    norm_changed = 0
    for col, mapping in NORMALIZE_MAP.items():
        if col not in df.columns:
            continue
        for dirty, clean in mapping.items():
            mask  = df[col] == dirty
            count = mask.sum()
            if count > 0:
                df.loc[mask, col] = clean
                norm_changed += count
                print(f"  [NORMALIZE] {col}: '{dirty}' → '{clean}' ({count}건)")

    print(f"  [INFO] 백필 {total_changed}개 + 정규화 {norm_changed}개 셀 변경")

    # 전체 시트 읽고 Products_DB만 교체하여 저장
    all_sheets = pd.read_excel(TARGET_FILE, sheet_name=None, engine='openpyxl')
    all_sheets[DB_SHEET] = df
    with pd.ExcelWriter(TARGET_FILE, engine='openpyxl') as writer:
        for sheet_name, sheet_df in all_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("  [OK] Products_DB 저장 완료\n")


# ─────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────
def _hdr(ws, row, n_col):
    for c in range(1, n_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CTR
        cell.border    = TK_BORDER


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx  = 0
        ltr = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[ltr].width = min(max(mx + 3, min_w), max_w)


def _del(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def _fmt_ton(val):
    try:
        v = float(val)
        return f"{v:.1f}T" if v % 1 else f"{int(v)}T"
    except Exception:
        return str(val) if val else '-'


def _pct_str(val):
    try:
        v = float(val)
        return f"{v*100:.0f}%" if v <= 1 else f"{v:.0f}%"
    except Exception:
        return '-'


# ─────────────────────────────────────────────
# 데이터 로드 (Price용)
# ─────────────────────────────────────────────
def load_price_data():
    df = pd.read_excel(TARGET_FILE, sheet_name=DB_SHEET, engine='openpyxl')

    COL_RENAME = {
        'Final Promotion Price': 'Final_Promo_Price',
        'Al Ahli Bank Promotion': 'AlAhli_Price',
    }
    df.rename(columns=COL_RENAME, inplace=True)
    df['Scraped_At'] = pd.to_datetime(df['Scraped_At'], errors='coerce')

    for col in ['Original_Price', 'Promo_Price', 'Discount_Pct',
                'Capacity_Ton', 'BTU', 'Warranty_Yr',
                'Compressor_Warranty_Yr', 'Stock',
                'Final_Promo_Price', 'AlAhli_Price']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    for col in ['Brand', 'Category', 'Function', 'Compressor_Type',
                'Country', 'Color', 'Has_Offer']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()

    df = df[df['Category'].isin(CAT_ORDER)].copy()

    if 'AlAhli_Price' in df.columns and df['AlAhli_Price'].notna().any():
        df['Final_Price'] = (df['AlAhli_Price']
                             .fillna(df['Final_Promo_Price'] if 'Final_Promo_Price' in df.columns else pd.NA)
                             .fillna(df['Promo_Price'])
                             .fillna(df['Original_Price']))
    elif 'Final_Promo_Price' in df.columns and df['Final_Promo_Price'].notna().any():
        df['Final_Price'] = (df['Final_Promo_Price']
                             .fillna(df['Promo_Price'])
                             .fillna(df['Original_Price']))
    else:
        df['Final_Price'] = df['Promo_Price'].fillna(df['Original_Price'])

    return df


# 데이터 로드 (Stock용)
def load_stock_data():
    df = pd.read_excel(TARGET_FILE, sheet_name=DB_SHEET, engine='openpyxl')
    df['Scraped_At'] = pd.to_datetime(df['Scraped_At'], errors='coerce')

    dates = sorted(df['Scraped_At'].dropna().unique())
    print(f"[INFO] Found {len(dates)} collection dates:")
    for date in dates:
        count = len(df[df['Scraped_At'] == date])
        print(f"  - {date.strftime('%Y-%m-%d')}: {count} products")

    if len(dates) < 2:
        print("[ERROR] Need at least 2 collection dates")
        sys.exit(1)

    latest_date = dates[-1]
    prev_date   = dates[-2]

    df_latest = df[df['Scraped_At'] == latest_date].copy()
    df_prev   = df[df['Scraped_At'] == prev_date].copy()

    merged = pd.merge(
        df_latest[['SKU', 'Brand', 'Model', 'Product_Name', 'Category', 'Capacity_Ton', 'Stock']],
        df_prev[['SKU', 'Stock']],
        on='SKU', how='left', suffixes=('_now', '_prev')
    )
    merged['Stock_prev']   = merged['Stock_prev'].fillna(0)
    merged['Stock_Change'] = merged['Stock_now'] - merged['Stock_prev']

    print(f"[INFO] Merged data: {len(merged)} SKUs")
    return merged, latest_date, prev_date


# ═════════════════════════════════════════════
# PRICE DASHBOARD SHEETS (1~6)
# ═════════════════════════════════════════════

def build_summary(wb, df):
    sn = 'Dashboard_Summary'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].dropna().unique())
    latest = dates[-1]
    prev   = dates[-2] if len(dates) >= 2 else None
    dl     = df[df['Scraped_At'] == latest]
    dp     = df[df['Scraped_At'] == prev] if prev else pd.DataFrame()

    ws.merge_cells('A1:L1')
    ws['A1'] = 'Al Manea — AC Price Tracking Dashboard'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = (f"Latest: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
                f"Total SKUs: {len(dl)}  |  "
                f"Brands: {dl['Brand'].nunique()}")
    ws['A2'].font = SUB_FONT

    r = 4
    ws.merge_cells(f'A{r}:L{r}')
    ws[f'A{r}'] = '▶  Price KPI by Category / Compressor / Function / Capacity'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    kpi_hdrs = ['Category', 'Compressor Type', 'Function', 'Capacity',
                'SKUs', 'Avg Orig (SAR)', 'Avg Promo (SAR)', 'Avg Disc %',
                'LG SKUs', 'LG Avg Promo', 'LG vs Mkt (SAR)', 'LG vs Mkt %']
    for ci, h in enumerate(kpi_hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(kpi_hdrs))
    r += 1

    grand = 0
    for cat in CAT_ORDER:
        dc = dl[dl['Category'] == cat]
        if dc.empty:
            continue

        mkt_avg = round(dc['Final_Price'].mean()) if dc['Final_Price'].notna().any() else None
        lg_dc   = dc[dc['Brand'].str.upper() == 'LG']
        lg_avg  = round(lg_dc['Final_Price'].mean()) if not lg_dc.empty and lg_dc['Final_Price'].notna().any() else None
        gap_sar = (lg_avg - mkt_avg) if (lg_avg and mkt_avg) else None
        gap_pct = (gap_sar / mkt_avg) if (gap_sar is not None and mkt_avg) else None

        ws.cell(row=r, column=1, value=cat).font = BOLD_FONT
        for ci in range(1, len(kpi_hdrs) + 1):
            ws.cell(row=r, column=ci).fill = CAT_FILL
        ws.cell(row=r, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=r, column=5).alignment = CTR
        if dc['Original_Price'].notna().any():
            ws.cell(row=r, column=6, value=round(dc['Original_Price'].mean())).number_format = '#,##0'
            ws.cell(row=r, column=6).font = BOLD_FONT
        if mkt_avg:
            ws.cell(row=r, column=7, value=mkt_avg).number_format = '#,##0'
            ws.cell(row=r, column=7).font = BOLD_FONT
        ws.cell(row=r, column=8, value=_pct_str(dc['Discount_Pct'].mean())).font = BOLD_FONT
        ws.cell(row=r, column=8).alignment = CTR
        ws.cell(row=r, column=9, value=len(lg_dc)).font = BOLD_FONT
        ws.cell(row=r, column=9).alignment = CTR
        if lg_avg:
            ws.cell(row=r, column=10, value=lg_avg).number_format = '#,##0'
            ws.cell(row=r, column=10).font = BOLD_FONT
        if gap_sar is not None:
            c11 = ws.cell(row=r, column=11, value=gap_sar)
            c11.number_format = '#,##0'
            c11.font = UP_FONT if gap_sar > 0 else (DOWN_FONT if gap_sar < 0 else BOLD_FONT)
        if gap_pct is not None:
            c12 = ws.cell(row=r, column=12, value=gap_pct)
            c12.number_format = '0%'
            c12.font = UP_FONT if gap_pct > 0 else (DOWN_FONT if gap_pct < 0 else BOLD_FONT)
        for ci in [10, 11, 12]:
            ws.cell(row=r, column=ci).alignment = CTR
        grand += len(dc)
        r += 1

        grp_iter = dc.groupby(['Compressor_Type', 'Function', 'Capacity_Ton'], dropna=False)
        for (comp, func, ton), grp in sorted(
                grp_iter, key=lambda x: (str(x[0][0]), str(x[0][1]),
                                         x[0][2] if pd.notna(x[0][2]) else 9999)):
            mkt = round(grp['Final_Price'].mean()) if grp['Final_Price'].notna().any() else None
            lg_g = grp[grp['Brand'].str.upper() == 'LG']
            lg_m = round(lg_g['Final_Price'].mean()) if not lg_g.empty and lg_g['Final_Price'].notna().any() else None
            gs = (lg_m - mkt) if (lg_m and mkt) else None
            gp = (gs / mkt)   if (gs is not None and mkt) else None

            ws.cell(row=r, column=2, value=comp or '-').font = DATA_FONT
            ws.cell(row=r, column=3, value=func or '-').font = DATA_FONT
            ws.cell(row=r, column=4, value=_fmt_ton(ton)).font = DATA_FONT
            ws.cell(row=r, column=4).alignment = CTR
            ws.cell(row=r, column=5, value=len(grp)).font = NUM_FONT
            ws.cell(row=r, column=5).alignment = CTR
            if grp['Original_Price'].notna().any():
                ws.cell(row=r, column=6, value=round(grp['Original_Price'].mean())).number_format = '#,##0'
                ws.cell(row=r, column=6).font = NUM_FONT
            if mkt:
                ws.cell(row=r, column=7, value=mkt).number_format = '#,##0'
                ws.cell(row=r, column=7).font = NUM_FONT
            ws.cell(row=r, column=8, value=_pct_str(grp['Discount_Pct'].mean())).font = NUM_FONT
            ws.cell(row=r, column=8).alignment = CTR
            ws.cell(row=r, column=9, value=len(lg_g)).font = NUM_FONT
            ws.cell(row=r, column=9).alignment = CTR
            if lg_m:
                ws.cell(row=r, column=10, value=lg_m).number_format = '#,##0'
                ws.cell(row=r, column=10).font = NUM_FONT
            if gs is not None:
                cg = ws.cell(row=r, column=11, value=gs)
                cg.number_format = '#,##0'
                cg.font = UP_FONT if gs > 0 else (DOWN_FONT if gs < 0 else NUM_FONT)
            if gp is not None:
                cgp = ws.cell(row=r, column=12, value=gp)
                cgp.number_format = '0%'
                cgp.font = UP_FONT if gp > 0 else (DOWN_FONT if gp < 0 else NUM_FONT)
            for ci in [10, 11, 12]:
                ws.cell(row=r, column=ci).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(kpi_hdrs) + 1):
                    ws.cell(row=r, column=ci).fill = LIGHT_FILL
            r += 1

    for ci in range(1, len(kpi_hdrs) + 1):
        ws.cell(row=r, column=ci).fill = CAT_FILL
    ws.cell(row=r, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=r, column=5, value=grand).font = BOLD_FONT
    ws.cell(row=r, column=5).alignment = CTR
    r += 2

    ws.merge_cells(f'A{r}:L{r}')
    ws[f'A{r}'] = '▶  Category Distribution by Brand'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    brands_top = dl['Brand'].replace('', pd.NA).dropna().value_counts().head(8).index.tolist()
    if 'LG' not in [b.upper() for b in brands_top]:
        brands_top.append('LG')
    dist_hdrs = ['Category', 'Compressor Type', 'Function', 'Capacity', 'Total'] + brands_top
    for ci, h in enumerate(dist_hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(dist_hdrs))
    r += 1

    for cat in CAT_ORDER:
        dc = dl[dl['Category'] == cat]
        if dc.empty:
            continue
        for ci in range(1, len(dist_hdrs) + 1):
            ws.cell(row=r, column=ci).fill = CAT_FILL
        ws.cell(row=r, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=r, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=r, column=5).alignment = CTR
        for bi, brand in enumerate(brands_top):
            cnt = len(dc[dc['Brand'].str.upper() == brand.upper()])
            ws.cell(row=r, column=6 + bi, value=cnt if cnt else '-').font = BOLD_FONT
            ws.cell(row=r, column=6 + bi).alignment = CTR
        r += 1

        for (comp, func, ton), grp in sorted(
                dc.groupby(['Compressor_Type', 'Function', 'Capacity_Ton'], dropna=False),
                key=lambda x: (str(x[0][0]), str(x[0][1]),
                               x[0][2] if pd.notna(x[0][2]) else 9999)):
            ws.cell(row=r, column=2, value=comp or '-').font = DATA_FONT
            ws.cell(row=r, column=3, value=func or '-').font = DATA_FONT
            ws.cell(row=r, column=4, value=_fmt_ton(ton)).font = DATA_FONT
            ws.cell(row=r, column=4).alignment = CTR
            ws.cell(row=r, column=5, value=len(grp)).font = NUM_FONT
            ws.cell(row=r, column=5).alignment = CTR
            for bi, brand in enumerate(brands_top):
                cnt = len(grp[grp['Brand'].str.upper() == brand.upper()])
                ws.cell(row=r, column=6 + bi, value=cnt if cnt else '-').font = NUM_FONT
                ws.cell(row=r, column=6 + bi).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(dist_hdrs) + 1):
                    ws.cell(row=r, column=ci).fill = LIGHT_FILL
            r += 1

    for ci in range(1, len(dist_hdrs) + 1):
        ws.cell(row=r, column=ci).fill = CAT_FILL
    ws.cell(row=r, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=r, column=5, value=len(dl)).font = BOLD_FONT
    ws.cell(row=r, column=5).alignment = CTR
    for bi, brand in enumerate(brands_top):
        cnt = len(dl[dl['Brand'].str.upper() == brand.upper()])
        ws.cell(row=r, column=6 + bi, value=cnt).font = BOLD_FONT
        ws.cell(row=r, column=6 + bi).alignment = CTR
    r += 2

    if not dp.empty:
        ws.merge_cells(f'A{r}:L{r}')
        ws[f'A{r}'] = f'▶  Changes vs Previous ({pd.Timestamp(prev).strftime("%Y-%m-%d")})'
        ws[f'A{r}'].font = SEC_FONT
        r += 2

        prev_skus = set(dp['SKU'].astype(str))
        curr_skus = set(dl['SKU'].astype(str))
        mg = pd.merge(
            dp[['SKU', 'Final_Price']].astype({'SKU': str}),
            dl[['SKU', 'Final_Price']].astype({'SKU': str}),
            on='SKU', suffixes=('_p', '_c')
        )
        mg['Chg'] = mg['Final_Price_c'] - mg['Final_Price_p']

        items = [
            ('New SKUs',     len(curr_skus - prev_skus)),
            ('Removed SKUs', len(prev_skus - curr_skus)),
            ('Price Up ▲',   len(mg[mg['Chg'] > 0])),
            ('Price Down ▼', len(mg[mg['Chg'] < 0])),
            ('No Change',    len(mg[mg['Chg'] == 0])),
        ]
        ws.cell(row=r, column=1, value='Item')
        ws.cell(row=r, column=2, value='Count')
        _hdr(ws, r, 2)
        for i, (label, cnt) in enumerate(items):
            rr = r + 1 + i
            ws.cell(row=rr, column=1, value=label).font = DATA_FONT
            c = ws.cell(row=rr, column=2, value=cnt)
            c.alignment = CTR
            c.font = (UP_FONT if '▲' in label and cnt > 0 else
                      DOWN_FONT if '▼' in label and cnt > 0 else NUM_FONT)

    _auto_width(ws)
    ws.sheet_properties.tabColor = '1F4E79'


def build_price_alert(wb, df):
    sn = 'Price_Change_Alert'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = '※ 가격 변동 비교를 위해 최소 2회 이상 수집 데이터가 필요합니다.'
        ws['A1'].font = SUB_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev]
    dc = df[df['Scraped_At'] == latest]

    cols = ['SKU', 'Brand', 'Product_Name', 'Category',
            'Compressor_Type', 'Function', 'Capacity_Ton', 'Final_Price']
    mg = pd.merge(
        dp[cols].astype({'SKU': str}),
        dc[cols].astype({'SKU': str}),
        on='SKU', suffixes=('_p', '_c')
    )
    mg['Chg_SAR'] = mg['Final_Price_c'] - mg['Final_Price_p']
    mg['Chg_Pct'] = mg['Chg_SAR'] / mg['Final_Price_p']
    changed = mg[mg['Chg_SAR'] != 0].sort_values('Chg_Pct')

    ws.merge_cells('A1:K1')
    ws['A1'] = 'Price Change Alert'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = (f"{pd.Timestamp(prev).strftime('%Y-%m-%d')} → "
                f"{pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
                f"Changed: {len(changed)} SKUs")
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = ['Brand', 'SKU', 'Product Name', 'Category',
            'Compressor', 'Function', 'Capacity',
            'Prev Price', 'Curr Price', 'Change (SAR)', 'Change %']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(hdrs))

    for i, (_, row_data) in enumerate(changed.iterrows()):
        rr = r + 1 + i
        ws.cell(rr, 1, row_data['Brand_c']).font = DATA_FONT
        ws.cell(rr, 2, row_data['SKU']).font = DATA_FONT
        ws.cell(rr, 3, row_data['Product_Name_c']).font = DATA_FONT
        ws.cell(rr, 3).alignment = WRAP
        ws.cell(rr, 4, row_data['Category_c']).font = DATA_FONT
        ws.cell(rr, 5, row_data['Compressor_Type_c']).font = DATA_FONT
        ws.cell(rr, 5).alignment = CTR
        ws.cell(rr, 6, row_data['Function_c']).font = DATA_FONT
        ws.cell(rr, 7, _fmt_ton(row_data['Capacity_Ton_c'])).font = DATA_FONT
        ws.cell(rr, 7).alignment = CTR
        ws.cell(rr, 8, row_data['Final_Price_p']).number_format = '#,##0'
        ws.cell(rr, 8).font = NUM_FONT
        ws.cell(rr, 9, row_data['Final_Price_c']).number_format = '#,##0'
        ws.cell(rr, 9).font = NUM_FONT
        chg = row_data['Chg_SAR']
        cg = ws.cell(rr, 10, chg)
        cg.number_format = '#,##0'
        cg.font = DOWN_FONT if chg < 0 else UP_FONT
        cp = ws.cell(rr, 11, row_data['Chg_Pct'])
        cp.number_format = '0%'
        cp.font = DOWN_FONT if chg < 0 else UP_FONT
        if rr % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = LIGHT_FILL

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.sheet_properties.tabColor = 'FF6600'


def build_new_discontinued(wb, df):
    sn = 'New_Discontinued_SKUs'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = '※ 최소 2회 이상 수집 데이터가 필요합니다.'
        ws['A1'].font = SUB_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev]
    dc = df[df['Scraped_At'] == latest]
    new_skus  = set(dc['SKU'].astype(str)) - set(dp['SKU'].astype(str))
    disc_skus = set(dp['SKU'].astype(str)) - set(dc['SKU'].astype(str))

    ws.merge_cells('A1:H1')
    ws['A1'] = 'New & Discontinued SKUs'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:H2')
    ws['A2'] = (f"{pd.Timestamp(prev).strftime('%Y-%m-%d')} → "
                f"{pd.Timestamp(latest).strftime('%Y-%m-%d')}")
    ws['A2'].font = SUB_FONT

    hdrs = ['Brand', 'SKU', 'Product Name', 'Category',
            'Compressor', 'Function', 'Capacity', 'AlAhli Price']

    def _write_section(title, title_font, skus, src_df, fill, row_font, start):
        ws.merge_cells(f'A{start}:H{start}')
        ws[f'A{start}'] = title
        ws[f'A{start}'].font = title_font
        hr = start + 1
        for ci, h in enumerate(hdrs, 1):
            ws.cell(hr, ci, h)
        _hdr(ws, hr, len(hdrs))
        sub = src_df[src_df['SKU'].astype(str).isin(skus)]
        for i, (_, rd) in enumerate(sub.iterrows()):
            rr = hr + 1 + i
            ws.cell(rr, 1, rd.get('Brand', '')).font = row_font
            ws.cell(rr, 2, str(rd['SKU'])).font = row_font
            ws.cell(rr, 3, rd.get('Product_Name', '')).font = DATA_FONT
            ws.cell(rr, 3).alignment = WRAP
            ws.cell(rr, 4, rd.get('Category', '')).font = DATA_FONT
            ws.cell(rr, 5, rd.get('Compressor_Type', '')).font = DATA_FONT
            ws.cell(rr, 5).alignment = CTR
            ws.cell(rr, 6, rd.get('Function', '')).font = DATA_FONT
            ws.cell(rr, 7, _fmt_ton(rd.get('Capacity_Ton', ''))).font = DATA_FONT
            ws.cell(rr, 7).alignment = CTR
            alahli = rd.get('AlAhli_Price', '')
            c = ws.cell(rr, 8, int(alahli) if pd.notna(alahli) and alahli != '' else '')
            c.number_format = '#,##0'
            c.font = NUM_FONT
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = fill
        return hr + 1 + len(sub)

    r = 4
    end = _write_section(
        f'New SKUs ({len(new_skus)})',
        Font(name='Arial', bold=True, size=12, color='375623'),
        new_skus, dc, GREEN_FILL, NEW_FONT, r
    )
    _write_section(
        f'Discontinued SKUs ({len(disc_skus)})',
        Font(name='Arial', bold=True, size=12, color='C00000'),
        disc_skus, dp, RED_FILL, DISC_FONT, end + 2
    )

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.sheet_properties.tabColor = '00B050'


def build_brand_compare(wb, df):
    sn = 'Brand_Avg_Price_Compare'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].dropna().unique())
    latest = dates[-1]
    dl     = df[df['Scraped_At'] == latest]

    brands_top = dl['Brand'].replace('', pd.NA).dropna().value_counts().head(7).index.tolist()
    if 'LG' not in [b.upper() for b in brands_top]:
        brands_top.append('LG')

    ws.merge_cells('A1:L1')
    ws['A1'] = 'Brand Average Price Comparison by Category & Capacity'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = (f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
                f"Price basis: AlAhli Price (avg) / SAR")
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = ['Category', 'Compressor', 'Function', 'Capacity'] + brands_top + ['Market Avg', 'LG Min']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h)
    _hdr(ws, r, len(hdrs))
    r += 1

    for cat in CAT_ORDER:
        dc = dl[dl['Category'] == cat]
        if dc.empty:
            continue
        for ci in range(1, len(hdrs) + 1):
            ws.cell(r, ci).fill = CAT_FILL
        ws.cell(r, 1, cat).font = BOLD_FONT
        r += 1

        for (comp, func, ton), grp in sorted(
                dc.groupby(['Compressor_Type', 'Function', 'Capacity_Ton'], dropna=False),
                key=lambda x: (str(x[0][0]), str(x[0][1]),
                               x[0][2] if pd.notna(x[0][2]) else 9999)):
            ws.cell(r, 2, comp or '-').font = DATA_FONT
            ws.cell(r, 3, func or '-').font = DATA_FONT
            ws.cell(r, 4, _fmt_ton(ton)).font = DATA_FONT
            ws.cell(r, 4).alignment = CTR
            mkt = grp['Final_Price'].mean()
            for bi, brand in enumerate(brands_top):
                bg = grp[grp['Brand'].str.upper() == brand.upper()]
                if not bg.empty and bg['Final_Price'].notna().any():
                    c = ws.cell(r, 5 + bi, round(bg['Final_Price'].mean()))
                    c.number_format = '#,##0'
                    c.font = NUM_FONT
                else:
                    ws.cell(r, 5 + bi, '-').font = NUM_FONT
                ws.cell(r, 5 + bi).alignment = CTR
            mc = ws.cell(r, 5 + len(brands_top), round(mkt) if pd.notna(mkt) else '-')
            mc.number_format = '#,##0'
            mc.font = BOLD_FONT
            mc.alignment = CTR
            lg_g    = grp[grp['Brand'].str.upper() == 'LG']
            lc_col  = 6 + len(brands_top)
            if not lg_g.empty and lg_g['Final_Price'].notna().any():
                lm = ws.cell(r, lc_col, round(lg_g['Final_Price'].min()))
                lm.number_format = '#,##0'
                lm.font = Font(name='Arial', size=10, color='0070C0', bold=True)
            else:
                ws.cell(r, lc_col, '-').font = NUM_FONT
            ws.cell(r, lc_col).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(hdrs) + 1):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = 'FFC000'


def build_full_tracking(wb, df):
    sn = 'Full_Price_Tracking'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].dropna().unique())
    latest = dates[-1]
    prev   = dates[-2] if len(dates) >= 2 else None
    dc     = df[df['Scraped_At'] == latest].copy()
    prev_map = {}
    if prev is not None:
        dp = df[df['Scraped_At'] == prev]
        prev_map = dp.set_index(dp['SKU'].astype(str))['Final_Price'].to_dict()

    ws.merge_cells('A1:Q1')
    ws['A1'] = 'Full Price Tracking — Latest Snapshot'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:Q2')
    ws['A2'] = (f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
                f"Total: {len(dc)} SKUs")
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = ['Brand', 'SKU', 'Model', 'Product Name',
            'Category', 'Compressor', 'Function', 'Capacity', 'BTU',
            'Orig Price', 'AlAhli Price', 'Disc %',
            'Prev Price', 'Change', 'Has Offer', 'Offer Detail', 'Stock']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h)
    _hdr(ws, r, len(hdrs))

    dc_sorted = dc.sort_values(
        ['Category', 'Compressor_Type', 'Function', 'Capacity_Ton', 'Brand'],
        na_position='last'
    )

    for i, (_, rd) in enumerate(dc_sorted.iterrows()):
        rr  = r + 1 + i
        sku = str(rd['SKU'])

        ws.cell(rr, 1, rd.get('Brand', '')).font = DATA_FONT
        ws.cell(rr, 2, sku).font = DATA_FONT
        ws.cell(rr, 3, rd.get('Model', '')).font = DATA_FONT
        ws.cell(rr, 4, rd.get('Product_Name', '')).font = DATA_FONT
        ws.cell(rr, 4).alignment = WRAP
        ws.cell(rr, 5, rd.get('Category', '')).font = DATA_FONT
        ws.cell(rr, 6, rd.get('Compressor_Type', '')).font = DATA_FONT
        ws.cell(rr, 6).alignment = CTR
        ws.cell(rr, 7, rd.get('Function', '')).font = DATA_FONT
        ws.cell(rr, 8, _fmt_ton(rd.get('Capacity_Ton', ''))).font = DATA_FONT
        ws.cell(rr, 8).alignment = CTR
        btu = rd.get('BTU', '')
        ws.cell(rr, 9, int(btu) if pd.notna(btu) and btu != '' else '').font = NUM_FONT
        ws.cell(rr, 9).number_format = '#,##0'
        ws.cell(rr, 9).alignment = CTR

        orig   = rd.get('Original_Price', '')
        alahli = rd.get('AlAhli_Price', '')
        ws.cell(rr, 10, int(orig)   if pd.notna(orig)   and orig   != '' else '').number_format = '#,##0'
        ws.cell(rr, 10).font = NUM_FONT
        ws.cell(rr, 11, int(alahli) if pd.notna(alahli) and alahli != '' else '').number_format = '#,##0'
        ws.cell(rr, 11).font = NUM_FONT

        disc = rd.get('Discount_Pct', '')
        if pd.notna(disc) and disc != '':
            d_cell = ws.cell(rr, 12, float(disc))
            d_cell.number_format = '0%'
            d_cell.font = NUM_FONT
        ws.cell(rr, 12).alignment = CTR

        prev_price = prev_map.get(sku)
        curr_price = rd.get('Final_Price')
        if prev_price and pd.notna(prev_price):
            ws.cell(rr, 13, int(prev_price)).number_format = '#,##0'
            ws.cell(rr, 13).font = NUM_FONT
            if curr_price and pd.notna(curr_price):
                chg = curr_price - prev_price
                cg = ws.cell(rr, 14, int(chg))
                cg.number_format = '#,##0'
                cg.font = DOWN_FONT if chg < 0 else (UP_FONT if chg > 0 else NUM_FONT)
        else:
            ws.cell(rr, 13, 'NEW').font = NEW_FONT
            ws.cell(rr, 13).alignment = CTR

        ws.cell(rr, 15, rd.get('Has_Offer', '')).font = DATA_FONT
        ws.cell(rr, 15).alignment = CTR
        ws.cell(rr, 16, rd.get('Offer_Detail', '')).font = DATA_FONT
        ws.cell(rr, 16).alignment = WRAP
        stk = rd.get('Stock', '')
        ws.cell(rr, 17, int(stk) if pd.notna(stk) and stk != '' else '').font = NUM_FONT
        ws.cell(rr, 17).number_format = '#,##0'
        ws.cell(rr, 17).alignment = CTR

        if rr % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = LIGHT_FILL

    _auto_width(ws, max_w=50)
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['P'].width = 50
    ws.freeze_panes = 'A5'
    ws.sheet_properties.tabColor = '4472C4'


def build_offer_analysis(wb, df):
    sn = 'Offer_Analysis'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped_At'].dropna().unique())
    latest = dates[-1]
    dl     = df[df['Scraped_At'] == latest]

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Offer & Free Gift Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}"
    ws['A2'].font = SUB_FONT

    r = 4
    ws[f'A{r}'] = '▶  Offer Status'
    ws[f'A{r}'].font = SEC_FONT
    r += 2
    ws.cell(r, 1, 'Has Offer')
    ws.cell(r, 2, 'SKU Count')
    ws.cell(r, 3, '% of Total')
    _hdr(ws, r, 3)
    r += 1
    for val in ['Yes', 'No']:
        cnt = len(dl[dl['Has_Offer'] == val])
        ws.cell(r, 1, val).font = DATA_FONT
        ws.cell(r, 2, cnt).font = NUM_FONT
        ws.cell(r, 2).alignment = CTR
        ws.cell(r, 3, f"{cnt/len(dl)*100:.0f}%").font = NUM_FONT
        ws.cell(r, 3).alignment = CTR
        r += 1
    r += 1

    ws[f'A{r}'] = '▶  Offer by Brand'
    ws[f'A{r}'].font = SEC_FONT
    r += 2
    ws.cell(r, 1, 'Brand')
    ws.cell(r, 2, 'Total SKUs')
    ws.cell(r, 3, 'SKUs w/ Offer')
    ws.cell(r, 4, 'Offer Rate %')
    _hdr(ws, r, 4)
    r += 1
    brand_summary = (
        dl.groupby('Brand')
          .agg(total=('SKU', 'count'),
               with_offer=('Has_Offer', lambda x: (x == 'Yes').sum()))
          .reset_index()
          .sort_values('total', ascending=False)
    )
    for _, brd in brand_summary.iterrows():
        ws.cell(r, 1, brd['Brand']).font = DATA_FONT
        ws.cell(r, 2, brd['total']).font = NUM_FONT
        ws.cell(r, 2).alignment = CTR
        ws.cell(r, 3, brd['with_offer']).font = NUM_FONT
        ws.cell(r, 3).alignment = CTR
        pct = brd['with_offer'] / brd['total'] if brd['total'] else 0
        ws.cell(r, 4, f"{pct*100:.0f}%").font = NUM_FONT
        ws.cell(r, 4).alignment = CTR
        if r % 2 == 0:
            for ci in range(1, 5):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1
    r += 1

    if 'Free_Gift' in dl.columns:
        has_gift = dl[dl['Free_Gift'].notna() & (dl['Free_Gift'] != '') & (dl['Free_Gift'] != 'N/A')]
        if not has_gift.empty:
            ws[f'A{r}'] = f'▶  Free Gift SKUs ({len(has_gift)}개)'
            ws[f'A{r}'].font = SEC_FONT
            r += 2
            ws.cell(r, 1, 'Brand')
            ws.cell(r, 2, 'SKU')
            ws.cell(r, 3, 'Product Name')
            ws.cell(r, 4, 'Free Gift')
            ws.cell(r, 5, 'AlAhli Price')
            _hdr(ws, r, 5)
            r += 1
            for _, rd in has_gift.iterrows():
                ws.cell(r, 1, rd.get('Brand', '')).font = DATA_FONT
                ws.cell(r, 2, str(rd['SKU'])).font = DATA_FONT
                ws.cell(r, 3, rd.get('Product_Name', '')).font = DATA_FONT
                ws.cell(r, 3).alignment = WRAP
                ws.cell(r, 4, rd.get('Free_Gift', '')).font = DATA_FONT
                alahli = rd.get('AlAhli_Price', '')
                c = ws.cell(r, 5, int(alahli) if pd.notna(alahli) and alahli != '' else '')
                c.number_format = '#,##0'
                c.font = NUM_FONT
                r += 1
            r += 1

    with_offer = dl[dl['Has_Offer'] == 'Yes'][['Brand', 'Product_Name', 'Category', 'Offer_Detail']].head(30)
    if not with_offer.empty:
        ws[f'A{r}'] = '▶  Offer Detail Sample (up to 30)'
        ws[f'A{r}'].font = SEC_FONT
        r += 2
        ws.cell(r, 1, 'Brand')
        ws.cell(r, 2, 'Category')
        ws.cell(r, 3, 'Product Name')
        ws.cell(r, 4, 'Offer Detail')
        _hdr(ws, r, 4)
        r += 1
        for _, rd in with_offer.iterrows():
            ws.cell(r, 1, rd.get('Brand', '')).font = DATA_FONT
            ws.cell(r, 2, rd.get('Category', '')).font = DATA_FONT
            ws.cell(r, 3, rd.get('Product_Name', '')).font = DATA_FONT
            ws.cell(r, 3).alignment = WRAP
            ws.cell(r, 4, rd.get('Offer_Detail', '')).font = DATA_FONT
            ws.cell(r, 4).alignment = WRAP
            if r % 2 == 0:
                for ci in range(1, 5):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 55
    ws.sheet_properties.tabColor = '7030A0'


# ═════════════════════════════════════════════
# STOCK DASHBOARD SHEETS (7~9)
# ═════════════════════════════════════════════

def build_stock_summary(wb, merged, latest_date, prev_date):
    sn = 'Stock_Dashboard_Summary'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    r = 1
    ws.cell(r, 1, 'Stock Dashboard - Executive Summary').font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, f"Analysis Period: {prev_date.strftime('%Y-%m-%d')} -> {latest_date.strftime('%Y-%m-%d')}").font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    ws.cell(r, 1, '▣ Key Performance Indicators').font = SEC_FONT
    r += 1

    total_stock_now  = merged['Stock_now'].sum()
    total_stock_prev = merged['Stock_prev'].sum()
    stock_change     = total_stock_now - total_stock_prev
    increased_skus   = len(merged[merged['Stock_Change'] > 0])
    decreased_skus   = len(merged[merged['Stock_Change'] < 0])
    out_of_stock     = len(merged[merged['Stock_now'] == 0])
    critical_stock   = len(merged[merged['Stock_now'].between(1, 50)])
    low_stock        = len(merged[merged['Stock_now'].between(51, 100)])
    ok_stock         = len(merged[merged['Stock_now'].between(101, 230)])
    high_stock       = len(merged[merged['Stock_now'] >= 231])

    kpi_data = [
        ('Total Stock',          int(total_stock_now), int(stock_change)),
        ('Out of Stock (0)',      out_of_stock,         None),
        ('Critical Stock (1~50)', critical_stock,       None),
        ('Low Stock (51~100)',    low_stock,             None),
        ('OK Stock (101~230)',    ok_stock,              None),
        ('High Stock (231+)',     high_stock,            None),
        ('Stock Increased',       increased_skus,        None),
        ('Stock Decreased',       decreased_skus,        None),
    ]

    ws.cell(r, 1, 'Metric').font = BOLD_FONT
    ws.cell(r, 2, 'Current').font = BOLD_FONT
    ws.cell(r, 3, 'Change').font = BOLD_FONT
    _hdr(ws, r, 3)
    r += 1

    for metric, current, change in kpi_data:
        ws.cell(r, 1, metric).font = DATA_FONT
        c2 = ws.cell(r, 2, current)
        c2.font = NUM_FONT; c2.alignment = CTR; c2.number_format = '#,##0'
        if change is not None:
            c3 = ws.cell(r, 3, change)
            c3.font = STOCK_UP if change > 0 else STOCK_DOWN if change < 0 else NUM_FONT
            c3.alignment = CTR; c3.number_format = '+#,##0;-#,##0;0'
        if r % 2 == 0:
            for ci in range(1, 4):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    r += 1
    ws.cell(r, 1, '▣ Stock by Category').font = SEC_FONT
    r += 1

    cat_summary = merged.groupby('Category').agg({
        'Stock_now': 'sum', 'Stock_prev': 'sum', 'SKU': 'count'
    }).reset_index()
    cat_summary['Change'] = cat_summary['Stock_now'] - cat_summary['Stock_prev']
    cat_summary = cat_summary.sort_values('Stock_now', ascending=False)

    ws.cell(r, 1, 'Category').font = BOLD_FONT
    ws.cell(r, 2, 'Current Stock').font = BOLD_FONT
    ws.cell(r, 3, 'Previous Stock').font = BOLD_FONT
    ws.cell(r, 4, 'Change').font = BOLD_FONT
    ws.cell(r, 5, 'SKU Count').font = BOLD_FONT
    _hdr(ws, r, 5)
    r += 1

    for _, row in cat_summary.iterrows():
        ws.cell(r, 1, row['Category']).font = DATA_FONT
        c2 = ws.cell(r, 2, int(row['Stock_now']))
        c2.font = NUM_FONT; c2.alignment = CTR; c2.number_format = '#,##0'
        c3 = ws.cell(r, 3, int(row['Stock_prev']))
        c3.font = NUM_FONT; c3.alignment = CTR; c3.number_format = '#,##0'
        change_val = int(row['Change'])
        c4 = ws.cell(r, 4, change_val)
        c4.font = STOCK_UP if change_val > 0 else STOCK_DOWN if change_val < 0 else NUM_FONT
        c4.alignment = CTR; c4.number_format = '+#,##0;-#,##0;0'
        c5 = ws.cell(r, 5, int(row['SKU']))
        c5.font = NUM_FONT; c5.alignment = CTR; c5.number_format = '#,##0'
        if r % 2 == 0:
            for ci in range(1, 6):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    r += 1
    ws.cell(r, 1, 'Critical Stock Alert (Stock 1~50)').font = SEC_FONT
    r += 1

    critical = merged[merged['Stock_now'].between(1, 50)].sort_values('Stock_now')

    if len(critical) == 0:
        ws.cell(r, 1, 'No critical stock items').font = DATA_FONT
        r += 1
    else:
        ws.cell(r, 1, 'Brand').font = BOLD_FONT
        ws.cell(r, 2, 'Model').font = BOLD_FONT
        ws.cell(r, 3, 'Category').font = BOLD_FONT
        ws.cell(r, 4, 'Current').font = BOLD_FONT
        ws.cell(r, 5, 'Previous').font = BOLD_FONT
        ws.cell(r, 6, 'Change').font = BOLD_FONT
        _hdr(ws, r, 6)
        r += 1
        for _, row in critical.iterrows():
            ws.cell(r, 1, row['Brand']).font = DATA_FONT
            ws.cell(r, 2, row['Model']).font = DATA_FONT
            ws.cell(r, 3, row['Category']).font = DATA_FONT
            c4 = ws.cell(r, 4, int(row['Stock_now']))
            c4.font = NUM_FONT; c4.alignment = CTR; c4.number_format = '#,##0'
            c5 = ws.cell(r, 5, int(row['Stock_prev']))
            c5.font = NUM_FONT; c5.alignment = CTR; c5.number_format = '#,##0'
            change_val = int(row['Stock_Change'])
            c6 = ws.cell(r, 6, change_val)
            c6.font = STOCK_DOWN if change_val < 0 else NUM_FONT
            c6.alignment = CTR; c6.number_format = '+#,##0;-#,##0;0'
            for ci in range(1, 7):
                ws.cell(r, ci).fill = CRITICAL_FILL
            r += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = 'FF6B6B'


def build_stock_brand(wb, merged, latest_date, prev_date):
    sn = 'Stock_Dashboard_Brand'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    r = 1
    ws.cell(r, 1, 'Stock Dashboard - Manager View').font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    ws.cell(r, 1, f"Analysis Period: {prev_date.strftime('%Y-%m-%d')} -> {latest_date.strftime('%Y-%m-%d')}").font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    for cat in CAT_ORDER:
        cat_data = merged[merged['Category'] == cat]
        if len(cat_data) == 0:
            continue

        ws.cell(r, 1, f'▣ {cat}').font = SEC_FONT
        ws.cell(r, 1).fill = CAT_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1

        brand_summary = cat_data.groupby('Brand').agg({
            'Stock_now': 'sum', 'Stock_prev': 'sum', 'SKU': 'count'
        }).reset_index()
        brand_summary['Change'] = brand_summary['Stock_now'] - brand_summary['Stock_prev']
        brand_summary = brand_summary.sort_values('Stock_now', ascending=False)

        cat_total_stock_now  = brand_summary['Stock_now'].sum()
        cat_total_stock_prev = brand_summary['Stock_prev'].sum()

        ws.cell(r, 1, 'Brand').font = BOLD_FONT
        ws.cell(r, 2, 'SKU Count').font = BOLD_FONT
        ws.cell(r, 3, 'Current Stock').font = BOLD_FONT
        ws.cell(r, 4, 'Previous Stock').font = BOLD_FONT
        ws.cell(r, 5, 'Change').font = BOLD_FONT
        ws.cell(r, 6, 'Avg Stock/SKU').font = BOLD_FONT
        ws.cell(r, 7, 'Current %').font = BOLD_FONT
        ws.cell(r, 8, 'Previous %').font = BOLD_FONT
        _hdr(ws, r, 8)
        r += 1

        for _, row in brand_summary.iterrows():
            ws.cell(r, 1, row['Brand']).font = DATA_FONT

            c2 = ws.cell(r, 2, int(row['SKU']))
            c2.font = NUM_FONT; c2.alignment = CTR; c2.number_format = '#,##0'

            c3 = ws.cell(r, 3, int(row['Stock_now']))
            c3.font = NUM_FONT; c3.alignment = CTR; c3.number_format = '#,##0'

            c4 = ws.cell(r, 4, int(row['Stock_prev']))
            c4.font = NUM_FONT; c4.alignment = CTR; c4.number_format = '#,##0'

            change_val = int(row['Change'])
            c5 = ws.cell(r, 5, change_val)
            c5.font = STOCK_UP if change_val > 0 else STOCK_DOWN if change_val < 0 else NUM_FONT
            c5.alignment = CTR; c5.number_format = '+#,##0;-#,##0;0'

            avg_stock = row['Stock_now'] / row['SKU'] if row['SKU'] > 0 else 0
            c6 = ws.cell(r, 6, round(avg_stock, 1))
            c6.font = NUM_FONT; c6.alignment = CTR; c6.number_format = '#,##0.0'

            curr_pct = row['Stock_now'] / cat_total_stock_now if cat_total_stock_now > 0 else 0
            c7 = ws.cell(r, 7, round(curr_pct, 4))
            c7.font = NUM_FONT; c7.alignment = CTR; c7.number_format = '0.0%'

            prev_pct = row['Stock_prev'] / cat_total_stock_prev if cat_total_stock_prev > 0 else 0
            c8 = ws.cell(r, 8, round(prev_pct, 4))
            c8.font = NUM_FONT; c8.alignment = CTR; c8.number_format = '0.0%'

            if r % 2 == 0:
                for ci in range(1, 9):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

        r += 1

    r += 1
    ws.cell(r, 1, '▣ LG vs Competitors Stock Comparison').font = SEC_FONT
    ws.cell(r, 1).fill = GREEN_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

    lg_data = merged[merged['Brand'] == 'LG']
    competitor_brands = ['Gree', 'Samsung']
    comparison = []
    for cat in CAT_ORDER:
        lg_cat   = lg_data[lg_data['Category'] == cat]
        lg_stock = lg_cat['Stock_now'].sum()
        lg_skus  = len(lg_cat)
        for comp_brand in competitor_brands:
            comp_data  = merged[(merged['Brand'] == comp_brand) & (merged['Category'] == cat)]
            comp_stock = comp_data['Stock_now'].sum()
            comp_skus  = len(comp_data)
            if lg_skus > 0 or comp_skus > 0:
                comparison.append({
                    'Category': cat, 'Competitor': comp_brand,
                    'LG_Stock': lg_stock, 'LG_SKUs': lg_skus,
                    'Comp_Stock': comp_stock, 'Comp_SKUs': comp_skus,
                    'Stock_Diff': lg_stock - comp_stock
                })

    ws.cell(r, 1, 'Category').font = BOLD_FONT
    ws.cell(r, 2, 'Competitor').font = BOLD_FONT
    ws.cell(r, 3, 'LG Stock').font = BOLD_FONT
    ws.cell(r, 4, 'LG SKUs').font = BOLD_FONT
    ws.cell(r, 5, 'Comp Stock').font = BOLD_FONT
    ws.cell(r, 6, 'Comp SKUs').font = BOLD_FONT
    ws.cell(r, 7, 'Difference').font = BOLD_FONT
    _hdr(ws, r, 7)
    r += 1

    for item in comparison:
        ws.cell(r, 1, item['Category']).font = DATA_FONT
        ws.cell(r, 2, item['Competitor']).font = DATA_FONT
        c3 = ws.cell(r, 3, int(item['LG_Stock']))
        c3.font = NUM_FONT; c3.alignment = CTR; c3.number_format = '#,##0'
        c4 = ws.cell(r, 4, item['LG_SKUs'])
        c4.font = NUM_FONT; c4.alignment = CTR; c4.number_format = '#,##0'
        c5 = ws.cell(r, 5, int(item['Comp_Stock']))
        c5.font = NUM_FONT; c5.alignment = CTR; c5.number_format = '#,##0'
        c6 = ws.cell(r, 6, item['Comp_SKUs'])
        c6.font = NUM_FONT; c6.alignment = CTR; c6.number_format = '#,##0'
        diff = int(item['Stock_Diff'])
        c7 = ws.cell(r, 7, diff)
        c7.font = STOCK_UP if diff > 0 else STOCK_DOWN if diff < 0 else NUM_FONT
        c7.alignment = CTR; c7.number_format = '+#,##0;-#,##0;0'
        if r % 2 == 0:
            for ci in range(1, 8):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = '4ECDC4'


def build_stock_detail(wb, merged, latest_date, prev_date):
    sn = 'Stock_Dashboard_Detail'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    merged_sorted = merged.sort_values(['Category', 'Stock_Change'], ascending=[True, False])

    r = 1
    ws.cell(r, 1, 'Stock Dashboard - Detail View (Full Data)').font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    ws.cell(r, 1, f"Analysis Period: {prev_date.strftime('%Y-%m-%d')} -> {latest_date.strftime('%Y-%m-%d')}").font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = ['Category', 'Brand', 'Model', 'Product Name', 'Capacity',
               'Current Stock', 'Previous Stock', 'Change', 'Change %', 'Status']
    for idx, hdr in enumerate(headers, 1):
        ws.cell(r, idx, hdr).font = BOLD_FONT
    _hdr(ws, r, len(headers))
    r += 1

    for _, row in merged_sorted.iterrows():
        ws.cell(r, 1, row['Category']).font = DATA_FONT
        ws.cell(r, 2, row['Brand']).font = DATA_FONT
        ws.cell(r, 3, row['Model']).font = DATA_FONT
        ws.cell(r, 4, row['Product_Name']).font = DATA_FONT
        ws.cell(r, 4).alignment = WRAP
        ws.cell(r, 5, _fmt_ton(row['Capacity_Ton'])).font = DATA_FONT
        ws.cell(r, 5).alignment = CTR

        stock_now  = int(row['Stock_now'])
        stock_prev = int(row['Stock_prev'])
        change     = int(row['Stock_Change'])

        c6 = ws.cell(r, 6, stock_now)
        c6.font = NUM_FONT; c6.alignment = CTR; c6.number_format = '#,##0'
        c7 = ws.cell(r, 7, stock_prev)
        c7.font = NUM_FONT; c7.alignment = CTR; c7.number_format = '#,##0'
        c8 = ws.cell(r, 8, change)
        c8.font = STOCK_UP if change > 0 else STOCK_DOWN if change < 0 else NUM_FONT
        c8.alignment = CTR; c8.number_format = '+#,##0;-#,##0;0'

        if stock_prev > 0:
            change_pct = (change / stock_prev) * 100
            c9 = ws.cell(r, 9, round(change_pct, 1) / 100)
            c9.font = STOCK_UP if change_pct > 0 else STOCK_DOWN if change_pct < 0 else NUM_FONT
            c9.alignment = CTR; c9.number_format = '+0.0%;-0.0%;0%'
        else:
            ws.cell(r, 9, "NEW" if stock_now > 0 else "-").font = NEW_FONT if stock_now > 0 else DATA_FONT
        ws.cell(r, 9).alignment = CTR

        if stock_now == 0:
            status      = "OUT OF STOCK"
            status_font = DISC_FONT
            for ci in range(1, len(headers) + 1):
                ws.cell(r, ci).fill = OUT_OF_STOCK_FILL
        elif stock_now <= 50:
            status      = "CRITICAL"
            status_font = Font(name='Arial', size=10, color='C00000', bold=True)
            for ci in range(1, len(headers) + 1):
                ws.cell(r, ci).fill = RED_FILL
        elif stock_now <= 100:
            status      = "LOW"
            status_font = Font(name='Arial', size=10, color='FF9800')
        elif stock_now <= 230:
            status      = "OK"
            status_font = DATA_FONT
        else:
            status      = "HIGH"
            status_font = HIGH_FONT
            for ci in range(1, len(headers) + 1):
                ws.cell(r, ci).fill = HIGH_FILL

        ws.cell(r, 10, status).font = status_font
        ws.cell(r, 10).alignment = CTR

        if r % 2 == 0 and 101 <= stock_now <= 230:
            for ci in range(1, len(headers) + 1):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1

    _auto_width(ws, max_w=60)
    ws.column_dimensions['D'].width = 50
    ws.sheet_properties.tabColor = '95A5A6'


# ═════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Almanea AC Master Dashboard Builder v2.0")
    print("  (Price Dashboard + Stock Dashboard 통합)")
    print("=" * 60)

    if not os.path.exists(TARGET_FILE):
        print(f"[ERROR] 파일 없음: {TARGET_FILE}")
        sys.exit(1)

    print(f"\n[INFO] 파일 로드: {TARGET_FILE}")

    # STEP 0: Products_DB 속성 백필 & 정규화 (대시보드 빌드 전 선행 실행)
    fill_product_attributes()

    # 데이터 로드
    df_price = load_price_data()
    merged, latest_date, prev_date = load_stock_data()

    wb = load_workbook(TARGET_FILE)

    # Price Dashboard (6개)
    price_steps = [
        ("Dashboard_Summary",       lambda: build_summary(wb, df_price)),
        ("Price_Change_Alert",      lambda: build_price_alert(wb, df_price)),
        ("New_Discontinued_SKUs",   lambda: build_new_discontinued(wb, df_price)),
        ("Brand_Avg_Price_Compare", lambda: build_brand_compare(wb, df_price)),
        ("Full_Price_Tracking",     lambda: build_full_tracking(wb, df_price)),
        ("Offer_Analysis",          lambda: build_offer_analysis(wb, df_price)),
    ]

    # Stock Dashboard (3개)
    stock_steps = [
        ("Stock_Dashboard_Summary", lambda: build_stock_summary(wb, merged, latest_date, prev_date)),
        ("Stock_Dashboard_Brand",   lambda: build_stock_brand(wb, merged, latest_date, prev_date)),
        ("Stock_Dashboard_Detail",  lambda: build_stock_detail(wb, merged, latest_date, prev_date)),
    ]

    all_steps = price_steps + stock_steps
    total = len(all_steps)

    print(f"\n[INFO] 대시보드 생성 중... (총 {total}개 시트)")
    for i, (label, func) in enumerate(all_steps, 1):
        print(f"  [{i}/{total}] {label}...")
        try:
            func()
            print(f"        [OK]")
        except Exception as e:
            import traceback
            print(f"        [WARN] 오류 발생: {e}")
            traceback.print_exc()

    # 시트 순서 정렬
    desired_order = (
        [s for s, _ in price_steps] +
        [DB_SHEET] +
        [s for s, _ in stock_steps]
    )
    # 기타 기존 시트는 뒤에 유지
    remaining = [s for s in wb.sheetnames if s not in desired_order]
    final_order = desired_order + remaining

    current = list(wb.sheetnames)
    for i, name in enumerate(final_order):
        if name in current:
            wb.move_sheet(name, offset=i - current.index(name))
            current = list(wb.sheetnames)

    # 구버전 시트 정리
    for old in ['Sheet1', 'Sheet2', 'Products']:
        if old in wb.sheetnames:
            del wb[old]

    wb.save(TARGET_FILE)
    print(f"\n[SUCCESS] 완료! -> {TARGET_FILE}")
    print("=" * 60)
    print("\n생성된 시트 순서:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i:2d}. {name}")
    print("=" * 60)

    try:
        if sys.stdin.isatty():
            input("\n  엔터를 누르면 종료...")
    except (EOFError, OSError):
        pass


if __name__ == "__main__":
    main()
