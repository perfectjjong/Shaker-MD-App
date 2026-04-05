#!/usr/bin/env python3
"""
Extra.com AC - Dashboard Auto Builder v2
- 세분화: Category > Compressor > Cold_or_HC > Ton
- Brand Price Comparison: 평균가, Ton 기준
- Full_Price_Tracking: Ton 기준 정렬
Author: 핍쫑이
"""

import os, sys, re

try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
except ImportError as e:
    print(f"❌ 필수 패키지 설치 필요: {e}")
    print("py -m pip install pandas openpyxl numpy --break-system-packages")
    sys.exit(1)

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__)) if os.path.dirname(os.path.abspath(__file__)) else os.getcwd()
INPUT_FILE = os.path.join(CURRENT_DIR, "extra_ac_Prices_Tracking_Master.xlsx")

# 스타일
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=10, color='666666')
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E79')
DATA_FONT = Font(name='Arial', size=10)
NUM_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
UP_FONT = Font(name='Arial', size=10, color='FF0000', bold=True)
DOWN_FONT = Font(name='Arial', size=10, color='008000', bold=True)
NEW_FONT = Font(name='Arial', size=10, color='0070C0', bold=True)
DISC_FONT = Font(name='Arial', size=10, color='999999')
BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))
THICK_BORDER = Border(bottom=Side(style='medium', color='1F4E79'))
CENTER = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
LIGHT_FILL = PatternFill('solid', fgColor='F2F7FB')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
RED_FILL = PatternFill('solid', fgColor='FCE4EC')
CAT_FILL = PatternFill('solid', fgColor='D6E4F0')


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THICK_BORDER


def auto_width(ws, min_width=10, max_width=35):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def delete_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ── Category Validation ──────────────────────────────────────────────────────
NON_AC_PATTERNS = [r'^SM-[A-Z]\d', r'^SM[A-Z]\d', r'^(iPhone|iPad|Galaxy)',
                   r'^(WM|WF|WW)\d', r'^(RF|RS|RT|RR)\d']
WINDOW_KEYWORDS = ['WINDOW', 'WDV', 'WINDOW AC', 'WINDOW AIR']
WINDOW_MODEL_PREFIXES = ['WDV', 'GJC', 'H182EH', 'H242EH', 'W18', 'W24']
SPLIT_KEYWORDS = ['SPLIT', 'WALL MOUNT', 'WALL-MOUNT']
SPLIT_MODEL_PREFIXES = ['CLW', 'NS', 'NT', 'ND', 'NF', 'LA']
FREESTANDING_KEYWORDS = ['FLOOR STANDING', 'FREE STANDING', 'FREESTANDING', 'FLOOR-STANDING']
FREESTANDING_MODEL_PREFIXES = ['APW', 'APQ', 'FS']


def _is_non_ac(pname, model, brand):
    model_u = (str(model) if pd.notna(model) else '').upper()
    for pat in NON_AC_PATTERNS:
        if re.match(pat, model_u):
            return True
    name_u = (str(pname) if pd.notna(pname) else '').upper()
    ac_kw = ['AC','AIR CONDITIONER','BTU','TON','SPLIT','WINDOW','INVERTER','ROTARY',
             'COMPRESSOR','COOLING','COLD','FLOOR STANDING','FREE STANDING']
    brand_u = (str(brand) if pd.notna(brand) else '').upper()
    if brand_u in ['APPLE','XIAOMI','HUAWEI','OPPO','VIVO','REALME','HONOR'] and not any(k in name_u for k in ac_kw):
        return True
    return False


def _fix_category(cat, pname, model):
    n = (str(pname) if pd.notna(pname) else '').upper()
    m = (str(model) if pd.notna(model) else '').upper()
    is_w = any(k in n for k in WINDOW_KEYWORDS) or any(m.startswith(p) for p in WINDOW_MODEL_PREFIXES)
    is_s = any(k in n for k in SPLIT_KEYWORDS) or any(m.startswith(p) for p in SPLIT_MODEL_PREFIXES)
    is_f = any(k in n for k in FREESTANDING_KEYWORDS) or any(m.startswith(p) for p in FREESTANDING_MODEL_PREFIXES)
    if is_w and not is_s and cat != 'Window Air Conditioner':
        return 'Window Air Conditioner'
    if is_s and not is_w and not is_f and cat != 'Split Air Conditioner':
        return 'Split Air Conditioner'
    if is_f and not is_s and cat != 'Free Standing Air Conditioner':
        return 'Free Standing Air Conditioner'
    return cat


def load_data(filepath):
    df = pd.read_excel(filepath, sheet_name='Prices DB', engine='openpyxl')
    df['Scraped_At'] = pd.to_datetime(df['Scraped_At'])
    for col in ['Standard_Price', 'Sale_Price', 'Jood_Gold_Price']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    df['BTU'] = pd.to_numeric(df['BTU'], errors='coerce')
    df['Cooling_Capacity_Ton'] = pd.to_numeric(df['Cooling_Capacity_Ton'], errors='coerce')
    df['Compressor_Type'] = df['Compressor_Type'].fillna('N/A').astype(str).str.strip()
    df['Compressor_Type'] = df['Compressor_Type'].replace('', 'N/A')
    if 'Discount_Rate' in df.columns:
        df['Discount_Rate'] = df['Discount_Rate'].astype(str).str.replace('%','').str.strip()
        df['Discount_Rate'] = pd.to_numeric(df['Discount_Rate'], errors='coerce')
        # 값이 1 이상이면 퍼센트 단위로 입력된 것 → 소수로 변환
        df.loc[df['Discount_Rate'] >= 1, 'Discount_Rate'] = df.loc[df['Discount_Rate'] >= 1, 'Discount_Rate'] / 100
    df['Promo_Code'] = df['Promo_Code'].fillna('').astype(str).str.strip()
    has_extra10 = df['Promo_Code'].str.lower() == 'extra10'
    df['Final_Sale_Price'] = np.where(has_extra10, np.round(df['Sale_Price'] * 0.9, 0), df['Sale_Price'])
    df['Final_Jood_Gold_Price'] = np.where(has_extra10, np.round(df['Jood_Gold_Price'] * 0.9, 0), df['Jood_Gold_Price'])

    # ── 카테고리 검증/교정 + 비AC 제품 필터링 ──
    non_ac = df.apply(lambda r: _is_non_ac(r.get('Product_Name'), r.get('Model_No'), r.get('Brand')), axis=1)
    removed = non_ac.sum()
    if removed > 0:
        print(f"  🗑️ Removed {removed} non-AC products")
        df = df[~non_ac].copy()
    corrected = 0
    for idx in df.index:
        old_cat = df.at[idx, 'Category']
        new_cat = _fix_category(old_cat, df.at[idx, 'Product_Name'], df.at[idx, 'Model_No'])
        if new_cat != old_cat:
            df.at[idx, 'Category'] = new_cat
            corrected += 1
    if corrected > 0:
        print(f"  🔄 Corrected {corrected} category misclassifications")

    return df


# ============================================================
# 1. Dashboard_Summary
# ============================================================
def build_dashboard_summary(wb, df):
    sn = 'Dashboard_Summary'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn, 0)

    dates = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    prev = dates[-2] if len(dates) >= 2 else None
    dl = df[df['Scraped_At'] == latest]
    dp = df[df['Scraped_At'] == prev] if prev is not None else pd.DataFrame()

    ws.merge_cells('A1:J1')
    ws['A1'] = '🛒 Extra.com AC Price Tracking Dashboard'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:J2')
    ws['A2'] = f'Last Updated: {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A2'].font = SUBTITLE_FONT

    # --- KPI: 세분화 테이블 ---
    r = 4
    ws.merge_cells(f'A{r}:M{r}')
    ws[f'A{r}'] = '📊 Key Performance Indicators (Detailed)'
    ws[f'A{r}'].font = SECTION_FONT

    r = 6
    headers = ['Category', 'Compressor', 'Cold/HC', 'Ton', 'SKU Count', 'Avg Std Price', 'Avg Sale Price', 'Avg Jood Price', 'Avg Discount %', 'Exclusive', 'LG Avg', 'LG Gap', 'LG Gap %']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    cat_order = ['Split Air Conditioner', 'Window Air Conditioner', 'Free Standing Air Conditioner']
    row_idx = r + 1
    grand_total = 0

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue

        # Category 소계 행
        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=row_idx, column=5).alignment = CENTER
        ws.cell(row=row_idx, column=6, value=round(dc['Standard_Price'].mean())).font = BOLD_FONT
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        mkt_avg_cat = round(dc['Final_Sale_Price'].mean())
        ws.cell(row=row_idx, column=7, value=mkt_avg_cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=7).number_format = '#,##0'
        ws.cell(row=row_idx, column=8, value=round(dc['Final_Jood_Gold_Price'].mean())).font = BOLD_FONT
        ws.cell(row=row_idx, column=8).number_format = '#,##0'
        disc_avg = dc['Discount_Rate'].mean()
        ws.cell(row=row_idx, column=9, value=f'{disc_avg*100:.0f}%' if disc_avg < 1 else f'{disc_avg:.0f}%').font = BOLD_FONT
        ws.cell(row=row_idx, column=9).alignment = CENTER
        ws.cell(row=row_idx, column=10, value=len(dc[dc['eXtra_Exclusive']=='Yes'])).font = BOLD_FONT
        ws.cell(row=row_idx, column=10).alignment = CENTER
        # LG vs Market (Category 소계)
        lg_cat = dc[dc['Brand']=='LG']
        if len(lg_cat) > 0:
            lg_avg_cat = round(lg_cat['Final_Sale_Price'].mean())
            gap_sar = lg_avg_cat - mkt_avg_cat
            gap_pct = gap_sar / mkt_avg_cat if mkt_avg_cat else 0
            ws.cell(row=row_idx, column=11, value=lg_avg_cat).font = BOLD_FONT
            ws.cell(row=row_idx, column=11).number_format = '#,##0'
            c_gap = ws.cell(row=row_idx, column=12, value=gap_sar)
            c_gap.number_format = '#,##0'
            c_gap.font = UP_FONT if gap_sar > 0 else DOWN_FONT if gap_sar < 0 else BOLD_FONT
            c_pct = ws.cell(row=row_idx, column=13, value=gap_pct)
            c_pct.number_format = '0%'
            c_pct.font = UP_FONT if gap_sar > 0 else DOWN_FONT if gap_sar < 0 else BOLD_FONT
        else:
            ws.cell(row=row_idx, column=11, value='-').font = BOLD_FONT
            ws.cell(row=row_idx, column=12, value='-').font = BOLD_FONT
            ws.cell(row=row_idx, column=13, value='-').font = BOLD_FONT
        for ci in range(11, 14):
            ws.cell(row=row_idx, column=ci).alignment = CENTER
        for c in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1
        grand_total += len(dc)

        if cat == 'Free Standing Air Conditioner':
            groups = dc.groupby(['Cooling_Capacity_Ton'])
            for ton, grp in sorted(groups, key=lambda x: x[0]):
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
                ws.cell(row=row_idx, column=5).alignment = CENTER
                ws.cell(row=row_idx, column=6, value=round(grp['Standard_Price'].mean())).font = NUM_FONT
                ws.cell(row=row_idx, column=6).number_format = '#,##0'
                mkt_avg = round(grp['Final_Sale_Price'].mean())
                ws.cell(row=row_idx, column=7, value=mkt_avg).font = NUM_FONT
                ws.cell(row=row_idx, column=7).number_format = '#,##0'
                ws.cell(row=row_idx, column=8, value=round(grp['Final_Jood_Gold_Price'].mean())).font = NUM_FONT
                ws.cell(row=row_idx, column=8).number_format = '#,##0'
                d = grp['Discount_Rate'].mean()
                ws.cell(row=row_idx, column=9, value=f'{d*100:.0f}%' if d < 1 else f'{d:.0f}%').font = NUM_FONT
                ws.cell(row=row_idx, column=9).alignment = CENTER
                ws.cell(row=row_idx, column=10, value=len(grp[grp['eXtra_Exclusive']=='Yes'])).font = NUM_FONT
                ws.cell(row=row_idx, column=10).alignment = CENTER
                # LG vs Market
                lg_grp = grp[grp['Brand']=='LG']
                if len(lg_grp) > 0:
                    lg_avg = round(lg_grp['Final_Sale_Price'].mean())
                    gap_s = lg_avg - mkt_avg
                    gap_p = gap_s / mkt_avg if mkt_avg else 0
                    ws.cell(row=row_idx, column=11, value=lg_avg).font = NUM_FONT
                    ws.cell(row=row_idx, column=11).number_format = '#,##0'
                    cg = ws.cell(row=row_idx, column=12, value=gap_s)
                    cg.number_format = '#,##0'
                    cg.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
                    cp = ws.cell(row=row_idx, column=13, value=gap_p)
                    cp.number_format = '0%'
                    cp.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
                else:
                    ws.cell(row=row_idx, column=11, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=12, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=13, value='-').font = NUM_FONT
                for ci in range(11, 14):
                    ws.cell(row=row_idx, column=ci).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1
        else:
            groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
            for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2])):
                ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
                ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
                ws.cell(row=row_idx, column=5).alignment = CENTER
                ws.cell(row=row_idx, column=6, value=round(grp['Standard_Price'].mean())).font = NUM_FONT
                ws.cell(row=row_idx, column=6).number_format = '#,##0'
                mkt_avg = round(grp['Final_Sale_Price'].mean())
                ws.cell(row=row_idx, column=7, value=mkt_avg).font = NUM_FONT
                ws.cell(row=row_idx, column=7).number_format = '#,##0'
                ws.cell(row=row_idx, column=8, value=round(grp['Final_Jood_Gold_Price'].mean())).font = NUM_FONT
                ws.cell(row=row_idx, column=8).number_format = '#,##0'
                d = grp['Discount_Rate'].mean()
                ws.cell(row=row_idx, column=9, value=f'{d*100:.0f}%' if d < 1 else f'{d:.0f}%').font = NUM_FONT
                ws.cell(row=row_idx, column=9).alignment = CENTER
                ws.cell(row=row_idx, column=10, value=len(grp[grp['eXtra_Exclusive']=='Yes'])).font = NUM_FONT
                ws.cell(row=row_idx, column=10).alignment = CENTER
                # LG vs Market
                lg_grp = grp[grp['Brand']=='LG']
                if len(lg_grp) > 0:
                    lg_avg = round(lg_grp['Final_Sale_Price'].mean())
                    gap_s = lg_avg - mkt_avg
                    gap_p = gap_s / mkt_avg if mkt_avg else 0
                    ws.cell(row=row_idx, column=11, value=lg_avg).font = NUM_FONT
                    ws.cell(row=row_idx, column=11).number_format = '#,##0'
                    cg = ws.cell(row=row_idx, column=12, value=gap_s)
                    cg.number_format = '#,##0'
                    cg.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
                    cp = ws.cell(row=row_idx, column=13, value=gap_p)
                    cp.number_format = '0%'
                    cp.font = UP_FONT if gap_s > 0 else DOWN_FONT if gap_s < 0 else NUM_FONT
                else:
                    ws.cell(row=row_idx, column=11, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=12, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=13, value='-').font = NUM_FONT
                for ci in range(11, 14):
                    ws.cell(row=row_idx, column=ci).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1

    # TOTAL
    ws.cell(row=row_idx, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=row_idx, column=5, value=grand_total).font = BOLD_FONT
    ws.cell(row=row_idx, column=5).alignment = CENTER
    for c in range(1, len(headers)+1):
        ws.cell(row=row_idx, column=c).fill = CAT_FILL
    row_idx += 1

    # --- Category Distribution (세분화) ---
    row_idx += 1
    ws.merge_cells(f'A{row_idx}:J{row_idx}')
    ws[f'A{row_idx}'] = '📦 Category Distribution by Brand'
    ws[f'A{row_idx}'].font = SECTION_FONT
    row_idx += 2

    brands_top = dl['Brand'].value_counts().head(8).index.tolist()
    if 'SAMSUNG' not in brands_top:
        brands_top.append('SAMSUNG')
    dist_headers = ['Category', 'Compressor', 'Cold/HC', 'Ton', 'Total'] + brands_top
    for i, h in enumerate(dist_headers):
        ws.cell(row=row_idx, column=i+1, value=h)
    style_header_row(ws, row_idx, len(dist_headers))
    row_idx += 1

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue

        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=row_idx, column=5, value=len(dc)).font = BOLD_FONT
        ws.cell(row=row_idx, column=5).alignment = CENTER
        for bi, brand in enumerate(brands_top):
            cnt = len(dc[dc['Brand']==brand])
            ws.cell(row=row_idx, column=6+bi, value=cnt if cnt > 0 else '-').font = BOLD_FONT
            ws.cell(row=row_idx, column=6+bi).alignment = CENTER
        for c in range(1, len(dist_headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1

        if cat == 'Free Standing Air Conditioner':
            groups = dc.groupby(['Cooling_Capacity_Ton'])
            for ton, grp in sorted(groups, key=lambda x: x[0]):
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
                ws.cell(row=row_idx, column=5).alignment = CENTER
                for bi, brand in enumerate(brands_top):
                    cnt = len(grp[grp['Brand']==brand])
                    ws.cell(row=row_idx, column=6+bi, value=cnt if cnt > 0 else '-').font = NUM_FONT
                    ws.cell(row=row_idx, column=6+bi).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(dist_headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1
        else:
            groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
            for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2])):
                ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
                ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                ws.cell(row=row_idx, column=5, value=len(grp)).font = NUM_FONT
                ws.cell(row=row_idx, column=5).alignment = CENTER
                for bi, brand in enumerate(brands_top):
                    cnt = len(grp[grp['Brand']==brand])
                    ws.cell(row=row_idx, column=6+bi, value=cnt if cnt > 0 else '-').font = NUM_FONT
                    ws.cell(row=row_idx, column=6+bi).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(dist_headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1

    ws.cell(row=row_idx, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=row_idx, column=5, value=len(dl)).font = BOLD_FONT
    ws.cell(row=row_idx, column=5).alignment = CENTER
    for bi, brand in enumerate(brands_top):
        ws.cell(row=row_idx, column=6+bi, value=len(dl[dl['Brand']==brand])).font = BOLD_FONT
        ws.cell(row=row_idx, column=6+bi).alignment = CENTER
    for c in range(1, len(dist_headers)+1):
        ws.cell(row=row_idx, column=c).fill = CAT_FILL
    row_idx += 1

    # --- Changes vs Previous ---
    if not dp.empty:
        row_idx += 1
        ws.merge_cells(f'A{row_idx}:J{row_idx}')
        ws[f'A{row_idx}'] = f'🔄 Changes vs Previous ({pd.Timestamp(prev).strftime("%Y-%m-%d")})'
        ws[f'A{row_idx}'].font = SECTION_FONT
        row_idx += 2

        prev_skus = set(dp['SKU'].astype(str))
        curr_skus = set(dl['SKU'].astype(str))
        merged = pd.merge(
            dp[['SKU','Final_Sale_Price']].astype({'SKU':str}),
            dl[['SKU','Final_Sale_Price']].astype({'SKU':str}),
            on='SKU', suffixes=('_prev','_curr')
        )
        merged['Change'] = merged['Final_Sale_Price_curr'] - merged['Final_Sale_Price_prev']

        items = [
            ('New SKUs', len(curr_skus - prev_skus)),
            ('Removed SKUs', len(prev_skus - curr_skus)),
            ('Price Increased ▲', len(merged[merged['Change']>0])),
            ('Price Decreased ▼', len(merged[merged['Change']<0])),
            ('No Change', len(merged[merged['Change']==0])),
        ]
        ws.cell(row=row_idx, column=1, value='Metric')
        ws.cell(row=row_idx, column=2, value='Count')
        style_header_row(ws, row_idx, 2)
        for i, (m, cnt) in enumerate(items):
            rr = row_idx + 1 + i
            ws.cell(row=rr, column=1, value=m).font = DATA_FONT
            cell = ws.cell(row=rr, column=2, value=cnt)
            cell.alignment = CENTER
            if '▲' in m and cnt > 0: cell.font = UP_FONT
            elif '▼' in m and cnt > 0: cell.font = DOWN_FONT
            else: cell.font = NUM_FONT

    auto_width(ws)
    ws.sheet_properties.tabColor = '1F4E79'


# ============================================================
# 2&3. Price_Change_Alert / Jood_Price_Change_Alert (공통)
# ============================================================
def _build_price_alert(wb, df, sheet_name, price_col, title, tab_color):
    delete_sheet_if_exists(wb, sheet_name)
    ws = wb.create_sheet(sheet_name)

    dates = sorted(df['Scraped_At'].unique())
    if len(dates) < 2:
        ws['A1'] = '⚠️ 최소 2회 이상 스크래핑 데이터가 필요합니다.'
        ws['A1'].font = SUBTITLE_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev].copy()
    dc = df[df['Scraped_At'] == latest].copy()

    cols = ['SKU','Brand','Product_Name','Category','BTU','Cold_or_HC','Compressor_Type','Cooling_Capacity_Ton', price_col]
    merged = pd.merge(dp[cols].astype({'SKU':str}), dc[cols].astype({'SKU':str}), on='SKU', suffixes=('_prev','_curr'))
    merged['Change_SAR'] = merged[f'{price_col}_curr'] - merged[f'{price_col}_prev']
    merged['Change_Pct'] = merged['Change_SAR'] / merged[f'{price_col}_prev']
    changed = merged[merged['Change_SAR'] != 0].sort_values('Change_Pct')

    ws.merge_cells('A1:K1')
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = f'Price changes: {pd.Timestamp(prev).strftime("%Y-%m-%d")} → {pd.Timestamp(latest).strftime("%Y-%m-%d")} ({len(changed)} models)'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Brand','SKU','Product Name','Category','Compressor','Cold/HC','Ton','Prev Price','Curr Price','Change (SAR)','Change %']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    for idx, (_, rd) in enumerate(changed.iterrows()):
        row = r + 1 + idx
        ws.cell(row=row, column=1, value=rd['Brand_curr']).font = DATA_FONT
        ws.cell(row=row, column=2, value=rd['SKU']).font = DATA_FONT
        ws.cell(row=row, column=3, value=rd['Product_Name_curr']).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=4, value=rd['Category_curr']).font = DATA_FONT
        ws.cell(row=row, column=5, value=rd['Compressor_Type_curr']).font = DATA_FONT
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=6, value=rd['Cold_or_HC_curr']).font = DATA_FONT
        ton = rd['Cooling_Capacity_Ton_curr']
        ws.cell(row=row, column=7, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=8, value=rd[f'{price_col}_prev']).font = NUM_FONT
        ws.cell(row=row, column=8).number_format = '#,##0'
        ws.cell(row=row, column=9, value=rd[f'{price_col}_curr']).font = NUM_FONT
        ws.cell(row=row, column=9).number_format = '#,##0'
        chg = rd['Change_SAR']
        c = ws.cell(row=row, column=10, value=chg)
        c.number_format = '#,##0'
        c.font = DOWN_FONT if chg < 0 else UP_FONT
        p = ws.cell(row=row, column=11, value=rd['Change_Pct'])
        p.number_format = '0%'
        p.font = DOWN_FONT if chg < 0 else UP_FONT
        if row % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=row, column=c).fill = LIGHT_FILL

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = tab_color


def build_price_change_alert(wb, df):
    _build_price_alert(wb, df, 'Price_Change_Alert', 'Final_Sale_Price',
                       '🔔 Price Change Alert (Final Sale Price)', 'FF6600')

def build_jood_price_change_alert(wb, df):
    _build_price_alert(wb, df, 'Jood_Price_Change_Alert', 'Final_Jood_Gold_Price',
                       '🔔 Jood Gold Price Change Alert', 'FF9900')


# ============================================================
# 4. New_Discontinued_SKUs (+ Compressor_Type)
# ============================================================
def build_new_discontinued(wb, df):
    sn = 'New_Discontinued_SKUs'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    if len(dates) < 2:
        ws['A1'] = '⚠️ 최소 2회 이상 스크래핑 데이터가 필요합니다.'
        ws['A1'].font = SUBTITLE_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped_At'] == prev]
    dc = df[df['Scraped_At'] == latest]
    prev_skus = set(dp['SKU'].astype(str))
    curr_skus = set(dc['SKU'].astype(str))
    new_skus = curr_skus - prev_skus
    disc_skus = prev_skus - curr_skus
    price_col = 'Final_Sale_Price'

    ws.merge_cells('A1:I1')
    ws['A1'] = '🆕 New & Discontinued SKUs'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:I2')
    ws['A2'] = f'{pd.Timestamp(prev).strftime("%Y-%m-%d")} → {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A2'].font = SUBTITLE_FONT

    headers = ['Brand','SKU','Product Name','Category','Compressor','Cold/HC','Ton','Sale Price']

    def write_section(title, title_font, skus, source_df, fill, font, start_row):
        ws.merge_cells(f'A{start_row}:I{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = title_font
        hr = start_row + 1
        for i, h in enumerate(headers):
            ws.cell(row=hr, column=i+1, value=h)
        style_header_row(ws, hr, len(headers))
        df_sub = source_df[source_df['SKU'].astype(str).isin(skus)]
        for idx, (_, rd) in enumerate(df_sub.iterrows()):
            row = hr + 1 + idx
            ws.cell(row=row, column=1, value=rd['Brand']).font = font
            ws.cell(row=row, column=2, value=str(rd['SKU'])).font = font
            ws.cell(row=row, column=3, value=rd['Product_Name']).font = DATA_FONT
            ws.cell(row=row, column=3).alignment = WRAP
            ws.cell(row=row, column=4, value=rd['Category']).font = DATA_FONT
            ws.cell(row=row, column=5, value=rd['Compressor_Type']).font = DATA_FONT
            ws.cell(row=row, column=5).alignment = CENTER
            ws.cell(row=row, column=6, value=rd['Cold_or_HC']).font = DATA_FONT
            ton = rd['Cooling_Capacity_Ton']
            ws.cell(row=row, column=7, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
            ws.cell(row=row, column=7).alignment = CENTER
            c = ws.cell(row=row, column=8, value=rd.get(price_col, ''))
            c.font = NUM_FONT
            c.number_format = '#,##0'
            for ci in range(1, len(headers)+1):
                ws.cell(row=row, column=ci).fill = fill
        return hr + 1 + len(df_sub)

    r = 4
    end = write_section(f'🟢 New SKUs ({len(new_skus)})',
                        Font(name='Arial', bold=True, size=12, color='008000'),
                        new_skus, dc, GREEN_FILL, NEW_FONT, r)

    r = end + 2
    write_section(f'🔴 Discontinued SKUs ({len(disc_skus)})',
                  Font(name='Arial', bold=True, size=12, color='CC0000'),
                  disc_skus, dp, RED_FILL, DISC_FONT, r)

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = '00B050'


# ============================================================
# 5. Brand_Avg_Price_Comparison (평균가, Ton 기준)
# ============================================================
def build_brand_price_comparison(wb, df):
    sn = 'Brand_Avg_Price_Comparison'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    dl = df[df['Scraped_At'] == latest].copy()
    price_col = 'Final_Sale_Price'
    brands_top = dl['Brand'].value_counts().head(7).index.tolist()
    if 'SAMSUNG' not in brands_top:
        brands_top.append('SAMSUNG')

    ws.merge_cells('A1:L1')
    ws['A1'] = '💰 Brand Avg Price Comparison by Category & Ton'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Price: Final Sale Price (Avg)'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Category', 'Compressor', 'Cold/HC', 'Ton'] + brands_top + ['Market Avg', 'LG Min']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    cat_order = ['Split Air Conditioner', 'Window Air Conditioner', 'Free Standing Air Conditioner']
    row_idx = r + 1

    for cat in cat_order:
        dc = dl[dl['Category'] == cat]
        if len(dc) == 0:
            continue

        ws.cell(row=row_idx, column=1, value=cat).font = BOLD_FONT
        for c in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=c).fill = CAT_FILL
        row_idx += 1

        if cat == 'Free Standing Air Conditioner':
            groups = dc.groupby(['Cooling_Capacity_Ton'])
            for ton, grp in sorted(groups, key=lambda x: x[0]):
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                mkt_avg = grp[price_col].mean()
                for bi, brand in enumerate(brands_top):
                    bg = grp[grp['Brand']==brand]
                    if len(bg) > 0:
                        c = ws.cell(row=row_idx, column=5+bi, value=round(bg[price_col].mean()))
                        c.number_format = '#,##0'
                        c.font = NUM_FONT
                    else:
                        ws.cell(row=row_idx, column=5+bi, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=5+bi).alignment = CENTER
                mc = ws.cell(row=row_idx, column=5+len(brands_top), value=round(mkt_avg))
                mc.number_format = '#,##0'
                mc.font = BOLD_FONT
                mc.alignment = CENTER
                # LG Min
                lg_grp = grp[grp['Brand']=='LG']
                lg_min_col = 6 + len(brands_top)
                if len(lg_grp) > 0:
                    lm = ws.cell(row=row_idx, column=lg_min_col, value=round(lg_grp[price_col].min()))
                    lm.number_format = '#,##0'
                    lm.font = Font(name='Arial', size=10, color='0070C0', bold=True)
                else:
                    ws.cell(row=row_idx, column=lg_min_col, value='-').font = NUM_FONT
                ws.cell(row=row_idx, column=lg_min_col).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1
        else:
            groups = dc.groupby(['Compressor_Type', 'Cold_or_HC', 'Cooling_Capacity_Ton'])
            for (comp, hc, ton), grp in sorted(groups, key=lambda x: (x[0][0], x[0][1], x[0][2])):
                ws.cell(row=row_idx, column=2, value=comp).font = DATA_FONT
                ws.cell(row=row_idx, column=3, value=hc).font = DATA_FONT
                ws.cell(row=row_idx, column=4, value=f'{ton}T').font = DATA_FONT
                ws.cell(row=row_idx, column=4).alignment = CENTER
                mkt_avg = grp[price_col].mean()
                for bi, brand in enumerate(brands_top):
                    bg = grp[grp['Brand']==brand]
                    if len(bg) > 0:
                        c = ws.cell(row=row_idx, column=5+bi, value=round(bg[price_col].mean()))
                        c.number_format = '#,##0'
                        c.font = NUM_FONT
                    else:
                        ws.cell(row=row_idx, column=5+bi, value='-').font = NUM_FONT
                    ws.cell(row=row_idx, column=5+bi).alignment = CENTER
                mc = ws.cell(row=row_idx, column=5+len(brands_top), value=round(mkt_avg))
                mc.number_format = '#,##0'
                mc.font = BOLD_FONT
                mc.alignment = CENTER
                # LG Min
                lg_grp = grp[grp['Brand']=='LG']
                lg_min_col = 6 + len(brands_top)
                if len(lg_grp) > 0:
                    lm = ws.cell(row=row_idx, column=lg_min_col, value=round(lg_grp[price_col].min()))
                    lm.number_format = '#,##0'
                    lm.font = Font(name='Arial', size=10, color='0070C0', bold=True)
                else:
                    ws.cell(row=row_idx, column=lg_min_col, value='-').font = NUM_FONT
                ws.cell(row=row_idx, column=lg_min_col).alignment = CENTER
                if row_idx % 2 == 0:
                    for c in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=c).fill = LIGHT_FILL
                row_idx += 1

    auto_width(ws)
    ws.sheet_properties.tabColor = 'FFC000'


# ============================================================
# 6. Full_Price_Tracking (Ton 기준 정렬)
# ============================================================
def build_full_tracking(wb, df):
    sn = 'Full_Price_Tracking'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    prev = dates[-2] if len(dates) >= 2 else None
    dc = df[df['Scraped_At'] == latest].copy()
    price_col = 'Final_Sale_Price'

    if prev is not None:
        dp = df[df['Scraped_At'] == prev]
        prev_map = dp.set_index(dp['SKU'].astype(str))[price_col].to_dict()
    else:
        prev_map = {}

    ws.merge_cells('A1:P1')
    ws['A1'] = '📋 Full Price Tracking (Latest Snapshot)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:P2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")} | Total: {len(dc)} SKUs'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    headers = ['Brand','SKU','Model No','Product Name','Category','Compressor','Cold/HC','Ton',
               'Std Price','Sale Price','Jood Gold','Discount %','Prev Price','Change (SAR)','Promo']
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_header_row(ws, r, len(headers))

    dc_sorted = dc.sort_values(['Category','Compressor_Type','Cold_or_HC','Cooling_Capacity_Ton','Brand'])

    for idx, (_, rd) in enumerate(dc_sorted.iterrows()):
        row = r + 1 + idx
        sku_str = str(rd['SKU'])
        ws.cell(row=row, column=1, value=rd['Brand']).font = DATA_FONT
        ws.cell(row=row, column=2, value=sku_str).font = DATA_FONT
        ws.cell(row=row, column=3, value=rd.get('Model_No','')).font = DATA_FONT
        ws.cell(row=row, column=4, value=rd['Product_Name']).font = DATA_FONT
        ws.cell(row=row, column=4).alignment = WRAP
        ws.cell(row=row, column=5, value=rd['Category']).font = DATA_FONT
        ws.cell(row=row, column=6, value=rd['Compressor_Type']).font = DATA_FONT
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=7, value=rd['Cold_or_HC']).font = DATA_FONT
        ton = rd['Cooling_Capacity_Ton']
        ws.cell(row=row, column=8, value=f'{ton}T' if pd.notna(ton) else '').font = DATA_FONT
        ws.cell(row=row, column=8).alignment = CENTER

        for ci, col_name in enumerate(['Standard_Price', price_col, 'Final_Jood_Gold_Price'], start=9):
            c = ws.cell(row=row, column=ci, value=rd.get(col_name,''))
            c.number_format = '#,##0'
            c.font = NUM_FONT

        disc = rd.get('Discount_Rate','')
        if pd.notna(disc) and disc != '':
            dc2 = ws.cell(row=row, column=12, value=disc)
            if isinstance(disc, (int, float)):
                dc2.number_format = '0%'
            dc2.font = NUM_FONT
        ws.cell(row=row, column=12).alignment = CENTER

        prev_price = prev_map.get(sku_str)
        curr_price = rd.get(price_col)
        if prev_price and pd.notna(prev_price):
            ws.cell(row=row, column=13, value=prev_price).font = NUM_FONT
            ws.cell(row=row, column=13).number_format = '#,##0'
            if curr_price and pd.notna(curr_price):
                change = curr_price - prev_price
                chg = ws.cell(row=row, column=14, value=change)
                chg.number_format = '#,##0'
                if change < 0: chg.font = DOWN_FONT
                elif change > 0: chg.font = UP_FONT
                else: chg.font = NUM_FONT
        else:
            ws.cell(row=row, column=13, value='NEW').font = NEW_FONT
            ws.cell(row=row, column=13).alignment = CENTER

        ws.cell(row=row, column=15, value=rd.get('Promo_Code','')).font = DATA_FONT

        if row % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=row, column=c).fill = LIGHT_FILL

    auto_width(ws, max_width=45)
    ws.column_dimensions['D'].width = 45
    ws.sheet_properties.tabColor = '4472C4'


# ============================================================
# 7. Promo_Analysis
# ============================================================
def build_promo_analysis(wb, df):
    sn = 'Promo_Analysis'
    delete_sheet_if_exists(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped_At'].unique())
    latest = dates[-1]
    dl = df[df['Scraped_At'] == latest]

    ws.merge_cells('A1:F1')
    ws['A1'] = '🏷️ Promotion & Offer Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Date: {pd.Timestamp(latest).strftime("%Y-%m-%d")}'
    ws['A2'].font = SUBTITLE_FONT

    r = 4
    ws[f'A{r}'] = '📌 Promo Code Distribution'
    ws[f'A{r}'].font = SECTION_FONT
    r = 6
    promo = dl[dl['Promo_Code'].notna() & (dl['Promo_Code']!='')]['Promo_Code'].value_counts()
    ws.cell(row=r, column=1, value='Promo Code')
    ws.cell(row=r, column=2, value='SKU Count')
    style_header_row(ws, r, 2)
    for idx, (code, cnt) in enumerate(promo.items()):
        ws.cell(row=r+1+idx, column=1, value=code).font = DATA_FONT
        ws.cell(row=r+1+idx, column=2, value=cnt).font = NUM_FONT
        ws.cell(row=r+1+idx, column=2).alignment = CENTER

    r2 = r + 2 + len(promo)
    ws[f'A{r2}'] = '⭐ eXtra Exclusive by Brand'
    ws[f'A{r2}'].font = SECTION_FONT
    r2 += 2
    excl = dl[dl['eXtra_Exclusive']=='Yes']['Brand'].value_counts()
    ws.cell(row=r2, column=1, value='Brand')
    ws.cell(row=r2, column=2, value='Exclusive Count')
    style_header_row(ws, r2, 2)
    for idx, (brand, cnt) in enumerate(excl.items()):
        ws.cell(row=r2+1+idx, column=1, value=brand).font = DATA_FONT
        ws.cell(row=r2+1+idx, column=2, value=cnt).font = NUM_FONT
        ws.cell(row=r2+1+idx, column=2).alignment = CENTER

    auto_width(ws)
    ws.sheet_properties.tabColor = '7030A0'


# ============================================================
# 8. Prices DB B~J열 누락 정보 보완 (실행일 데이터 기준)
# ============================================================
def fill_missing_info(wb):
    """실행일(최신) 데이터의 B~J열 공란을 이전 날짜 데이터의 동일 SKU 정보로 채움.
    매핑 키: E열(SKU). 채움 대상: B,C,D,F,G,H,I,J열 (Brand~Compressor_Type).
    A열(Scraped_At) 기준으로 최신 날짜 행을 판별하고, 그 이전 행들에서 값을 가져옴.
    """
    ws = wb['Prices DB']

    # 헤더 행(1행)에서 열 인덱스 파악
    header = {cell.value: cell.column for cell in ws[1]}
    # 필요한 열 번호 (1-based)
    col_scraped = header.get('Scraped_At', 1)   # A
    col_sku     = header.get('SKU', 5)           # E
    # B~J에 해당하는 채움 대상 컬럼
    fill_cols = []
    for name in ['Brand', 'Product_Name', 'Model_No', 'Category',
                 'Cold_or_HC', 'Cooling_Capacity_Ton', 'BTU', 'Compressor_Type']:
        if name in header:
            fill_cols.append(header[name])

    if not fill_cols:
        print("  ⚠️ fill_missing_info: 대상 컬럼을 찾을 수 없음")
        return

    # 최신 날짜 파악
    latest_date = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(v is None for v in row):
            break
        val = row[col_scraped - 1]
        if val is not None:
            if latest_date is None or val > latest_date:
                latest_date = val

    if latest_date is None:
        print("  ⚠️ fill_missing_info: 날짜 데이터 없음")
        return

    # 이전 날짜 행에서 SKU별 가장 최신 값 수집 (채움 대상 컬럼만)
    # {sku: {col_idx: value, ...}}
    sku_info = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            break
        scraped_val = row[col_scraped - 1].value
        if scraped_val is None or scraped_val >= latest_date:
            continue  # 최신 날짜 행은 건너뜀
        sku_val = row[col_sku - 1].value
        if sku_val is None:
            continue
        sku_str = str(sku_val).strip()
        if sku_str not in sku_info:
            sku_info[sku_str] = {}
        for c in fill_cols:
            v = row[c - 1].value
            if v is not None and str(v).strip() != '':
                sku_info[sku_str][c] = v  # 더 최신 값이 있으면 덮어씀

    if not sku_info:
        print("  ℹ️ fill_missing_info: 이전 데이터 없음, 건너뜀")
        return

    # 최신 날짜 행에서 공란 채우기
    filled_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            break
        scraped_val = row[col_scraped - 1].value
        if scraped_val is None or scraped_val < latest_date:
            continue  # 최신 날짜 행만 대상
        sku_val = row[col_sku - 1].value
        if sku_val is None:
            continue
        sku_str = str(sku_val).strip()
        if sku_str not in sku_info:
            continue
        prev_data = sku_info[sku_str]
        for c in fill_cols:
            cell = row[c - 1]
            if (cell.value is None or str(cell.value).strip() == '') and c in prev_data:
                cell.value = prev_data[c]
                filled_count += 1

    print(f"  ✅ fill_missing_info: {filled_count}개 셀 보완 완료 (SKU {len(sku_info)}개 참조)")


# ============================================================
# 9. Prices DB Z/AA열 채우기
# ============================================================
def fill_prices_db_zaa(wb):
    """Prices DB 시트 Z열(Final_Sale_Price), AA열(Final_Jood_Gold_Price) 자동 채우기
    - P열에 extra10 있으면: L/M 가격에서 10% 할인
    - P열에 extra10 없으면: L/M 가격 그대로
    """
    ws = wb['Prices DB']
    # 헤더 기록 (1행)
    ws.cell(row=1, column=26).value = 'Final_Sale_Price'
    ws.cell(row=1, column=27).value = 'Final_Jood_Gold_Price'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            break
        l_val = row[11].value   # L열: Sale_Price
        m_val = row[12].value   # M열: Jood_Gold_Price
        p_val = row[15].value   # P열: Promo_Code

        has_extra10 = isinstance(p_val, str) and p_val.strip().lower() == 'extra10'

        if has_extra10:
            row[25].value = round(l_val * 0.9) if isinstance(l_val, (int, float)) else None  # Z열
            row[26].value = round(m_val * 0.9) if isinstance(m_val, (int, float)) else None  # AA열
        else:
            row[25].value = l_val  # Z열
            row[26].value = m_val  # AA열


# ============================================================
# MAIN
# ============================================================
def main():
    # Windows 터미널 UTF-8 출력 설정
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    print("=" * 60)
    print("📊 Extra.com AC Dashboard Builder v2")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ 파일 없음: {INPUT_FILE}")
        sys.exit(1)

    print(f"📂 파일 로드: {INPUT_FILE}")
    df = load_data(INPUT_FILE)

    dates = sorted(df['Scraped_At'].unique())
    print(f"📅 스크래핑 일자: {len(dates)}회")
    for d in dates:
        print(f"   - {pd.Timestamp(d).strftime('%Y-%m-%d')}: {len(df[df['Scraped_At']==d])}개 SKU")

    wb = load_workbook(INPUT_FILE)

    print("\n🔧 [1단계] Prices DB B~J열 누락 정보 보완 중...")
    fill_missing_info(wb)

    print("\n🔧 [2단계] Prices DB Z/AA열 채우는 중...")
    fill_prices_db_zaa(wb)

    print("\n🔧 대시보드 시트 생성 중...")
    print("  1/7 Dashboard_Summary...")
    build_dashboard_summary(wb, df)
    print("  2/7 Price_Change_Alert...")
    build_price_change_alert(wb, df)
    print("  3/7 Jood_Price_Change_Alert...")
    build_jood_price_change_alert(wb, df)
    print("  4/7 New_Discontinued_SKUs...")
    build_new_discontinued(wb, df)
    print("  5/7 Brand_Avg_Price_Comparison...")
    build_brand_price_comparison(wb, df)
    print("  6/7 Full_Price_Tracking...")
    build_full_tracking(wb, df)
    print("  7/7 Promo_Analysis...")
    build_promo_analysis(wb, df)

    desired_order = [
        'Dashboard_Summary','Price_Change_Alert','Jood_Price_Change_Alert',
        'New_Discontinued_SKUs','Brand_Avg_Price_Comparison','Full_Price_Tracking',
        'Promo_Analysis','Prices DB','Model_Master'
    ]
    current = wb.sheetnames
    for i, name in enumerate(desired_order):
        if name in current:
            wb.move_sheet(name, offset=i - current.index(name))
            current = wb.sheetnames

    # 이전 버전 시트 정리
    for old in ['Brand_Price_Comparison']:
        if old in wb.sheetnames:
            del wb[old]

    wb.save(INPUT_FILE)
    print(f"\n✅ 대시보드 업데이트 완료!")
    print(f"📁 파일: {INPUT_FILE}")
    print("=" * 60)

    try:
        if sys.stdin.isatty():
            input("\n 엔터를 누르면 종료...")
    except (EOFError, OSError):
        pass

if __name__ == "__main__":
    main()
