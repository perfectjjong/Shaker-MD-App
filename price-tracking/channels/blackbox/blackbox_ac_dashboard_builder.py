#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Blackbox AC Dashboard Builder
==============================
대상 파일 : Black Box_AC_Price tracking_Master.xlsx
대상 시트 : Product_DB

Product_DB 컬럼 (A~X):
  A  AC Type       B  Brand         C  Name           D  Model Code
  E  BTU           F  Ton           G  Mode            H  Compressor
  I  Original Price  J  Sale Price  K  Discount %      L  Extra Disc %
  M  Effective Price N  BP Price    O  Effective BP
  P  Free Install    Q  Install SAR R  +10% Regular    S  +10% BP Only
  T  Sale Ends       U  In Stock    V  Stock Qty
  W  URL             X  Scraped At

생성 시트:
  1. Dashboard_Summary       - KPI / AC Type 분포 / 브랜드 점유 / 이전 대비 변동
  2. Price_Change_Alert       - 가격 변동 SKU 목록
  3. New_Discontinued_SKUs    - 신규 / 단종 모델
  4. Brand_Price_Compare      - 브랜드별 평균가 비교 (AC Type × Ton)
  5. Promo_Analysis           - 프로모션 분석 (Free Install / Extra Disc / +10% / Sale Ends / BP)
  6. Full_Price_Tracking      - 전체 제품 최신 스냅샷
"""

import os
import sys

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[ERROR] Missing package: {e}")
    print("  >> pip install pandas openpyxl numpy")
    sys.exit(1)

# ── 경로 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
TARGET_FILE = os.path.join(SCRIPT_DIR, "Black Box_AC_Price tracking_Master.xlsx")
DB_SHEET    = "Product_DB"

# ── 스타일 ────────────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill('solid', fgColor='2F5496')
HDR_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT  = Font(name='Arial', bold=True, size=13, color='2F5496')
SUB_FONT    = Font(name='Arial', bold=True, size=10, color='595959')
SEC_FONT    = Font(name='Arial', bold=True, size=11, color='2F5496')
DATA_FONT   = Font(name='Arial', size=10)
NUM_FONT    = Font(name='Arial', size=10)
BOLD_FONT   = Font(name='Arial', bold=True, size=10)
UP_FONT     = Font(name='Arial', size=10, color='C00000', bold=True)   # 가격 상승 ▲
DOWN_FONT   = Font(name='Arial', size=10, color='375623', bold=True)   # 가격 하락 ▼
NEW_FONT    = Font(name='Arial', size=10, color='0070C0', bold=True)
DISC_FONT   = Font(name='Arial', size=10, color='999999')
BP_FONT     = Font(name='Arial', size=10, color='7030A0', bold=True)   # Blackbox+ 전용
TK_BORDER   = Border(bottom=Side(style='medium', color='2F5496'))
CTR         = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')
WRAP        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
LIGHT_FILL      = PatternFill('solid', fgColor='EEF3FB')
CAT_FILL        = PatternFill('solid', fgColor='D6E4F0')
GREEN_FILL      = PatternFill('solid', fgColor='E2EFDA')
RED_FILL        = PatternFill('solid', fgColor='FCE4EC')
BP_FILL         = PatternFill('solid', fgColor='EDE7F6')
CRITICAL_FILL   = PatternFill('solid', fgColor='FCE4EC')
OOS_FILL        = PatternFill('solid', fgColor='CCCCCC')
HIGH_FILL       = PatternFill('solid', fgColor='D6E4F0')

# Stock-specific fonts
STOCK_UP_FONT   = Font(name='Arial', size=10, color='375623', bold=True)
STOCK_DOWN_FONT = Font(name='Arial', size=10, color='C00000', bold=True)
HIGH_FONT       = Font(name='Arial', size=10, color='1F4E79', bold=True)

AC_TYPE_ORDER = ['Split', 'Window', 'Floor Standing', 'Cassette', 'Portable']


# ── 공통 유틸 ─────────────────────────────────────────────────────────────────
def _hdr(ws, row, n_col):
    for c in range(1, n_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CTR
        cell.border    = TK_BORDER


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx  = 0
        ltr = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[ltr].width = min(max(mx + 3, min_w), max_w)


def _del(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def _fmt_ton(val):
    try:
        v = float(val)
        return f"{v:.1f}T" if v % 1 else f"{int(v)}T"
    except Exception:
        return str(val) if val else '-'


def _pct_str(val):
    try:
        v = float(val)
        return f"{v:.0f}%"
    except Exception:
        return '-'


# ── 데이터 로드 ───────────────────────────────────────────────────────────────
def load_data() -> pd.DataFrame:
    df = pd.read_excel(TARGET_FILE, sheet_name=DB_SHEET, engine='openpyxl')

    # Scraped At → datetime (날짜 부분만 사용해 일별 스냅샷 비교)
    df['Scraped At'] = pd.to_datetime(df['Scraped At'], errors='coerce').dt.normalize()

    # 숫자 컬럼
    num_cols = [
        'BTU', 'Ton', 'Original Price', 'Sale Price', 'Discount %',
        'Extra Disc %', 'Effective Price', 'BP Price', 'Effective BP',
        'Install SAR', 'Stock Qty',
    ]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # 불리언 컬럼
    for col in ['Free Install', '+10% Regular', '+10% BP Only', 'In Stock']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().isin(['true', '1', 'yes'])

    # 문자 컬럼 정리
    for col in ['AC Type', 'Brand', 'Name', 'Model Code', 'Mode', 'Compressor']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()

    # 최종 판매가: Effective Price → Sale Price → Original Price
    df['Final Price'] = (
        df.get('Effective Price', pd.Series(dtype=float))
          .fillna(df.get('Sale Price', pd.Series(dtype=float)))
          .fillna(df.get('Original Price', pd.Series(dtype=float)))
    )

    # AC Type 정규화 (대소문자 통일)
    def _norm_type(t):
        t = str(t).strip()
        for std in AC_TYPE_ORDER:
            if std.lower() in t.lower():
                return std
        return t if t else 'Other'

    df['AC Type'] = df['AC Type'].apply(_norm_type)

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Dashboard_Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_summary(wb, df: pd.DataFrame):
    sn = 'Dashboard_Summary'
    _del(wb, sn)
    ws = wb.create_sheet(sn, 0)

    dates  = sorted(df['Scraped At'].dropna().unique())
    latest = dates[-1]
    prev   = dates[-2] if len(dates) >= 2 else None
    dl     = df[df['Scraped At'] == latest]
    dp     = df[df['Scraped At'] == prev] if prev else pd.DataFrame()

    # 제목
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Blackbox.com.sa — AC Price Tracking Dashboard'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:N2')
    ws['A2'] = (
        f"Latest: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
        f"Total SKUs: {len(dl)}  |  "
        f"Brands: {dl['Brand'].replace('', pd.NA).dropna().nunique()}  |  "
        f"In Stock: {dl['In Stock'].sum()}"
    )
    ws['A2'].font = SUB_FONT

    # ── Section 1: KPI by AC Type × Compressor × Mode × Ton ──────────────────
    r = 4
    ws.merge_cells(f'A{r}:N{r}')
    ws[f'A{r}'] = '▶  Price KPI by AC Type / Compressor / Mode / Capacity'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    kpi_hdrs = [
        'AC Type', 'Compressor', 'Mode', 'Capacity',
        'SKUs', 'Avg Orig (SAR)', 'Avg Sale (SAR)', 'Avg Disc %',
        'BP SKUs', 'Avg BP (SAR)',
        'LG SKUs', 'LG Avg Sale', 'LG vs Mkt (SAR)', 'LG vs Mkt %',
    ]
    for ci, h in enumerate(kpi_hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(kpi_hdrs))
    r += 1

    grand = 0
    for ac_type in AC_TYPE_ORDER:
        dc = dl[dl['AC Type'] == ac_type]
        if dc.empty:
            continue

        mkt_avg = round(dc['Final Price'].mean()) if dc['Final Price'].notna().any() else None
        lg_dc   = dc[dc['Brand'].str.upper() == 'LG']
        lg_avg  = round(lg_dc['Final Price'].mean()) if not lg_dc.empty and lg_dc['Final Price'].notna().any() else None
        gap_sar = (lg_avg - mkt_avg) if (lg_avg and mkt_avg) else None
        gap_pct = (gap_sar / mkt_avg) if (gap_sar is not None and mkt_avg) else None
        bp_skus = int(dc['BP Price'].notna().sum())
        bp_avg  = round(dc['Effective BP'].mean()) if dc['Effective BP'].notna().any() else None

        for ci in range(1, len(kpi_hdrs) + 1):
            ws.cell(row=r, column=ci).fill = CAT_FILL

        ws.cell(r, 1, ac_type).font = BOLD_FONT
        ws.cell(r, 5, len(dc)).font = BOLD_FONT
        ws.cell(r, 5).alignment = CTR
        if dc['Original Price'].notna().any():
            c = ws.cell(r, 6, round(dc['Original Price'].mean()))
            c.number_format = '#,##0'; c.font = BOLD_FONT
        if mkt_avg:
            c = ws.cell(r, 7, mkt_avg)
            c.number_format = '#,##0'; c.font = BOLD_FONT
        ws.cell(r, 8, _pct_str(dc['Discount %'].mean())).font = BOLD_FONT
        ws.cell(r, 8).alignment = CTR
        ws.cell(r, 9, bp_skus).font = BOLD_FONT; ws.cell(r, 9).alignment = CTR
        if bp_avg:
            c = ws.cell(r, 10, bp_avg)
            c.number_format = '#,##0'; c.font = BP_FONT
        ws.cell(r, 11, len(lg_dc)).font = BOLD_FONT; ws.cell(r, 11).alignment = CTR
        if lg_avg:
            c = ws.cell(r, 12, lg_avg)
            c.number_format = '#,##0'; c.font = BOLD_FONT
        if gap_sar is not None:
            c = ws.cell(r, 13, gap_sar)
            c.number_format = '#,##0'
            c.font = UP_FONT if gap_sar > 0 else (DOWN_FONT if gap_sar < 0 else BOLD_FONT)
        if gap_pct is not None:
            c = ws.cell(r, 14, gap_pct)
            c.number_format = '0%'
            c.font = UP_FONT if gap_pct > 0 else (DOWN_FONT if gap_pct < 0 else BOLD_FONT)
        for ci in [10, 12, 13, 14]:
            ws.cell(r, ci).alignment = CTR
        grand += len(dc)
        r += 1

        # 세분화 행
        group_cols = ['Compressor', 'Mode', 'Ton']
        grp_iter = dc.groupby(group_cols, dropna=False)
        for (comp, mode, ton), grp in sorted(
                grp_iter,
                key=lambda x: (str(x[0][0]), str(x[0][1]),
                               x[0][2] if pd.notna(x[0][2]) else 9999)):
            mkt  = round(grp['Final Price'].mean()) if grp['Final Price'].notna().any() else None
            lg_g = grp[grp['Brand'].str.upper() == 'LG']
            lg_m = round(lg_g['Final Price'].mean()) if not lg_g.empty and lg_g['Final Price'].notna().any() else None
            gs   = (lg_m - mkt) if (lg_m and mkt) else None
            gp   = (gs / mkt)   if (gs is not None and mkt) else None
            b_s  = int(grp['BP Price'].notna().sum())
            b_a  = round(grp['Effective BP'].mean()) if grp['Effective BP'].notna().any() else None

            ws.cell(r, 2, comp or '-').font = DATA_FONT
            ws.cell(r, 3, mode or '-').font = DATA_FONT
            ws.cell(r, 4, _fmt_ton(ton)).font = DATA_FONT; ws.cell(r, 4).alignment = CTR
            ws.cell(r, 5, len(grp)).font = NUM_FONT; ws.cell(r, 5).alignment = CTR
            if grp['Original Price'].notna().any():
                c = ws.cell(r, 6, round(grp['Original Price'].mean()))
                c.number_format = '#,##0'; c.font = NUM_FONT
            if mkt:
                c = ws.cell(r, 7, mkt)
                c.number_format = '#,##0'; c.font = NUM_FONT
            ws.cell(r, 8, _pct_str(grp['Discount %'].mean())).font = NUM_FONT
            ws.cell(r, 8).alignment = CTR
            ws.cell(r, 9, b_s).font = NUM_FONT; ws.cell(r, 9).alignment = CTR
            if b_a:
                c = ws.cell(r, 10, b_a)
                c.number_format = '#,##0'; c.font = BP_FONT; c.alignment = CTR
            ws.cell(r, 11, len(lg_g)).font = NUM_FONT; ws.cell(r, 11).alignment = CTR
            if lg_m:
                c = ws.cell(r, 12, lg_m)
                c.number_format = '#,##0'; c.font = NUM_FONT
            if gs is not None:
                c = ws.cell(r, 13, gs)
                c.number_format = '#,##0'
                c.font = UP_FONT if gs > 0 else (DOWN_FONT if gs < 0 else NUM_FONT)
            if gp is not None:
                c = ws.cell(r, 14, gp)
                c.number_format = '0%'
                c.font = UP_FONT if gp > 0 else (DOWN_FONT if gp < 0 else NUM_FONT)
            for ci in [12, 13, 14]:
                ws.cell(r, ci).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(kpi_hdrs) + 1):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    for ci in range(1, len(kpi_hdrs) + 1):
        ws.cell(r, ci).fill = CAT_FILL
    ws.cell(r, 1, 'TOTAL').font = BOLD_FONT
    ws.cell(r, 5, grand).font  = BOLD_FONT; ws.cell(r, 5).alignment = CTR
    r += 2

    # ── Section 2: AC Type × Brand 분포 ──────────────────────────────────────
    ws.merge_cells(f'A{r}:N{r}')
    ws[f'A{r}'] = '▶  SKU Distribution by AC Type & Brand'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    brands_top = (
        dl['Brand'].replace('', pd.NA).dropna()
        .value_counts().head(8).index.tolist()
    )
    if not any(b.upper() == 'LG' for b in brands_top):
        brands_top.append('LG')

    dist_hdrs = ['AC Type', 'Compressor', 'Mode', 'Capacity', 'Total'] + brands_top
    for ci, h in enumerate(dist_hdrs, 1):
        ws.cell(row=r, column=ci, value=h)
    _hdr(ws, r, len(dist_hdrs))
    r += 1

    for ac_type in AC_TYPE_ORDER:
        dc = dl[dl['AC Type'] == ac_type]
        if dc.empty:
            continue
        for ci in range(1, len(dist_hdrs) + 1):
            ws.cell(r, ci).fill = CAT_FILL
        ws.cell(r, 1, ac_type).font = BOLD_FONT
        ws.cell(r, 5, len(dc)).font  = BOLD_FONT; ws.cell(r, 5).alignment = CTR
        for bi, brand in enumerate(brands_top):
            cnt = len(dc[dc['Brand'].str.upper() == brand.upper()])
            ws.cell(r, 6 + bi, cnt if cnt else '-').font = BOLD_FONT
            ws.cell(r, 6 + bi).alignment = CTR
        r += 1

        grp_iter2 = dc.groupby(['Compressor', 'Mode', 'Ton'], dropna=False)
        for (comp, mode, ton), grp in sorted(
                grp_iter2,
                key=lambda x: (str(x[0][0]), str(x[0][1]),
                               x[0][2] if pd.notna(x[0][2]) else 9999)):
            ws.cell(r, 2, comp or '-').font = DATA_FONT
            ws.cell(r, 3, mode or '-').font = DATA_FONT
            ws.cell(r, 4, _fmt_ton(ton)).font = DATA_FONT; ws.cell(r, 4).alignment = CTR
            ws.cell(r, 5, len(grp)).font = NUM_FONT; ws.cell(r, 5).alignment = CTR
            for bi, brand in enumerate(brands_top):
                cnt = len(grp[grp['Brand'].str.upper() == brand.upper()])
                ws.cell(r, 6 + bi, cnt if cnt else '-').font = NUM_FONT
                ws.cell(r, 6 + bi).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(dist_hdrs) + 1):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    for ci in range(1, len(dist_hdrs) + 1):
        ws.cell(r, ci).fill = CAT_FILL
    ws.cell(r, 1, 'TOTAL').font = BOLD_FONT
    ws.cell(r, 5, len(dl)).font  = BOLD_FONT; ws.cell(r, 5).alignment = CTR
    for bi, brand in enumerate(brands_top):
        cnt = len(dl[dl['Brand'].str.upper() == brand.upper()])
        ws.cell(r, 6 + bi, cnt).font = BOLD_FONT; ws.cell(r, 6 + bi).alignment = CTR
    r += 2

    # ── Section 3: 이전 대비 변동 요약 ───────────────────────────────────────
    if not dp.empty:
        ws.merge_cells(f'A{r}:N{r}')
        ws[f'A{r}'] = f"▶  Changes vs Previous ({pd.Timestamp(prev).strftime('%Y-%m-%d')})"
        ws[f'A{r}'].font = SEC_FONT
        r += 2

        prev_codes = set(dp['Model Code'].astype(str))
        curr_codes = set(dl['Model Code'].astype(str))
        mg = pd.merge(
            dp[['Model Code', 'Final Price']].astype({'Model Code': str}),
            dl[['Model Code', 'Final Price']].astype({'Model Code': str}),
            on='Model Code', suffixes=('_p', '_c')
        )
        mg['Chg'] = mg['Final Price_c'] - mg['Final Price_p']

        items = [
            ('New Models',      len(curr_codes - prev_codes)),
            ('Removed Models',  len(prev_codes - curr_codes)),
            ('Price Up ▲',      len(mg[mg['Chg'] > 0])),
            ('Price Down ▼',    len(mg[mg['Chg'] < 0])),
            ('No Change',       len(mg[mg['Chg'] == 0])),
        ]
        ws.cell(r, 1, 'Item').font = BOLD_FONT
        ws.cell(r, 2, 'Count').font = BOLD_FONT
        _hdr(ws, r, 2)
        for i, (label, cnt) in enumerate(items):
            rr = r + 1 + i
            ws.cell(rr, 1, label).font = DATA_FONT
            c = ws.cell(rr, 2, cnt)
            c.alignment = CTR
            c.font = (UP_FONT   if '▲' in label and cnt > 0 else
                      DOWN_FONT if '▼' in label and cnt > 0 else NUM_FONT)

    _auto_width(ws)
    ws.sheet_properties.tabColor = '2F5496'


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Price_Change_Alert
# ═══════════════════════════════════════════════════════════════════════════════
def build_price_alert(wb, df: pd.DataFrame):
    sn = 'Price_Change_Alert'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Price comparison requires at least 2 collection dates.'
        ws['A1'].font = SUB_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped At'] == prev]
    dc = df[df['Scraped At'] == latest]

    cols = ['Model Code', 'Brand', 'Name', 'AC Type', 'Compressor', 'Mode', 'Ton', 'Final Price']
    mg = pd.merge(
        dp[cols].astype({'Model Code': str}),
        dc[cols].astype({'Model Code': str}),
        on='Model Code', suffixes=('_p', '_c')
    )
    mg['Chg_SAR'] = mg['Final Price_c'] - mg['Final Price_p']
    mg['Chg_Pct'] = mg['Chg_SAR'] / mg['Final Price_p'].replace(0, np.nan)
    changed = mg[mg['Chg_SAR'] != 0].sort_values('Chg_Pct')

    ws.merge_cells('A1:K1')
    ws['A1'] = 'Price Change Alert'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = (
        f"{pd.Timestamp(prev).strftime('%Y-%m-%d')} → "
        f"{pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
        f"Changed: {len(changed)} models"
    )
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = ['Brand', 'Model Code', 'Product Name', 'AC Type',
            'Compressor', 'Mode', 'Capacity',
            'Prev Price', 'Curr Price', 'Change (SAR)', 'Change %']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h)
    _hdr(ws, r, len(hdrs))

    for i, (_, rd) in enumerate(changed.iterrows()):
        rr = r + 1 + i
        ws.cell(rr, 1,  rd.get('Brand_c', '')).font = DATA_FONT
        ws.cell(rr, 2,  rd['Model Code']).font = DATA_FONT
        ws.cell(rr, 3,  rd.get('Name_c', '')).font = DATA_FONT
        ws.cell(rr, 3).alignment = WRAP
        ws.cell(rr, 4,  rd.get('AC Type_c', '')).font = DATA_FONT
        ws.cell(rr, 5,  rd.get('Compressor_c', '')).font = DATA_FONT
        ws.cell(rr, 5).alignment = CTR
        ws.cell(rr, 6,  rd.get('Mode_c', '')).font = DATA_FONT
        ws.cell(rr, 7,  _fmt_ton(rd.get('Ton_c', ''))).font = DATA_FONT
        ws.cell(rr, 7).alignment = CTR
        ws.cell(rr, 8,  rd['Final Price_p']).number_format = '#,##0'
        ws.cell(rr, 8).font = NUM_FONT
        ws.cell(rr, 9,  rd['Final Price_c']).number_format = '#,##0'
        ws.cell(rr, 9).font = NUM_FONT
        chg = rd['Chg_SAR']
        c   = ws.cell(rr, 10, chg)
        c.number_format = '#,##0'
        c.font = DOWN_FONT if chg < 0 else UP_FONT
        cp = ws.cell(rr, 11, rd['Chg_Pct'])
        cp.number_format = '0%'
        cp.font = DOWN_FONT if chg < 0 else UP_FONT
        if rr % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = LIGHT_FILL

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.sheet_properties.tabColor = 'FF6600'


# ═══════════════════════════════════════════════════════════════════════════════
# 3. New_Discontinued_SKUs
# ═══════════════════════════════════════════════════════════════════════════════
def build_new_discontinued(wb, df: pd.DataFrame):
    sn = 'New_Discontinued_SKUs'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Requires at least 2 collection dates.'
        ws['A1'].font = SUB_FONT
        return

    latest, prev = dates[-1], dates[-2]
    dp = df[df['Scraped At'] == prev]
    dc = df[df['Scraped At'] == latest]
    new_codes  = set(dc['Model Code'].astype(str)) - set(dp['Model Code'].astype(str))
    disc_codes = set(dp['Model Code'].astype(str)) - set(dc['Model Code'].astype(str))

    ws.merge_cells('A1:I1')
    ws['A1'] = 'New & Discontinued Models'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:I2')
    ws['A2'] = (
        f"{pd.Timestamp(prev).strftime('%Y-%m-%d')} → "
        f"{pd.Timestamp(latest).strftime('%Y-%m-%d')}"
    )
    ws['A2'].font = SUB_FONT

    hdrs = ['Brand', 'Model Code', 'Product Name', 'AC Type',
            'Compressor', 'Mode', 'Capacity', 'Effective Price', 'BP Price']

    def _write_section(title, title_font, codes, src_df, fill, row_font, start):
        ws.merge_cells(f'A{start}:I{start}')
        ws[f'A{start}'] = title
        ws[f'A{start}'].font = title_font
        hr = start + 1
        for ci, h in enumerate(hdrs, 1):
            ws.cell(hr, ci, h)
        _hdr(ws, hr, len(hdrs))
        sub = src_df[src_df['Model Code'].astype(str).isin(codes)]
        for i, (_, rd) in enumerate(sub.iterrows()):
            rr = hr + 1 + i
            ws.cell(rr, 1, rd.get('Brand', '')).font = row_font
            ws.cell(rr, 2, rd.get('Model Code', '')).font = row_font
            ws.cell(rr, 3, rd.get('Name', '')).font = DATA_FONT
            ws.cell(rr, 3).alignment = WRAP
            ws.cell(rr, 4, rd.get('AC Type', '')).font = DATA_FONT
            ws.cell(rr, 5, rd.get('Compressor', '')).font = DATA_FONT
            ws.cell(rr, 5).alignment = CTR
            ws.cell(rr, 6, rd.get('Mode', '')).font = DATA_FONT
            ws.cell(rr, 7, _fmt_ton(rd.get('Ton', ''))).font = DATA_FONT
            ws.cell(rr, 7).alignment = CTR
            ep = rd.get('Effective Price', '')
            c = ws.cell(rr, 8, int(ep) if pd.notna(ep) and ep != '' else '')
            c.number_format = '#,##0'; c.font = NUM_FONT
            bp = rd.get('BP Price', '')
            c = ws.cell(rr, 9, int(bp) if pd.notna(bp) and bp != '' else '')
            c.number_format = '#,##0'; c.font = BP_FONT
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = fill
        return hr + 1 + len(sub)

    r = 4
    end = _write_section(
        f'New Models ({len(new_codes)})',
        Font(name='Arial', bold=True, size=12, color='375623'),
        new_codes, dc, GREEN_FILL, NEW_FONT, r
    )
    _write_section(
        f'Discontinued Models ({len(disc_codes)})',
        Font(name='Arial', bold=True, size=12, color='C00000'),
        disc_codes, dp, RED_FILL, DISC_FONT, end + 2
    )

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.sheet_properties.tabColor = '00B050'


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Brand_Price_Compare
# ═══════════════════════════════════════════════════════════════════════════════
def build_brand_compare(wb, df: pd.DataFrame):
    sn = 'Brand_Price_Compare'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped At'].dropna().unique())
    latest = dates[-1]
    dl     = df[df['Scraped At'] == latest]

    brands_top = (
        dl['Brand'].replace('', pd.NA).dropna()
        .value_counts().head(7).index.tolist()
    )
    if not any(b.upper() == 'LG' for b in brands_top):
        brands_top.append('LG')

    ws.merge_cells('A1:L1')
    ws['A1'] = 'Brand Average Price Comparison by AC Type & Capacity'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:L2')
    ws['A2'] = (
        f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
        "Price basis: Effective Price (avg) / SAR"
    )
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = ['AC Type', 'Compressor', 'Mode', 'Capacity'] + brands_top + ['Market Avg', 'LG Min']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h)
    _hdr(ws, r, len(hdrs))
    r += 1

    for ac_type in AC_TYPE_ORDER:
        dc = dl[dl['AC Type'] == ac_type]
        if dc.empty:
            continue
        for ci in range(1, len(hdrs) + 1):
            ws.cell(r, ci).fill = CAT_FILL
        ws.cell(r, 1, ac_type).font = BOLD_FONT
        r += 1

        grp_iter = dc.groupby(['Compressor', 'Mode', 'Ton'], dropna=False)
        for (comp, mode, ton), grp in sorted(
                grp_iter,
                key=lambda x: (str(x[0][0]), str(x[0][1]),
                               x[0][2] if pd.notna(x[0][2]) else 9999)):
            ws.cell(r, 2, comp or '-').font = DATA_FONT
            ws.cell(r, 3, mode or '-').font = DATA_FONT
            ws.cell(r, 4, _fmt_ton(ton)).font = DATA_FONT
            ws.cell(r, 4).alignment = CTR
            mkt = grp['Final Price'].mean()
            for bi, brand in enumerate(brands_top):
                bg = grp[grp['Brand'].str.upper() == brand.upper()]
                if not bg.empty and bg['Final Price'].notna().any():
                    c = ws.cell(r, 5 + bi, round(bg['Final Price'].mean()))
                    c.number_format = '#,##0'; c.font = NUM_FONT
                else:
                    ws.cell(r, 5 + bi, '-').font = NUM_FONT
                ws.cell(r, 5 + bi).alignment = CTR
            mc = ws.cell(r, 5 + len(brands_top), round(mkt) if pd.notna(mkt) else '-')
            mc.number_format = '#,##0'; mc.font = BOLD_FONT; mc.alignment = CTR
            lc_col = 6 + len(brands_top)
            lg_g = grp[grp['Brand'].str.upper() == 'LG']
            if not lg_g.empty and lg_g['Final Price'].notna().any():
                lm = ws.cell(r, lc_col, round(lg_g['Final Price'].min()))
                lm.number_format = '#,##0'
                lm.font = Font(name='Arial', size=10, color='0070C0', bold=True)
            else:
                ws.cell(r, lc_col, '-').font = NUM_FONT
            ws.cell(r, lc_col).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, len(hdrs) + 1):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = 'FFC000'


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Promo_Analysis
# ═══════════════════════════════════════════════════════════════════════════════
def build_promo_analysis(wb, df: pd.DataFrame):
    sn = 'Promo_Analysis'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped At'].dropna().unique())
    latest = dates[-1]
    dl     = df[df['Scraped At'] == latest]

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Promotion Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  Total SKUs: {len(dl)}"
    ws['A2'].font = SUB_FONT

    r = 4

    # ── Section A: Promo Overview ─────────────────────────────────────────────
    ws[f'A{r}'] = '▶  Promotion Overview'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    total = len(dl)
    fi_cnt   = int(dl['Free Install'].sum())
    ed_cnt   = int(dl['Extra Disc %'].notna().sum())
    reg_cnt  = int(dl['+10% Regular'].sum())
    bp_cnt   = int(dl['+10% BP Only'].sum())
    bp_p_cnt = int(dl['BP Price'].notna().sum())
    sale_cnt = int(dl['Sale Ends'].notna().sum()) if 'Sale Ends' in dl.columns else 0

    overview = [
        ('Free Installation',     fi_cnt,   f"{fi_cnt/total*100:.0f}%"),
        ('Extra Discount',         ed_cnt,   f"{ed_cnt/total*100:.0f}%"),
        ('+10% Regular Promo',    reg_cnt,  f"{reg_cnt/total*100:.0f}%"),
        ('+10% BP Only Promo',    bp_cnt,   f"{bp_cnt/total*100:.0f}%"),
        ('Blackbox+ Price (BP)',  bp_p_cnt, f"{bp_p_cnt/total*100:.0f}%"),
        ('Active Sale (end date)', sale_cnt, f"{sale_cnt/total*100:.0f}%"),
    ]
    ws.cell(r, 1, 'Promo Type').font = BOLD_FONT
    ws.cell(r, 2, 'SKU Count').font  = BOLD_FONT
    ws.cell(r, 3, '% of Total').font = BOLD_FONT
    _hdr(ws, r, 3)
    r += 1
    for label, cnt, pct in overview:
        ws.cell(r, 1, label).font = DATA_FONT
        ws.cell(r, 2, cnt).font   = NUM_FONT; ws.cell(r, 2).alignment = CTR
        ws.cell(r, 3, pct).font   = NUM_FONT; ws.cell(r, 3).alignment = CTR
        if r % 2 == 0:
            for ci in range(1, 4):
                ws.cell(r, ci).fill = LIGHT_FILL
        r += 1
    r += 1

    # ── Section B: Free Install by Brand ─────────────────────────────────────
    ws[f'A{r}'] = '▶  Free Install by Brand'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    fi_df = dl[dl['Free Install'] == True]
    if fi_df.empty:
        ws.cell(r, 1, 'No free install items found.').font = DATA_FONT
        r += 2
    else:
        brand_fi = (
            fi_df.groupby('Brand')
            .agg(fi_skus=('Model Code', 'count'),
                 free_skus=('Install SAR', lambda x: (x == 0).sum()),
                 paid_skus=('Install SAR', lambda x: (x > 0).sum()))
            .reset_index()
            .sort_values('fi_skus', ascending=False)
        )
        ws.cell(r, 1, 'Brand').font          = BOLD_FONT
        ws.cell(r, 2, 'Free Install SKUs').font = BOLD_FONT
        ws.cell(r, 3, 'Free (0 SAR)').font   = BOLD_FONT
        ws.cell(r, 4, 'Paid Install').font   = BOLD_FONT
        _hdr(ws, r, 4)
        r += 1
        for _, brd in brand_fi.iterrows():
            ws.cell(r, 1, brd['Brand']).font         = DATA_FONT
            ws.cell(r, 2, int(brd['fi_skus'])).font  = NUM_FONT; ws.cell(r, 2).alignment = CTR
            ws.cell(r, 3, int(brd['free_skus'])).font = NUM_FONT; ws.cell(r, 3).alignment = CTR
            ws.cell(r, 4, int(brd['paid_skus'])).font = NUM_FONT; ws.cell(r, 4).alignment = CTR
            if r % 2 == 0:
                for ci in range(1, 5):
                    ws.cell(r, ci).fill = LIGHT_FILL
            r += 1
    r += 1

    # ── Section C: Blackbox+ Price Summary ───────────────────────────────────
    ws[f'A{r}'] = '▶  Blackbox+ (BP) Price by AC Type & Brand'
    ws[f'A{r}'].font = SEC_FONT
    r += 2

    bp_df = dl[dl['BP Price'].notna()]
    if bp_df.empty:
        ws.cell(r, 1, 'No Blackbox+ price found.').font = DATA_FONT
        r += 2
    else:
        ws.cell(r, 1, 'AC Type').font        = BOLD_FONT
        ws.cell(r, 2, 'Brand').font          = BOLD_FONT
        ws.cell(r, 3, 'BP SKUs').font        = BOLD_FONT
        ws.cell(r, 4, 'Avg Sale Price').font = BOLD_FONT
        ws.cell(r, 5, 'Avg BP Price').font   = BOLD_FONT
        ws.cell(r, 6, 'Avg BP Saving').font  = BOLD_FONT
        ws.cell(r, 7, 'Avg Saving %').font   = BOLD_FONT
        _hdr(ws, r, 7)
        r += 1
        bp_grp = bp_df.groupby(['AC Type', 'Brand'])
        for (ac, brand), grp in sorted(bp_grp, key=lambda x: x[0]):
            avg_sale = grp['Sale Price'].mean()
            avg_bp   = grp['Effective BP'].mean()
            saving   = avg_sale - avg_bp if pd.notna(avg_sale) and pd.notna(avg_bp) else None
            saving_p = (saving / avg_sale) if (saving and avg_sale) else None
            ws.cell(r, 1, ac).font    = DATA_FONT
            ws.cell(r, 2, brand).font = DATA_FONT
            ws.cell(r, 3, len(grp)).font = NUM_FONT; ws.cell(r, 3).alignment = CTR
            if pd.notna(avg_sale):
                c = ws.cell(r, 4, round(avg_sale)); c.number_format = '#,##0'; c.font = NUM_FONT
            if pd.notna(avg_bp):
                c = ws.cell(r, 5, round(avg_bp)); c.number_format = '#,##0'; c.font = BP_FONT
            if saving:
                c = ws.cell(r, 6, round(saving)); c.number_format = '#,##0'
                c.font = DOWN_FONT
            if saving_p:
                c = ws.cell(r, 7, saving_p); c.number_format = '0%'; c.font = DOWN_FONT
            for ci in range(1, 8):
                ws.cell(r, ci).alignment = CTR if ci >= 3 else LEFT
            if r % 2 == 0:
                for ci in range(1, 8):
                    ws.cell(r, ci).fill = BP_FILL
            r += 1
    r += 1

    # ── Section D: Sale Ends 임박 목록 ────────────────────────────────────────
    if 'Sale Ends' in dl.columns:
        sale_df = dl[dl['Sale Ends'].notna()].copy()
        if not sale_df.empty:
            ws[f'A{r}'] = f'▶  Active Sale SKUs with End Date ({len(sale_df)} items)'
            ws[f'A{r}'].font = SEC_FONT
            r += 2

            ws.cell(r, 1, 'Brand').font        = BOLD_FONT
            ws.cell(r, 2, 'Model Code').font   = BOLD_FONT
            ws.cell(r, 3, 'Product Name').font = BOLD_FONT
            ws.cell(r, 4, 'AC Type').font      = BOLD_FONT
            ws.cell(r, 5, 'Sale Price').font   = BOLD_FONT
            ws.cell(r, 6, 'Disc %').font       = BOLD_FONT
            ws.cell(r, 7, 'Sale Ends').font    = BOLD_FONT
            _hdr(ws, r, 7)
            r += 1
            sale_sorted = sale_df.sort_values('Sale Ends')
            for _, rd in sale_sorted.iterrows():
                ws.cell(r, 1, rd.get('Brand', '')).font      = DATA_FONT
                ws.cell(r, 2, rd.get('Model Code', '')).font = DATA_FONT
                ws.cell(r, 3, rd.get('Name', '')).font       = DATA_FONT
                ws.cell(r, 3).alignment = WRAP
                ws.cell(r, 4, rd.get('AC Type', '')).font    = DATA_FONT
                sp = rd.get('Sale Price', '')
                c = ws.cell(r, 5, int(sp) if pd.notna(sp) and sp != '' else '')
                c.number_format = '#,##0'; c.font = NUM_FONT
                ws.cell(r, 6, _pct_str(rd.get('Discount %', ''))).font = NUM_FONT
                ws.cell(r, 6).alignment = CTR
                ws.cell(r, 7, str(rd['Sale Ends'])).font = BOLD_FONT
                ws.cell(r, 7).alignment = CTR
                if r % 2 == 0:
                    for ci in range(1, 8):
                        ws.cell(r, ci).fill = LIGHT_FILL
                r += 1

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.sheet_properties.tabColor = '7030A0'


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Full_Price_Tracking
# ═══════════════════════════════════════════════════════════════════════════════
def build_full_tracking(wb, df: pd.DataFrame):
    sn = 'Full_Price_Tracking'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates  = sorted(df['Scraped At'].dropna().unique())
    latest = dates[-1]
    prev   = dates[-2] if len(dates) >= 2 else None
    dc     = df[df['Scraped At'] == latest].copy()

    prev_map = {}
    if prev is not None:
        dp = df[df['Scraped At'] == prev]
        dp_dedup = dp.drop_duplicates('Model Code')
        prev_map = (
            dp_dedup.set_index(dp_dedup['Model Code'].astype(str))['Final Price']
              .to_dict()
        )

    ws.merge_cells('A1:R1')
    ws['A1'] = 'Full Price Tracking — Latest Snapshot'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:R2')
    ws['A2'] = (
        f"Date: {pd.Timestamp(latest).strftime('%Y-%m-%d')}  |  "
        f"Total: {len(dc)} SKUs"
    )
    ws['A2'].font = SUB_FONT

    r = 4
    hdrs = [
        'Brand', 'Model Code', 'Product Name', 'AC Type',
        'Compressor', 'Mode', 'Capacity', 'BTU',
        'Orig Price', 'Sale Price', 'Disc %', 'Effective Price',
        'BP Price', 'Effective BP',
        'Prev Price', 'Change',
        'Free Install', 'In Stock', 'Stock Qty',
    ]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h)
    _hdr(ws, r, len(hdrs))

    dc_sorted = dc.sort_values(
        ['AC Type', 'Compressor', 'Mode', 'Ton', 'Brand'],
        na_position='last'
    )

    for i, (_, rd) in enumerate(dc_sorted.iterrows()):
        rr = r + 1 + i
        mc = str(rd.get('Model Code', ''))

        ws.cell(rr, 1,  rd.get('Brand', '')).font = DATA_FONT
        ws.cell(rr, 2,  mc).font = DATA_FONT
        ws.cell(rr, 3,  rd.get('Name', '')).font = DATA_FONT
        ws.cell(rr, 3).alignment = WRAP
        ws.cell(rr, 4,  rd.get('AC Type', '')).font = DATA_FONT
        ws.cell(rr, 5,  rd.get('Compressor', '')).font = DATA_FONT
        ws.cell(rr, 5).alignment = CTR
        ws.cell(rr, 6,  rd.get('Mode', '')).font = DATA_FONT
        ws.cell(rr, 7,  _fmt_ton(rd.get('Ton', ''))).font = DATA_FONT
        ws.cell(rr, 7).alignment = CTR
        btu = rd.get('BTU', '')
        ws.cell(rr, 8, int(btu) if pd.notna(btu) and btu != '' else '').font = NUM_FONT
        ws.cell(rr, 8).number_format = '#,##0'; ws.cell(rr, 8).alignment = CTR

        for col_idx, col_key in [
            (9,  'Original Price'),
            (10, 'Sale Price'),
            (12, 'Effective Price'),
            (13, 'BP Price'),
            (14, 'Effective BP'),
        ]:
            val = rd.get(col_key, '')
            c = ws.cell(rr, col_idx, int(val) if pd.notna(val) and val != '' else '')
            c.number_format = '#,##0'
            c.font = BP_FONT if col_idx in (13, 14) else NUM_FONT

        disc = rd.get('Discount %', '')
        if pd.notna(disc) and disc != '':
            c = ws.cell(rr, 11, float(disc) / 100)
            c.number_format = '0%'; c.font = NUM_FONT
        ws.cell(rr, 11).alignment = CTR

        prev_price = prev_map.get(mc)
        curr_price = rd.get('Final Price')
        if prev_price and pd.notna(prev_price):
            ws.cell(rr, 15, int(prev_price)).number_format = '#,##0'
            ws.cell(rr, 15).font = NUM_FONT
            if curr_price and pd.notna(curr_price):
                chg = curr_price - prev_price
                cg  = ws.cell(rr, 16, int(chg))
                cg.number_format = '#,##0'
                cg.font = DOWN_FONT if chg < 0 else (UP_FONT if chg > 0 else NUM_FONT)
        else:
            ws.cell(rr, 15, 'NEW').font = NEW_FONT
            ws.cell(rr, 15).alignment  = CTR

        fi = rd.get('Free Install', False)
        ws.cell(rr, 17, 'Yes' if fi else '-').font = (
            Font(name='Arial', size=10, color='375623', bold=True) if fi else DATA_FONT
        )
        ws.cell(rr, 17).alignment = CTR

        in_stk = rd.get('In Stock', False)
        ws.cell(rr, 18, 'Yes' if in_stk else 'No').font = (
            Font(name='Arial', size=10, color='375623', bold=True) if in_stk
            else Font(name='Arial', size=10, color='C00000')
        )
        ws.cell(rr, 18).alignment = CTR

        qty = rd.get('Stock Qty', '')
        ws.cell(rr, 19, int(qty) if pd.notna(qty) and qty != '' else '').font = NUM_FONT
        ws.cell(rr, 19).alignment = CTR

        if rr % 2 == 0:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(rr, ci).fill = LIGHT_FILL

    _auto_width(ws, max_w=50)
    ws.column_dimensions['C'].width = 55
    ws.freeze_panes = 'A5'
    ws.sheet_properties.tabColor = '4472C4'


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Stock_Dashboard_Summary  (Executive)
# ═══════════════════════════════════════════════════════════════════════════════
def build_stock_summary(wb, df: pd.DataFrame):
    sn = 'Stock_Dashboard_Summary'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Stock comparison requires at least 2 collection dates.'
        ws['A1'].font = SUB_FONT
        return

    latest_date = dates[-1]
    prev_date   = dates[-2]
    dl = df[df['Scraped At'] == latest_date].copy()
    dp = df[df['Scraped At'] == prev_date].copy()

    # Model Code 기준 stock merge
    merged = pd.merge(
        dl[['Model Code', 'Brand', 'Name', 'AC Type', 'Ton', 'In Stock', 'Stock Qty']],
        dp[['Model Code', 'Stock Qty']].rename(columns={'Stock Qty': 'Stock_prev'}),
        on='Model Code', how='left'
    )
    merged['Stock_prev']  = pd.to_numeric(merged['Stock_prev'],  errors='coerce').fillna(0)
    merged['Stock Qty']   = pd.to_numeric(merged['Stock Qty'],   errors='coerce').fillna(0)
    merged['Stock_Change'] = merged['Stock Qty'] - merged['Stock_prev']

    # ── 제목 ──
    r = 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, 'Stock Dashboard — Executive Summary').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1,
        f"Period: {pd.Timestamp(prev_date).strftime('%Y-%m-%d')} -> "
        f"{pd.Timestamp(latest_date).strftime('%Y-%m-%d')}"
    ).font = SUB_FONT
    r += 2

    # ── KPI ──
    ws.cell(r, 1, '▣ Key Performance Indicators').font = SEC_FONT
    r += 1

    total_now  = int(merged['Stock Qty'].sum())
    total_prev = int(merged['Stock_prev'].sum())
    stk_chg    = total_now - total_prev
    oos        = int((merged['Stock Qty'] == 0).sum())
    critical   = int(merged['Stock Qty'].between(1,  10).sum())
    low        = int(merged['Stock Qty'].between(11, 20).sum())
    ok         = int(merged['Stock Qty'].between(21, 50).sum())
    high       = int((merged['Stock Qty'] >= 51).sum())
    inc        = int((merged['Stock_Change'] > 0).sum())
    dec        = int((merged['Stock_Change'] < 0).sum())

    kpi_rows = [
        ('Total Stock (units)', total_now,  stk_chg,  '#,##0', '+#,##0;-#,##0;0'),
        ('Out of Stock (0)',    oos,        None,      '#,##0', None),
        ('Critical  (1~10)',   critical,   None,      '#,##0', None),
        ('Low       (11~20)',  low,        None,      '#,##0', None),
        ('OK        (21~50)',  ok,         None,      '#,##0', None),
        ('High      (51+)',    high,       None,      '#,##0', None),
        ('Stock Increased',    inc,        None,      '#,##0', None),
        ('Stock Decreased',    dec,        None,      '#,##0', None),
    ]
    ws.cell(r, 1, 'Metric').font  = BOLD_FONT
    ws.cell(r, 2, 'Current').font = BOLD_FONT
    ws.cell(r, 3, 'Change').font  = BOLD_FONT
    _hdr(ws, r, 3)
    r += 1
    for metric, current, change, cur_fmt, chg_fmt in kpi_rows:
        ws.cell(r, 1, metric).font = DATA_FONT
        c2 = ws.cell(r, 2, current); c2.font = NUM_FONT; c2.number_format = cur_fmt; c2.alignment = CTR
        if change is not None:
            c3 = ws.cell(r, 3, change); c3.font = NUM_FONT; c3.number_format = chg_fmt; c3.alignment = CTR
        if r % 2 == 0:
            for ci in range(1, 4): ws.cell(r, ci).fill = LIGHT_FILL
        r += 1
    r += 1

    # ── AC Type별 재고 ──
    ws.cell(r, 1, '▣ Stock by AC Type').font = SEC_FONT
    r += 1
    cat_sum = (
        merged.groupby('AC Type')
        .agg(stock_now=('Stock Qty', 'sum'), stock_prev=('Stock_prev', 'sum'), skus=('Model Code', 'count'))
        .reset_index()
    )
    cat_sum['Change'] = cat_sum['stock_now'] - cat_sum['stock_prev']
    cat_sum = cat_sum.sort_values('stock_now', ascending=False)

    for hdr_text, col in [('AC Type', 1), ('Current Stock', 2), ('Previous Stock', 3),
                           ('Change', 4), ('SKU Count', 5)]:
        ws.cell(r, col, hdr_text).font = BOLD_FONT
    _hdr(ws, r, 5)
    r += 1
    for _, row in cat_sum.iterrows():
        ws.cell(r, 1, row['AC Type']).font  = DATA_FONT
        c2 = ws.cell(r, 2, int(row['stock_now']));  c2.font = NUM_FONT; c2.number_format = '#,##0'; c2.alignment = CTR
        c3 = ws.cell(r, 3, int(row['stock_prev'])); c3.font = NUM_FONT; c3.number_format = '#,##0'; c3.alignment = CTR
        chg = int(row['Change'])
        c4 = ws.cell(r, 4, chg); c4.number_format = '+#,##0;-#,##0;0'
        c4.font = STOCK_UP_FONT if chg > 0 else (STOCK_DOWN_FONT if chg < 0 else NUM_FONT); c4.alignment = CTR
        ws.cell(r, 5, int(row['skus'])).font = NUM_FONT; ws.cell(r, 5).alignment = CTR
        if r % 2 == 0:
            for ci in range(1, 6): ws.cell(r, ci).fill = LIGHT_FILL
        r += 1
    r += 1

    # ── Critical Stock Alert (1~50) ──
    ws.cell(r, 1, 'Critical Stock Alert  (Stock 1~10)').font = SEC_FONT
    r += 1
    critical_df = merged[merged['Stock Qty'].between(1, 10)].sort_values('Stock Qty')
    if critical_df.empty:
        ws.cell(r, 1, 'No critical stock items.').font = DATA_FONT
        r += 1
    else:
        for hdr_text, col in [('Brand',1),('Model Code',2),('Product Name',3),
                               ('AC Type',4),('Current',5),('Previous',6),('Change',7)]:
            ws.cell(r, col, hdr_text).font = BOLD_FONT
        _hdr(ws, r, 7)
        r += 1
        for _, row in critical_df.iterrows():
            ws.cell(r, 1, row['Brand']).font = DATA_FONT
            ws.cell(r, 2, row['Model Code']).font = DATA_FONT
            ws.cell(r, 3, row['Name']).font = DATA_FONT; ws.cell(r, 3).alignment = WRAP
            ws.cell(r, 4, row['AC Type']).font = DATA_FONT
            c5 = ws.cell(r, 5, int(row['Stock Qty']));  c5.font = NUM_FONT; c5.number_format = '#,##0'; c5.alignment = CTR
            c6 = ws.cell(r, 6, int(row['Stock_prev'])); c6.font = NUM_FONT; c6.number_format = '#,##0'; c6.alignment = CTR
            chg = int(row['Stock_Change'])
            c7 = ws.cell(r, 7, chg); c7.number_format = '+#,##0;-#,##0;0'
            c7.font = STOCK_DOWN_FONT if chg < 0 else NUM_FONT; c7.alignment = CTR
            for ci in range(1, 8): ws.cell(r, ci).fill = CRITICAL_FILL
            r += 1

    _auto_width(ws)
    ws.column_dimensions['C'].width = 45
    ws.sheet_properties.tabColor = 'FF6B6B'


# ═══════════════════════════════════════════════════════════════════════════════
# 8. Stock_Dashboard_Brand  (Manager)
# ═══════════════════════════════════════════════════════════════════════════════
def build_stock_brand(wb, df: pd.DataFrame):
    sn = 'Stock_Dashboard_Brand'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Stock comparison requires at least 2 collection dates.'
        ws['A1'].font = SUB_FONT
        return

    latest_date = dates[-1]
    prev_date   = dates[-2]
    dl = df[df['Scraped At'] == latest_date].copy()
    dp = df[df['Scraped At'] == prev_date].copy()

    merged = pd.merge(
        dl[['Model Code', 'Brand', 'AC Type', 'Ton', 'Stock Qty']],
        dp[['Model Code', 'Stock Qty']].rename(columns={'Stock Qty': 'Stock_prev'}),
        on='Model Code', how='left'
    )
    merged['Stock_prev']   = pd.to_numeric(merged['Stock_prev'],  errors='coerce').fillna(0)
    merged['Stock Qty']    = pd.to_numeric(merged['Stock Qty'],   errors='coerce').fillna(0)
    merged['Stock_Change'] = merged['Stock Qty'] - merged['Stock_prev']

    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r, 1, 'Stock Dashboard — Manager View (Brand)').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r, 1,
        f"Period: {pd.Timestamp(prev_date).strftime('%Y-%m-%d')} -> "
        f"{pd.Timestamp(latest_date).strftime('%Y-%m-%d')}"
    ).font = SUB_FONT
    r += 2

    # ── AC Type × Brand ──
    for ac_type in AC_TYPE_ORDER:
        cat_data = merged[merged['AC Type'] == ac_type]
        if cat_data.empty:
            continue
        ws.merge_cells(f'A{r}:H{r}')
        ws.cell(r, 1, f'▣ {ac_type}').font = SEC_FONT
        for ci in range(1, 9): ws.cell(r, ci).fill = CAT_FILL
        r += 1

        brand_sum = (
            cat_data.groupby('Brand')
            .agg(skus=('Model Code', 'count'),
                 stock_now=('Stock Qty', 'sum'),
                 stock_prev=('Stock_prev', 'sum'))
            .reset_index()
        )
        brand_sum['Change'] = brand_sum['stock_now'] - brand_sum['stock_prev']
        brand_sum = brand_sum.sort_values('stock_now', ascending=False)

        cat_total_now  = brand_sum['stock_now'].sum()
        cat_total_prev = brand_sum['stock_prev'].sum()

        for hdr_text, col in [('Brand',1),('SKU Count',2),('Current Stock',3),
                               ('Previous Stock',4),('Change',5),('Avg Stock/SKU',6),
                               ('Current %',7),('Previous %',8)]:
            ws.cell(r, col, hdr_text).font = BOLD_FONT
        _hdr(ws, r, 8)
        r += 1
        for _, row in brand_sum.iterrows():
            ws.cell(r, 1, row['Brand']).font = DATA_FONT
            ws.cell(r, 2, int(row['skus'])).font = NUM_FONT; ws.cell(r, 2).alignment = CTR
            c3 = ws.cell(r, 3, int(row['stock_now']));  c3.font = NUM_FONT; c3.number_format = '#,##0'; c3.alignment = CTR
            c4 = ws.cell(r, 4, int(row['stock_prev'])); c4.font = NUM_FONT; c4.number_format = '#,##0'; c4.alignment = CTR
            chg = int(row['Change'])
            c5 = ws.cell(r, 5, chg); c5.number_format = '+#,##0;-#,##0;0'
            c5.font = STOCK_UP_FONT if chg > 0 else (STOCK_DOWN_FONT if chg < 0 else NUM_FONT); c5.alignment = CTR
            avg = row['stock_now'] / row['skus'] if row['skus'] > 0 else 0
            c6 = ws.cell(r, 6, round(avg, 1)); c6.number_format = '#,##0.0'; c6.font = NUM_FONT; c6.alignment = CTR
            pct_now  = row['stock_now']  / cat_total_now  if cat_total_now  > 0 else 0
            pct_prev = row['stock_prev'] / cat_total_prev if cat_total_prev > 0 else 0
            c7 = ws.cell(r, 7, pct_now);  c7.number_format = '0.0%'; c7.font = NUM_FONT; c7.alignment = CTR
            c8 = ws.cell(r, 8, pct_prev); c8.number_format = '0.0%'; c8.font = NUM_FONT; c8.alignment = CTR
            if r % 2 == 0:
                for ci in range(1, 9): ws.cell(r, ci).fill = LIGHT_FILL
            r += 1
        r += 1

    r += 1

    # ── LG vs Top Competitors ──
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r, 1, '▣ LG vs Competitors Stock Comparison').font = SEC_FONT
    for ci in range(1, 9): ws.cell(r, ci).fill = GREEN_FILL
    r += 1

    # 상위 브랜드 중 LG 제외 top 2
    top_brands = (
        merged[merged['Brand'].str.upper() != 'LG']['Brand']
        .value_counts().head(2).index.tolist()
    )

    for hdr_text, col in [('AC Type',1),('Competitor',2),('LG Stock',3),('LG SKUs',4),
                           ('Comp Stock',5),('Comp SKUs',6),('Difference',7)]:
        ws.cell(r, col, hdr_text).font = BOLD_FONT
    _hdr(ws, r, 7)
    r += 1

    lg_data = merged[merged['Brand'].str.upper() == 'LG']
    for ac_type in AC_TYPE_ORDER:
        lg_cat = lg_data[lg_data['AC Type'] == ac_type]
        lg_stock = int(lg_cat['Stock Qty'].sum())
        lg_skus  = len(lg_cat)
        for comp in top_brands:
            cd = merged[(merged['Brand'] == comp) & (merged['AC Type'] == ac_type)]
            comp_stock = int(cd['Stock Qty'].sum())
            comp_skus  = len(cd)
            if lg_skus == 0 and comp_skus == 0:
                continue
            diff = lg_stock - comp_stock
            ws.cell(r, 1, ac_type).font = DATA_FONT
            ws.cell(r, 2, comp).font    = DATA_FONT
            c3 = ws.cell(r, 3, lg_stock);   c3.font = NUM_FONT; c3.number_format = '#,##0'; c3.alignment = CTR
            ws.cell(r, 4, lg_skus).font   = NUM_FONT; ws.cell(r, 4).alignment = CTR
            c5 = ws.cell(r, 5, comp_stock); c5.font = NUM_FONT; c5.number_format = '#,##0'; c5.alignment = CTR
            ws.cell(r, 6, comp_skus).font  = NUM_FONT; ws.cell(r, 6).alignment = CTR
            c7 = ws.cell(r, 7, diff); c7.number_format = '+#,##0;-#,##0;0'
            c7.font = STOCK_UP_FONT if diff > 0 else (STOCK_DOWN_FONT if diff < 0 else NUM_FONT); c7.alignment = CTR
            if r % 2 == 0:
                for ci in range(1, 8): ws.cell(r, ci).fill = LIGHT_FILL
            r += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = '4ECDC4'


# ═══════════════════════════════════════════════════════════════════════════════
# 9. Stock_Dashboard_Detail  (Full SKU view)
# ═══════════════════════════════════════════════════════════════════════════════
def build_stock_detail(wb, df: pd.DataFrame):
    sn = 'Stock_Dashboard_Detail'
    _del(wb, sn)
    ws = wb.create_sheet(sn)

    dates = sorted(df['Scraped At'].dropna().unique())
    if len(dates) < 2:
        ws['A1'] = 'Stock comparison requires at least 2 collection dates.'
        ws['A1'].font = SUB_FONT
        return

    latest_date = dates[-1]
    prev_date   = dates[-2]
    dl = df[df['Scraped At'] == latest_date].copy()
    dp = df[df['Scraped At'] == prev_date].copy()

    merged = pd.merge(
        dl[['Model Code', 'Brand', 'Name', 'AC Type', 'Compressor', 'Mode', 'Ton', 'Stock Qty', 'In Stock']],
        dp[['Model Code', 'Stock Qty']].rename(columns={'Stock Qty': 'Stock_prev'}),
        on='Model Code', how='left'
    )
    merged['Stock_prev']   = pd.to_numeric(merged['Stock_prev'],  errors='coerce').fillna(0)
    merged['Stock Qty']    = pd.to_numeric(merged['Stock Qty'],   errors='coerce').fillna(0)
    merged['Stock_Change'] = merged['Stock Qty'] - merged['Stock_prev']
    merged = merged.sort_values(['AC Type', 'Stock_Change'], ascending=[True, False])

    r = 1
    ws.merge_cells(f'A{r}:J{r}')
    ws.cell(r, 1, 'Stock Dashboard — Detail View (Full SKU List)').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:J{r}')
    ws.cell(r, 1,
        f"Period: {pd.Timestamp(prev_date).strftime('%Y-%m-%d')} -> "
        f"{pd.Timestamp(latest_date).strftime('%Y-%m-%d')}  |  "
        f"Total SKUs: {len(merged)}"
    ).font = SUB_FONT
    r += 2

    hdrs = ['AC Type', 'Brand', 'Model Code', 'Product Name', 'Capacity',
            'Current Stock', 'Previous Stock', 'Change', 'Change %', 'Status']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(r, ci, h).font = BOLD_FONT
    _hdr(ws, r, len(hdrs))
    r += 1

    for _, row in merged.iterrows():
        stock_now  = int(row['Stock Qty'])
        stock_prev = int(row['Stock_prev'])
        change     = int(row['Stock_Change'])

        ws.cell(r, 1, row['AC Type']).font  = DATA_FONT
        ws.cell(r, 2, row['Brand']).font    = DATA_FONT
        ws.cell(r, 3, row['Model Code']).font = DATA_FONT
        ws.cell(r, 4, row['Name']).font     = DATA_FONT; ws.cell(r, 4).alignment = WRAP
        ws.cell(r, 5, _fmt_ton(row['Ton'])).font = DATA_FONT; ws.cell(r, 5).alignment = CTR
        c6 = ws.cell(r, 6, stock_now);  c6.font = NUM_FONT; c6.number_format = '#,##0'; c6.alignment = CTR
        c7 = ws.cell(r, 7, stock_prev); c7.font = NUM_FONT; c7.number_format = '#,##0'; c7.alignment = CTR
        c8 = ws.cell(r, 8, change); c8.number_format = '+#,##0;-#,##0;0'
        c8.font = STOCK_UP_FONT if change > 0 else (STOCK_DOWN_FONT if change < 0 else NUM_FONT); c8.alignment = CTR

        if stock_prev > 0:
            pct = change / stock_prev
            c9 = ws.cell(r, 9, pct); c9.number_format = '+0.0%;-0.0%;0.0%'
            c9.font = STOCK_UP_FONT if pct > 0 else (STOCK_DOWN_FONT if pct < 0 else NUM_FONT)
        else:
            ws.cell(r, 9, 'NEW' if stock_now > 0 else '-').font = NEW_FONT if stock_now > 0 else DATA_FONT
        ws.cell(r, 9).alignment = CTR

        # Status + 배경색
        if stock_now == 0:
            status       = 'OUT OF STOCK'
            status_font  = DISC_FONT
            row_fill     = OOS_FILL
        elif stock_now <= 10:
            status       = 'CRITICAL'
            status_font  = Font(name='Arial', size=10, color='C00000', bold=True)
            row_fill     = CRITICAL_FILL
        elif stock_now <= 20:
            status       = 'LOW'
            status_font  = Font(name='Arial', size=10, color='FF9800', bold=True)
            row_fill     = None
        elif stock_now <= 50:
            status       = 'OK'
            status_font  = DATA_FONT
            row_fill     = None
        else:
            status       = 'HIGH'
            status_font  = HIGH_FONT
            row_fill     = HIGH_FILL

        ws.cell(r, 10, status).font = status_font; ws.cell(r, 10).alignment = CTR

        if row_fill:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(r, ci).fill = row_fill
        elif r % 2 == 0 and 21 <= stock_now <= 50:
            for ci in range(1, len(hdrs) + 1):
                ws.cell(r, ci).fill = LIGHT_FILL

        r += 1

    _auto_width(ws, max_w=50)
    ws.column_dimensions['D'].width = 50
    ws.freeze_panes = 'A5'
    ws.sheet_properties.tabColor = '95A5A6'


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Blackbox AC Dashboard Builder")
    print("=" * 60)

    if not os.path.exists(TARGET_FILE):
        print(f"[ERROR] File not found: {TARGET_FILE}")
        print("Run blackbox_ac_scraper.py first to collect data.")
        sys.exit(1)

    print(f"  Loading: {TARGET_FILE}")
    df = load_data()

    dates = sorted(df['Scraped At'].dropna().unique())
    print(f"  Collection dates: {len(dates)}")
    for d in dates:
        n = len(df[df['Scraped At'] == d])
        print(f"    {pd.Timestamp(d).strftime('%Y-%m-%d')}: {n} SKUs")

    wb = load_workbook(TARGET_FILE)

    steps = [
        ("Dashboard_Summary",       build_summary),
        ("Price_Change_Alert",      build_price_alert),
        ("New_Discontinued_SKUs",   build_new_discontinued),
        ("Brand_Price_Compare",     build_brand_compare),
        ("Promo_Analysis",          build_promo_analysis),
        ("Full_Price_Tracking",     build_full_tracking),
        ("Stock_Dashboard_Summary", build_stock_summary),
        ("Stock_Dashboard_Brand",   build_stock_brand),
        ("Stock_Dashboard_Detail",  build_stock_detail),
    ]

    print("\n  Building dashboards...")
    for i, (label, func) in enumerate(steps, 1):
        print(f"  [{i}/{len(steps)}] {label}...")
        try:
            func(wb, df)
        except Exception as e:
            import traceback
            print(f"    [WARN] {label} error: {e}")
            traceback.print_exc()

    # 시트 순서: 대시보드 → Product_DB
    desired = [s for s, _ in steps] + [DB_SHEET]
    current = list(wb.sheetnames)
    for i, name in enumerate(desired):
        if name in current:
            wb.move_sheet(name, offset=i - current.index(name))
            current = list(wb.sheetnames)

    # 구버전 시트 정리
    for old in ['Sheet1', 'Sheet']:
        if old in wb.sheetnames:
            del wb[old]

    wb.save(TARGET_FILE)
    print(f"\n  Done! -> {TARGET_FILE}")
    print("=" * 60)

    try:
        if sys.stdin.isatty():
            input("\n  Press Enter to exit...")
    except (EOFError, OSError):
        pass


if __name__ == "__main__":
    main()
