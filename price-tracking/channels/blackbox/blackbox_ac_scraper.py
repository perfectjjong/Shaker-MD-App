#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Blackbox.com.sa Air Conditioner Price Scraper
==============================================
Collection method:
  1. Load category page via Playwright -> extract handshake JWT cookie
  2. Call backend API directly (pageSize=300 -> all products in one request)
  3. Parse JSON -> append to Product_DB sheet in Master XLSX

Accumulation logic:
  - Fixed file: Black Box_AC_Price tracking_Master.xlsx
  - New rows appended below existing data in Product_DB sheet
  - Cols A-H (AC Type~Compressor) of new rows are mapped from existing rows by Model Code (col D)
  - New models not found in existing data use scraped values as-is

API : https://api.ops.blackbox.com.sa/api/v1/search/products/facets/category/{id}
Auth: Authorization: Bearer <handshake_token>
"""

import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import asyncio
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# == 설정 ======================================================================
BASE_URL    = "https://www.blackbox.com.sa"
API_BASE    = "https://api.ops.blackbox.com.sa/api/v1"
CATEGORY_ID = 657    # Air Conditioners (parent) - 전체 수집
PAGE_SIZE   = 300
OUTPUT_DIR  = Path(__file__).parent
MASTER_FILE = OUTPUT_DIR / "Black Box_AC_Price tracking_Master.xlsx"
SHEET_NAME  = "Product_DB"

# XLSX 컬럼 정의: (헤더 표시명, dict 키)
COLUMNS = [
    ("AC Type",          "ac_type"),
    ("Brand",            "brand"),
    ("Name",             "name"),
    ("Model Code",       "model_code"),
    ("BTU",              "btu"),
    ("Ton",              "ton"),
    ("Mode",             "mode"),
    ("Compressor",       "compressor_type"),
    ("Original Price",   "original_price"),
    ("Sale Price",       "sale_price"),
    ("Discount %",       "discount_pct"),
    ("Extra Disc %",     "extra_discount_pct"),
    ("Effective Price",  "effective_price"),
    ("BP Price",         "blackbox_plus_price"),
    ("Effective BP",     "effective_bp_price"),
    ("Free Install",     "free_install"),
    ("Install SAR",      "free_install_sar"),
    ("+10% Regular",     "extra_10pct_regular"),
    ("+10% BP Only",     "extra_10pct_bp"),
    ("Sale Ends",        "sale_ends"),
    ("In Stock",         "in_stock"),
    ("Stock Qty",        "stock_qty"),
    ("URL",              "url"),
    ("Scraped At",       "scraped_at"),
]

# A~H열 (인덱스 0~7): 기존 데이터로 매핑할 컬럼
MAPPABLE_COLS = [
    "ac_type",
    "brand",
    "name",
    "model_code",
    "btu",
    "ton",
    "mode",
    "compressor_type",
]

COL_WIDTHS = [
    12, 18, 65, 25, 8, 6, 12, 16,
    14, 12, 11, 11, 14, 11, 11,
    12, 11, 13, 13,
    12, 10, 10,
    70, 20,
]


# == 파싱 함수 ==================================================================

def parse_btu_ton(raw: str):
    """'18000/ 1.5Ton' -> (18000, 1.5)"""
    if not raw:
        return None, None
    btu_m = re.search(r'(\d+)\s*/', raw)
    ton_m = re.search(r'([\d.]+)\s*[Tt]on', raw)
    btu = int(btu_m.group(1)) if btu_m else None
    ton = float(ton_m.group(1)) if ton_m else None
    return btu, ton


def parse_model_code(name: str) -> str:
    """제품명 끝의 모델 코드 추출.
    예: '... - GES-18LQJ13/R2(T3)' -> 'GES-18LQJ13/R2(T3)'
    """
    m = re.search(r'\s*-\s*([A-Z0-9][A-Z0-9\-\.\/()\\ ]+)\s*$', name)
    return m.group(1).strip() if m else ''


def parse_compressor_type(product_attributes: list, name: str = '') -> str:
    """
    product_attributes 에서 Compressor Type 추출 + 이름으로 Dual Inverter 구분.

    API 실측값:
      'Inverter' / 'INVERTER'        -> 'Inverter' (이름에 'Dual Inverter' 있으면 'Dual Inverter')
      'Energy Saver' / 'Energy saver' -> 'Energy Saver'
      속성 없음 (non-inverter)        -> 'Rotary'
    """
    for attr_row in (product_attributes or []):
        for val in attr_row.values():
            if isinstance(val, dict) and val.get('label') == 'Compressor Type':
                ct = val.get('value', '').strip()
                if ct:
                    ct_lower = ct.lower()
                    if 'inverter' in ct_lower:
                        if 'dual inverter' in name.lower():
                            return 'Dual Inverter'
                        return 'Inverter'
                    elif 'energy saver' in ct_lower:
                        return 'Energy Saver'
                    else:
                        return ct.title()
    return 'Rotary'


def parse_labels(label_list: list) -> dict:
    """
    label[] -> 구조화된 프로모션 필드.
      'free install EN'        -> free_install=True, free_install_sar=0
      'free install EN 200-RS' -> free_install=True, free_install_sar=200
      'free install EN 300-RS' -> free_install=True, free_install_sar=300
      '10% EN AC'              -> extra_10pct_regular=True
      '10% EN AC (BP)'         -> extra_10pct_bp=True
    """
    names = [lbl.get('name', '') for lbl in (label_list or [])]

    free_install     = False
    free_install_sar = 0
    for n in names:
        if 'free install EN' in n:
            free_install = True
            m = re.search(r'(\d+)-RS', n)
            if m:
                free_install_sar = int(m.group(1))

    return {
        'free_install':        free_install,
        'free_install_sar':    free_install_sar if free_install else None,
        'extra_10pct_regular': any(n == '10% EN AC' for n in names),
        'extra_10pct_bp':      any(n == '10% EN AC (BP)' for n in names),
    }


def parse_cart_rules(cart_rule_list: list):
    if not cart_rule_list:
        return 0.0, None
    max_discount = max(float(r.get('discount_amount', 0)) for r in cart_rule_list)
    to_date      = cart_rule_list[0].get('to_date') if cart_rule_list else None
    return max_discount, to_date


def parse_product(p: dict) -> dict:
    btu_raw = (p.get('option_text_cooling_capacity_btu') or [''])[0]
    btu, ton = parse_btu_ton(btu_raw)

    name           = (p.get('name') or [''])[0]
    pwt            = p.get('prices_with_tax', {})
    original_price = pwt.get('original_price')
    sale_price     = pwt.get('price') or p.get('display_price')
    discount_pct   = round(float((p.get('discount_percentage') or [0])[0]), 1)
    blackbox_plus  = float(p.get('blackbox_plus_price') or 0)

    sale_ends = None
    if pwt.get('discounted_price_to'):
        sale_ends = pwt['discounted_price_to'].split(' ')[0]

    stock                        = p.get('stock', {})
    lbl                          = parse_labels(p.get('label'))
    extra_discount_pct, _        = parse_cart_rules(p.get('cart_rule'))
    compressor_type              = parse_compressor_type(p.get('product_attributes'), name)

    effective_price = (
        int(round(sale_price * (1 - extra_discount_pct / 100), 0))
        if extra_discount_pct else sale_price
    )
    effective_bp = (
        int(round(blackbox_plus * (1 - extra_discount_pct / 100), 0))
        if (extra_discount_pct and blackbox_plus) else (int(blackbox_plus) if blackbox_plus else None)
    )

    return {
        'ac_type':            (p.get('option_text_a_ac_type') or [''])[0],
        'brand':              (p.get('option_text_a_brand') or [''])[0],
        'name':               name,
        'model_code':         parse_model_code(name),
        'btu':                btu,
        'ton':                ton,
        'mode':               (p.get('option_text_a_cold_hold') or [''])[0],
        'compressor_type':    compressor_type,
        'original_price':     original_price,
        'sale_price':         sale_price,
        'discount_pct':       discount_pct,
        'extra_discount_pct': extra_discount_pct if extra_discount_pct else None,
        'effective_price':    effective_price,
        'blackbox_plus_price':int(blackbox_plus) if blackbox_plus > 0 else None,
        'effective_bp_price': effective_bp,
        'free_install':       lbl['free_install'],
        'free_install_sar':   lbl['free_install_sar'],
        'extra_10pct_regular':lbl['extra_10pct_regular'],
        'extra_10pct_bp':     lbl['extra_10pct_bp'],
        'sale_ends':          sale_ends,
        'in_stock':           stock.get('is_in_stock', False),
        'stock_qty':          stock.get('qty', 0),
        'url':                f"{BASE_URL}/en/product/{p.get('rewrite_url', '')}",
        'scraped_at':         datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


# == Product_DB 누적 저장 ======================================================

def _make_styles():
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align   = Alignment(vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", fgColor="EEF3FB")
    return header_font, header_fill, header_align, data_align, border, alt_fill


def _build_col_index(header_row) -> dict:
    """Return {header_name: 0-based_index} from header row."""
    return {cell.value: i for i, cell in enumerate(header_row) if cell.value}


def _load_existing_mapping(ws) -> dict:
    """
    Read existing Product_DB sheet and build a mapping:
      { model_code: {col_key: value, ...}, ... }
    Keyed on Model Code (col D / header 'Model Code').
    Only cols A-H (MAPPABLE_COLS) are stored.
    Later rows overwrite earlier rows for the same model code.
    """
    if ws.max_row < 2:
        return {}

    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    col_idx    = _build_col_index(header_row)

    model_code_col = col_idx.get("Model Code")
    if model_code_col is None:
        model_code_col = 2  # fallback: col D (0-based index 3) -> col index 3, but D=3 so index=3

    mappable_headers    = [label for label, _ in COLUMNS[:8]]
    mappable_col_indices = [col_idx.get(h) for h in mappable_headers]

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= model_code_col:
            continue
        model_code = row[model_code_col]
        if not model_code:
            continue
        model_code = str(model_code).strip()

        row_data = {}
        for col_key, col_i in zip(MAPPABLE_COLS, mappable_col_indices):
            if col_i is not None and col_i < len(row):
                row_data[col_key] = row[col_i]
            else:
                row_data[col_key] = None
        mapping[model_code] = row_data

    return mapping


def _apply_mapping(products: list, existing_mapping: dict) -> tuple:
    """
    Apply existing A-H values to scraped products by Model Code.
    - Found in mapping  -> replace A-H cols with existing values
    - Not found         -> keep scraped values as-is
    Returns: (products, new_count, mapped_count)
    """
    mapped_count = 0
    new_count    = 0

    for p in products:
        mc = (p.get('model_code') or '').strip()
        if mc and mc in existing_mapping:
            existing = existing_mapping[mc]
            for col_key in MAPPABLE_COLS:
                existing_val = existing.get(col_key)
                if existing_val is not None and existing_val != '':
                    p[col_key] = existing_val
            mapped_count += 1
        else:
            new_count += 1

    return products, new_count, mapped_count


def _write_header(ws):
    """Write header row to Product_DB sheet."""
    header_font, header_fill, header_align, _, border, _ = _make_styles()
    for col_idx, (col_label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 28


def _append_products(ws, products: list, start_row: int):
    """Append products to Product_DB sheet starting at start_row."""
    _, _, _, data_align, border, alt_fill = _make_styles()

    for row_offset, product in enumerate(products):
        row_idx = start_row + row_offset
        is_alt  = (row_idx % 2 == 0)
        for col_idx, (_, key) in enumerate(COLUMNS, 1):
            val  = product.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border    = border
            if is_alt:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 16


def save_to_master(products: list, master_path: Path):
    """
    Append scraped products to Product_DB sheet in Master XLSX.

    1. Create file + header if not exists
    2. Open existing file; create Product_DB sheet if missing
    3. Build Model Code -> A-H mapping from existing rows
    4. Apply mapping to new products
    5. Append new rows below existing data
    """
    if master_path.exists():
        print(f"  [Open] {master_path.name}")
        wb = load_workbook(master_path)
    else:
        print(f"  [Create] {master_path.name}")
        wb = Workbook()
        if wb.active:
            wb.active.title = SHEET_NAME

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        existing_row_count = ws.max_row
        print(f"  [Existing rows] {max(0, existing_row_count - 1)}")
    else:
        ws = wb.create_sheet(title=SHEET_NAME, index=0)
        existing_row_count = 0
        print(f"  [Product_DB sheet created]")

    if existing_row_count < 1 or ws.cell(row=1, column=1).value is None:
        _write_header(ws)
        existing_row_count = 1

    print(f"  [Mapping] Building Model Code -> A-H map...")
    existing_mapping = _load_existing_mapping(ws)
    print(f"  [Mapping] {len(existing_mapping)} existing models found")

    products, new_count, mapped_count = _apply_mapping(products, existing_mapping)
    print(f"  [Mapping] Mapped: {mapped_count}  |  New models: {new_count}")

    next_row = existing_row_count + 1
    _append_products(ws, products, start_row=next_row)

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    total_rows = next_row + len(products) - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{total_rows}"
    ws.freeze_panes    = "A2"

    wb.save(master_path)
    print(f"  [Saved] {len(products)} rows appended -> total {total_rows - 1} rows (excl. header)")


# == 메인 스크래퍼 ==============================================================

async def scrape():
    print("=" * 60)
    print("  Blackbox.com.sa AC Scraper  (Accumulation Mode)")
    print("=" * 60)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        page = await context.new_page()

        # Step 1: handshake cookie
        print("\n[1/4] Loading Blackbox.com.sa to obtain handshake token...")
        await page.goto(
            f"{BASE_URL}/en/air-conditioners-accessories-c-657",
            wait_until='domcontentloaded',
            timeout=60_000,
        )

        # Poll until handshake cookie appears (up to 15 seconds)
        handshake = None
        for attempt in range(15):
            await page.wait_for_timeout(1000)
            cookies   = await context.cookies()
            handshake = next((c['value'] for c in cookies if c['name'] == 'handshake'), None)
            if handshake:
                print(f"  [OK] handshake token found after {attempt + 1}s ({len(handshake)} chars)")
                break
            print(f"  Waiting for handshake cookie... ({attempt + 1}s)")

        if not handshake:
            print("  ERROR: handshake cookie not found after 15s.")
            all_names = [c['name'] for c in cookies]
            print(f"  Available cookies: {all_names}")
            await browser.close()
            return None

        # Step 2: API call
        print(f"\n[2/4] Fetching all products (category {CATEGORY_ID})...")
        api_url = (
            f"{API_BASE}/search/products/facets/category/{CATEGORY_ID}"
            f"?pageSize={PAGE_SIZE}&pageNo=0&sortBy=position&sortDir=ASC"
        )

        raw_data = await page.evaluate(
            """async ([url, token]) => {
                const res = await fetch(url, {
                    headers: { 'Authorization': `Bearer ${token}` }
                });
                if (!res.ok) return { error: res.status };
                return await res.json();
            }""",
            [api_url, handshake],
        )
        await browser.close()

        if 'error' in raw_data:
            print(f"  ERROR: API status {raw_data['error']}")
            return None

        products_raw   = raw_data['data']['products']
        total_products = raw_data['data']['totalProducts']
        print(f"  [OK] {len(products_raw)} products fetched (total: {total_products})")

    # Step 3: parse
    print("\n[3/4] Parsing products...")
    products_all = [parse_product(p) for p in products_raw]

    # exclude accessories / services (no ac_type and no btu)
    products = [p for p in products_all if p['ac_type'] or p['btu']]
    skipped  = len(products_all) - len(products)
    if skipped:
        print(f"  (Skipped {skipped} accessories / services)")

    comp_dist     = Counter(p['compressor_type'] for p in products)
    in_stock_cnt  = sum(1 for p in products if p['in_stock'])
    free_inst_cnt = sum(1 for p in products if p['free_install'])
    extra_cnt     = sum(1 for p in products if p['extra_10pct_regular'] or p['extra_10pct_bp'])
    co_cnt        = sum(1 for p in products if 'Cold Only' in (p['mode'] or ''))
    hc_cnt        = sum(1 for p in products if 'Hot' in (p['mode'] or ''))

    print(f"  Total       : {len(products)}")
    print(f"  Compressor  : " + " / ".join(f"{k} {v}" for k, v in comp_dist.most_common()))
    print(f"  Mode        : Cold Only {co_cnt} / Hot&Cold {hc_cnt}")
    print(f"  In stock    : {in_stock_cnt}")
    print(f"  Free install: {free_inst_cnt}  |  Extra 10%: {extra_cnt}")

    # Step 4: append to Master XLSX
    print("\n[4/4] Saving to Master XLSX (Product_DB append)...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    save_to_master(products, MASTER_FILE)
    print(f"\n  >> {MASTER_FILE}")
    print("\n" + "=" * 60)

    return products, MASTER_FILE


if __name__ == '__main__':
    asyncio.run(scrape())
