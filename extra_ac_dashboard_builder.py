"""
eXtra AC Dashboard Builder
===========================
Prices DB 시트 데이터를 읽어 Dashboard 시트를 업데이트합니다.

수정된 주요 에러:
  1. xlrd 2.x 는 .xlsx 를 지원하지 않으므로 openpyxl 엔진 명시
  2. 시트명 대소문자/공백 불일치 방어
  3. 숫자 컬럼 강제 변환(read-only 모드에서 문자열로 읽히는 문제)
  4. Dashboard 시트가 없을 때 자동 생성
  5. 차트 참조 범위를 데이터 행 수에 맞게 동적 계산
"""

import os
import logging
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── 설정 ──────────────────────────────────────────────────────────────────────
EXCEL_FILE = "eXtra_ac_Prices_Tracking_Master.xlsx"
PRICES_DB_SHEET = "Prices DB"
DASHBOARD_SHEET = "Dashboard"

# Prices DB 예상 컬럼 (순서가 다를 경우 이름으로 탐색)
COL_DATE = "스크랩 날짜"
COL_CATEGORY = "카테고리"
COL_BRAND = "브랜드"
COL_MODEL = "모델명"
COL_SKU = "SKU"
COL_CURRENT_PRICE = "현재가 (SAR)"
COL_ORIGINAL_PRICE = "원가 (SAR)"
COL_DISCOUNT = "할인율 (%)"
COL_STOCK = "재고 상태"
COL_URL = "URL"

# ── 스타일 상수 ────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "1F3864"     # 진한 파랑
COLOR_HEADER_FONT = "FFFFFF"   # 흰색
COLOR_SUBHEADER_BG = "2E75B6"  # 중간 파랑
COLOR_ACCENT = "D6E4F0"        # 연한 파랑 (짝수행)
COLOR_HIGHLIGHT = "FFF2CC"     # 노랑 (강조)
COLOR_GREEN = "E2EFDA"         # 연한 초록
COLOR_RED = "FCE4D6"           # 연한 빨강
COLOR_TITLE_BG = "16375C"      # 타이틀 배경


def thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_style(ws, cell, bg: str = COLOR_HEADER_BG, font_size: int = 11) -> None:
    cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=font_size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def cell_style(ws, cell, bg: Optional[str] = None, bold: bool = False,
               align: str = "center") -> None:
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()


# ── 데이터 로드 ────────────────────────────────────────────────────────────────

def load_prices_db() -> pd.DataFrame:
    """
    Prices DB 시트를 DataFrame으로 읽어옴.
    - openpyxl 엔진 사용 (xlrd 2.x 호환 문제 방지)
    - 시트명을 찾지 못하면 명확한 에러 메시지 출력
    """
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(
            f"Excel 파일을 찾을 수 없습니다: '{EXCEL_FILE}'\n"
            "먼저 eXtra_ac_scaper_v4.py 를 실행하세요."
        )

    # ── 시트명 검증 ────────────────────────────────────────────────────────────
    wb_check = load_workbook(EXCEL_FILE, read_only=True)
    available_sheets = wb_check.sheetnames
    wb_check.close()

    sheet_name = _find_sheet(available_sheets, PRICES_DB_SHEET)
    if sheet_name is None:
        raise ValueError(
            f"'{PRICES_DB_SHEET}' 시트를 찾을 수 없습니다.\n"
            f"현재 시트 목록: {available_sheets}"
        )

    logger.info(f"Prices DB 시트 로드 중: '{sheet_name}'")

    # ── openpyxl 엔진으로 읽기 ─────────────────────────────────────────────────
    df = pd.read_excel(
        EXCEL_FILE,
        sheet_name=sheet_name,
        engine="openpyxl",
        dtype=str,          # 모든 컬럼을 문자열로 읽어 타입 오류 방지
    )

    if df.empty:
        raise ValueError(
            f"'{sheet_name}' 시트에 데이터가 없습니다.\n"
            "먼저 eXtra_ac_scaper_v4.py 를 실행하세요."
        )

    # ── 컬럼명 정리 ────────────────────────────────────────────────────────────
    df.columns = df.columns.str.strip()

    # 필수 컬럼 존재 확인
    required = [COL_DATE, COL_BRAND, COL_MODEL, COL_CURRENT_PRICE]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Prices DB 시트에 필수 컬럼이 없습니다: {missing}\n"
            f"현재 컬럼: {list(df.columns)}"
        )

    # ── 숫자 컬럼 변환 ─────────────────────────────────────────────────────────
    for col in [COL_CURRENT_PRICE, COL_ORIGINAL_PRICE, COL_DISCOUNT]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ── 날짜 컬럼 변환 ─────────────────────────────────────────────────────────
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")

    # NaT/NaN 행 제거
    df = df.dropna(subset=[COL_DATE, COL_MODEL])
    df = df[df[COL_CURRENT_PRICE] > 0]

    logger.info(f"  → {len(df)}개 행 로드 완료")
    return df


def _find_sheet(sheet_list: list, target: str) -> Optional[str]:
    """대소문자·공백 무시 시트 탐색"""
    target_clean = target.strip().lower()
    for s in sheet_list:
        if s.strip().lower() == target_clean:
            return s
    return None


# ── 분석 함수 ──────────────────────────────────────────────────────────────────

def compute_summary(df: pd.DataFrame) -> dict:
    """전체 요약 통계 계산"""
    latest_date = df[COL_DATE].max()
    latest_df = df[df[COL_DATE] == latest_date].copy()

    total_products = len(latest_df)
    avg_price = latest_df[COL_CURRENT_PRICE].mean()
    min_price = latest_df[COL_CURRENT_PRICE].min()
    max_price = latest_df[COL_CURRENT_PRICE].max()
    on_sale = latest_df[latest_df[COL_DISCOUNT] > 0]
    avg_discount = on_sale[COL_DISCOUNT].mean() if not on_sale.empty else 0

    return {
        "latest_date": latest_date,
        "total_products": total_products,
        "avg_price": avg_price,
        "min_price": min_price,
        "max_price": max_price,
        "on_sale_count": len(on_sale),
        "avg_discount": avg_discount,
        "latest_df": latest_df,
    }


def compute_brand_stats(latest_df: pd.DataFrame) -> pd.DataFrame:
    """브랜드별 통계"""
    brand_col = COL_BRAND if COL_BRAND in latest_df.columns else COL_MODEL
    stats = (
        latest_df.groupby(brand_col, as_index=False)
        .agg(
            상품수=(COL_CURRENT_PRICE, "count"),
            평균가=(COL_CURRENT_PRICE, "mean"),
            최저가=(COL_CURRENT_PRICE, "min"),
            최고가=(COL_CURRENT_PRICE, "max"),
        )
        .sort_values("상품수", ascending=False)
        .reset_index(drop=True)
    )
    stats.rename(columns={brand_col: "브랜드"}, inplace=True)
    stats["평균가"] = stats["평균가"].round(0)
    stats["최저가"] = stats["최저가"].round(0)
    stats["최고가"] = stats["최고가"].round(0)
    return stats


def compute_category_stats(latest_df: pd.DataFrame) -> pd.DataFrame:
    """카테고리별 통계"""
    if COL_CATEGORY not in latest_df.columns:
        return pd.DataFrame()
    stats = (
        latest_df.groupby(COL_CATEGORY, as_index=False)
        .agg(
            상품수=(COL_CURRENT_PRICE, "count"),
            평균가=(COL_CURRENT_PRICE, "mean"),
            최저가=(COL_CURRENT_PRICE, "min"),
        )
        .sort_values("상품수", ascending=False)
        .reset_index(drop=True)
    )
    stats.rename(columns={COL_CATEGORY: "카테고리"}, inplace=True)
    stats["평균가"] = stats["평균가"].round(0)
    stats["최저가"] = stats["최저가"].round(0)
    return stats


def compute_top_deals(latest_df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
    """할인율 상위 N개 상품"""
    if COL_DISCOUNT not in latest_df.columns:
        return pd.DataFrame()
    deals = (
        latest_df[latest_df[COL_DISCOUNT] > 0]
        .sort_values(COL_DISCOUNT, ascending=False)
        .head(n)
        .reset_index(drop=True)
    )
    cols = [COL_BRAND, COL_MODEL, COL_CURRENT_PRICE, COL_ORIGINAL_PRICE, COL_DISCOUNT]
    cols = [c for c in cols if c in deals.columns]
    return deals[cols].copy()


def compute_cheapest(latest_df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
    """최저가 상위 N개 상품"""
    cheap = (
        latest_df.sort_values(COL_CURRENT_PRICE)
        .head(n)
        .reset_index(drop=True)
    )
    cols = [COL_BRAND, COL_MODEL, COL_CURRENT_PRICE, COL_DISCOUNT]
    cols = [c for c in cols if c in cheap.columns]
    return cheap[cols].copy()


def compute_price_trend(df: pd.DataFrame) -> pd.DataFrame:
    """날짜별 평균가 추이"""
    trend = (
        df.groupby(df[COL_DATE].dt.date)[COL_CURRENT_PRICE]
        .mean()
        .reset_index()
    )
    trend.columns = ["날짜", "평균가 (SAR)"]
    trend["평균가 (SAR)"] = trend["평균가 (SAR)"].round(0)
    trend = trend.sort_values("날짜").reset_index(drop=True)
    return trend


# ── 대시보드 쓰기 ──────────────────────────────────────────────────────────────

def clear_dashboard(wb) -> None:
    """Dashboard 시트 초기화 (없으면 생성)"""
    sheet_name = _find_sheet(wb.sheetnames, DASHBOARD_SHEET)
    if sheet_name:
        del wb[sheet_name]
        logger.info(f"  기존 '{sheet_name}' 시트 삭제")
    ws = wb.create_sheet(DASHBOARD_SHEET)
    logger.info(f"  '{DASHBOARD_SHEET}' 시트 생성")
    return ws


def write_title(ws, summary: dict) -> int:
    """타이틀 행 작성, 다음 시작 행 반환"""
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "eXtra 에어컨 가격 트래킹 대시보드"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    update_date = summary["latest_date"]
    date_str = update_date.strftime("%Y-%m-%d") if pd.notna(update_date) else "N/A"
    sub_cell.value = f"최종 데이터 기준일: {date_str}  |  업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font = Font(size=10, color="FFFFFF", italic=True)
    sub_cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER_BG)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    return 4


def write_kpi_row(ws, summary: dict, start_row: int) -> int:
    """KPI 카드 행 작성"""
    kpis = [
        ("총 상품 수", f"{summary['total_products']:,}개"),
        ("평균 가격", f"{summary['avg_price']:,.0f} SAR"),
        ("최저 가격", f"{summary['min_price']:,.0f} SAR"),
        ("최고 가격", f"{summary['max_price']:,.0f} SAR"),
        ("할인 상품 수", f"{summary['on_sale_count']:,}개"),
        ("평균 할인율", f"{summary['avg_discount']:.1f}%"),
    ]

    # KPI 레이블 행
    col_step = 2  # 2칸씩 사용
    for i, (label, value) in enumerate(kpis):
        col_start = 1 + i * col_step
        col_end = col_start + col_step - 1

        ws.merge_cells(
            start_row=start_row,
            start_column=col_start,
            end_row=start_row,
            end_column=min(col_end, 12),
        )
        lbl_cell = ws.cell(row=start_row, column=col_start, value=label)
        lbl_cell.font = Font(bold=True, size=9, color=COLOR_HEADER_FONT)
        lbl_cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER_BG)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.border = thin_border()

        ws.merge_cells(
            start_row=start_row + 1,
            start_column=col_start,
            end_row=start_row + 1,
            end_column=min(col_end, 12),
        )
        val_cell = ws.cell(row=start_row + 1, column=col_start, value=value)
        val_cell.font = Font(bold=True, size=13, color="1F3864")
        val_cell.fill = PatternFill("solid", fgColor="EBF3FB")
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = thin_border()

    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row + 1].height = 30
    return start_row + 3


def write_table(ws, title: str, df: pd.DataFrame, start_row: int,
                start_col: int = 1, col_widths: Optional[list] = None) -> int:
    """
    범용 테이블 쓰기 함수.
    - 섹션 제목, 헤더, 데이터 행을 작성
    - 데이터가 없으면 "데이터 없음" 표시
    - 시작 row를 반환
    """
    n_cols = len(df.columns) if not df.empty else 1

    # 섹션 제목
    end_col = start_col + n_cols - 1
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=end_col,
    )
    sec_cell = ws.cell(row=start_row, column=start_col, value=title)
    sec_cell.font = Font(bold=True, size=11, color="FFFFFF")
    sec_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 22
    start_row += 1

    if df.empty:
        ws.cell(row=start_row, column=start_col, value="데이터 없음").font = Font(italic=True)
        return start_row + 2

    # 헤더 행
    for ci, col_name in enumerate(df.columns, start=start_col):
        c = ws.cell(row=start_row, column=ci, value=col_name)
        header_style(ws, c, bg=COLOR_SUBHEADER_BG, font_size=10)
    ws.row_dimensions[start_row].height = 20
    start_row += 1

    # 데이터 행
    for ri, (_, row_data) in enumerate(df.iterrows()):
        bg = COLOR_ACCENT if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, start=start_col):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            cell_style(ws, c, bg=bg, align="center")
            # 숫자 포맷
            if isinstance(val, float):
                c.number_format = "#,##0.0"
            elif isinstance(val, int):
                c.number_format = "#,##0"
    start_row += len(df)

    # 컬럼 너비 설정
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w

    return start_row + 1


def add_bar_chart(ws, data_start_row: int, data_end_row: int,
                  label_col: int, value_col: int,
                  anchor_cell: str, title: str, x_title: str, y_title: str) -> None:
    """브랜드별 바 차트 추가"""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.height = 12
    chart.width = 18

    data_ref = Reference(
        ws,
        min_col=value_col,
        min_row=data_start_row - 1,  # 헤더 포함
        max_col=value_col,
        max_row=data_end_row,
    )
    cats_ref = Reference(
        ws,
        min_col=label_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"

    ws.add_chart(chart, anchor_cell)


def add_line_chart(ws, data_start_row: int, data_end_row: int,
                   date_col: int, price_col: int, anchor_cell: str) -> None:
    """가격 추이 라인 차트 추가"""
    if data_end_row < data_start_row:
        return

    chart = LineChart()
    chart.style = 10
    chart.title = "평균 가격 추이 (SAR)"
    chart.y_axis.title = "평균가 (SAR)"
    chart.x_axis.title = "날짜"
    chart.height = 12
    chart.width = 22

    data_ref = Reference(
        ws,
        min_col=price_col,
        min_row=data_start_row - 1,
        max_col=price_col,
        max_row=data_end_row,
    )
    cats_ref = Reference(
        ws,
        min_col=date_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, anchor_cell)


# ── 메인 대시보드 빌더 ─────────────────────────────────────────────────────────

def build_dashboard(df: pd.DataFrame) -> None:
    """Dashboard 시트 생성 및 업데이트"""
    logger.info("대시보드 생성 시작...")

    # 분석
    summary = compute_summary(df)
    latest_df = summary.pop("latest_df")
    brand_stats = compute_brand_stats(latest_df)
    category_stats = compute_category_stats(latest_df)
    top_deals = compute_top_deals(latest_df, n=10)
    cheapest = compute_cheapest(latest_df, n=10)
    trend = compute_price_trend(df)

    # 워크북 열기
    wb = load_workbook(EXCEL_FILE)
    ws = clear_dashboard(wb)

    # 컬럼 기본 너비
    for col_idx in range(1, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    # ── 타이틀 ────────────────────────────────────────────────────────────────
    row = write_title(ws, summary)

    # ── KPI 카드 ──────────────────────────────────────────────────────────────
    row = write_kpi_row(ws, summary, row)

    # ── 브랜드별 통계 테이블 ──────────────────────────────────────────────────
    brand_table_start = row + 1
    row = write_table(
        ws,
        title="브랜드별 통계 (최신 데이터 기준)",
        df=brand_stats,
        start_row=brand_table_start,
        start_col=1,
    )
    brand_data_start = brand_table_start + 2   # 섹션 제목 + 헤더
    brand_data_end = brand_data_start + len(brand_stats) - 1

    # 브랜드 바 차트 (우측에 배치)
    if not brand_stats.empty:
        add_bar_chart(
            ws,
            data_start_row=brand_data_start,
            data_end_row=brand_data_end,
            label_col=1,    # 브랜드
            value_col=3,    # 평균가
            anchor_cell=f"F{brand_table_start}",
            title="브랜드별 평균 가격 (SAR)",
            x_title="브랜드",
            y_title="평균가 (SAR)",
        )

    # ── 카테고리별 통계 ────────────────────────────────────────────────────────
    if not category_stats.empty:
        cat_table_start = row + 1
        row = write_table(
            ws,
            title="카테고리별 통계",
            df=category_stats,
            start_row=cat_table_start,
            start_col=1,
        )

    # ── 최고 할인 상품 ─────────────────────────────────────────────────────────
    deals_start = row + 1
    row = write_table(
        ws,
        title="할인율 TOP 10 상품",
        df=top_deals,
        start_row=deals_start,
        start_col=1,
    )

    # ── 최저가 상품 ────────────────────────────────────────────────────────────
    cheap_start = row + 1
    row = write_table(
        ws,
        title="최저가 TOP 10 상품",
        df=cheapest,
        start_row=cheap_start,
        start_col=1,
    )

    # ── 가격 추이 테이블 + 라인 차트 ──────────────────────────────────────────
    if not trend.empty:
        trend_start = row + 1
        row = write_table(
            ws,
            title="날짜별 평균 가격 추이",
            df=trend,
            start_row=trend_start,
            start_col=1,
        )
        trend_data_start = trend_start + 2
        trend_data_end = trend_data_start + len(trend) - 1

        add_line_chart(
            ws,
            data_start_row=trend_data_start,
            data_end_row=trend_data_end,
            date_col=1,
            price_col=2,
            anchor_cell=f"E{trend_start}",
        )

    # ── 시트 고정 및 탭 색상 ──────────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "2E75B6"

    # Dashboard 시트를 맨 앞으로 이동
    idx = wb.sheetnames.index(DASHBOARD_SHEET)
    wb.move_sheet(DASHBOARD_SHEET, offset=-idx)

    # 저장
    wb.save(EXCEL_FILE)
    wb.close()
    logger.info(f"대시보드 저장 완료 → '{EXCEL_FILE}' > '{DASHBOARD_SHEET}' 시트")


# ── 진입점 ────────────────────────────────────────────────────────────────────

def main() -> None:
    logger.info("=" * 60)
    logger.info("eXtra AC 대시보드 빌더 시작")
    logger.info("=" * 60)

    try:
        df = load_prices_db()
        build_dashboard(df)
        logger.info("\n[완료] Dashboard 시트 업데이트 성공!")

    except FileNotFoundError as e:
        logger.error(f"\n[파일 오류] {e}")
        raise SystemExit(1)
    except ValueError as e:
        logger.error(f"\n[데이터 오류] {e}")
        raise SystemExit(1)
    except Exception as e:
        logger.exception(f"\n[예상치 못한 오류] {e}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
